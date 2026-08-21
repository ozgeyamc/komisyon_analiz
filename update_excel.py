"""
Çekilen komisyon/ücret verilerini Excel dosyasına yazan/güncelleyen modül.

Ek olarak final Excel yayımlanmadan hemen önce EK KAYNAK açıklamalarını kullanıcıya
okunabilir hale getirir. Teknik supplemental işaretleri karşılaştırma hesaplanırken
korunur; yalnız KARŞILAŞTIRMA oluşturulduktan sonra KOMİSYONLAR sayfasındaki görünen
açıklamalar temizlenir.
"""

from __future__ import annotations

import os
import re
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from scraper import UcretSatiri


EXCEL_WRITER_VERSION = "2026-08-21-v2-clean-supplemental-display"
EXCEL_DOSYA_ADI = "komisyonlar_guncel.xlsx"
SHEET_ADI = "KOMISYONLAR"

BASLIKLAR = [
    "BANKA",
    "KATEGORİ",
    "MASRAF",
    "ASGARİ TUTAR",
    "ASGARİ ORAN",
    "AZAMİ TUTAR",
    "AZAMİ ORAN",
    "AÇIKLAMA",
    "KOMİSYON GÜNCELLEME TARİHİ",
    "ÇALIŞTIRILMA TARİHİ",
]

COL_BANKA = 1
COL_KATEGORI = 2
COL_MASRAF = 3
COL_ASGARI_TUTAR = 4
COL_ASGARI_ORAN = 5
COL_AZAMI_TUTAR = 6
COL_AZAMI_ORAN = 7
COL_ACIKLAMA = 8
COL_SITE_TARIHI = 9
COL_CALISMA_TARIHI = 10

BASLIK_FILL = PatternFill(
    start_color="1F3864",
    end_color="1F3864",
    fill_type="solid",
)
BASLIK_FONT = Font(color="FFFFFF", bold=True)

BANKA_RENKLER = {
    "GARANTİ": {"bg": "00B050", "fg": "FFFFFF"},
    "ZİRAAT": {"bg": "C00000", "fg": "FFFFFF"},
    "HALKBANK": {"bg": "7030A0", "fg": "FFFFFF"},
    "AKBANK": {"bg": "FF0000", "fg": "FFFF00"},
    "YAPIKREDI": {"bg": "003087", "fg": "FFD700"},
    "VAKIFBANK": {"bg": "FFD700", "fg": "000000"},
    "QNB": {"bg": "6B0F8E", "fg": "FFFFFF"},
    "DENİZBANK": {"bg": "003DA5", "fg": "FFFFFF"},
    "TEB": {"bg": "00539B", "fg": "FFFFFF"},
    "İŞBANKASI": {"bg": "012169", "fg": "FFFFFF"},
}

STATUS_AVAILABLE = "[SUPPLEMENTAL][AVAILABLE_NO_SEPARATE_FEE]"
STATUS_EMPTY = "[SUPPLEMENTAL][PUBLISHED_EMPTY]"
STATUS_NUMERIC = "[SUPPLEMENTAL][OFFICIAL_FEE]"
STATUS_NOT_APPLICABLE = "[SUPPLEMENTAL][NOT_APPLICABLE]"
STATUS_BACKFILLED = "[SUPPLEMENTAL][FEE_BACKFILLED_FROM_PRIMARY]"


# ---------------------------------------------------------------------------
# EXCEL YAZIMI
# ---------------------------------------------------------------------------


def _bugun_tarih_str() -> str:
    turkey_tz = timezone(timedelta(hours=3))
    return datetime.now(turkey_tz).strftime("%d.%m.%Y %H:%M")


def _yeni_workbook_olustur() -> Tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_ADI

    for idx, baslik in enumerate(BASLIKLAR, start=1):
        cell = ws.cell(row=1, column=idx, value=baslik)
        cell.fill = BASLIK_FILL
        cell.font = BASLIK_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    genislikler = [15, 40, 40, 14, 14, 14, 14, 60, 26, 22]
    for idx, genislik in enumerate(genislikler, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = genislik

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:J1"

    return wb, ws


def _satirlari_yaz(
    ws: Worksheet,
    satirlar: List[UcretSatiri],
    banka_adi: str,
    calisma_tarihi: str,
) -> int:
    renk = BANKA_RENKLER.get(
        banka_adi,
        {"bg": "808080", "fg": "FFFFFF"},
    )

    banka_fill = PatternFill(
        start_color=renk["bg"],
        end_color=renk["bg"],
        fill_type="solid",
    )
    banka_font = Font(color=renk["fg"], bold=True)

    satirlar = sorted(
        satirlar,
        key=lambda s: (s.kategori, s.masraf),
    )

    for satir in satirlar:
        yeni_row = ws.max_row + 1

        banka_cell = ws.cell(
            row=yeni_row,
            column=COL_BANKA,
            value=banka_adi,
        )
        banka_cell.fill = banka_fill
        banka_cell.font = banka_font
        banka_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        ws.cell(row=yeni_row, column=COL_KATEGORI, value=satir.kategori)
        ws.cell(row=yeni_row, column=COL_MASRAF, value=satir.masraf)
        ws.cell(row=yeni_row, column=COL_ASGARI_TUTAR, value=satir.asgari_tutar)
        ws.cell(row=yeni_row, column=COL_ASGARI_ORAN, value=satir.asgari_oran)
        ws.cell(row=yeni_row, column=COL_AZAMI_TUTAR, value=satir.azami_tutar)
        ws.cell(row=yeni_row, column=COL_AZAMI_ORAN, value=satir.azami_oran)

        aciklama_cell = ws.cell(
            row=yeni_row,
            column=COL_ACIKLAMA,
            value=satir.aciklama,
        )
        aciklama_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        ws.cell(
            row=yeni_row,
            column=COL_SITE_TARIHI,
            value=satir.site_guncelleme_tarihi,
        )
        ws.cell(
            row=yeni_row,
            column=COL_CALISMA_TARIHI,
            value=calisma_tarihi,
        )

    return len(satirlar)


def excel_guncelle_coklu(
    banka_verileri: Dict[str, List[UcretSatiri]],
    dosya_yolu: str = EXCEL_DOSYA_ADI,
) -> Dict[str, int]:
    if os.path.exists(dosya_yolu):
        os.remove(dosya_yolu)

    wb, ws = _yeni_workbook_olustur()
    calisma_tarihi = _bugun_tarih_str()

    toplam = 0
    for banka_adi, satirlar in banka_verileri.items():
        toplam += _satirlari_yaz(
            ws,
            satirlar,
            banka_adi,
            calisma_tarihi,
        )

    wb.save(dosya_yolu)

    return {
        "eklendi": toplam,
        "guncellendi": 0,
        "degismedi": 0,
    }


def excel_guncelle(
    satirlar: List[UcretSatiri],
    dosya_yolu: str = EXCEL_DOSYA_ADI,
) -> Dict[str, int]:
    return excel_guncelle_coklu(
        {"GARANTİ": satirlar},
        dosya_yolu,
    )


# ---------------------------------------------------------------------------
# FINAL GÖRÜNÜM TEMİZLİĞİ
# ---------------------------------------------------------------------------
#
# ÖNEMLİ:
# Bu fonksiyon excel_guncelle_coklu() içinde çağrılmaz. Çünkü teknik
# SERVICE/CHANNEL/BAND marker'ları update_comparison.py tarafından kullanılır.
# main.py önce KARŞILAŞTIRMA'yı oluşturur, sonra bu fonksiyonu çağırır.
# Böylece eşleştirme doğruluğu korunurken final KOMİSYONLAR sayfası temiz görünür.
# ---------------------------------------------------------------------------


def _collapse(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def _source_url(aciklama: str) -> str:
    match = re.search(
        r"Resm[îi]\s+ek\s+kaynak\s*:\s*(https?://[^\s|]+)",
        aciklama or "",
        flags=re.I,
    )
    if not match:
        return ""

    # Sadece görünüm için URL'i metinden çıkarıyoruz. Parantez gibi URL içinde
    # geçerli karakterleri korur, yalnız yapışmış cümle sonu noktalamasını temizler.
    return match.group(1).rstrip(".,;")


def _meta_value(aciklama: str, key: str) -> str:
    match = re.search(
        rf"(?:^|[;|]\s*){re.escape(key)}\s*=\s*([^;|]+)",
        aciklama or "",
        flags=re.I | re.S,
    )
    if not match:
        return ""
    return _collapse(match.group(1))


def _backfill_source_name(aciklama: str) -> str:
    patterns = (
        r"ana\s+resm[îi]\s+ücret\s+tablosundaki\s+'([^']+)'",
        r'ana\s+resm[îi]\s+ücret\s+tablosundaki\s+"([^"]+)"',
    )

    for pattern in patterns:
        match = re.search(
            pattern,
            aciklama or "",
            flags=re.I,
        )
        if match:
            return _collapse(match.group(1))

    return ""


def _human_extra_text(aciklama: str) -> str:
    """
    Teknik marker, URL ve meta alanlarını atıp yalnız okunabilir ek notları korur.

    Özellikle BSMV, değerli kâğıt bedeli, aynı tarife hücresinin iki kalemi
    kapsaması gibi finansal açıdan önemli notlar kaybolmaz.
    """
    text = str(aciklama or "")

    text = re.sub(
        r"\[SUPPLEMENTAL\]\[[^\]]+\]",
        " ",
        text,
        flags=re.I,
    )

    text = re.sub(
        r"Resm[îi]\s+ek\s+kaynak\s*:\s*https?://[^\s|]+",
        " ",
        text,
        flags=re.I,
    )

    # Backfill açıklaması final metinde ayrı ve daha kısa cümleyle verilecek.
    text = re.sub(
        r"Tutar/oran\s+bankan[ıi]n\s+ana\s+resm[îi]\s+ücret\s+tablosundaki.*?"
        r"(?:yeni\s+ücret\s+hesaplanmad[ıi]\.?|$)",
        " ",
        text,
        flags=re.I | re.S,
    )

    pieces = []

    for pipe_part in re.split(r"\s*\|\s*", text):
        for part in re.split(r"\s*;\s*", pipe_part):
            clean = _collapse(part)
            if not clean:
                continue

            if re.match(
                r"^(SERVICE|CHANNEL|BAND|DISPLAY_TEXT)\s*=",
                clean,
                flags=re.I,
            ):
                continue

            pieces.append(clean)

    # Aynı not tekrarlanmışsa tek kez göster.
    unique = []
    seen = set()
    for piece in pieces:
        key = piece.casefold()
        if key in seen:
            continue
        seen.add(key)
        unique.append(piece)

    return " ".join(unique).strip()


def _clean_supplemental_description(
    kategori: str,
    masraf: str,
    aciklama: str,
) -> Tuple[str, str]:
    """
    Teknik supplemental açıklamasını final Excel'de gösterilecek kısa metne çevirir.

    İkinci dönüş değeri kaynak URL'sidir. URL hücrenin görünen metnine yazılmaz;
    Excel hyperlink olarak tutulur.
    """
    kategori_text = _collapse(kategori)
    aciklama_text = str(aciklama or "")
    url = _source_url(aciklama_text)

    display_text = _meta_value(
        aciklama_text,
        "DISPLAY_TEXT",
    )
    display_text = (
        display_text
        .replace("\\n", " ")
        .replace("\n", " ")
        .replace("TRY", "TL")
    )
    display_text = _collapse(display_text)

    is_service_status = (
        kategori_text.casefold()
        == "EK KAYNAK - Hizmet Durumu".casefold()
    )

    # 1) Primary resmî ücret tablosundan güvenli numeric tarife taşınmış.
    if STATUS_BACKFILLED in aciklama_text:
        source_name = _backfill_source_name(aciklama_text)

        if source_name:
            visible = (
                "Resmî ek kaynakta hizmet doğrulandı. "
                f"Ücret, bankanın ana resmî ücret tablosundaki “{source_name}” "
                "tarifesinden eşleştirildi."
            )
        else:
            visible = (
                "Resmî ek kaynakta hizmet doğrulandı. "
                "Ücret, bankanın ana resmî ücret tablosundaki karşılık gelen "
                "tarifeden eşleştirildi."
            )

        return visible, url

    # 2) İşlem/kanal uygulanmıyor.
    if STATUS_NOT_APPLICABLE in aciklama_text:
        if display_text:
            return display_text, url

        return (
            "Resmî kaynağa göre bu işlem veya kanal uygulanmıyor.",
            url,
        )

    # 3) Kalem yayımlanmış ama numeric tutar yayımlanmamış.
    if STATUS_EMPTY in aciklama_text:
        if display_text:
            return display_text, url

        extra = _human_extra_text(aciklama_text)
        base = (
            "Resmî kaynakta kalem yayımlanıyor ancak ayrı ücret tutarı "
            "belirtilmemiş."
        )

        if extra:
            return f"{base} {extra}", url

        return base, url

    # 4) Hizmet varlığı doğrulanmış ancak ayrı tarife yok.
    if STATUS_AVAILABLE in aciklama_text:
        if display_text:
            return display_text, url

        return (
            "Hizmet resmî ek kaynakta doğrulandı; ayrı bir ücret tarifesi "
            "yayımlanmıyor.",
            url,
        )

    # 5) Numeric tarife doğrudan ek resmî kaynaktan alınmış.
    if STATUS_NUMERIC in aciklama_text:
        extra = _human_extra_text(aciklama_text)
        base = "Resmî ek kaynaktaki güncel tarife."

        if extra:
            return f"{base} {extra}", url

        return base, url

    # Marker olmayan EK KAYNAK satırlarında teknik alanları yine temizle,
    # fakat finansal açıklamayı kaybetme.
    if kategori_text.casefold().startswith("ek kaynak"):
        extra = _human_extra_text(aciklama_text)

        if extra:
            return extra, url

        return "Resmî ek kaynak üzerinden doğrulandı.", url

    # Normal primary satırlar ASLA değiştirilmez.
    return aciklama_text, ""


def final_excel_gorunumunu_temizle(
    dosya_yolu: str = EXCEL_DOSYA_ADI,
) -> Dict[str, int]:
    """
    Final Excel'deki yalnız EK KAYNAK açıklamalarını temizler.

    - KARŞILAŞTIRMA sayfasına dokunmaz.
    - Primary banka satırlarının açıklamasını değiştirmez.
    - Teknik marker'ları görünür metinden kaldırır.
    - Resmî kaynak URL'sini görünür metin yerine hücrenin hyperlink'i olarak saklar.
    """
    wb = load_workbook(dosya_yolu)

    if SHEET_ADI not in wb.sheetnames:
        raise RuntimeError(
            f"{SHEET_ADI} sayfası bulunamadı; final görünüm temizlenemedi."
        )

    ws = wb[SHEET_ADI]

    cleaned = 0
    hyperlinks = 0

    for row_idx in range(2, ws.max_row + 1):
        kategori = _collapse(
            ws.cell(row=row_idx, column=COL_KATEGORI).value
        )

        if not kategori.casefold().startswith("ek kaynak"):
            continue

        masraf = _collapse(
            ws.cell(row=row_idx, column=COL_MASRAF).value
        )
        cell = ws.cell(
            row=row_idx,
            column=COL_ACIKLAMA,
        )
        old_value = str(cell.value or "")

        new_value, url = _clean_supplemental_description(
            kategori,
            masraf,
            old_value,
        )

        if new_value != old_value:
            cell.value = new_value
            cleaned += 1

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        if url:
            # URL görünürde uzun bir metin olarak yer kaplamaz; açıklama hücresi
            # yine de tıklanabilir resmî kaynak bağlantısını taşır.
            cell.hyperlink = url
            hyperlinks += 1

    wb.save(dosya_yolu)

    return {
        "temizlenen_aciklama": cleaned,
        "kaynak_linki": hyperlinks,
    }
