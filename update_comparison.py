"""
komisyonlar_guncel.xlsx içindeki banka ücretlerini ortak bir masraf sözlüğüne
çevirip KARŞILAŞTIRMA sayfasını otomatik üretir.

Temel fark:
- Bankaların MASRAF adlarını birebir eşleştirmez.
- Önce hizmeti (EFT, Havale, FAST, Fatura vb.), kanalı (Mobil/Şube/ATM)
  ve tutar aralığını semantik kurallarla çıkarır.
- Aynı hizmet + aynı kanal + aynı tutar aralığı olan farklı banka satırlarını
  tek bir ORTAK MASRAF ADI altında toplar.

Örnek:
  Garanti : "Elektronik Fon Transferi (EFT) Ücreti - EFT Garanti BBVA
             Mobil / İnternet (0-8300 TL)"
  Akbank  : "FAST / Akbank Mobil'den, İnternet'ten ... EFT ...
             (8.300 TL ve Altı)"
  YKB     : "EFT Gönderimi - İnternet/Mobil – 0 - 8.300 TL"
  İşbank  : "EFT Gönderilmesi - 0-8.300 TL"

hepsi şu anahtara bağlanır:
  EFT Gönderimi | Mobil | 0 TRY - 8.300 TRY

Ana Excel her gün güncellendiğinde KARŞILAŞTIRMA sayfası da aynı veriyle
tekrar üretilir. NOTLAR sütunundaki manuel notlar ortak satır anahtarına göre
korunur.
"""

import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Set, Tuple

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


COMPARISON_VERSION = "2026-08-26-v30-fast-source-limit-note"
COMPARISON_SHEET = "KARŞILAŞTIRMA"
PREVIEW_LAYOUT_SIGNATURE = "4BANKS|A:I_FILTERED|J:L_EMPTY|M_GLOSSARY_LIST|FAIL_CLOSED_V30|PRIMARY_FIRST|CHANNEL_LABELS|CHEQUE_CURRENT|FAST_SOURCE_LIMITS|RESEARCH_SECURITIES"

STATUS_AVAILABLE = "[SUPPLEMENTAL][AVAILABLE_NO_SEPARATE_FEE]"
STATUS_EMPTY = "[SUPPLEMENTAL][PUBLISHED_EMPTY]"
STATUS_NUMERIC = "[SUPPLEMENTAL][OFFICIAL_FEE]"
STATUS_NOT_APPLICABLE = "[SUPPLEMENTAL][NOT_APPLICABLE]"


# Eski denemelerde oluşmuş karşılaştırma sayfaları kullanıcıyı yanıltmasın.
# v9 çalıştığında KARŞILAŞTIRMA ile başlayan eski sayfaların tamamı silinir
# ve sadece preview ile aynı kompakt sayfa yeniden oluşturulur.

# Kullanıcının karşılaştırma şablonundaki banka seti.
BANKS = [
    "GARANTİ",
    "İŞBANKASI",
    "AKBANK",
    "YAPIKREDI",
]

DISPLAY_BANKS = {
    "GARANTİ": "Garanti BBVA",
    "İŞBANKASI": "İş Bankası",
    "AKBANK": "Akbank",
    "YAPIKREDI": "Yapı ve Kredi Bankası",
}

BANK_COLORS = {
    "GARANTİ": "70AD47",
    "İŞBANKASI": "2E75B6",
    "AKBANK": "E31E24",
    "YAPIKREDI": "17365D",
}


PRIMARY_SOURCE_URLS = {
    "GARANTİ": "https://www.garantibbva.com.tr/urun-ve-hizmet-ucretleri",
    "İŞBANKASI": "https://www.isbank.com.tr/urun-ve-hizmet-ucretleri",
    "AKBANK": "https://www.akbank.com/urun-ve-hizmet-ucretleri",
    "YAPIKREDI": "https://www.yapikredi.com.tr/bireysel-bankacilik/hesaplama-araclari/bireysel-urun-ve-hizmet-ucretleri",
}

HEADER_ALIASES = {
    "banka": "BANKA",
    "kategori": "KATEGORİ",
    "masraf": "MASRAF",
    "asgari tutar": "ASGARİ TUTAR",
    "asgari oran": "ASGARİ ORAN",
    "azami tutar": "AZAMİ TUTAR",
    "azami oran": "AZAMİ ORAN",
    "aciklama": "AÇIKLAMA",
    "site guncelleme tarihi": "GÜNCELLEME TARİHİ",
    "komisyon guncelleme tarihi": "GÜNCELLEME TARİHİ",
    "son kontrol": "SON KONTROL",
    "calistirilma tarihi": "SON KONTROL",
}


@dataclass(frozen=True)
class FeeRow:
    banka: str
    kategori: str
    masraf: str
    asgari_tutar: str = ""
    asgari_oran: str = ""
    azami_tutar: str = ""
    azami_oran: str = ""
    aciklama: str = ""
    guncelleme_tarihi: str = ""

    @property
    def text(self) -> str:
        return " | ".join(
            x for x in (self.kategori, self.masraf, self.aciklama) if x
        )


@dataclass(frozen=True)
class Band:
    low: Optional[float]
    high: Optional[float]


@dataclass(frozen=True)
class RowSpec:
    """Karşılaştırmada görünen ORTAK satır tanımı."""

    label: str
    service: str
    band_key: Optional[str] = None
    detail: Optional[str] = None
    split_channel: bool = True


# ---------------------------------------------------------------------------
# ORTAK MASRAF SÖZLÜĞÜ
# ---------------------------------------------------------------------------
# Buradaki label'lar artık banka sitesindeki ham MASRAF adı değildir.
# Tüm bankalar bu ortak isimlere eşleştirilir.
#
# SECTION satırı: ("SECTION", "başlık")
# ROW satırı    : ("ROW", RowSpec(...))
# ---------------------------------------------------------------------------

LAYOUT = [
    # Kullanıcı denetimi sonrası güvenlik öncelikli karşılaştırma sözlüğü.
    # Temel ilke: KOMİSYONLAR sayfasında var olan doğru satırı kullan;
    # belirsiz/başka ürün olan veriyi zorla eşleştirme.

    ("SECTION", "EFT"),
    ("ROW", RowSpec("0 TRY - 8.300 TRY", "EFT", "TRANSFER_1")),
    ("ROW", RowSpec("8.300,01 TRY - 399.000 TRY", "EFT", "TRANSFER_2")),
    ("ROW", RowSpec("399.000,01 TRY -", "EFT", "TRANSFER_3")),

    ("SECTION", "ŞANS OYUNU"),
    ("ROW", RowSpec("Şans Oyunu Ödemeleri", "SANS_OYUNU")),

    ("SECTION", "PARA ÇEKME / ATM"),
    ("ROW", RowSpec("Günlük Limit Üzeri Para Çekme", "LIMIT_UZERI_PARA_CEKME", split_channel=False)),
    ("ROW", RowSpec("Ortak ATM Para Çekme", "ORTAK_ATM_PARA_CEKME", split_channel=False)),
    ("ROW", RowSpec("Ortak ATM Bakiye Sorgulama", "BAKIYE_ATM_YURTICI", split_channel=False)),
    ("ROW", RowSpec("Ortak ATM Bakiye Sorgulama – Yurtdışı", "BAKIYE_ATM_YURTDISI", split_channel=False)),

    ("SECTION", "FATURA / KURUM ÖDEMELERİ"),
    ("ROW", RowSpec("Hesaptan Fatura / Kurum Ödemesi", "FATURA", detail="HESAPTAN")),
    ("ROW", RowSpec("Kredi Kartından Fatura / Kurum Ödemesi", "FATURA", detail="KREDI_KARTI")),
    ("ROW", RowSpec("Nakit Fatura / Kurum Ödemesi", "FATURA", detail="NAKIT")),

    ("SECTION", "HAVALE"),
    ("ROW", RowSpec("0 TRY - 8.300 TRY", "HAVALE", "TRANSFER_1")),
    ("ROW", RowSpec("8.300,01 TRY - 399.000 TRY", "HAVALE", "TRANSFER_2")),
    ("ROW", RowSpec("399.000,01 TRY -", "HAVALE", "TRANSFER_3")),

    ("SECTION", "FAST"),
    ("ROW", RowSpec("0 TRY - 8.300 TRY", "FAST", "TRANSFER_1")),
    ("ROW", RowSpec("8.300,01 TRY - 399.000 TRY", "FAST", "TRANSFER_2")),

    # İlk sayfada ürün adı olarak bulunmayan sentetik Visa Direct / Global Fast
    # karşılaştırma satırları şimdilik kaldırıldı. Yalnız gerçek SWIFT ailesi kalır.
    ("SECTION", "YURT DIŞI TRANSFERLER"),
    ("ROW", RowSpec("SWIFT - Gelen", "SWIFT_GELEN", split_channel=False)),
    ("ROW", RowSpec("SWIFT - Giden", "SWIFT_GIDEN")),

    # Bankaların eşikleri aynı olmadığı için 100 ve 150 TRY kırılımları korunur.
    ("SECTION", "SGK TAHSİLAT"),
    ("ROW", RowSpec("0 TRY - 99,99 TRY", "SGK", "SGK_LOW")),
    ("ROW", RowSpec("100 TRY - 150 TRY", "SGK", "SGK_MID")),
    ("ROW", RowSpec("150,01 TRY -", "SGK", "SGK_HIGH")),

    ("SECTION", "KİRALIK KASA"),
    ("ROW", RowSpec("Büyük Kasa", "KASA", detail="BUYUK", split_channel=False)),
    ("ROW", RowSpec("Orta Kasa", "KASA", detail="ORTA", split_channel=False)),
    ("ROW", RowSpec("Küçük Kasa", "KASA", detail="KUCUK", split_channel=False)),
    ("ROW", RowSpec("Özel / Süper Kasa", "KASA", detail="OZEL", split_channel=False)),

    # Bankaların çek kırılımları birebir aynı değil. Aynı hizmet ailesindeki
    # gerçek alt tarifeler tek hücrede etiketli gösterilir; zorla tek fiyat seçilmez.
    ("SECTION", "ÇEK"),
    ("ROW", RowSpec("Çek Defteri / Çek Karnesi", "CEK_DEFTERI_GRUP", split_channel=False)),
    ("ROW", RowSpec("Bloke / Keşide Çeki Düzenleme", "CEK_DUZENLEME_STANDART_GRUP", split_channel=False)),
    ("ROW", RowSpec("Dövizli / Özel Nitelikli Çek Düzenleme", "CEK_DUZENLEME_OZEL_GRUP", split_channel=False)),
    ("ROW", RowSpec("Çek İade", "CEK_IADE_GRUP", split_channel=False)),
    ("ROW", RowSpec("Çek Tahsil - Yurtiçi", "CEK_TAHSIL_YURTICI_GRUP", split_channel=False)),
    ("ROW", RowSpec("Çek Tahsil - Dövizli / YP", "CEK_TAHSIL_DOVIZ_GRUP", split_channel=False)),
    ("ROW", RowSpec("Karşılıksız Çek / Düzeltme", "CEK_KARSILIKSIZ_GRUP", split_channel=False)),

    ("SECTION", "KKB / RİSK RAPORLARI"),
    ("ROW", RowSpec("KKB Risk Raporu", "KREDI_RISK", split_channel=False)),
    ("ROW", RowSpec("KKB Çek Bilgileri / Çek Risk Raporu", "CEK_RISK", split_channel=False)),

    ("SECTION", "HGS"),
    ("ROW", RowSpec("HGS Etiket Bedeli", "HGS_ETIKET", split_channel=False)),
    ("ROW", RowSpec("HGS Kart Bedeli", "HGS_KART", split_channel=False)),

    ("SECTION", "VERGİ / DEVLET ÖDEMELERİ"),
    ("ROW", RowSpec("Vergi Ödemeleri", "VERGI")),

    ("SECTION", "SENET"),
    ("ROW", RowSpec("Senet İade", "SENET_IADE", split_channel=False)),
    ("ROW", RowSpec("Senet Protesto", "SENET_PROTESTO", split_channel=False)),
    ("ROW", RowSpec("Senet Protesto Kaldırma", "SENET_PROTESTO_KALDIRMA", split_channel=False)),
    ("ROW", RowSpec("Senet Tahsil", "SENET_TAHSIL_GRUP", split_channel=False)),

    ("SECTION", "DÜZENLİ TRANSFERLER"),
    ("ROW", RowSpec("Düzenli EFT - 0 TRY - 8.300 TRY", "DUZENLI_EFT", "TRANSFER_1")),
    ("ROW", RowSpec("Düzenli EFT - 8.300,01 TRY - 399.000 TRY", "DUZENLI_EFT", "TRANSFER_2")),
    ("ROW", RowSpec("Düzenli EFT - 399.000,01 TRY -", "DUZENLI_EFT", "TRANSFER_3")),
    ("ROW", RowSpec("Düzenli Havale - 0 TRY - 8.300 TRY", "DUZENLI_HAVALE", "TRANSFER_1")),
    ("ROW", RowSpec("Düzenli Havale - 8.300,01 TRY - 399.000 TRY", "DUZENLI_HAVALE", "TRANSFER_2")),
    ("ROW", RowSpec("Düzenli Havale - 399.000,01 TRY -", "DUZENLI_HAVALE", "TRANSFER_3")),

    # Kullanıcı denetimine göre yapay 149,99/150 kırılımı kaldırıldı.
    # Kaynağın yayımladığı gerçek aralık tek hücrede gösterilir.
    ("SECTION", "AİDAT ÖDEMELERİ"),
    ("ROW", RowSpec("Aidat Ödemeleri", "AIDAT")),

    ("SECTION", "ÖZEL OKUL ÖDEME"),
    ("ROW", RowSpec("Özel Okul Ödemeleri", "OZEL_OKUL")),

    ("SECTION", "TELEFON ÖDEMELERİ"),
    ("ROW", RowSpec("Telefon / Cep Telefonu Faturası Ödemeleri", "TELEFON")),

    # Belge/ekstre/araştırma satırları genel HESAP_OZETI veya ARSIV etiketiyle
    # çözülmez. Her satır aşağıdaki ayrı hizmet koduyla banka + tam ürün adına
    # bağlanır; böylece ticari/bireysel, kart/mevduat ve posta/şube tarifeleri
    # birbirine taşınmaz.
    ("SECTION", "BELGE / EKSTRE / ARAŞTIRMA"),
    ("ROW", RowSpec("Vadesiz Mevduat Hesabı Ekstresi – TL", "DOC_VADESIZ_TL")),
    ("ROW", RowSpec("Vadesiz Mevduat Hesabı Ekstresi – YP", "DOC_VADESIZ_YP")),
    ("ROW", RowSpec("Bireysel KMH Hesap Özeti – Basılı", "DOC_KMH_BIREYSEL", split_channel=False)),
    ("ROW", RowSpec("Kredi Kartı Geçmiş Dönem Hesap Özeti", "DOC_KK_GECMIS", split_channel=False)),
    ("ROW", RowSpec("Arşiv Araştırma – Genel Belge", "DOC_ARSIV_GENEL", split_channel=False)),
    ("ROW", RowSpec("Mevduat Araştırma", "DOC_DEPOSIT_RESEARCH", split_channel=False)),
    ("ROW", RowSpec("1 Yıldan Eski Dekont/Ekstre Verilmesi", "DOC_ESKI_DEKONT_EKSTRE", split_channel=False)),
    ("ROW", RowSpec("Aylık Hesap Özeti – Posta Gönderimi", "DOC_AYLIK_POSTA", split_channel=False)),
    ("ROW", RowSpec("Arşiv/Araştırma – Şube Bazında", "DOC_ARSIV_SUBE", split_channel=False)),
    ("ROW", RowSpec("Arşiv/Araştırma – Banka Bazında", "DOC_ARSIV_BANKA", split_channel=False)),

    ("SECTION", "BANKAYA ÖZEL EKSTRE TARİFELERİ"),
    ("ROW", RowSpec("Ticari KMH Hesap Özeti – Basılı", "DOC_KMH_TICARI", split_channel=False)),
    ("ROW", RowSpec("Ticari Kart Geçmiş Dönem Ekstresi", "DOC_TICARI_KART_GECMIS", split_channel=False)),
    ("ROW", RowSpec("Yatırım Hesabı Ekstresi – Gönderim", "DOC_YATIRIM_EKSTRE", split_channel=False)),
    ("ROW", RowSpec("Ticari Müşteri Aylık Hesap Özeti", "DOC_TICARI_AYLIK", split_channel=False)),
    ("ROW", RowSpec("Ticari Kart – Basılı Ekstre Ücreti", "DOC_TICARI_KART_BASILI", split_channel=False)),
    ("ROW", RowSpec("Ticari Kart – Ekstre Gönderim Ücreti", "DOC_TICARI_KART_GONDERIM", split_channel=False)),
    ("ROW", RowSpec("Ticari Hesap – Basılı Ekstre Gönderimi", "DOC_TICARI_HESAP_BASILI", split_channel=False)),
    ("ROW", RowSpec("Üye İşyeri/POS – Ekstre Gönderimi", "DOC_POS_EKSTRE", split_channel=False)),

    # Menkul kıymet tarifeleri tek bir genel rakama indirgenemez. Hisse,
    # virman ve saklama aileleri ayrı gösterilir; banka yalnız açıklama/formül
    # yayımlıyorsa bu durum rakam uydurulmadan hücrede açıkça belirtilir.
    ("SECTION", "YURTİÇİ MENKUL KIYMET İŞLEMLERİ"),
    ("ROW", RowSpec("Hisse Senedi Alım-Satım Komisyonu", "SEC_DOMESTIC_STOCK_TRADE", split_channel=False)),
    ("ROW", RowSpec("Hisse Senedi Virmanı", "SEC_DOMESTIC_STOCK_TRANSFER", split_channel=False)),
    ("ROW", RowSpec("Bono/Tahvil Virmanı", "SEC_DOMESTIC_BOND_TRANSFER", split_channel=False)),
    ("ROW", RowSpec("Yatırım Hesabı / Saklama Ücreti", "SEC_DOMESTIC_CUSTODY", split_channel=False)),

    ("SECTION", "YURTDIŞI MENKUL KIYMET İŞLEMLERİ"),
    ("ROW", RowSpec("Yurtdışı Pay Alım-Satım Komisyonu", "SEC_FOREIGN_STOCK_TRADE", split_channel=False)),
    ("ROW", RowSpec("Yurtdışı Pay Saklama Ücreti", "SEC_FOREIGN_STOCK_CUSTODY", split_channel=False)),
    ("ROW", RowSpec("Eurobond Virmanı", "SEC_EUROBOND_TRANSFER", split_channel=False)),
    ("ROW", RowSpec("Eurobond Saklama Ücreti", "SEC_EUROBOND_CUSTODY", split_channel=False)),
]

# ---------------------------------------------------------------------------
# NORMALİZASYON
# ---------------------------------------------------------------------------

@lru_cache(maxsize=65536)
def _norm(value) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\u0307", "").replace("\xa0", " ")
    text = " ".join(text.split())
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.translate(str.maketrans({
        "ı": "i", "İ": "i", "ğ": "g", "Ğ": "g",
        "ü": "u", "Ü": "u", "ş": "s", "Ş": "s",
        "ö": "o", "Ö": "o", "ç": "c", "Ç": "c",
    }))
    return text.lower().strip()


@lru_cache(maxsize=65536)
def _has_word(text: str, word: str) -> bool:
    """Kısa hizmet kodlarını başka kelimelerin içinden yanlış yakalamaz."""
    t = _norm(text)
    w = _norm(word)
    return re.search(rf"(?<![a-z0-9]){re.escape(w)}(?![a-z0-9])", t) is not None


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_number(raw: str) -> Optional[float]:
    s = re.sub(r"[^0-9.,]", "", raw or "")
    if not s:
        return None

    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        parts = s.split(".")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]):
            s = "".join(parts)

    try:
        return float(s)
    except ValueError:
        return None


def _parse_band(text: str) -> Optional[Band]:
    """MASRAF adından işlem TUTARI aralığını çıkarır; ücret kolonlarına bakmaz."""
    t = _norm(text)

    # 8.300 TL ve altı
    m = re.search(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*(?:ve\s*)?(?:alti|altinda)",
        t,
    )
    if m:
        high = _parse_number(m.group(1))
        if high is not None:
            return Band(0.0, high)

    # 399.000,01 TL ve üzeri / üstü
    m = re.search(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*(?:ve\s*)?(?:uzeri|ustu)",
        t,
    )
    if m:
        low = _parse_number(m.group(1))
        if low is not None:
            return Band(low, None)

    # 8.300,01 TL - 399.000 TL
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*[-–—]\s*"
        r"(\d[\d.,]*)\s*(?:tl|try)?",
        t,
    ):
        local = t[max(0, m.start() - 6): min(len(t), m.end() + 6)]
        if ":" in local:  # saat aralığı (16:00-17:15) olmasın
            continue

        low = _parse_number(m.group(1))
        high = _parse_number(m.group(2))

        if low is None or high is None or high < low:
            continue

        return Band(low, high)

    return None



@lru_cache(maxsize=32768)
def _all_band_keys(text: str) -> Set[str]:
    """
    MASRAF + AÇIKLAMA içinde geçen bütün standart tutar bantlarını bulur.

    Böylece:
      - "149,99 TL'ye kadar ... 150 TL ve üzeri ..."
      - "0-150 TL ... 150,01 TL ve üzeri"
      - "100 TL altı ... 100 TL ve üzeri"
    gibi tek satır içinde iki farklı kademe yayınlayan bankalar da
    iki karşılaştırma satırına bağlanabilir.
    """
    t = _norm(text)
    bands: List[Band] = []

    # "8.300 TL ve altı", "100 TL altında"
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*(?:ve\s*)?(?:alti|altinda)",
        t,
    ):
        high = _parse_number(m.group(1))
        if high is not None:
            bands.append(Band(0.0, high))

    # "149,99 TL'ye kadar"
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?(?:'?[yea]*)?\s*kadar",
        t,
    ):
        high = _parse_number(m.group(1))
        if high is not None:
            bands.append(Band(0.0, high))

    # "399.000,01 TL ve üzeri / üstü"
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*(?:ve\s*)?(?:uzeri|ustu)",
        t,
    ):
        low = _parse_number(m.group(1))
        if low is not None:
            bands.append(Band(low, None))

    # Normal aralıklar: 0-8.300 / 8.300,01-399.000
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*[-–—]\s*"
        r"(\d[\d.,]*)\s*(?:tl|try)?(?:\s*arasi)?",
        t,
    ):
        local = t[max(0, m.start() - 6): min(len(t), m.end() + 6)]

        if ":" in local:
            continue

        low = _parse_number(m.group(1))
        high = _parse_number(m.group(2))

        if low is None or high is None or high < low:
            continue

        bands.append(Band(low, high))

    keys: Set[str] = set()

    for band in bands:
        key = _band_key(band)

        if key and not key.startswith("RAW:"):
            keys.add(key)

    return keys


def _close(a: Optional[float], b: Optional[float], tolerance: float = 1.01) -> bool:
    if a is None or b is None:
        return a is b
    return abs(a - b) <= tolerance


def _band_key(band: Optional[Band]) -> Optional[str]:
    if band is None:
        return None

    low, high = band.low, band.high

    # EFT / Havale / FAST mevzuat dilimleri.
    if low is not None and low <= 1.0 and _close(high, 8300.0):
        return "TRANSFER_1"
    if _close(low, 8300.01) and _close(high, 399000.0):
        return "TRANSFER_2"
    if low is not None and abs(low - 399000.01) <= 1.01 and high is None:
        return "TRANSFER_3"

    # Fatura / Aidat / Okul örnek şablon dilimleri.
    if low is not None and low <= 1.0 and _close(high, 149.99):
        return "FATURA_1"
    if low is not None and abs(low - 150.0) <= 1.01 and high is None:
        return "FATURA_2"

    # SGK örnek şablon dilimleri.
    if low is not None and low <= 1.0 and _close(high, 99.99):
        return "SGK_1"
    if low is not None and abs(low - 100.0) <= 1.01 and high is None:
        return "SGK_2"

    return f"RAW:{low}:{high}"


# ---------------------------------------------------------------------------
# HİZMET / KANAL SINIFLANDIRMA
# ---------------------------------------------------------------------------

@lru_cache(maxsize=16384)
def _service_tags(row: FeeRow) -> Set[str]:
    """Farklı bankaların adlandırmalarını ortak hizmet kimliklerine çevirir."""
    cat = _norm(row.kategori)
    mas = _norm(row.masraf)
    short = f"{cat} | {mas}"
    full = _norm(row.text)

    tags: Set[str] = set()

    # Supplemental satırları kendi kanonik hizmetini açıkça taşır.
    marker = re.search(r"service\s*=\s*([a-z0-9_]+)", full, flags=re.I)
    if marker:
        tags.add(marker.group(1).upper())

    # ---------------- PARA TRANSFERİ ----------------
    international = any(x in full for x in (
        "swift", "uluslararasi fon transfer", "yurt disi fast",
        "global fast", "western union", "fast uluslararasi",
    ))
    package = any(x in short for x in ("paket", "kota"))
    card_cash = any(x in short for x in ("nakit avans", "faiz orani"))

    if ((_has_word(full, "eft") or "elektronik fon transfer" in full)
            and not international and not package and "altin eft" not in full):
        if not card_cash or any(x in short for x in (
            "eft ucreti", "eft gonder", "elektronik fon transfer",
        )):
            tags.add("EFT")

    if ((_has_word(full, "fast") or "fonlarin anlik ve surekli transferi" in full)
            and not international and not package):
        tags.add("FAST")

    if (any(x in full for x in ("havale", "hesaptan hesaba"))
            and not international and not package):
        if not card_cash or any(x in short for x in (
            "havale ucreti", "havale gonder", "hesaptan hesaba",
        )):
            tags.add("HAVALE")

    # Referans karşılaştırma tablosundaki ek transfer aileleri.
    if any(x in full for x in ("swift", "uluslararasi fon transfer")):
        if any(x in full for x in ("gelen swift", "gelen doviz", "gelen uluslararasi", "gelen fon transfer")):
            tags.add("SWIFT_GELEN")
        if any(x in full for x in ("giden swift", "giden doviz", "doviz havale", "hesaptan giden", "gonderim")):
            tags.add("SWIFT_GIDEN")

    # Yurt dışı hızlı transferleri ürün markasına göre değil hedefe göre ayır:
    # 1) hesaba transfer, 2) karta transfer. Böylece Visa Direct / MoneySend /
    # Global Fast Karta gibi aynı işi farklı isimle sunan ürünler eşleşebilir.
    is_global_fast_card = (
        "global fast" in full
        and any(x in full for x in ("karta para gonder", "karta gonderim"))
    )

    is_fast_account = (
        any(x in full for x in (
            "yurt disi fast", "yurtdisi fast",
            "fast uluslararasi", "fast uluslararasi turkiye disina",
            "akbank fast uluslararasi",
        ))
        or ("global fast" in full and not is_global_fast_card)
        # Visa ile Yurt Dışı Para Transferi ayrı bir ürün ailesidir.
        # "Hızlı hesaba transfer" satırında yalnız bankaların FAST/Global FAST
        # benzeri kanonik ürünleri karşılaştırılır; aksi halde aynı hücreye iki
        # farklı ürün tarifesi yığılabiliyor.
    )
    if is_fast_account:
        tags.add("YURT_DISI_FAST")

    is_card_transfer = (
        # MasterCard kart ağı üzerinden yurt dışı karta gönderim.
        (
            "moneysend" in full
            and "alici" not in full
            and (
                any(x in full for x in (
                    "gonderici",
                    "yurtdisi banka kartina",
                    "yurt disi banka kartina",
                    "yurtdisi kart",
                    "yurt disi kart",
                ))
                # İş Bankası ücret tablosunda Bankamatik Kartından satırları
                # kategori altında "Moneysend" diye yayınlanır; MASRAF adında
                # "yurtdışı" kelimesi her zaman tekrar edilmez.
                or "bankamatik kartindan" in full
            )
        )
        # Yapı Kredi aynı işi Global Fast - Karta adıyla yayımlıyor.
        or is_global_fast_card
        # İş Bankası sözleşmelerinde aynı aile Moneysend / VISA Direct olarak geçiyor.
        or (
            "visa direct" in full
            and "alici" not in full
            and "hesaba" not in full
        )
    )
    if is_card_transfer:
        tags.add("KART_YURTDISI_TRANSFER")

    if "duzenli" in full and (_has_word(full, "eft") or "elektronik fon transfer" in full):
        tags.add("DUZENLI_EFT")
    if "duzenli" in full and "havale" in full:
        tags.add("DUZENLI_HAVALE")

    # ---------------- ATM / NAKİT ----------------
    if any(x in full for x in ("gunluk limit uzeri para cekme", "limit uzeri para cekme", "gunluk limit ustu para cekme", "limit ustu para cekme")):
        tags.add("LIMIT_UZERI_PARA_CEKME")

    if (
        "para cekme" in full
        and any(x in full for x in ("ortak atm", "diger banka atm", "baska kurulus", "baska banka atm"))
        and "limit uzeri" not in full
        and "nakit avans" not in full
    ):
        tags.add("ORTAK_ATM_PARA_CEKME")

    if (
        "para yatirma" in full
        and any(x in full for x in ("ortak atm", "diger banka atm", "baska kurulus", "baska banka atm", "atm"))
        and "nakit avans" not in mas
    ):
        tags.add("PARA_YATIRMA")

    if (
        any(x in mas for x in (
            "altin transfer",
            "ats ile altin gonderimi",
            "kiymetli maden transferi ucreti - altin",
            "kiymetli maden transfer - altin",
        ))
        and not any(x in mas for x in (
            "fiziki", "teslim", "kulce altin cekme",
            "western union", "eft", "havale", "fast",
        ))
    ):
        tags.add("ALTIN_TRANSFER")

    # ---------------- KASA / RAPOR ----------------
    if any(x in full for x in ("kiralik kasa", "kasa kiralama", "kasa ucreti")):
        tags.add("KASA")

    if (
        any(x in full for x in (
            "kiymetli maden teslim", "fiziki altin teslim", "altin teslim",
            "kulce altin cekme", "kulce altin teslim", "fiziki altin cekme",
            "musteriye fiziki altin teslim",
        ))
        and not any(x in full for x in (
            "musteriden fiziki altin kabul",
            "fiziki altin kabul",
        ))
    ):
        tags.add("KIYMETLI_MADEN_TESLIM")

    # Garanti gibi bazı bankalar "KKB Çek / Risk Raporu"nu tek satırda
    # yayımlar. Böyle bir satır hem çek risk hem genel risk raporunu temsil eder.
    combined_risk = any(x in full for x in (
        "kkb cek / risk raporu", "kkb cek/risk raporu", "cek / risk raporu",
    ))
    if combined_risk:
        tags.update({"CEK_RISK", "KREDI_RISK"})
    else:
        if any(x in full for x in (
            "cek risk raporu", "cek bilgileri raporu", "kkb cek", "findeks cek raporu",
            "cek sorgu raporu",
        )):
            tags.add("CEK_RISK")
        if any(x in full for x in (
            "risk raporu", "kkb risk", "kredi risk", "findeks risk", "risk merkezi raporu",
        )):
            tags.add("KREDI_RISK")

    # ---------------- TAHSİLAT / ÖDEME ----------------
    if (("fatura" in short or "fatura / kurum" in short or "fatura/kurum" in short
         or "fatura tahsil" in short or "kurum tahsil" in short)
            and "e-fatura" not in short):
        tags.add("FATURA")

    if "sgk" in short or "sosyal guvenlik" in full:
        tags.add("SGK")

    # HGS Etiket ve HGS Kart farklı ürünlerdir.
    # "HGS" kelimesi tek başına yeterli değildir; ürün tipi MASRAF/KATEGORİ
    # alanında açıkça görünmelidir.
    hgs_structural = short
    if "hgs" in hgs_structural:
        has_etiket = "etiket" in hgs_structural
        has_kart = bool(re.search(r"\bkart(?:i|ı|lari|ları|ucreti|ücreti)?\b", hgs_structural))
        if has_etiket and not has_kart:
            tags.add("HGS_ETIKET")
        if has_kart and not has_etiket:
            tags.add("HGS_KART")

    if "sans oyun" in full or any(x in full for x in (
        "bilyoner", "nesine", "tuttur", "oley", "misli", "sisal sans", "tjk",
    )):
        tags.add("SANS_OYUNU")

    # Aidat ödeme hizmetini kredi kartı "Aidatsız Kart" / kart aidatı kavramından ayır.
    has_aidat_word = bool(re.search(r"\baidat\b", short)) or (
        bool(re.search(r"\baidat\b", full))
        and any(x in full for x in ("fatura", "tahsilat", "odeme", "site", "apartman"))
    )
    if has_aidat_word and not any(x in full for x in (
        "aidatsiz kart", "kart aidati", "yillik kart ucreti", "uyelik ucreti - kart",
    )):
        tags.add("AIDAT")

    # Özel okul ÖDEME hizmetini "Vize ve Özel Okullar İçin Düzenlenen Mektup"
    # gibi belge hizmetlerinden kesin olarak ayır.
    if (any(x in full for x in (
        "ozel okul", "okul odeme", "okul taksiti", "egitim odeme", "egitim kurumu odeme",
    )) and not any(x in full for x in (
        "mektup", "vize", "konsolosluk", "referans yazisi", "referans mektubu",
    ))):
        tags.add("OZEL_OKUL")

    telefon_candidate = (
        any(x in short for x in (
            "telefon odeme", "telefon fatur", "cep telefonu fatur",
            "telefon operatorleri odemelerine aracilik",
            "gsm odeme", "telekom odeme", "turkcell", "vodafone",
        ))
        or (
            any(x in full for x in ("turkcell", "vodafone", "superonline", "tellcom", "turk telekom"))
            and any(x in short for x in ("fatura", "kurum odeme"))
        )
    )
    if (
        telefon_candidate
        and "tl/paket yukleme" not in mas
        and "paket yukleme" not in mas
        and "otomatik fatura odeme" not in mas
        and "alisveris faiz" not in mas
    ):
        tags.add("TELEFON")

    if ("vergi" in short and any(x in short for x in (
        "vergi tahsil", "vergi odeme", "fatura/vergi/sgk", "fatura / vergi / sgk", "mtv",
    )) and not any(x in short for x in ("vergi numarasi", "vergi yazisi", "kredi"))):
        tags.add("VERGI")

    # ---------------- BELGE / HESAP ----------------
    if (any(x in full for x in (
        "arsiv arastirma", "gecmis donem ekstre arsiv", "gecmis donem bankacilik islemleri bildirimi",
        "arsivden belge", "dokuman arastirma", "dokuman talebi",
        "1 yildan eski islemlere ait gecmise yonelik dekont", "gecmise yonelik dekont",
        "gecmis donem ekstre", "dekont masrafi - gecmise yonelik",
    )) or ("arsiv" in full and "arastirma" in full)):
        tags.add("ARSIV")

    if (any(x in full for x in ("mevduat arastirma", "mevduat hesap arastirma"))
            or ("arsiv" in cat and "arastirma" in cat and mas.startswith("merkezden"))):
        tags.add("MEVDUAT_ARASTIRMA")

    if any(x in short for x in (
        "referans mektubu", "referans yazisi", "banka referans",
        "itibar mektubu", "itibar /niyet/referans", "itibar/niyet/referans",
    )):
        tags.add("REFERANS_MEKTUBU")

    if any(x in full for x in (
        "vize icin", "konsolosluk icin mektup", "konsolosluk icin",
        "ozel okullar icin duzenlenen mektup", "ozel okul icin duzenlenen mektup",
    )):
        tags.add("VIZE_MEKTUBU")

    postal_statement = any(x in full for x in (
        "posta ile aylik hesap ozeti", "basili ekstre", "basili hesap ozeti",
        "ekstre gonderim", "hesap ozeti gonderimi", "hesap ozeti posta", "ekstre posta",
        "gecmis donem kredi karti hesap ozeti gonderimi",
    ))

    if postal_statement:
        tags.add("HESAP_OZETI_POSTA")
    elif (
        "hesap ozeti" in short
        or "ekstre masraf" in short
        or "ekstre ucret" in short
        or "ekstre veril" in short
        or "ekstre - " in short
    ):
        tags.add("HESAP_OZETI")

    if any(x in full for x in ("hesap arastirma", "hesap tespit", "hesap bulma")):
        tags.add("HESAP_ARASTIRMA")
    if any(x in full for x in ("borcu yoktur", "borc yoktur", "borcsuzluk")):
        tags.add("BORCU_YOKTUR")

    # ---------------- ATM BAKİYE ----------------
    if (any(x in full for x in ("bakiye sorma", "bakiye sorgu", "bakiye goruntule"))
            and (any(x in full for x in ("atm", "bankamatik"))
                 or "baska kurulus araciligiyla yapilan islemler" in full)):
        if any(x in full for x in ("yurt disi", "yurtdisi")):
            tags.add("BAKIYE_ATM_YURTDISI")
        else:
            tags.add("BAKIYE_ATM_YURTICI")

    # ---------------- ÇEK ----------------
    # Çek tarafında yanlış eşleştirme riski yüksek olduğu için mümkün olduğunca
    # MASRAF + KATEGORİ gibi yapısal alanlar kullanılır; açıklama tek başına
    # ürün kimliği belirlemez.
    cek_struct = short

    # Çek defteri: yaprak başı ve 10 yapraklı karne ayrı karşılaştırılır.
    if (
        any(x in cek_struct for x in ("cek defteri", "cek karnesi", "cek yaprak", "cek kitabi"))
        and any(x in cek_struct for x in ("yaprak basi", "yaprak başi", "yaprak başı", "50-350 yaprakli", "351 yaprak"))
    ):
        tags.add("CEK_DEFTERI_YAPRAK")

    if (
        any(x in cek_struct for x in ("cek defteri", "cek karnesi"))
        and any(x in cek_struct for x in ("10 yaprakli", "10'luk", "10luk"))
    ):
        tags.add("CEK_KARNESI_10")

    # Standart bloke/keşide çeki düzenleme. Hediye/armağan, seyahat,
    # dövizli/DTH ve karşılıklı çekler bu aileye giremez.
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "bloke cek duzenleme",
            "bloke/ keside ceki duzenleme",
            "bloke/keside ceki duzenleme",
            "keside ceki / bloke cek duzenleme",
            "keside cek duzenleme",
            "duzenleme - tp/yp",
        ))
        and not any(x in cek_struct for x in (
            "hediye", "armagan", "seyahat", "doviz", "dovizi natik",
            "dth", "karsilikli", "odeme", "durdurma",
        ))
    ):
        tags.add("CEK_DUZENLEME_STANDART")

    # Dövizli / DTH çek düzenleme.
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "dovizli cek duzenleme",
            "dovizi natik cek duzenleme",
            "dth'dan cek duzenlenmesi",
            "dth dan cek duzenlenmesi",
        ))
        and not any(x in cek_struct for x in ("odeme", "durdurma", "tahsil"))
    ):
        tags.add("CEK_DUZENLEME_DOVIZ")

    # İşlemsiz/muamelesiz çek iadesi. Senet iadesi veya başka iade türleri alınmaz.
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "cek muamelesiz iade",
            "cekin islemsiz iades",
            "ceklerin islemsiz iades",
            "tahsile verilen cekin islemsiz iades",
            "cek iade ucreti",
        ))
        and "senet" not in cek_struct
    ):
        tags.add("CEK_IADE")

    # Çek tahsilatı yalnız açık tahsil/tahsile alma ifadelerinden üretilir.
    # Gişeden çek ödeme / bloke çek ödeme ayrı işlemdir ve tahsilata girmez.
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "cek tahsil",
            "tahsile alinan cek",
            "tahsile alinan bankamiz ceki",
            "tahsile alinan diger banka ceki",
            "bankamiz ceki (tl-yp) tahsile alma",
            "baska banka ceki (tl-yp) tahsile alma",
            "yurtici banka ceki",
            "ykb ceki",
            "yp cek tahsilati",
            "dovizli cek - tahsile alinan",
            "tahsile alinan dovizli cek",
        ))
        and not any(x in cek_struct for x in (
            "gişeden cek odeme", "giseden cek odeme", "bloke cek odeme",
            "seyahat ceki odeme", "karsiliksiz cek elden odeme",
            "cek odeme -", "odemeyi durdurma",
        ))
    ):
        tags.add("CEK_TAHSIL")

    if (
        "karsiliksiz cek" in cek_struct
        and any(x in cek_struct for x in ("belgelendirme", "elden odeme"))
    ):
        tags.add("CEK_KARSILIKSIZ")

    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "duzeltme hakki",
            "duzeltme ucreti",
        ))
    ):
        tags.add("CEK_DUZELTME_HAKKI")

    # ---------------- SENET ----------------
    # "Çekler ve Senetler" kategori başlığı yüzünden çek satırlarının senet
    # olarak etiketlenmesini engelle. Öncelik MASRAF adıdır; MASRAF kısa ise
    # yalnız açık "Senet Tahsil" kategorisi fallback olarak kullanılabilir.
    if "senet" in mas and "iade" in mas:
        tags.add("SENET_IADE")

    protesto_name = (
        "senet" in mas
        and "protesto" in mas
        and "protestosuz" not in mas
        and "iskonto" not in mas
        and "istira" not in mas
    )
    if protesto_name and "kaldir" not in mas:
        tags.add("SENET_PROTESTO")

    if protesto_name and "kaldir" in mas:
        tags.add("SENET_PROTESTO_KALDIRMA")

    senet_tahsil_context = (
        ("senet" in mas and any(x in mas for x in ("tahsil", "tahsile alma")))
        or (
            "senet tahsil" in cat
            and "cek" not in mas
            and any(x in mas for x in ("ayni sube", "diger sube", "tp/yp"))
        )
    )
    if senet_tahsil_context:
        tags.add("SENET_TAHSIL")

    return tags


@lru_cache(maxsize=16384)
def _channels(row: FeeRow) -> Set[str]:
    """
    Bir satırın geçerli olduğu kanal kümesini döndürür.

    Kanal önce KATEGORİ + MASRAF gibi yapısal alanlardan çıkarılır.
    "İnternet Şube" fiziksel şube değildir. Özellikle İş Bankası'nın
    "İnternet Şube, İşCep, Çözüm Merkezi" birleşik dijital tarifesi yalnız
    MOBİL/DİJİTAL olarak sınıflandırılır; fiziksel "... - Şube" ayrı kalır.
    """
    cat = _norm(row.kategori)
    mas = _norm(row.masraf)
    desc = _norm(row.aciklama)
    full = _norm(row.text)

    explicit = re.search(r"channel\s*=\s*([a-z0-9_]+)", full, flags=re.I)
    if explicit:
        channel = explicit.group(1).upper()
        if channel == "GENEL":
            return {"GENEL"}
        return {channel}

    structural = f"{cat} | {mas}"

    # İş Bankası'nın dijital birleşik adı: "İnternet Şube, İşCep, Çözüm Merkezi".
    # Buradaki "Şube" ve "Çözüm Merkezi" fiziksel şube tarifesi anlamına gelmez.
    is_isbank_digital_bundle = (
        row.banka == "İŞBANKASI"
        and (
            "internet sube" in structural
            or "iscep" in structural
        )
        and "fatura odemeleri" in structural
    )
    if is_isbank_digital_bundle:
        return {"MOBIL"}

    structural2 = (
        structural
        .replace("internet subesi", "internet")
        .replace("internet sube", "internet")
    )

    if "tum kanal" in structural2:
        return {"MOBIL", "SUBE", "ATM"}

    channels: Set[str] = set()

    if any(x in structural2 for x in (
        "mobil", "internet", "dijital", "cepteteb", "iscep",
        "asistan", "sgk.gov.tr", "web",
    )):
        channels.add("MOBIL")

    if any(x in structural2 for x in (
        "sube", "subeden", "musteri iletisim merkezi",
        "cozum merkezi", "telefon subesi", "kasadan",
    )):
        channels.add("SUBE")

    if any(x in structural2 for x in (
        "atm", "btm", "kiosk", "bankamatik",
    )):
        channels.add("ATM")

    if channels:
        return channels

    # Açıklama fallback'i yalnız açık işlem-kanalı ifadelerinde kullanılır.
    desc2 = (
        desc
        .replace("internet subesi", "internet")
        .replace("internet sube", "internet")
    )

    if any(x in desc2 for x in (
        "mobil uzerinden", "mobil'den", "mobilden",
        "internet uzerinden", "dijital kanaldan",
        "yalniz mobil", "sadece mobil",
    )):
        channels.add("MOBIL")

    if any(x in desc2 for x in (
        "subeden yapilan", "subeden gerceklestir",
        "sube kanalindan", "yalniz sube", "sadece sube",
        "musteri iletisim merkezi'nden", "musteri iletisim merkezinden",
    )):
        channels.add("SUBE")

    if any(x in desc2 for x in (
        "atm'den", "atmden", "atm uzerinden",
        "bankamatikten", "bankamatik uzerinden",
    )):
        channels.add("ATM")

    return channels or {"GENEL"}

def _detail_match(row: FeeRow, detail: Optional[str]) -> bool:
    if not detail:
        return True

    text = _norm(row.text)

    tests = {
        "BUYUK": lambda: (
            any(x in text for x in ("buyuk", "large"))
            and not any(x in text for x in ("ozel", "super", "extra buyuk", "xl"))
        ),
        "ORTA": lambda: (
            any(x in text for x in ("orta", "medium"))
            and not any(x in text for x in ("ozel", "super", "extra buyuk", "xl"))
        ),
        "KUCUK": lambda: (
            any(x in text for x in ("kucuk", "small"))
            and not any(x in text for x in ("ozel", "super", "extra buyuk", "xl"))
        ),
        "OZEL": lambda: any(
            x in text
            for x in (
                "ozel",
                "super",
                "extra buyuk",
                "xl",
            )
        ),
        "AYNI": lambda: any(
            x in text
            for x in (
                "ayni banka",
                "ayni sube",
                "bankamiza ait",
                "bankamiz cek",
                "bankamiz ceki",
                "ykb ceki",
                "gb subesinden",
                "kendi bankasi",
            )
        ),
        "DIGER": lambda: any(
            x in text
            for x in (
                "diger banka",
                "diger sube",
                "farkli sube",
                "muhabir banka",
                "muhabirden",
                "baska banka",
                "yurtici banka ceki",
            )
        ),
        "DOVIZ": lambda: any(
            x in text
            for x in (
                "doviz cek",
                "dovizli cek",
                "dovizi natik",
                "yp cek",
                "yabanci banka",
                "yurt disi",
            )
        ),
    }

    fn = tests.get(detail)
    return fn() if fn else True



# ---------------------------------------------------------------------------
# EŞLEŞME PUANI
# ---------------------------------------------------------------------------

def _candidate_score(row: FeeRow, spec: RowSpec, wanted_channel: str) -> int:
    tags = _service_tags(row)

    if spec.service not in tags:
        return -10_000

    score = 100
    mas = _norm(row.masraf)
    cat = _norm(row.kategori)
    full = _norm(row.text)

    # Tekil transfer karşılaştırmasına KOBİ/paket/kota fiyatı asla girmez.
    # Eski bir çıktıda "Kobi Giden Swift Paketi 250" tek SWIFT ücreti gibi
    # seçilip 685.714,29 TL gösterilmişti. Bu artık hard-reject.
    transfer_services = {
        "EFT", "HAVALE", "FAST", "SWIFT_GELEN", "SWIFT_GIDEN",
        "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER",
        "DUZENLI_EFT", "DUZENLI_HAVALE",
    }
    if spec.service in transfer_services:
        if (
            any(x in mas for x in ("paket", "kobi", "kota"))
            or "urun ve hizmet paket" in cat
            or "ek kaynak - akbank ticari" in cat
        ):
            return -10_000

    # Kaynak önceliği:
    # - Bireysel ana Ürün/Hizmet Ücretleri satırı temel referanstır.
    # - Ek resmî kaynak yalnız ana kaynakta eksik kalan kalemi tamamlar.
    # - Ticari paket/ücret, bireysel karşılaştırmaya taşınmaz.
    is_supplemental = cat.startswith("ek kaynak")
    if not is_supplemental:
        score += 38
    if "ek kaynak - akbank ticari" in cat:
        score -= 55

    if STATUS_NUMERIC.lower() in row.aciklama.lower():
        score += 18
    if STATUS_AVAILABLE.lower() in row.aciklama.lower():
        score -= 55
    if STATUS_EMPTY.lower() in row.aciklama.lower():
        score -= 45
    if STATUS_NOT_APPLICABLE.lower() in row.aciklama.lower():
        score -= 80

    # ---------------- BAND ----------------
    if spec.band_key:
        masraf_keys = _all_band_keys(row.masraf)
        all_keys = _all_band_keys(
            f"{row.masraf} | {row.aciklama}"
        )

        if spec.band_key in all_keys:
            score += 40

            # Band doğrudan MASRAF adındaysa en güvenilir eşleşme.
            if spec.band_key in masraf_keys:
                score += 22

        else:
            # Bazı bankalar Fatura / SGK gibi hizmetleri tutar bandı
            # belirtmeden "tüm tutarlara %X" olarak yayınlıyor.
            # Bu durumda iki ortak banda da aynı tarife uygulanabilir.
            if spec.service in {
                "FATURA",
                "SGK",
                "AIDAT",
                "OZEL_OKUL",
            }:
                score -= 12
            else:
                return -10_000

    # ---------------- DETAY ----------------
    # Çek tahsilatında "aynı şube çek ödeme" ile "bankanın kendi çekini
    # tahsile alma" aynı işlem değildir. Yanlış rakam riskini azaltmak için
    # çek tahsil detayları fail-closed / açık ürün adı ile eşleşir.
    if spec.service == "CEK_TAHSIL":
        structural = f"{cat} | {mas}"

        if spec.detail == "AYNI":
            same_bank = any(x in structural for x in (
                "bankamiz ceki",
                "bankamiza ait",
                "tahsile alinan bankamiz ceki",
                "ykb ceki",
                "garanti bankasi ceki",
                "kendi bankasi ceki",
            ))
            if not same_bank:
                return -10_000

        elif spec.detail == "DIGER":
            other_bank = any(x in structural for x in (
                "diger banka ceki",
                "baska banka ceki",
                "tahsile alinan diger banka ceki",
                "yurtici banka ceki",
            ))
            if not other_bank:
                return -10_000

        elif spec.detail == "DOVIZ":
            fx_cheque = any(x in structural for x in (
                "yp cek tahsil",
                "dovizli cek - tahsile alinan",
                "tahsile alinan dovizli cek",
                "tahsile alinan yp cekler",
                "diger banka yp - tahsile alinan",
                "yurt disi yabanci banka doviz ceki tahsile alinmasi",
            ))
            if not fx_cheque:
                return -10_000

        # Açık aynı/diğer/döviz tahsil ifadesi bulundu: yüksek güven.
        score += 80
        if "tahsile alinan" in structural or "tahsile alma" in structural:
            score += 25
        if "teminata alinan" in structural:
            score -= 30

    elif not _detail_match(row, spec.detail):
        generic_detail_fallback = False

        # Senette bazı bankalar aynı/diğer ayrımı yapmadan tek genel tarife
        # yayımlar. Yalnız senet tarafında düşük öncelikli fallback korunur.
        if spec.service == "SENET_TAHSIL":
            generic_detail_fallback = (
                "senet tahsil" in full
                and not any(
                    x in full
                    for x in (
                        "ayni sube",
                        "diger sube",
                        "muhabir",
                        "baska banka",
                    )
                )
            )

        if generic_detail_fallback:
            score -= 18
        else:
            return -10_000

    # ---------------- KANAL ----------------
    if spec.split_channel:
        channels = _channels(row)

        if wanted_channel in channels:
            score += 60
        elif "GENEL" in channels:
            score += 15
        else:
            return -10_000

    # Hizmet MASRAF adında açıkça geçiyorsa açıklama fallback'ından üstündür.
    if spec.service == "EFT" and _has_word(mas, "eft"):
        score += 30
    elif spec.service == "FAST" and _has_word(mas, "fast"):
        score += 35
    elif spec.service == "HAVALE" and "havale" in mas:
        score += 30
    elif spec.service in _service_tags(
        FeeRow(
            banka=row.banka,
            kategori=row.kategori,
            masraf=row.masraf,
        )
    ):
        score += 20

    if spec.service in {"SWIFT_GELEN", "SWIFT_GIDEN"}:
        if "swift" in mas:
            score += 35
        if spec.service == "SWIFT_GELEN" and "gelen" in full:
            score += 45
        if spec.service == "SWIFT_GIDEN" and any(x in full for x in ("giden", "gonder")):
            score += 45

    if spec.service == "YURT_DISI_FAST" and (
        any(x in full for x in (
            "yurt disi fast", "yurtdisi fast", "fast uluslararasi",
            "akbank fast uluslararasi",
        ))
        or ("global fast" in full and "karta para gonder" not in full)
        or "visa ile yurt disi para transferi" in full
        or "visa ile yurtdisi para transferi" in full
    ):
        score += 65
    if spec.service == "KART_YURTDISI_TRANSFER" and any(x in full for x in (
        "moneysend", "visa direct", "global fast - karta", "karta para gonderim",
    )):
        score += 65
    if spec.service == "LIMIT_UZERI_PARA_CEKME" and any(x in full for x in ("limit uzeri", "limit ustu")):
        score += 60
    if spec.service == "ORTAK_ATM_PARA_CEKME" and "para cekme" in full:
        score += 45
    if spec.service == "PARA_YATIRMA" and "para yatirma" in full:
        score += 45
    if spec.service == "DUZENLI_EFT" and "duzenli" in full and _has_word(full, "eft"):
        score += 55
    if spec.service == "DUZENLI_HAVALE" and "duzenli" in full and "havale" in full:
        score += 55
    if spec.service == "ALTIN_TRANSFER" and "altin" in full:
        score += 45

    # Arşiv kategorisi birçok farklı özel raporu içeriyor.
    # "Borcu Yoktur", Risk Raporu, Vize Mektubu gibi alt hizmetlerin
    # Arşiv Araştırma satırına yanlış düşmesini engelle.
    if spec.service == "ARSIV":
        specialized = {
            "BORCU_YOKTUR",
            "KREDI_RISK",
            "CEK_RISK",
            "VIZE_MEKTUBU",
            "REFERANS_MEKTUBU",
            "MEVDUAT_ARASTIRMA",
        }

        if tags & specialized:
            score -= 90

        if any(
            x in mas
            for x in (
                "gecmis donem bankacilik islemleri bildirimi",
                "sozlesme, dekont",
                "dokuman talebi",
                "1 yildan eski",
                "gecmise yonelik",
                "arsiv arastirma",
            )
        ):
            score += 65

    if spec.service == "HESAP_OZETI":
        if any(x in full for x in (
            "posta ile", "basili ekstre", "basili hesap ozeti",
            "ekstre gonderim", "hesap ozeti gonderimi",
        )):
            return -10_000
        if any(x in mas for x in ("ekstre", "hesap ozeti")):
            score += 25

    if spec.service == "SENET_TAHSIL":
        if any(x in mas for x in ("iskonto", "istira", "teminat")):
            return -10_000
        if "cek" in mas:
            return -10_000

    if spec.service == "HESAP_OZETI_POSTA":
        if "kktc" in full:
            score -= 40

        if any(
            x in full
            for x in (
                "posta ile",
                "basili ekstre",
                "ekstre gonderim",
                "hesap ozeti gonderimi",
            )
        ):
            score += 55

    if spec.service == "BAKIYE_ATM_YURTDISI":
        if "kktc" in full:
            score -= 90
        if any(x in full for x in ("turkiye kart", "turkiye bireysel", "t.c. kart")):
            score += 70

    # Standart işlem ücretini özel varyantlardan öne al.
    if any(
        x in mas
        for x in (
            "gonderimi",
            "gonderilmesi",
            "gonderme",
        )
    ):
        score += 18

    if any(
        x in mas
        for x in (
            "gec eft",
            "gec havale",
            "gec fast",
        )
    ):
        score -= 35

    if (
        spec.service in {"EFT", "HAVALE", "FAST", "SWIFT_GELEN", "SWIFT_GIDEN", "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER", "DUZENLI_EFT", "DUZENLI_HAVALE"}
        and any(
            x in mas
            for x in (
                "duzenli",
                "talimat",
                "supurme",
            )
        )
    ):
        score -= 12

    if any(
        x in mas
        for x in (
            "cebe",
            "kartsiz",
            "kartli para yatirma",
        )
    ):
        score -= 24

    if any(
        x in cat
        for x in (
            "kampanyali",
            "urun ve hizmet paket",
        )
    ):
        score -= 45

    # Template'teki Fatura satırları kart ödemesi odaklı.
    if spec.service == "FATURA":
        if (
            "kredi kart" in full
            or "kartindan" in full
        ):
            score += 25

        if "nakit" in mas:
            score -= 20

    if (
        spec.service == "SGK"
        and "kredi kart" in full
    ):
        score += 20

    if spec.service == "CEK_RISK":
        if any(x in mas for x in ("cek risk", "cek bilgileri", "kkb cek", "cek sorgu")):
            score += 70

    if spec.service == "KREDI_RISK":
        if "gercek kisiler" in full:
            score += 120
        if "ek kaynak - akbank ticari" in cat or "urun ve hizmet paketleri" in cat:
            score -= 110

    if spec.service == "PARA_YATIRMA":
        if "banka kartiyla" in mas and "cari hesaba para yatirma" in mas:
            score += 85
        if "t.c. subelerinden" in mas:
            score += 35
        if "standart atm" in mas:
            score += 28
        if "k.k.t.c." in mas:
            score -= 55
        if "kredi kartiyla" in mas or "maxipara" in mas:
            score -= 45
        if "ek kaynak - is bankasi banka karti sozlesmesi" in cat:
            score += 125

    if spec.service == "TELEFON":
        if any(x in mas for x in ("tl/paket", "paket yukleme", "otomatik fatura odeme faizi")):
            return -10_000

    if spec.service == "SENET_IADE":
        if (
            "senet" not in mas
            or "iade" not in mas
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            return -10_000
        score += 70

    if spec.service == "SENET_PROTESTO":
        if (
            "senet" not in mas
            or "protesto" not in mas
            or "protestosuz" in mas
            or "kaldir" in mas
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            return -10_000
        score += 80

    if spec.service == "SENET_PROTESTO_KALDIRMA":
        if (
            not ("senet" in mas and "protesto" in mas and "kaldir" in mas)
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            return -10_000
        score += 80

    if spec.service == "HESAP_OZETI_POSTA":
        if any(x in full for x in ("kredi kart", "ticari kart", "business kart")):
            return -10_000

    if spec.service == "KASA":
        if "depozito" in full:
            score -= 80

        if "aylik" in full:
            score -= 35

        if "yillik" in full:
            score += 45

    # ---------------- YÜKSEK RİSKLİ HİZMETLER ----------------
    # Bu grupta "yakın isim" yeterli değildir. Ürünün yapısal adında açık
    # eşleşme yoksa rakam göstermek yerine kaynak boşluğu bırakılır.
    structural = f"{cat} | {mas}"

    if spec.service == "HGS_ETIKET":
        if not ("hgs" in structural and "etiket" in structural and "kart" not in mas):
            return -10_000
        score += 120

    if spec.service == "HGS_KART":
        if not ("hgs" in structural and "kart" in structural and "etiket" not in mas):
            return -10_000
        score += 120

    if spec.service == "CEK_DEFTERI_YAPRAK":
        if not (
            any(x in structural for x in ("cek defteri", "cek karnesi"))
            and any(x in structural for x in (
                "yaprak basi", "50-350 yaprakli", "351 yaprak",
            ))
        ):
            return -10_000
        if any(x in mas for x in ("10 yaprakli", "10'luk", "25 yaprakli", "25'lik")):
            return -10_000
        score += 110

    if spec.service == "CEK_KARNESI_10":
        if not (
            any(x in structural for x in ("cek defteri", "cek karnesi"))
            and any(x in structural for x in ("10 yaprakli", "10'luk", "10luk"))
        ):
            return -10_000
        score += 110

    if spec.service == "CEK_DUZENLEME_STANDART":
        if not any(x in structural for x in (
            "bloke cek duzenleme",
            "bloke/ keside ceki duzenleme",
            "bloke/keside ceki duzenleme",
            "keside ceki / bloke cek duzenleme",
            "duzenleme - tp/yp",
        )):
            return -10_000
        if any(x in structural for x in (
            "hediye", "armagan", "seyahat", "doviz", "dovizi natik",
            "dth", "karsilikli", "odeme", "durdurma",
        )):
            return -10_000
        score += 120

    if spec.service == "CEK_DUZENLEME_DOVIZ":
        if not any(x in structural for x in (
            "dovizli cek duzenleme",
            "dovizi natik cek duzenleme",
            "dth'dan cek duzenlenmesi",
            "dth dan cek duzenlenmesi",
        )):
            return -10_000
        if any(x in structural for x in ("odeme", "durdurma", "tahsil")):
            return -10_000
        score += 120

    if spec.service == "CEK_IADE":
        if not (
            "cek" in structural
            and any(x in structural for x in (
                "muamelesiz iade",
                "islemsiz iade",
                "cek iade ucreti",
            ))
        ):
            return -10_000
        score += 110

    if spec.service == "CEK_KARSILIKSIZ":
        if not (
            "karsiliksiz cek" in structural
            and any(x in structural for x in ("belgelendirme", "elden odeme"))
        ):
            return -10_000
        score += 100

    if spec.service == "CEK_DUZELTME_HAKKI":
        if not (
            "cek" in structural
            and any(x in structural for x in ("duzeltme hakki", "duzeltme ucreti"))
        ):
            return -10_000
        score += 100

    return score



_AMBIGUITY_LOGGED: Set[Tuple[str, str, str, str]] = set()


def _row_fee_signature(row: FeeRow) -> Tuple[str, str, str, str]:
    """Adayların gerçekten aynı tarifeyi taşıyıp taşımadığını karşılaştırır."""
    return tuple(
        _clean(getattr(row, attr, ""))
        for attr in ("asgari_tutar", "asgari_oran", "azami_tutar", "azami_oran")
    )


def _best_match(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
) -> Optional[FeeRow]:
    """
    En iyi adayı seçer; belirsiz durumda FAIL-CLOSED davranır.

    Önceki sürüm her zaman en yüksek puanlı ilk satırı seçiyordu. İki farklı
    tarife birbirine çok yakın puan alırsa bu yaklaşım yanlış rakam
    üretebiliyordu. Artık yakın puanlı ve farklı ücretli aday varsa hiçbirini
    seçmiyoruz; hücre "Kontrol gerekli" / kaynak boşluğu olarak kalıyor.
    """
    candidates: List[Tuple[int, FeeRow]] = []

    for row in rows:
        if row.banka != bank:
            continue

        score = _candidate_score(row, spec, wanted_channel)
        if score <= -10_000:
            continue

        candidates.append((score, row))

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: (
            item[0],
            -len(_norm(item[1].masraf)),
        ),
        reverse=True,
    )

    best_score, best = candidates[0]

    # Çok düşük güvenli eşleşmeyi yayımlama.
    min_score = 150
    exact_high_risk = {
        "HGS_ETIKET", "HGS_KART",
        "CEK_DEFTERI_YAPRAK", "CEK_KARNESI_10",
        "CEK_DUZENLEME_STANDART", "CEK_DUZENLEME_DOVIZ",
        "CEK_IADE", "CEK_TAHSIL", "CEK_KARSILIKSIZ",
        "CEK_DUZELTME_HAKKI",
    }
    transfer_high_risk = {
        "SWIFT_GELEN", "SWIFT_GIDEN",
        "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER",
    }
    high_risk = exact_high_risk | transfer_high_risk
    if spec.service in exact_high_risk:
        min_score = 170
    elif spec.service in transfer_high_risk:
        min_score = 185

    if best_score < min_score:
        return None

    best_fee = _row_fee_signature(best)

    # Aynı puanlı farklı ücret: kesin belirsizlik.
    conflicting_same_score = [
        row for score, row in candidates[1:]
        if score == best_score and _row_fee_signature(row) != best_fee
    ]

    # Yüksek riskli hizmetlerde puanı çok yakın farklı ücret de belirsizliktir.
    close_margin = 12 if spec.service in high_risk else 6
    conflicting_close = [
        row for score, row in candidates[1:]
        if best_score - score <= close_margin
        and _row_fee_signature(row) != best_fee
    ]

    if conflicting_same_score or conflicting_close:
        key = (bank, spec.service, spec.label, wanted_channel)
        if key not in _AMBIGUITY_LOGGED:
            _AMBIGUITY_LOGGED.add(key)
            conflict = (conflicting_same_score or conflicting_close)[0]
            print(
                "[comparison][AMBIGUOUS] "
                f"{spec.service} | {spec.label} | {bank} | {wanted_channel}: "
                f"'{best.masraf}' ({best_score}) ile "
                f"'{conflict.masraf}' arasında güvenli seçim yapılamadı. "
                "Yanlış tutar yazmak yerine hücre boşluğu bırakıldı."
            )
        return None

    return best


# ---------------------------------------------------------------------------
# EXCEL OKUMA / GÖSTERİM
# ---------------------------------------------------------------------------

def _find_source_sheet(wb):
    for ws in wb.worksheets:
        if ws.title == COMPARISON_SHEET:
            continue

        for row_idx in range(1, min(ws.max_row, 20) + 1):
            vals = [
                _norm(ws.cell(row=row_idx, column=col).value)
                for col in range(1, min(ws.max_column, 30) + 1)
            ]
            if "banka" in vals and "masraf" in vals:
                return ws, row_idx

    raise RuntimeError("BANKA ve MASRAF kolonlarını içeren ana veri sayfası bulunamadı.")


def _header_map(ws, header_row: int) -> Dict[str, int]:
    result: Dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        key = _norm(ws.cell(row=header_row, column=col).value)
        if key in HEADER_ALIASES:
            result[HEADER_ALIASES[key]] = col

    required = {"BANKA", "KATEGORİ", "MASRAF"}
    missing = required - set(result)
    if missing:
        raise RuntimeError("Eksik kolon: " + ", ".join(sorted(missing)))

    return result


def _read_rows(ws, header_row: int, cols: Mapping[str, int]) -> List[FeeRow]:
    def get(row_idx: int, name: str) -> str:
        col = cols.get(name)
        return _clean(ws.cell(row=row_idx, column=col).value) if col else ""

    result: List[FeeRow] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        bank = get(row_idx, "BANKA")
        masraf = get(row_idx, "MASRAF")
        if not bank or not masraf:
            continue

        if bank not in BANKS:
            continue

        result.append(FeeRow(
            banka=bank,
            kategori=get(row_idx, "KATEGORİ"),
            masraf=masraf,
            asgari_tutar=get(row_idx, "ASGARİ TUTAR"),
            asgari_oran=get(row_idx, "ASGARİ ORAN"),
            azami_tutar=get(row_idx, "AZAMİ TUTAR"),
            azami_oran=get(row_idx, "AZAMİ ORAN"),
            aciklama=get(row_idx, "AÇIKLAMA"),
            guncelleme_tarihi=get(row_idx, "GÜNCELLEME TARİHİ"),
        ))

    return result


def _percent(value: str) -> str:
    value = _clean(value)
    if not value or value in ("-", "0", "0.0", "0.00"):
        return ""

    if "%" in value:
        return value

    try:
        number = float(value.replace(",", "."))
        # API bazı bankalarda 0,7'yi "0.7", bazılarında 0.007 gibi verebilir.
        if 0 < abs(number) < 0.1:
            number *= 100
        return (f"{number:g}%").replace(".", ",")
    except ValueError:
        return value


def _display_amount(value: str) -> str:
    value = _clean(value)
    if not value or value == "-":
        return ""

    # Birimi standartlaştır. Sayısal nokta ondalıksa Türkçe virgüle çevir.
    value = re.sub(r"\s*TL\b", " TRY", value, flags=re.I)
    value = re.sub(r"\s*TRY\b", " TRY", value, flags=re.I)
    value = re.sub(r"\s*(USD|EUR)\b", r" \1", value, flags=re.I)

    m = re.fullmatch(r"\s*([0-9.,]+)\s*(TRY|USD|EUR)?\s*", value, flags=re.I)
    if not m:
        return value

    raw, currency = m.group(1), (m.group(2) or "").upper()

    # 1234.56 -> 1.234,56 ; 1.234,56 zaten Türkçe formattadır.
    number = _parse_number(raw)
    if number is None:
        return value

    if abs(number - round(number)) < 1e-9:
        formatted = f"{int(round(number)):,}".replace(",", ".")
    else:
        formatted = f"{number:,.2f}"
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
        formatted = formatted.rstrip("0").rstrip(",")

    return f"{formatted} {currency}".strip()



def _fee_value_compact(row: Optional[FeeRow]) -> str:
    """Tutar/oranı BSMV notu olmadan tek satırda özetler."""
    if row is None:
        return ""

    min_amount = _display_amount(row.asgari_tutar)
    max_amount = _display_amount(row.azami_tutar)
    min_rate = _percent(row.asgari_oran)
    max_rate = _percent(row.azami_oran)

    if min_amount and max_amount:
        amount = (
            min_amount
            if _norm(min_amount) == _norm(max_amount)
            else f"{min_amount} - {max_amount}"
        )
    else:
        amount = max_amount or min_amount

    if min_rate and max_rate:
        rate = (
            min_rate
            if _norm(min_rate) == _norm(max_rate)
            else f"{min_rate} - {max_rate}"
        )
    else:
        rate = max_rate or min_rate

    if amount and rate:
        return f"{amount} / {rate}"

    return amount or rate


def _extract_deposit_from_description(row: Optional[FeeRow]) -> str:
    """
    Akbank gibi depozitoyu ayrı satır yerine yıllık kira açıklamasında
    yayımlayan bankalar için açıklamadan depozito tutarını çıkarır.
    """
    if row is None:
        return ""

    raw = _clean(row.aciklama)

    patterns = (
        r"depozito\s+bedeli\s*[:\-]?\s*([0-9][0-9.\s]*(?:,[0-9]+)?)\s*(TL|TRY)",
        r"depozito\s+ucreti\s*[:\-]?\s*([0-9][0-9.\s]*(?:,[0-9]+)?)\s*(TL|TRY)",
    )

    norm_raw = _norm(raw)

    # _norm Türkçe karakterleri sadeleştirir fakat rakam biçimini korur.
    for pattern in patterns:
        m = re.search(pattern, norm_raw, flags=re.I)
        if not m:
            continue

        amount = _display_amount(f"{m.group(1).strip()} {m.group(2)}")
        if amount:
            return amount

    return ""


def _deposit_match_score(row: FeeRow, spec: RowSpec) -> int:
    full = _norm(row.text)
    short = _norm(
        f"{row.kategori} | {row.masraf}"
    )

    # Açıklamasında "depozito bedeli ..." geçen yıllık kira satırını
    # ayrı depozito kaydı sanma. Ayrı satır fallback'i yalnız kategori
    # veya MASRAF adında depozito açıkça yazıyorsa çalışır.
    if "depozito" not in short:
        return -10_000

    if "KASA" not in _service_tags(row):
        return -10_000

    score = 100

    if spec.detail == "BUYUK":
        if "buyuk" in full:
            score += 50
        elif "standart kasa depozito" in full:
            score += 15
        elif any(x in full for x in ("orta", "kucuk", "ozel")):
            return -10_000

    elif spec.detail == "ORTA":
        if "orta" in full:
            score += 50
        elif "standart kasa depozito" in full:
            score += 15
        elif any(x in full for x in ("buyuk", "kucuk", "ozel")):
            return -10_000

    elif spec.detail == "KUCUK":
        if "kucuk" in full:
            score += 50
        elif "standart kasa depozito" in full:
            score += 15
        elif any(x in full for x in ("buyuk", "orta", "ozel")):
            return -10_000

    elif spec.detail == "OZEL":
        if "ozel" in full:
            score += 60
        else:
            return -10_000

    # Kasa24 tipleri boy adı olmayan ayrı sınıflar; yanlış boyla
    # eşleştirmemek için standart büyük/orta/küçük satırlarda geriye at.
    if "kasa24" in full and spec.detail in {"BUYUK", "ORTA", "KUCUK"}:
        score -= 40

    return score


def _best_deposit_match(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    annual_row: Optional[FeeRow],
) -> Tuple[str, str]:
    """(depozito_değeri, BSMV_durumu) döndürür."""
    candidates = []
    for row in rows:
        if row.banka != bank:
            continue
        score = _deposit_match_score(row, spec)
        if score <= -10_000:
            continue
        candidates.append((score, row))

    if candidates:
        candidates.sort(key=lambda item: (item[0], -len(_norm(item[1].masraf))), reverse=True)
        dep_row = candidates[0][1]
        value = _fee_value_compact(dep_row)
        return value, _bsmv_label(dep_row)

    # Akbank gibi depozitoyu yıllık satır açıklamasında veren kaynaklar.
    value = _extract_deposit_from_description(annual_row)
    if not value or annual_row is None:
        return "", ""

    desc = _norm(annual_row.aciklama)
    # Depozito cümlesinin yakınında BSMV bilgisi varsa onu kullan; yoksa tahmin etme.
    m = re.search(r"depozito.{0,160}?(bsmv\s+(?:dahil|haric))", desc, flags=re.I)
    if not m:
        m = re.search(r"(bsmv\s+(?:dahil|haric)).{0,160}?depozito", desc, flags=re.I)
    tax = ""
    if m:
        tax = "BSMV dahil" if "dahil" in _norm(m.group(1)) else "BSMV hariç"
    return value, tax


def _bsmv_label(row: Optional[FeeRow]) -> str:
    if row is None:
        return ""
    desc = _norm(row.aciklama)
    if "bsmv dahil" in desc:
        return "BSMV dahil"
    if "bsmv haric" in desc or "bsmv'den haric" in desc:
        return "BSMV hariç"
    if "bsiv dahil" in desc:
        return "BSİV dahil"
    if "bsiv haric" in desc:
        return "BSİV hariç"
    return ""


def _status_meta(row: Optional[FeeRow]) -> Dict[str, str]:
    if row is None:
        return {}
    desc = row.aciklama or ""
    meta: Dict[str, str] = {}
    for key in ("SERVICE", "CHANNEL", "BAND", "DISPLAY_TEXT"):
        m = re.search(rf"(?:^|[;|]\s*){key}\s*=\s*([^;|]+)", desc, flags=re.I)
        if m:
            meta[key] = _clean(m.group(1)).replace("\\n", "\n")
    return meta


def _status_kind(row: Optional[FeeRow]) -> str:
    if row is None:
        return ""
    desc = row.aciklama or ""
    if STATUS_NOT_APPLICABLE in desc:
        return "NOT_APPLICABLE"
    if STATUS_EMPTY in desc:
        return "PUBLISHED_EMPTY"
    if STATUS_AVAILABLE in desc:
        return "AVAILABLE"
    if STATUS_NUMERIC in desc:
        return "OFFICIAL_FEE"
    return ""


def _has_numeric_fee(row: Optional[FeeRow]) -> bool:
    if row is None:
        return False
    return any(_clean(getattr(row, attr, "")) not in ("", "-", "0", "0.0", "0.00")
               for attr in ("asgari_tutar", "asgari_oran", "azami_tutar", "azami_oran"))


def _fee_text(row: Optional[FeeRow], spec: Optional[RowSpec] = None) -> str:
    if row is None:
        return "N/A"

    status = _status_kind(row)
    meta = _status_meta(row)
    display_text = meta.get("DISPLAY_TEXT", "").strip()
    if display_text and status in {"NOT_APPLICABLE", "PUBLISHED_EMPTY", "AVAILABLE"}:
        return display_text

    if status == "NOT_APPLICABLE" and not _has_numeric_fee(row):
        return "Uygulanmıyor"
    if status == "PUBLISHED_EMPTY" and not _has_numeric_fee(row):
        return "Ücret ilan edilmemiş"
    if status == "AVAILABLE" and not _has_numeric_fee(row):
        return "Hizmet var\nAyrı ücret ilan edilmemiş"

    min_amount = _display_amount(row.asgari_tutar)
    max_amount = _display_amount(row.azami_tutar)
    min_rate = _percent(row.asgari_oran)
    max_rate = _percent(row.azami_oran)

    amount = ""
    if min_amount and max_amount:
        amount = min_amount if _norm(min_amount) == _norm(max_amount) else f"min {min_amount}\nmax {max_amount}"
    else:
        amount = max_amount or min_amount

    rate = ""
    if min_rate and max_rate:
        rate = min_rate if _norm(min_rate) == _norm(max_rate) else f"min {min_rate}\nmax {max_rate}"
    else:
        rate = max_rate or min_rate

    if spec and spec.band_key in {"FATURA_1", "SGK_1"}:
        keys = _all_band_keys(f"{row.masraf} | {row.aciklama}")
        upper_key = "FATURA_2" if spec.band_key == "FATURA_1" else "SGK_2"
        if spec.band_key in keys and upper_key in keys and amount:
            result = amount
        elif amount and rate:
            result = f"{amount}\n{rate}"
        else:
            result = amount or rate
    elif spec and spec.band_key in {"FATURA_2", "SGK_2"}:
        keys = _all_band_keys(f"{row.masraf} | {row.aciklama}")
        lower_key = "FATURA_1" if spec.band_key == "FATURA_2" else "SGK_1"
        if spec.band_key in keys and lower_key in keys and rate:
            result = rate
        else:
            result = rate or amount
    else:
        parts = [x for x in (amount, rate) if x]
        result = "\n".join(parts)

    tax = _bsmv_label(row)
    if tax:
        result = f"{result}\n{tax}" if result else tax
    return result or "Ücret bilgisi açıklamada"



def _service_status_row(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
    *,
    kinds: Optional[Set[str]] = None,
) -> Optional[FeeRow]:
    candidates = []
    allowed = kinds or {"AVAILABLE", "PUBLISHED_EMPTY", "NOT_APPLICABLE"}

    for row in rows:
        if row.banka != bank or spec.service not in _service_tags(row):
            continue

        status = _status_kind(row)
        if status not in allowed:
            continue

        meta = _status_meta(row)
        row_band = meta.get("BAND", "").upper()

        # Status belirli bir banda özel ise yalnız o bandı çözsün.
        if row_band and spec.band_key and row_band != spec.band_key:
            continue
        if row_band and not spec.band_key:
            continue

        channels = _channels(row)
        if wanted_channel in channels:
            score = 120
        elif "GENEL" in channels:
            score = 75
        else:
            continue

        if row_band and spec.band_key and row_band == spec.band_key:
            score += 80
        if status == "NOT_APPLICABLE":
            score += 30

        candidates.append((score, row))

    return max(candidates, key=lambda x: x[0])[1] if candidates else None


def _fast_limit_note(
    rows: Sequence[FeeRow], bank: str, wanted_channel: str,
) -> str:
    """FAST limit notunu bankanın resmî supplemental statüsünden üretir."""
    upper_band = RowSpec(
        label="FAST üst limit durumu",
        service="FAST",
        band_key="TRANSFER_3",
    )
    status_row = _service_status_row(
        rows,
        bank,
        upper_band,
        wanted_channel,
        kinds={"NOT_APPLICABLE"},
    )
    if status_row is None:
        return ""

    display_text = _status_meta(status_row).get("DISPLAY_TEXT", "")
    match = re.search(r"FAST\s+limiti\s+([^\n]+)", display_text, flags=re.I)
    if not match:
        return ""
    return f"FAST işlemleri {match.group(1).strip()}'ye kadar yapılmaktadır."



def _generic_institution_fee(
    rows: Sequence[FeeRow], bank: str, wanted_channel: str,
) -> Optional[FeeRow]:
    """Aidat/okul/telefon için yalnız gerçek Fatura/Kurum tahsilat tarifesini seçer."""
    candidates = []

    for row in rows:
        if row.banka != bank or not _has_numeric_fee(row):
            continue

        full = _norm(row.text)
        mas = _norm(row.masraf)
        cat = _norm(row.kategori)

        if not any(x in full for x in (
            "fatura/kurum", "fatura / kurum", "fatura ve anlasmali kurum",
            "fatura odemeleri", "kurum tahsilat", "kurum odeme",
        )):
            continue

        # Faiz / kredi ürünü / SGK / şans / vergi / telefon yükleme gibi
        # komşu fakat farklı ücretleri genel kurum tarifesine sokma.
        if any(x in mas for x in (
            "faiz", "otomatik fatura odeme", "talimatli fatura odeme islem faizi",
            "alisveris faiz", "sgk", "sans oyun", "vergi", "tl/paket yukleme",
            "paket yukleme", "nakit avans", "konsolosluk", "vize randevu",
        )):
            continue

        channels = _channels(row)
        score = 100

        if wanted_channel in channels:
            score += 70
        elif "GENEL" in channels:
            score += 28
        else:
            continue

        # Bankaların ana, karşılaştırılabilir kurum tarifeleri.
        if bank == "YAPIKREDI" and "fatura ve anlasmali kurum odemeleri" in mas:
            score += 220
        if bank == "AKBANK" and "fatura / kurum tahsil" in mas:
            score += 190
        if bank == "GARANTİ" and "fatura/kurum odemesi" in mas:
            score += 190
        if bank == "İŞBANKASI" and mas == "fatura odemeleri":
            score += 190

        # Kart faiz başlıklarını ve ticari paketleri geriye at.
        if "kredi karti faiz" in cat:
            score -= 200
        if "paket" in mas or "ticari" in cat:
            score -= 90

        candidates.append((score, row))

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: (item[0], -len(_norm(item[1].masraf))),
        reverse=True,
    )
    return candidates[0][1]



def _fast_branch_publication_status(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[str]:
    if spec.service != "FAST" or wanted_channel != "SUBE":
        return None
    mobile = _best_match(rows, bank, spec, "MOBIL")
    if mobile is None:
        return None
    # Bankanın kaynak satırlarında bu band için şube FAST kaydı yoksa bunu eşleşme
    # hatası gibi N/A göstermek yerine kaynakta ayrı tarife bulunmadığını belirt.
    for row in rows:
        if row.banka != bank or "FAST" not in _service_tags(row):
            continue
        if spec.band_key and spec.band_key not in _all_band_keys(f"{row.masraf} | {row.aciklama}"):
            continue
        if "SUBE" in _channels(row):
            return None
    return "Şube FAST tarifesi\nyayımlanmıyor"



def _compact_hint(row: FeeRow) -> str:
    """Ücret hücresinde kısa bant/varyant etiketi üretir."""
    raw = _clean(row.masraf)
    parens = re.findall(r"\(([^()]*(?:TL|TRY|gr|USD|EUR)[^()]*)\)", raw, flags=re.I)
    if parens:
        hint = parens[-1].strip()
        if len(hint) <= 48:
            return hint
    parts = [p.strip() for p in re.split(r"\s+-\s+", raw) if p.strip()]
    if parts:
        tail = parts[-1]
        if len(tail) <= 48:
            return tail
    return ""


def _card_transfer_description_fee(row: FeeRow) -> str:
    """
    Numeric kolonları boş olup ücretini açıklamada yayımlayan kart-transfer
    satırlarını güvenli biçimde özetler. Şimdilik bunun gerekli olduğu resmî
    primary örnek Garanti MoneySend (Gönderici) satırıdır.
    """
    full = _norm(row.text)
    if "moneysend" not in full or "gonderici" not in full:
        return ""

    text = _norm(row.aciklama)

    p1 = re.search(
        r"([0-9][0-9.,]*)\s*tl(?:'?[a-z]+)?\s*kadar\s*%\s*([0-9.,]+)\s*\+\s*([0-9][0-9.,]*)\s*tl",
        text,
        flags=re.I,
    )
    p2 = re.search(
        r"([0-9][0-9.,]*)\s*[-–—]\s*([0-9][0-9.,]*)\s*tl\s*arasinda\s*%\s*([0-9.,]+)\s*\+\s*([0-9][0-9.,]*)\s*tl",
        text,
        flags=re.I,
    )
    p3 = re.search(
        r"([0-9][0-9.,]*)\s*tl\s*(?:uzerinde|ve uzeri)\s*%\s*([0-9.,]+)\s*\+\s*([0-9][0-9.,]*)\s*tl",
        text,
        flags=re.I,
    )

    if not (p1 and p2 and p3):
        return ""

    hi1, rate1, fixed1 = p1.groups()
    low2, hi2, rate2, fixed2 = p2.groups()
    low3, rate3, fixed3 = p3.groups()

    def amount(raw: str) -> str:
        return _display_amount(f"{raw} TRY")

    def rate(raw: str) -> str:
        return _percent(raw)

    lines = [
        f"0-{amount(hi1)}: {amount(fixed1)} + {rate(rate1)}",
        f"{amount(low2)}-{amount(hi2)}: {amount(fixed2)} + {rate(rate2)}",
        f"{amount(low3)}+: {amount(fixed3)} + {rate(rate3)}",
    ]

    if "bsmv" in text:
        lines[-1] += " (BSMV eklenir)"

    return "\n".join(lines)


def _transfer_band(row: FeeRow) -> Optional[Band]:
    """Uluslararası transfer satırının işlem tutarı bandını çıkarır."""
    return _parse_band(f"{row.masraf} | {row.aciklama}")


def _transfer_band_label(row: FeeRow) -> str:
    """0–12.000 TRY / 12.000 TRY+ gibi kısa ve okunabilir bant etiketi."""
    band = _transfer_band(row)
    if band is None:
        return ""

    def fmt(value: Optional[float]) -> str:
        if value is None:
            return ""
        return _display_amount(f"{value} TRY")

    low = band.low
    high = band.high

    if low is not None and low <= 1.0 and high is not None:
        return f"0–{fmt(high)}"
    if low is not None and high is not None:
        return f"{fmt(low)}–{fmt(high)}"
    if low is not None and high is None:
        return f"{fmt(low)}+"
    if low is None and high is not None:
        return f"≤{fmt(high)}"
    return ""


def _canonical_transfer_candidate(
    row: FeeRow,
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
) -> bool:
    """
    Yurt dışı transferlerde yalnız gerçekten aynı ürünü karşılaştır.

    Bu kurallar bilinçli olarak banka bazında dar tutulur. Amaç hücreyi
    doldurmak değil, İsme/Kasaya, araştırma/iade, farklı valör, kart/hesap,
    paket/toplu ödeme gibi başka ürünleri yanlışlıkla aynı hücreye sokmamaktır.
    """
    mas = _norm(row.masraf)
    cat = _norm(row.kategori)
    full = _norm(row.text)

    if any(x in mas for x in ("paket", "kobi", "kota")):
        return False
    if "ek kaynak - akbank ticari" in cat:
        return False

    if spec.service == "SWIFT_GELEN":
        # Araştırma/iade ve özel banka/isme-kasaya varyantları standart
        # "hesaba gelen" karşılaştırmasına dahil edilmez.
        if any(x in mas for x in (
            "arastirma", "iade", "isme", "kasaya",
            "garanti bbva international", "gb international",
            "isbank ag",
        )):
            return False

        if bank == "GARANTİ":
            return (
                "diger bankadan gelen doviz havale" in mas
                or "diger bankadan gelen doviz havalesi" in mas
            )

        if bank == "İŞBANKASI":
            return (
                "diger bankadan gelen doviz havalesi - odeme" in mas
                or "diger bankadan gelen doviz havale - odeme" in mas
            )

        if bank == "AKBANK":
            return (
                "gelen doviz havalesi - hesaba" in mas
                or "gelen doviz havale - hesaba" in mas
            )

        if bank == "YAPIKREDI":
            return (
                "gelen doviz havale" in mas
                and "arastirma" not in mas
                and "iade" not in mas
            )

        return False

    if spec.service == "SWIFT_GIDEN":
        if any(x in full for x in (
            "gelen swift",
            "gelen doviz",
            "yurtdisindan",
            "yurt disindan",
            "arastirma/iade",
            "arastirma / iade",
            "nydo toplu",
            "toplu odeme",
        )):
            return False

        if bank == "GARANTİ":
            if "diger bankaya giden doviz havale" not in mas:
                return False
            if wanted_channel == "MOBIL":
                return "mobil/internet" in mas or "mobil / internet" in mas
            if wanted_channel == "SUBE":
                return "sube" in mas
            return False

        if bank == "İŞBANKASI":
            if mas != "hesaptan giden swift":
                return False
            # Kanal İş Bankası'nda kategori başlığında taşınıyor.
            if wanted_channel == "MOBIL":
                return any(x in cat for x in ("internet/iscep", "internet / iscep"))
            if wanted_channel == "SUBE":
                return "sube kanali" in cat
            return False

        if bank == "AKBANK":
            if wanted_channel == "MOBIL":
                # Mobil ve İnternet aynı tarifeyi iki ayrı satırla yayınlıyor.
                # Tek bir kanonik kaynak olarak Mobil satırını kullan.
                return (
                    "akbank mobil baska bankaya doviz transferi" in mas
                    and "internet" not in mas
                )
            if wanted_channel == "SUBE":
                # Kanal yazmayan "Giden - 30.000..." satırı bankanın standart
                # şube/genel tarifesidir. Mobil/İnternet satırlarını alma.
                return (
                    "swift - giden -" in mas
                    and "akbank mobil" not in mas
                    and "akbank internet" not in mas
                )
            return False

        if bank == "YAPIKREDI":
            if "doviz havale gonderimi" not in mas:
                return False
            # Yapı Kredi aynı hizmet için İleri Gün / 1 Gün / Aynı Gün
            # valörlü üç gerçek tarife yayımlıyor. Üçü de aynı hücrede
            # etiketli olarak gösterilir; hiçbiri kaybedilmez.
            valor_ok = any(x in mas for x in (
                "ileri gun valorlu",
                "1 gun valorlu",
                "ayni gun valorlu",
            ))
            if not valor_ok:
                return False
            if wanted_channel == "MOBIL":
                return any(x in mas for x in (
                    "internet / mobil / telefon",
                    "internet/mobil/telefon",
                ))
            if wanted_channel == "SUBE":
                return "subeden" in mas
            return False

        return False

    if spec.service == "YURT_DISI_FAST":
        # Aynı hücreye Visa transferi, Western Union vb. eklenmez.
        if any(x in full for x in (
            "moneysend",
            "western union",
            "karta para gonder",
            "visa ile yurt disi para transferi",
            "visa ile yurtdisi para transferi",
        )):
            return False

        if bank == "GARANTİ":
            return "yurt disi fast" in mas or "yurtdisi fast" in mas

        if bank == "İŞBANKASI":
            # Güncel ücret tablosunda diğer üç bankadaki FAST/Global FAST
            # hesaba gönderimle birebir karşılaştırılabilir ayrı bir ürün yok.
            return False

        if bank == "AKBANK":
            return (
                "akbank fast uluslararasi" in full
                and "turkiye disina para gonderim" in full
            )

        if bank == "YAPIKREDI":
            return (
                "global fast" in full
                and "hesaba para gonderim" in full
                and "karta" not in full
            )

        return False

    if spec.service == "KART_YURTDISI_TRANSFER":
        if "alici" in full or "kktc" in full:
            return False

        if bank == "GARANTİ":
            return "moneysend" in full and "gonderici" in full

        if bank == "İŞBANKASI":
            # Aynı ücret kredi kartı/ön ödemeli/banka kartı varyantlarında
            # tekrarlanabiliyor. Karşılaştırma için debit kart (Bankamatik)
            # + Türkiye Kartları + TL tarifesi kanonik kaynak seçilir.
            return (
                "moneysend" in full
                and "bankamatik kartindan" in mas
                and "turkiye kartlari" in mas
                and " tl" in mas
            )

        if bank == "AKBANK":
            return False

        if bank == "YAPIKREDI":
            return (
                "global fast" in full
                and "karta para gonderim" in full
            )

        return False

    return True


def _canonical_transfer_rows(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
) -> List[Tuple[int, FeeRow]]:
    """Kanonik transfer adaylarını puanlayıp tekrarsız sırala."""
    lookup_channel = wanted_channel if spec.split_channel else "GENEL"
    candidates: List[Tuple[int, FeeRow]] = []

    for row in rows:
        if row.banka != bank:
            continue

        score = _candidate_score(row, spec, lookup_channel)
        special_card_fee = (
            _card_transfer_description_fee(row)
            if spec.service == "KART_YURTDISI_TRANSFER"
            else ""
        )

        if score <= -10_000:
            continue
        if not _has_numeric_fee(row) and not special_card_fee:
            continue
        if not _canonical_transfer_candidate(row, bank, spec, wanted_channel):
            continue

        candidates.append((score, row))

    # Aynı bant + aynı ücret farklı bir satırda tekrar yayınlanmışsa tekilleştir.
    unique: Dict[Tuple[str, Tuple[str, str, str, str]], Tuple[int, FeeRow]] = {}

    for score, row in candidates:
        band_label = _transfer_band_label(row)
        fee_signature = _row_fee_signature(row)
        key = (_norm(band_label), fee_signature)

        existing = unique.get(key)
        if existing is None or score > existing[0]:
            unique[key] = (score, row)

    result = list(unique.values())

    def sort_key(item: Tuple[int, FeeRow]):
        score, row = item
        band = _transfer_band(row)
        low = band.low if band and band.low is not None else -1.0
        high = band.high if band and band.high is not None else float("inf")
        return (low, high, -score, _norm(row.masraf))

    result.sort(key=sort_key)
    return result


def _common_tax_label(rows: Sequence[FeeRow]) -> str:
    labels = [_bsmv_label(row) for row in rows]
    nonempty = [x for x in labels if x]
    if not nonempty:
        return ""
    normalized = {_norm(x) for x in nonempty}
    if len(normalized) == 1 and len(nonempty) == len(rows):
        return nonempty[0]
    return ""


def _fixed_transfer_surcharge(row: FeeRow) -> str:
    """
    Bazı hızlı yurt dışı transfer tarifelerinde yüzde/min-max ücretine ek
    sabit komisyon açıklama alanında yayınlanır (Akbank/Yapı Kredi örneği).
    Bu ek tutarı kaybetmeden hücre özetine taşır.
    """
    text = _norm(row.aciklama)

    patterns = (
        r"sabit komisyon\s*([0-9][0-9.,]*)\s*tl",
        r"islem ucreti\s*([0-9][0-9.,]*)\s*tl\s*ek",
        r"([0-9][0-9.,]*)\s*tl\s*ek olarak tahsil",
    )

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return _display_amount(f"{match.group(1)} TRY")

    return ""


def _compact_transfer_blocks(
    rows: Sequence[FeeRow],
    spec: RowSpec,
) -> str:
    """
    Aynı kanonik SWIFT ürününün gerçek tutar bantlarını kaybetmeden gösterir.

    Önceki sürüm 4+ bandı "x-y (3 kademe)" diye sıkıştırıyordu; bu görünüm
    gerçek eşikleri gizlediği için kaldırıldı. Artık her resmî bant ayrı satırdır.
    Yapı Kredi'de İleri Gün / 1 Gün / Aynı Gün valörleri de ayrı etiketlenir.
    """
    if not rows:
        return ""

    common_tax = _common_tax_label(rows)
    lines: List[str] = []
    seen: Set[str] = set()

    def variant_label(row: FeeRow) -> str:
        text = _norm(row.masraf)
        if "ileri gun valorlu" in text:
            return "İleri Gün"
        if "ayni gun valorlu" in text:
            return "Aynı Gün"
        if "1 gun valorlu" in text:
            return "1 Gün"
        return ""

    for row in rows:
        fee = _fee_value_compact(row)
        if not fee:
            continue

        band_label = _transfer_band_label(row)
        variant = variant_label(row)

        labels = [x for x in (variant, band_label) if x]
        line = f"{' | '.join(labels)}: {fee}" if labels else fee

        # Vergi bilgisi tüm satırlarda aynı değilse satır bazında koru.
        if not common_tax:
            tax = _bsmv_label(row)
            if tax:
                line += f" ({tax})"

        key = _norm(line)
        if key in seen:
            continue
        seen.add(key)
        lines.append(line)

    if common_tax:
        lines.append(common_tax)

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# KULLANICI DENETİMİ SONRASI YÜKSEK RİSKLİ ÖZEL ÇÖZÜCÜLER (V20)
# ---------------------------------------------------------------------------

def _audit_rows(
    rows: Sequence[FeeRow],
    bank: str,
    predicate,
    *,
    numeric_only: bool = True,
) -> List[FeeRow]:
    result: List[FeeRow] = []
    seen: Set[Tuple[str, str, str, str, str]] = set()
    for row in rows:
        if row.banka != bank:
            continue
        if numeric_only and not _has_numeric_fee(row):
            continue
        try:
            ok = predicate(row)
        except Exception:
            ok = False
        if not ok:
            continue
        key = (
            _norm(row.kategori), _norm(row.masraf),
            _norm(row.asgari_tutar), _norm(row.asgari_oran),
            _norm(row.azami_tutar + "|" + row.azami_oran),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def _is_supplemental_row(row: FeeRow) -> bool:
    """Ek kaynağı dataclass özelliği varsaymadan kategori adından belirler."""
    return _norm(row.kategori).startswith("ek kaynak")


def _primary_first(rows: Sequence[FeeRow]) -> List[FeeRow]:
    """Aynı hizmet için ana Ürün/Hizmet tablosunu ek kaynağın önüne alır."""
    return sorted(rows, key=lambda row: (1 if _is_supplemental_row(row) else 0))


def _audit_fee(row: FeeRow, *, prefix: str = "", suffix: str = "") -> str:
    fee = _fee_value_compact(row)
    if not fee:
        return ""
    value = f"{prefix}: {fee}" if prefix else fee
    if suffix:
        value += f" {suffix}"
    tax = _bsmv_label(row)
    if tax:
        value += f"\n{tax}"
    return value


def _audit_common_tax(rows: Sequence[FeeRow]) -> str:
    labels = {_bsmv_label(row) for row in rows if _bsmv_label(row)}
    return next(iter(labels)) if len(labels) == 1 else ""


def _audit_lines(rows: Sequence[FeeRow], labeler) -> Tuple[str, Optional[FeeRow]]:
    lines: List[str] = []
    seen: Set[str] = set()
    first: Optional[FeeRow] = None
    common_tax = _audit_common_tax(rows)

    for row in rows:
        fee = _fee_value_compact(row)
        if not fee:
            continue
        label = labeler(row)
        line = f"{label}: {fee}" if label else fee
        if not common_tax:
            tax = _bsmv_label(row)
            if tax:
                line += f" ({tax})"
        key = _norm(line)
        if key in seen:
            continue
        seen.add(key)
        first = first or row
        lines.append(line)

    if common_tax and lines:
        lines.append(common_tax)
    return "\n".join(lines), first


def _audit_status_text(row: Optional[FeeRow], fallback: str) -> str:
    if row is None:
        return fallback
    meta = _status_meta(row)
    display = meta.get("DISPLAY_TEXT", "").replace("\\n", "\n").strip()
    return display or fallback


def _audit_institution_range(row: FeeRow, *, label: str = "") -> str:
    value = _fee_value_compact(row)
    if not value:
        return ""
    desc = _norm(row.aciklama)
    if any(x in desc for x in ("kurum", "anlasma", "fatura turune gore", "islem bazinda")):
        value += "\n(kuruma/işleme göre)"
    tax = _bsmv_label(row)
    if tax:
        value += f"\n{tax}"
    if label:
        value = f"{label}: {value}"
    return value


def _audit_exact_status(
    rows: Sequence[FeeRow], bank: str, service: str, channel: str,
) -> Optional[FeeRow]:
    spec = RowSpec("audit", service)
    return _service_status_row(rows, bank, spec, channel)




def _audit_fatura_methods(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """
    Fatura/Kurum ödemelerini ödeme yöntemine göre ayırır.

    Kritik kural:
    - "Hesaptan", "kredi kartından" ve "nakit" tarifeleri birbirine taşınmaz.
    - Banka yalnız genel/gişe Fatura-Kurum tarifesi yayımlıyorsa hücre bunu
      açıkça "Genel tarife" diye belirtir; yöntem-spesifik ücretmiş gibi sunmaz.
    - İş Bankası kredi kartı kuralı resmî SSS sayfasındaki doğrulanmış
      supplemental status satırından alınır.
    - Yapı Kredi Mobil/İnternet vadesiz hesaptan fatura ödemesinin ücretsiz
      olduğu resmî fatura sayfasındaki supplemental status ile doğrulanır.
    """
    if spec.service != "FATURA":
        return None

    method = (spec.detail or "").upper()
    if method not in {"HESAPTAN", "KREDI_KARTI", "NAKIT"}:
        return None

    def first(pred) -> Optional[FeeRow]:
        candidates = _audit_rows(rows, bank, pred)
        return candidates[0] if candidates else None

    def amount_range(row: Optional[FeeRow]) -> str:
        if row is None:
            return ""
        lo = _display_amount(row.asgari_tutar)
        hi = _display_amount(row.azami_tutar)
        if lo and hi:
            return lo if _norm(lo) == _norm(hi) else f"{lo} - {hi}"
        return hi or lo

    def with_tax(lines: List[str], row: Optional[FeeRow]) -> str:
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return "\n".join(x for x in lines if x)

    # ------------------------------------------------------------------
    # GARANTİ BBVA
    # ------------------------------------------------------------------
    if bank == "GARANTİ":
        if method == "HESAPTAN":
            if wanted_channel == "SUBE":
                return (
                    "Ayrı hesaptan-Şube tarifesi yayımlanmıyor",
                    None,
                    "PUBLICATION_STATUS",
                )

            row = first(
                lambda r: (
                    "hesaptan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

            lines = [amount_range(row)]
            rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
            if rate:
                lines.append(f"azami {rate}")
            lines.append("Kuruma/işleme göre")
            return (with_tax(lines, row), row, "NUMERIC")

        if method == "KREDI_KARTI":
            if wanted_channel == "SUBE":
                return (
                    "Ayrı kredi kartı-Şube tarifesi yayımlanmıyor",
                    None,
                    "PUBLICATION_STATUS",
                )

            row = first(
                lambda r: (
                    "kredi kartindan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

            fixed = _display_amount(row.asgari_tutar)
            rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
            if not fixed or not rate:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

            return (
                with_tax(
                    [
                        f"149,99 TRY'ye kadar: {fixed}",
                        f"150 TRY ve üzeri: {rate}",
                    ],
                    row,
                ),
                row,
                "NUMERIC",
            )

        # NAKİT
        if wanted_channel == "MOBIL":
            return (
                "Mobil/İnternet kanalında nakit ödeme uygulanmıyor",
                None,
                "NOT_APPLICABLE",
            )

        row = first(
            lambda r: (
                "nakit fatura/kurum odemesi" in _norm(r.masraf)
                and "sube" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

        return (
            with_tax(
                [amount_range(row), "Kuruma/işleme göre"],
                row,
            ),
            row,
            "NUMERIC",
        )

    # ------------------------------------------------------------------
    # İŞ BANKASI
    # ------------------------------------------------------------------
    if bank == "İŞBANKASI":
        if method == "HESAPTAN":
            if wanted_channel == "MOBIL":
                row = first(
                    lambda r: (
                        _norm(r.masraf) == "fatura odemeleri"
                        and (
                            "internet sube" in _norm(r.kategori)
                            or "iscep" in _norm(r.kategori)
                        )
                        and "bankamatik" not in _norm(r.kategori)
                    )
                )
                if row is None:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                return (
                    with_tax(
                        [amount_range(row), "Kuruma/işleme göre"],
                        row,
                    ),
                    row,
                    "NUMERIC",
                )

            row = first(
                lambda r: (
                    _norm(r.masraf) == "fatura odemeleri"
                    and "fatura odemeleri - sube" in _norm(r.kategori)
                    and "internet sube" not in _norm(r.kategori)
                    and "bankamatik" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            return (
                with_tax(
                    [
                        "Genel gişe tarifesi:",
                        amount_range(row),
                        "Hesap/nakit ayrımı yayımlanmıyor",
                    ],
                    row,
                ),
                row,
                "GENERIC_TARIFF",
            )

        if method == "KREDI_KARTI":
            faq_rows = _audit_rows(
                rows,
                bank,
                lambda r: (
                    "service=fatura_kredi_karti" in _norm(r.aciklama)
                    and _status_kind(r) in {"AVAILABLE", "OFFICIAL_FEE"}
                ),
                numeric_only=False,
            )
            if not faq_rows:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            row = faq_rows[0]
            value = _audit_status_text(
                row,
                "Genel kredi kartı tarifesi:\n"
                "0-150 TRY: 5 TRY\n"
                "150 TRY üzeri: %3,50 + BSMV\n"
                "Kanal ayrımı yayımlanmıyor",
            )
            return (value, row, "NUMERIC")

        # NAKİT
        if wanted_channel == "MOBIL":
            return (
                "Mobil/İnternet kanalında nakit ödeme uygulanmıyor",
                None,
                "NOT_APPLICABLE",
            )

        row = first(
            lambda r: (
                _norm(r.masraf) == "fatura odemeleri"
                and "fatura odemeleri - sube" in _norm(r.kategori)
                and "internet sube" not in _norm(r.kategori)
                and "bankamatik" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (
            with_tax(
                [
                    "Genel gişe tarifesi:",
                    amount_range(row),
                    "Nakit/hesap ayrımı yayımlanmıyor",
                ],
                row,
            ),
            row,
            "GENERIC_TARIFF",
        )

    # ------------------------------------------------------------------
    # AKBANK
    # ------------------------------------------------------------------
    if bank == "AKBANK":
        if method == "HESAPTAN":
            if wanted_channel == "MOBIL":
                row = first(
                    lambda r: (
                        "akbank mobil'den fatura / kurum tahsilati" in _norm(r.masraf)
                        and "ek kaynak" not in _norm(r.kategori)
                    )
                )
                label = "Genel Mobil tarifesi:"
            else:
                row = first(
                    lambda r: (
                        "giseden fatura / kurum tahsilati" in _norm(r.masraf)
                        and "ek kaynak" not in _norm(r.kategori)
                    )
                )
                label = "Genel gişe tarifesi:"

            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

            return (
                with_tax(
                    [
                        label,
                        amount_range(row),
                        "Ödeme aracı ayrılmıyor",
                    ],
                    row,
                ),
                row,
                "GENERIC_TARIFF",
            )

        if method == "KREDI_KARTI":
            if wanted_channel == "MOBIL":
                row = first(
                    lambda r: (
                        "kredi kartindan anlik fatura odeme ucreti (akbank internet)" in _norm(r.masraf)
                        and "ek kaynak" not in _norm(r.kategori)
                    )
                )
                if row is None:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                max_fee = _display_amount(row.azami_tutar) or _display_amount(row.asgari_tutar)
                return (
                    with_tax(
                        [
                            "Mobil için ayrı tarife yayımlanmıyor",
                            f"Akbank İnternet: azami {max_fee}",
                            "Firmaya göre",
                        ],
                        row,
                    ),
                    row,
                    "GENERIC_TARIFF",
                )

            row = first(
                lambda r: (
                    "kredi kartindan anlik fatura odeme ucreti (sube)" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            max_fee = _display_amount(row.azami_tutar) or _display_amount(row.asgari_tutar)
            return (
                with_tax(
                    [f"Azami {max_fee}", "Firmaya göre"],
                    row,
                ),
                row,
                "NUMERIC",
            )

        # NAKİT
        if wanted_channel == "MOBIL":
            return (
                "Mobil/İnternet kanalında nakit ödeme uygulanmıyor",
                None,
                "NOT_APPLICABLE",
            )

        row = first(
            lambda r: (
                "giseden fatura / kurum tahsilati" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (
            with_tax(
                [
                    "Genel gişe tarifesi:",
                    amount_range(row),
                    "Nakit/hesap ayrımı yayımlanmıyor",
                ],
                row,
            ),
            row,
            "GENERIC_TARIFF",
        )

    # ------------------------------------------------------------------
    # YAPI KREDİ
    # ------------------------------------------------------------------
    if bank == "YAPIKREDI":
        general = first(
            lambda r: (
                "fatura ve anlasmali kurum odemeleri" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )

        if method == "HESAPTAN":
            if wanted_channel == "MOBIL":
                status_rows = _audit_rows(
                    rows,
                    bank,
                    lambda r: (
                        "service=fatura_hesaptan" in _norm(r.aciklama)
                        and "channel=mobil" in _norm(r.aciklama)
                        and _status_kind(r) in {"AVAILABLE", "OFFICIAL_FEE"}
                    ),
                    numeric_only=False,
                )
                if not status_rows:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                row = status_rows[0]
                return (
                    _audit_status_text(
                        row,
                        "Ücretsiz / 0 TRY\nVadesiz hesaptan",
                    ),
                    row,
                    "NUMERIC",
                )

            if general is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            return (
                with_tax(
                    [
                        "Genel tarife:",
                        amount_range(general),
                        "Kanal/ödeme aracı ayrılmıyor",
                    ],
                    general,
                ),
                general,
                "GENERIC_TARIFF",
            )

        if method == "KREDI_KARTI":
            card_rows = _audit_rows(
                rows,
                bank,
                lambda r: (
                    "anlasmali kurum fatura /sgk prim odemeleri" in _norm(r.masraf)
                    and "sube veya sgk.gov.tr" not in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                ),
            )
            low = next(
                (r for r in card_rows if "0 - 150" in _norm(r.masraf) or "0-150" in _norm(r.masraf)),
                None,
            )
            high = next(
                (r for r in card_rows if "150,01" in _norm(r.masraf) or "150.01" in _norm(r.masraf)),
                None,
            )
            if low is None or high is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

            fixed = _display_amount(low.azami_tutar) or _display_amount(low.asgari_tutar)
            rate = _percent(high.azami_oran) or _percent(high.asgari_oran)
            if not fixed or not rate:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

            return (
                with_tax(
                    [
                        "Genel kredi kartı tarifesi:",
                        f"0-150 TRY: {fixed}",
                        f"150,01 TRY ve üzeri: {rate}",
                        "Kanal ayrımı yayımlanmıyor",
                    ],
                    low,
                ),
                low,
                "GENERIC_TARIFF",
            )

        # NAKİT
        if wanted_channel == "MOBIL":
            return (
                "Mobil/İnternet kanalında nakit ödeme uygulanmıyor",
                None,
                "NOT_APPLICABLE",
            )

        if general is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (
            with_tax(
                [
                    "Genel tarife:",
                    amount_range(general),
                    "Nakit/hesap ayrımı yayımlanmıyor",
                ],
                general,
            ),
            general,
            "GENERIC_TARIFF",
        )

    return None


def _audit_atm(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service not in {
        "LIMIT_UZERI_PARA_CEKME",
        "ORTAK_ATM_PARA_CEKME",
        "BAKIYE_ATM_YURTICI",
        "BAKIYE_ATM_YURTDISI",
    }:
        return None

    # -----------------------------------------------------
    # İŞ BANKASI - ilk sayfadaki gerçek bireysel Türkiye tarifeleri
    # -----------------------------------------------------
    if bank == "İŞBANKASI":
        if spec.service == "BAKIYE_ATM_YURTDISI":
            candidates = _audit_rows(
                rows,
                bank,
                lambda r: (
                    "bankamatik karti - yurt disi atm bakiye sorgulama" in _norm(r.kategori)
                    and "turkiye bireysel kartlari" in _norm(r.masraf)
                    and "kktc" not in _norm(r.text)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

            eur = next((r for r in candidates if "eur" in _norm(r.asgari_tutar)), None)
            usd = next((r for r in candidates if "usd" in _norm(r.asgari_tutar)), None)
            if eur is None or usd is None:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            return (
                f"{_display_amount(eur.asgari_tutar)} / {_display_amount(usd.asgari_tutar)}\n"
                "Türkiye bireysel Bankamatik kartı\nBSMV hariç",
                eur,
                "NUMERIC",
            )

        if spec.service == "LIMIT_UZERI_PARA_CEKME":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    (
                        _norm(r.masraf) == "limit ustu para cekme"
                        or "mevduat hesabindan gunluk" in _norm(r.masraf)
                    )
                    and "turkiye" in _norm(r.text)
                    and "kktc" not in _norm(r.text)
                ),
            )
            # Aynı ücret Bankamatik/Kredi Kartı ana başlıklarında tekrar
            # yayımlanabiliyor. Türkiye bireysel mevduat/Bankamatik satırını
            # öncele.
            if not candidates:
                candidates = _audit_rows(
                    rows, bank,
                    lambda r: (
                        "limit ustu para cekme" in _norm(r.text)
                        and "kktc" not in _norm(r.text)
                        and "23,50" in _clean(r.asgari_tutar)
                    ),
                )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

            candidates.sort(
                key=lambda r: (
                    "bankamatik karti" not in _norm(r.kategori),
                    "mevduat hesaplari" not in _norm(r.kategori),
                    len(_norm(r.masraf)),
                )
            )
            row = candidates[0]
            amount = _display_amount(row.asgari_tutar) or _display_amount(row.azami_tutar)
            rate = _percent(row.asgari_oran) or _percent(row.azami_oran)
            pieces = []
            if amount:
                pieces.append(f"min {amount}")
            if rate:
                pieces.append(rate)
            value = " + ".join(pieces) if pieces else _fee_value_compact(row)
            tax = _bsmv_label(row)
            if tax:
                value += f"\n{tax}"
            return (value, row, "NUMERIC")

        if spec.service == "ORTAK_ATM_PARA_CEKME":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    "bankamatik karti - ortak atm" in _norm(r.kategori)
                    and "turkiye subelerinden verilmis kartlar" in _norm(r.masraf)
                    and "cari hesaptan para cekme" in _norm(r.masraf)
                    and "kibris" not in _norm(r.text)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

            def labeler(r: FeeRow) -> str:
                t = _norm(r.masraf)
                if "tek atm" in t:
                    return "Tek ATM"
                return "Standart ATM"

            candidates.sort(key=lambda r: 1 if "tek atm" in _norm(r.masraf) else 0)
            value, first = _audit_lines(candidates, labeler)
            return (value, first, "NUMERIC")

        if spec.service == "BAKIYE_ATM_YURTICI":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    " - bankamatik karti - " in _norm(r.kategori)
                    and "ortak atm" in _norm(r.kategori)
                    and "bakiye sorgulama" in _norm(r.text)
                    and "yurt ici diger banka atm" in _norm(r.masraf)
                    and "maxipara" not in _norm(r.kategori)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

            def labeler(r: FeeRow) -> str:
                t = _norm(r.masraf)
                if "tek atm" in t:
                    return "Tek ATM"
                return "Standart ATM"

            candidates.sort(key=lambda r: 1 if "tek atm" in _norm(r.masraf) else 0)
            value, first = _audit_lines(candidates, labeler)
            return (value, first, "NUMERIC")

    if spec.service == "BAKIYE_ATM_YURTDISI":
        return (
            "Karşılaştırılabilir ayrı yurtdışı ATM bakiye sorgulama tarifesi yayımlanmıyor",
            None,
            "PUBLICATION_STATUS",
        )

    # -----------------------------------------------------
    # YAPI KREDİ - Ortak ATM oranına ek yayımlanan maktu tutar
    # -----------------------------------------------------
    if bank == "YAPIKREDI" and spec.service == "ORTAK_ATM_PARA_CEKME":
        candidates = _audit_rows(
            rows,
            bank,
            lambda r: (
                "ortak atm" in _norm(r.masraf)
                and "para cekme" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

        row = candidates[0]
        rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
        fixed_match = re.search(
            r"ayrica\s+([0-9][0-9.,]*)\s*tl\s+maktu",
            _norm(row.aciklama),
        )
        fixed = (
            _display_amount(f"{fixed_match.group(1)} TL")
            if fixed_match
            else ""
        )
        if not fixed or not rate:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

        value = f"{fixed} + {rate}"
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")

    # -----------------------------------------------------
    # GARANTİ - yalnız yurtiçi TRY satırları
    # -----------------------------------------------------
    if bank != "GARANTİ":
        return None

    def pred(row: FeeRow) -> bool:
        if spec.service not in _service_tags(row):
            return False
        t = _norm(row.text)
        # Yalnız yurtiçi / TRY tarife. KKTÇ, yurtdışı ve dövizli alternatifleri reddet.
        if any(x in t for x in ("kktc", "yurt disi", "yurtdisi", "usd", "eur", "gbp", "sterlin")):
            return False
        if spec.service == "LIMIT_UZERI_PARA_CEKME":
            return "(tl)" in _norm(row.masraf) or "(try)" in _norm(row.masraf)
        return True

    candidates = _audit_rows(rows, bank, pred)
    if not candidates:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
    row = candidates[0]
    if spec.service == "LIMIT_UZERI_PARA_CEKME":
        amount = _display_amount(row.asgari_tutar) or _display_amount(row.azami_tutar)
        rate = _percent(row.asgari_oran) or _percent(row.azami_oran)
        lines = []
        if amount:
            lines.append(f"min {amount} (BSMV dahil)")
        if rate:
            lines.append(f"{rate} (BSMV hariç)")
        return ("\n".join(lines), row, "NUMERIC")
    return (_fee_text(row, spec), row, "NUMERIC")


def _audit_sans_oyunu(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Kanalsız genel şans oyunu tarifesini iki sütuna kopyalamaz."""
    if spec.service != "SANS_OYUNU":
        return None

    candidates = _primary_first(
        _audit_rows(
            rows,
            bank,
            lambda r: "SANS_OYUNU" in _service_tags(r),
        )
    )
    if not candidates:
        return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

    exact = [row for row in candidates if wanted_channel in _channels(row)]
    if exact:
        row = exact[0]
        return (_fee_text(row, None), row, "NUMERIC")

    general = [row for row in candidates if "GENEL" in _channels(row)]
    if general:
        row = general[0]
        if wanted_channel == "MOBIL":
            return (
                f"Genel / kanal belirtilmemiş:\n{_fee_text(row, None)}",
                row,
                "GENERIC_TARIFF",
            )
        return (
            "Şube için ayrı tarife\nyayımlanmıyor",
            row,
            "PUBLICATION_STATUS",
        )

    return (
        f"{wanted_channel.title()} için ayrı tarife yayımlanmıyor",
        None,
        "PUBLICATION_STATUS",
    )


def _audit_hgs(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """HGS ürün tipini ve kaynakta yayımlanan kanal bilgisini açık gösterir."""
    if spec.service not in {"HGS_ETIKET", "HGS_KART"}:
        return None

    candidates = _primary_first(
        _audit_rows(
            rows,
            bank,
            lambda r: spec.service in _service_tags(r),
        )
    )
    if not candidates:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

    row = candidates[0]
    if bank == "İŞBANKASI":
        channel_label = "Mobil / İnternet / Şube"
    else:
        channel_label = "Genel / kanal belirtilmemiş"
    return (
        f"{channel_label}:\n{_fee_text(row, None)}",
        row,
        "NUMERIC",
    )

def _audit_sgk(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "SGK":
        return None

    band = spec.band_key or ""

    if bank == "GARANTİ":
        candidates = _audit_rows(
            rows, bank,
            lambda r: _norm(r.masraf) == "sgk" and "ek kaynak" not in _norm(r.kategori),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        row = candidates[0]
        if band == "SGK_LOW":
            fee = _display_amount(row.asgari_tutar) or _display_amount(row.azami_tutar)
        else:
            fee = _percent(row.asgari_oran) or _percent(row.azami_oran)
        if not fee:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        value = (
            "Genel SGK tarifesi:\n"
            f"{fee}\n"
            "Kanal ayrımı yayımlanmıyor"
        )
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")

    if bank == "İŞBANKASI":
        if wanted_channel == "SUBE":
            status = _audit_exact_status(rows, bank, "SGK", "SUBE")
            return (
                _audit_status_text(status, "Şube için ayrı SGK kart ödeme tarifesi yayımlanmıyor"),
                status,
                "STATUS",
            )
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                "sgk prim odemesi" in _norm(r.text)
                and ("internet sube" in _norm(r.text) or "iscep" in _norm(r.text))
                and "ek kaynak" not in _norm(r.kategori)
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        row = candidates[0]
        # 0 TL minimum + %3,10 oran; sabit ücret aralığı gibi göstermiyoruz.
        rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
        value = rate or _fee_value_compact(row)
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")

    if bank == "AKBANK":
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                "anlik sgk prim odeme" in _norm(r.text)
                and "sgk.gov.tr" not in _norm(r.text)
                and "talimat" not in _norm(r.text)
                and "ticari" not in _norm(r.kategori)
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        # Kanal ayrımı yayımlanmayan genel Anlık SGK tarifesini kullan.
        # Min tutar + oran birlikte yayımlanan satır varsa onu tercih et.
        candidates.sort(
            key=lambda r: (
                bool(_display_amount(r.asgari_tutar) or _display_amount(r.azami_tutar)),
                bool(_percent(r.asgari_oran) or _percent(r.azami_oran)),
            ),
            reverse=True,
        )
        row = candidates[0]
        fee = _fee_value_compact(row)
        if not fee:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        value = (
            "Genel SGK tarifesi:\n"
            f"{fee}\n"
            "Kanal ayrımı yayımlanmıyor"
        )
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")

    if bank == "YAPIKREDI":
        # Fatura satırındaki 5 TRY / %3,5 SGK'ya taşınamaz. SGK için yayımlanan
        # tek doğru kredi kartı tarifesi Şube veya sgk.gov.tr satırıdır:
        # 0-100 TRY 3 TRY, 100,01 TRY ve üzeri %3, BSMV hariç.
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                "sube veya sgk.gov.tr" in _norm(r.text)
                and "sgk" in _norm(r.text)
                and "ek kaynak" not in _norm(r.kategori)
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

        row = candidates[0]
        fixed = _display_amount(row.asgari_tutar) or "3 TRY"
        rate = _percent(row.asgari_oran) or _percent(row.azami_oran) or "3%"
        channel_label = "Şube" if wanted_channel == "SUBE" else "sgk.gov.tr (referans)"

        if band == "SGK_LOW":
            value = f"{channel_label}:\n{fixed}"
        elif band == "SGK_MID":
            value = (
                f"{channel_label}:\n"
                f"100 TRY: {fixed}\n"
                f"100,01-150 TRY: {rate}"
            )
        else:
            value = f"{channel_label}:\n{rate}"
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")

    return None


def _audit_isbank_special_kasa(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if bank != "İŞBANKASI" or spec.service != "KASA" or spec.detail != "OZEL":
        return None

    annual = _audit_rows(
        rows, bank,
        lambda r: (
            "ozel kasa" in _norm(r.masraf)
            and "depozito" not in _norm(r.text)
            and "kasa(yillik)" in _norm(r.kategori)
        ),
    )
    deposit_rows = _audit_rows(
        rows, bank,
        lambda r: "ozel kasa" in _norm(r.text) and "depozito" in _norm(r.text),
    )
    if not annual:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

    order = {"kucuk": 0, "orta": 1, "buyuk": 2}
    annual.sort(key=lambda r: next((v for k, v in order.items() if k in _norm(r.masraf)), 9))
    deposit = _fee_value_compact(deposit_rows[0]) if deposit_rows else ""

    lines: List[str] = []
    for row in annual:
        t = _norm(row.masraf)
        if "kucuk" in t:
            label = "Küçük Özel"
        elif "orta" in t:
            label = "Orta Özel"
        elif "buyuk" in t:
            label = "Büyük Özel"
        else:
            continue
        line = f"{label}: Yıllık {_fee_value_compact(row)}"
        if deposit:
            line += f" | Depozito {deposit}"
        lines.append(line)

    tax = _audit_common_tax(annual)
    if tax:
        lines.append(tax)
    return ("\n".join(lines), annual[0], "NUMERIC")


def _audit_cheque(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    service = spec.service
    cheque_services = {
        "CEK_DEFTERI_GRUP", "CEK_DUZENLEME_STANDART_GRUP",
        "CEK_DUZENLEME_OZEL_GRUP", "CEK_IADE_GRUP",
        "CEK_TAHSIL_YURTICI_GRUP", "CEK_TAHSIL_DOVIZ_GRUP",
        "CEK_KARSILIKSIZ_GRUP",
    }
    if service not in cheque_services:
        return None

    def safe_cheque(row: FeeRow) -> bool:
        t = _norm(row.text)
        masraf = _norm(row.masraf)
        kategori = _norm(row.kategori)
        if any(x in t for x in (
            "hediye ceki", "armagan ceki", "seyahat ceki", "cek paketi",
            "kobi cek", "kota", "istihbarat", "kayip cek", "sahte cek",
            "odemeyi durdurma", "odeme durdurma",
        )):
            return False
        # Ana sayfadaki satırlarda "çek" masraf adında bulunur. İş Bankası BHS
        # ek kaynağında ise "Bankamız TP/YP - Tahsile Alınan" gibi satır adlarında
        # "çek" yazmayabilir; bu durumda kategori açıkça BHS-Çek olmalıdır.
        return (
            "cek" in masraf
            or "is bankasi bhs - cek" in kategori
            or "cek defteri" in kategori
        )

    # -----------------------------------------------------
    # ÇEK DEFTERİ / KARNESİ
    # -----------------------------------------------------
    if service == "CEK_DEFTERI_GRUP":
        if bank == "GARANTİ":
            candidates = _audit_rows(
                rows, bank,
                lambda r: safe_cheque(r) and "cek defteri teslimi" in _norm(r.masraf) and "yaprak" in _norm(r.masraf),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            row = candidates[0]
            lines = [f"Şube / yaprak: {_fee_value_compact(row)}"]
            desc = _clean(row.aciklama)
            m = re.search(r"dijital[^0-9]{0,120}([0-9][0-9.,]*)\s*TL", desc, flags=re.I | re.S)
            if m:
                lines.append(f"Dijital başvuru / yaprak: {_display_amount(m.group(1) + ' TL')}")
            tax = _bsmv_label(row)
            if tax:
                lines.append(tax)
            if "degerli kagit" in _norm(desc):
                lines.append("Değerli kâğıt bedeli ayrıca")
            return ("\n".join(lines), row, "NUMERIC")

        if bank == "İŞBANKASI":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    safe_cheque(r)
                    and "cek defteri" in _norm(r.text)
                    and any(x in _norm(r.text) for x in (
                        "10 yaprak", "25 yaprak", "50-350", "50 - 350",
                        "50 – 350", "351 yaprak",
                    ))
                    and "k.k.t.c" not in _norm(r.text)
                    and "kktc" not in _norm(r.text)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

            # Ana tarife ile BHS aynı kalemi tekrar edebiliyor. Aynı grup/kademe
            # için primary satır önde tutulur; BHS yalnız eksik kademeyi tamamlar.
            slots: Dict[Tuple[str, str], FeeRow] = {}
            for row in _primary_first(candidates):
                t = _norm(row.text)
                group = "Karekodlu + Logolu" if "logolu" in t else "Karekodlu"
                if "10 yaprak" in t:
                    tier = "10 yaprak"
                elif "25 yaprak" in t:
                    tier = "25 yaprak"
                elif "50" in t and "350" in t:
                    tier = "50–350"
                elif "351" in t:
                    tier = "351+"
                else:
                    continue
                slots.setdefault((group, tier), row)

            lines: List[str] = []
            first: Optional[FeeRow] = None
            tier_order = ("10 yaprak", "25 yaprak", "50–350", "351+")
            used_rows: List[FeeRow] = []
            for group_name in ("Karekodlu", "Karekodlu + Logolu"):
                bits: List[str] = []
                for tier in tier_order:
                    row = slots.get((group_name, tier))
                    if row is None:
                        continue
                    fee = _fee_value_compact(row)
                    if tier in {"50–350", "351+"}:
                        fee += " / yaprak"
                    bits.append(f"{tier}: {fee}")
                    first = first or row
                    used_rows.append(row)
                if not bits:
                    continue
                lines.append(group_name + ":\n" + "\n".join(bits))
            tax = _audit_common_tax(used_rows)
            if tax:
                lines.append(tax)
            lines.append("Değerli kâğıt bedeli ayrıca")
            return ("\n".join(lines), first, "NUMERIC")

        if bank == "AKBANK":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    safe_cheque(r)
                    and (
                        "cek defteri" in _norm(r.masraf)
                        or "cek karnesi" in _norm(r.masraf)
                    )
                    and "paket" not in _norm(r.kategori)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

            candidates = _primary_first(candidates)
            lines: List[str] = []
            first: Optional[FeeRow] = None
            for label, token in (
                ("10 yaprak", "10 yaprak"),
                ("25 yaprak", "25 yaprak"),
                ("Yaprak başı", "yaprak basi"),
            ):
                row = next((r for r in candidates if token in _norm(r.masraf)), None)
                if row is None:
                    continue
                lines.append(f"{label}: {_fee_value_compact(row)}")
                first = first or row

            if not lines or first is None:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            lines.append("Değerli kâğıt bedeli ayrıca")
            tax = _audit_common_tax(candidates)
            if tax:
                lines.append(tax)
            return ("\n".join(lines), first, "NUMERIC")

        if bank == "YAPIKREDI":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    safe_cheque(r)
                    and ("cek karnesi" in _norm(r.text) or "cek defteri" in _norm(r.text))
                    and any(x in _norm(r.text) for x in ("10 yaprak", "10'luk", "25 yaprak", "25'lik", "100", "500"))
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            def labeler(r: FeeRow) -> str:
                t = _norm(r.masraf)
                if "10 yaprak" in t or "10'luk" in t: return "10 yaprak"
                if "25 yaprak" in t or "25'lik" in t: return "25 yaprak"
                if "100" in t: return "100'lük sürekli form"
                if "500" in t: return "500'lük sürekli form"
                return _clean(r.masraf)
            value, first = _audit_lines(candidates, labeler)
            return (value or _source_gap_text(spec, bank, "GENEL"), first, "NUMERIC" if value else "SOURCE_GAP")

    # -----------------------------------------------------
    # STANDART BLOKE / KEŞİDE DÜZENLEME
    # -----------------------------------------------------
    if service == "CEK_DUZENLEME_STANDART_GRUP":
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                safe_cheque(r)
                and any(x in _norm(r.masraf) for x in ("bloke cek duzenleme", "keside ceki duzenleme", "bloke/keside", "bloke / keside"))
                and not any(x in _norm(r.masraf) for x in ("dovizli", "dovizi natik", "dth"))
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
        row = candidates[0]
        return (_fee_text(row, None), row, "NUMERIC")

    # -----------------------------------------------------
    # DÖVİZLİ / ÖZEL NİTELİKLİ DÜZENLEME
    # -----------------------------------------------------
    if service == "CEK_DUZENLEME_OZEL_GRUP":
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                safe_cheque(r)
                and any(x in _norm(r.masraf) for x in ("duzenleme", "duzenlen"))
                and any(x in _norm(r.masraf) for x in ("dovizli", "dovizi natik", "dth", "ozel nitelikli"))
            ),
        )
        if candidates:
            candidates = _primary_first(candidates)
            value, first = _audit_lines(candidates, lambda r: _clean(r.masraf))

            if bank == "AKBANK":
                published = _audit_rows(
                    rows,
                    bank,
                    lambda r: (
                        safe_cheque(r)
                        and "ozel nitelikli cek duzenleme" in _norm(r.masraf)
                    ),
                    numeric_only=False,
                )
                if published and not _has_numeric_fee(published[0]):
                    value += (
                        "\nÖzel Nitelikli Çek Düzenleme: "
                        "Ücret tutarı belirtilmemiş"
                    )
            return (value, first, "NUMERIC")

        # Akbank'ın resmî ticari tablosunda "Özel Nitelikli Çek Düzenleme"
        # kalemi mevcut ancak ücret alanları boş. Bunu kaynak boşluğu gibi
        # göstermeyip yayımlanmış-boş tarife olarak açıklıyoruz.
        if bank == "AKBANK":
            published = _audit_rows(
                rows, bank,
                lambda r: (
                    safe_cheque(r)
                    and "ozel nitelikli cek duzenleme" in _norm(r.masraf)
                ),
                numeric_only=False,
            )
            if published:
                return (
                    "Kalem yayımlanıyor\nAyrı ücret tutarı belirtilmemiş",
                    published[0],
                    "STATUS",
                )

        # Garanti'de hediye/seyahat çeki gibi farklı ürünleri bu alana
        # taşımıyoruz. Doğrudan karşılaştırılabilir ayrı tarife yok.
        if bank == "GARANTİ":
            return (
                "Ayrı dövizli / özel nitelikli çek düzenleme tarifesi yayımlanmıyor",
                None,
                "PUBLICATION_STATUS",
            )

        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

    # -----------------------------------------------------
    # ÇEK İADE
    # -----------------------------------------------------
    if service == "CEK_IADE_GRUP":
        candidates = _audit_rows(
            rows, bank,
            lambda r: safe_cheque(r) and "iade" in _norm(r.masraf),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

        # Aynı resmî tarife primary + supplemental olarak iki kez gelmişse tek göster.
        # Gerçekten farklı iade varyantları ücret imzası farklı olduğu için korunur.
        unique: List[FeeRow] = []
        seen_fee: Set[Tuple[str, str, str, str, str]] = set()
        for row in candidates:
            key = (*_row_fee_signature(row), _bsmv_label(row))
            if key in seen_fee:
                continue
            seen_fee.add(key)
            unique.append(row)

        def iade_label(r: FeeRow) -> str:
            t = _norm(r.masraf)
            if bank == "GARANTİ":
                if "takas gunu" in t:
                    return "Takas günü iade"
                return "Normal / işlemsiz iade"
            if bank == "İŞBANKASI":
                return "Muamelesiz / işlemsiz iade"
            if bank == "AKBANK":
                return "Çek iade"
            if bank == "YAPIKREDI":
                return "İşlemsiz iade"
            return _clean(r.masraf)

        value, first = _audit_lines(unique, iade_label)

        # Bazı bankalar Takas Günü iade ücretini ayrı satır yerine açıklamada yayımlar.
        desc = " ".join(_clean(r.aciklama) for r in candidates)
        m = re.search(
            r"takas\s+gunu.{0,120}?([0-9][0-9.,]*)\s*TL",
            _norm(desc),
            flags=re.I | re.S,
        )
        if m:
            extra = f"Takas günü: {_display_amount(m.group(1) + ' TL')}"
            if _norm(extra) not in _norm(value):
                value += f"\n{extra}"

        return (value, first, "NUMERIC")

    # -----------------------------------------------------
    # ÇEK TAHSİL YURTİÇİ
    # -----------------------------------------------------
    if service == "CEK_TAHSIL_YURTICI_GRUP":
        # Akbank'ta 4.7.3 üst satırı 380,95-4.000 TRY aralığı şeklinde
        # yayımlanıyor; ancak ilk sayfada bunun gerçek alt işlem tarifeleri
        # ayrıca bulunuyor. Üst satırı tek aralık diye göstermek yanıltıcıdır.
        # Bu nedenle doğrudan üç gerçek işlem satırını etiketli gösteriyoruz.
        if bank == "AKBANK":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    _norm(r.kategori) == "cekler ve senetler"
                    and "tahsile alinan cek" in _norm(r.masraf)
                    and "dovizli" not in _norm(r.masraf)
                    and "doviz" not in _norm(r.masraf)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

            order = {
                "hesaba yatan": 0,
                "nakit odenen": 1,
                "farkli subeden": 2,
            }
            def akbank_label(r: FeeRow) -> str:
                t = _norm(r.masraf)
                if "farkli subeden" in t:
                    return "Farklı şubeden nakit ödenen"
                if "hesaba yatan" in t:
                    return "Hesaba yatan"
                if "nakit odenen" in t:
                    return "Nakit ödenen"
                return _clean(r.masraf)

            def akbank_sort(r: FeeRow) -> int:
                t = _norm(r.masraf)
                if "hesaba yatan" in t:
                    return 0
                if "farkli subeden" in t:
                    return 2
                if "nakit odenen" in t:
                    return 1
                return 9

            candidates.sort(key=akbank_sort)
            value, first = _audit_lines(candidates, akbank_label)
            return (value, first, "NUMERIC")

        def yurtiçi_pred(r: FeeRow) -> bool:
            if not safe_cheque(r):
                return False
            m = _norm(r.masraf)
            if bank == "GARANTİ":
                return (
                    "cek tahsil" in m
                    and (
                        "bankamiz ceki" in m
                        or "baska banka ceki (tl-yp)" in m
                        or "baska banka ceki (tl / yp)" in m
                    )
                    and "takasa" not in m
                )
            if bank == "İŞBANKASI":
                return (
                    "tahsile alinan" in m
                    and (
                        "bankamiz tp/yp" in m
                        or "diger banka tp" in m
                    )
                    and "diger banka yp" not in m
                )
            if bank == "YAPIKREDI":
                return (
                    "cek tahsilati" in m
                    and "(tl)" in m
                )
            return False

        candidates = _audit_rows(rows, bank, yurtiçi_pred)
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

        def labeler(r: FeeRow) -> str:
            t = _norm(r.masraf)
            if any(x in t for x in ("bankamiz", "ykb", "ayni banka", "ayni sube")):
                return "Bankanın kendi çeki"
            if any(x in t for x in ("diger banka", "yurtici banka", "farkli sube", "baska banka")):
                return "Diğer banka çeki"
            return "Genel çek tahsil"

        value, first = _audit_lines(candidates, labeler)
        return (value, first, "NUMERIC")

    # -----------------------------------------------------
    # ÇEK TAHSİL DÖVİZLİ / YP
    # -----------------------------------------------------
    if service == "CEK_TAHSIL_DOVIZ_GRUP":
        if bank == "İŞBANKASI":
            # BHS'deki 1.170 TRY satırı, ana Ürün/Hizmet tablosunda daha sonra
            # 1.543 TRY olarak güncellenmiştir. Ana kaynak varken eski ek kaynak
            # hiçbir koşulda karşılaştırmaya giremez.
            candidates = _audit_rows(
                rows,
                bank,
                lambda r: (
                    not _is_supplemental_row(r)
                    and _norm(r.masraf) == "diger banka - tahsile alinan - yp"
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            row = candidates[0]
            return (
                _audit_fee(row, prefix="Diğer Banka YP - Tahsile Alınan"),
                row,
                "NUMERIC",
            )

        if bank == "AKBANK":
            # Güncel ana bireysel tabloda iki ayrı resmî kalem yayımlanıyor.
            # Eski ticari 571,43-12.952,38 TRY / %1 satırı bunların yerine geçmez.
            candidates = _audit_rows(
                rows,
                bank,
                lambda r: (
                    not _is_supplemental_row(r)
                    and (
                        _norm(r.masraf).startswith("baska banka yp cek")
                        or _norm(r.masraf) == "tahsile alinan dovizli cekler"
                    )
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            candidates.sort(
                key=lambda r: 0 if _norm(r.masraf).startswith("baska banka yp cek") else 1
            )
            value, first = _audit_lines(candidates, lambda r: _clean(r.masraf))
            return (value, first, "NUMERIC")

        def doviz_pred(r: FeeRow) -> bool:
            if not safe_cheque(r):
                return False
            m = _norm(r.masraf)
            if bank == "GARANTİ":
                return (
                    "yp cek tahsil" in m
                    or ("baska banka" in m and "yp cek" in m and "takasa" in m)
                )
            if bank == "YAPIKREDI":
                return (
                    ("cek tahsilati" in m and "(yp)" in m)
                    or "tahsile alinan yp cekler" in m
                )
            return False

        candidates = _audit_rows(rows, bank, doviz_pred)
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
        value, first = _audit_lines(candidates, lambda r: _clean(r.masraf))
        return (value, first, "NUMERIC")

    # -----------------------------------------------------
    # KARŞILIKSIZ ÇEK / DÜZELTME
    # -----------------------------------------------------
    if service == "CEK_KARSILIKSIZ_GRUP":
        numeric = _audit_rows(
            rows, bank,
            lambda r: (
                safe_cheque(r)
                and (
                    "karsiliksiz" in _norm(r.masraf)
                    or "duzeltme" in _norm(r.masraf)
                    or "elden odeme" in _norm(r.masraf)
                )
            ),
        )

        # Yapı Kredi'de düzeltme hakkı ile karşılıksız çek elden ödeme
        # birbirinden farklı ücretlerdir. TL/YP aynı ücretle tekrar eden satırları
        # tekilleştirip işlem tiplerini açıkça etiketliyoruz.
        if bank == "YAPIKREDI" and numeric:
            picked: List[Tuple[str, FeeRow]] = []
            seen: Set[Tuple[str, Tuple[str, str, str, str]]] = set()

            for row in numeric:
                t = _norm(row.masraf)
                if "duzeltme hakki" in t:
                    label = "Düzeltme Hakkı"
                    order = 0
                elif "ayni sube" in t and "karsiliksiz" in t and "elden odeme" in t:
                    label = "Aynı Şube Karşılıksız Çek Elden Ödeme"
                    order = 1
                elif "baska sube" in t and "karsiliksiz" in t and "elden odeme" in t:
                    label = "Başka Şube Karşılıksız Çek Elden Ödeme"
                    order = 2
                else:
                    continue

                key = (label, _row_fee_signature(row))
                if key in seen:
                    continue
                seen.add(key)
                picked.append((f"{order}|{label}", row))

            picked.sort(key=lambda item: int(item[0].split("|", 1)[0]))
            lines: List[str] = []
            first: Optional[FeeRow] = None
            common_tax = _audit_common_tax([row for _, row in picked])

            for encoded_label, row in picked:
                label = encoded_label.split("|", 1)[1]
                fee = _fee_value_compact(row)
                if not fee:
                    continue
                line = f"{label}: {fee}"
                if not common_tax:
                    tax = _bsmv_label(row)
                    if tax:
                        line += f" ({tax})"
                lines.append(line)
                first = first or row

            if common_tax and lines:
                lines.append(common_tax)

            if lines:
                return ("\n".join(lines), first, "NUMERIC")

        # İş Bankası BHS: belgelendirme ayrı ücret yayımlamıyor; düzeltme numeric.
        status_candidates = [
            r for r in rows
            if r.banka == bank
            and "karsiliksiz cek belgelendirme" in _norm(r.masraf)
            and _status_kind(r)
        ]
        lines: List[str] = []
        first: Optional[FeeRow] = None
        seen: Set[str] = set()
        for row in numeric:
            t = _norm(row.masraf)
            label = "Karşılıksız / elden ödeme" if "karsiliksiz" in t else "Düzeltme"
            line = f"{label}: {_fee_value_compact(row)}"
            tax = _bsmv_label(row)
            if tax:
                line += f" ({tax})"
            if _norm(line) not in seen:
                seen.add(_norm(line))
                lines.append(line)
                first = first or row
        if status_candidates:
            st = status_candidates[0]
            lines.insert(0, "Belgelendirme: Ayrı ücret yayımlanmıyor")
            first = first or st
        if not lines:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
        return ("\n".join(lines), first, "NUMERIC" if numeric else "STATUS")

    return None


def _audit_kkb(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service not in {"KREDI_RISK", "CEK_RISK"}:
        return None

    if bank == "İŞBANKASI" and spec.service == "KREDI_RISK":
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                "kredi risk raporu" in _norm(r.masraf)
                and any(x in _norm(r.text) for x in ("iscep", "internet", "bankamatik"))
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
        # Aynı 95 TL tarife farklı kanallarda tekrar eder; bir kez göster.
        row = candidates[0]
        return (_fee_text(row, None), row, "NUMERIC")

    if bank == "İŞBANKASI" and spec.service == "CEK_RISK":
        status = _audit_exact_status(rows, bank, "CEK_RISK", "MOBIL") or _audit_exact_status(rows, bank, "CEK_RISK", "GENEL")
        if status is not None:
            return (_audit_status_text(status, "Çek Raporu paket tarifeleri yayımlanıyor"), status, "STATUS")
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

    return None


def _audit_senet_tahsil(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "SENET_TAHSIL_GRUP":
        return None

    candidates = _audit_rows(
        rows, bank,
        lambda r: (
            "senet" in _norm(r.masraf)
            and "tahsil" in _norm(r.masraf)
            and not any(x in _norm(r.masraf) for x in ("iskonto", "istira", "teminat"))
        ),
    )
    if not candidates:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")

    # Yapı Kredi'de iki eş numeric satırı tekilleştir; açıklamadaki muhabir
    # masrafını ayrıca görünür kıl.
    unique_by_fee: List[FeeRow] = []
    seen_sig: Set[Tuple[str, str, str, str]] = set()
    for row in candidates:
        sig = _row_fee_signature(row)
        if sig in seen_sig and bank == "YAPIKREDI":
            continue
        seen_sig.add(sig)
        unique_by_fee.append(row)

    def labeler(r: FeeRow) -> str:
        t = _norm(r.masraf)
        if "muhabir" in t:
            return "Muhabir banka"
        if "farkli sube" in t:
            return "Farklı şube"
        if "ayni sube" in t or "ayni banka" in t:
            return "Aynı banka/şube"
        return "Senet tahsil"

    value, first = _audit_lines(unique_by_fee, labeler)
    if bank == "YAPIKREDI":
        desc = " ".join(_clean(r.aciklama) for r in candidates)
        m = re.search(r"binde\s*4.{0,80}?asgari\s*([0-9][0-9.,]*)\s*TL", desc, flags=re.I | re.S)
        if m:
            value += f"\nMuhabir masrafı: min {_display_amount(m.group(1) + ' TL')} + binde 4"
    return (value, first, "NUMERIC")


def _audit_normal_transfer_row(
    rows: Sequence[FeeRow], bank: str, service: str, band_key: str, channel: str,
) -> Optional[FeeRow]:
    """Normal EFT/Havale tarifesini ürün adına göre kesin seçer."""

    def band_ok(row: FeeRow) -> bool:
        return _band_key(_parse_band(row.masraf)) == band_key

    candidates: List[FeeRow] = []
    for row in rows:
        if row.banka != bank or not _has_numeric_fee(row) or not band_ok(row):
            continue
        mas = _norm(row.masraf)

        if bank == "YAPIKREDI":
            if service == "EFT":
                if "eft gonderimi" not in mas or "gec" in mas or "fast" in mas:
                    continue
                if channel == "MOBIL" and "internet/mobil" not in mas:
                    continue
                if channel == "SUBE" and "sube/diger" not in mas:
                    continue
            elif service == "HAVALE":
                if "havale gonderimi" not in mas:
                    continue
                if channel == "MOBIL" and "internet/mobil" not in mas:
                    continue
                if channel == "SUBE" and "sube/diger" not in mas:
                    continue
            else:
                continue
            candidates.append(row)
            continue

        if bank == "AKBANK" and service == "EFT" and channel == "MOBIL":
            if (
                "akbank mobil'den" in mas
                and "internet'ten" in mas
                and "eft" in mas
                and "gec eft" not in mas
            ):
                candidates.append(row)
            continue

        # Diğer durumlarda genel yüksek-güvenli skoru kullan; geç/paket ürünleri alma.
        if any(x in mas for x in ("gec eft", "paket", "kota")):
            continue
        base_spec = RowSpec("audit", service, band_key)
        score = _candidate_score(row, base_spec, channel)
        if score > -10_000:
            candidates.append(row)

    if not candidates:
        return None

    # Aynı ücret tekrarları varsa ilkini al; farklı ücretli birden fazla kesin aday
    # kalırsa otomatik seçim yapma.
    sigs = {_row_fee_signature(r) for r in candidates}
    if len(sigs) > 1 and bank not in {"YAPIKREDI", "AKBANK"}:
        return None
    return candidates[0]

def _audit_regular_transfer(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service not in {"DUZENLI_EFT", "DUZENLI_HAVALE"}:
        return None

    # Yapı Kredi: Mobil/İnternet düzenli talimatta normal dijital tarife
    # uygulanıyor. Şube için ayrı düzenli tarife yayımlanmadığından normal şube
    # ücretini düzenli ödeme ücreti gibi değil, yalnız açık bir referans olarak
    # gösteriyoruz.
    if bank == "YAPIKREDI":
        base_service = "EFT" if spec.service == "DUZENLI_EFT" else "HAVALE"
        row = _audit_normal_transfer_row(rows, bank, base_service, spec.band_key or "", wanted_channel)
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        fee = _fee_text(row, None)
        if wanted_channel == "MOBIL":
            value = f"Normal Mobil/İnternet tarifesi:\n{fee}"
            return (value, row, "NUMERIC")

        transfer_name = "EFT" if spec.service == "DUZENLI_EFT" else "Havale"
        value = (
            f"Normal Şube tarifesi (referans):\n{fee}\n"
            f"Düzenli {transfer_name} için Şube tarifesi ayrıca yayımlanmıyor"
        )
        return (value, row, "REFERENCE_TARIFF")

    # Akbank şubeden düzenli EFT talimatında kendi açıklamasına göre Mobil/İnternet
    # ücretleri uygulanır. Normal şube EFT tarifesini kullanmak yanlıştı.
    if bank == "AKBANK" and spec.service == "DUZENLI_EFT" and wanted_channel == "SUBE":
        row = _audit_normal_transfer_row(rows, bank, "EFT", spec.band_key or "", "MOBIL")
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (_fee_text(row, None), row, "NUMERIC")

    return None


def _audit_akbank_havale(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Akbank normal Havale'de cep telefonu/ATM transferini ana havale diye seçme."""
    if bank != "AKBANK" or spec.service != "HAVALE" or not spec.band_key:
        return None

    def pred(row: FeeRow) -> bool:
        if _band_key(_parse_band(row.masraf)) != spec.band_key:
            return False
        mas = _norm(row.masraf)
        if wanted_channel == "MOBIL":
            return (
                "akbank mobil'den" in mas
                and "hesaptan hesaba / isme tl ve yp havale" in mas
                and "cep telefonu numarasi" not in mas
            )
        if wanted_channel == "SUBE":
            return (
                "hesaptan hesaba / isme tl ve yp havale-subeden" in mas
                and "ileri vadeli" not in mas
            )
        return False

    candidates = _audit_rows(rows, bank, pred)
    if not candidates:
        return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
    row = candidates[0]
    return (_fee_text(row, None), row, "NUMERIC")


def _audit_backfill_primary_source(
    rows: Sequence[FeeRow], bank: str, supplemental_row: FeeRow,
) -> Optional[FeeRow]:
    """Supplemental satırın açıkça belirttiği primary kaynak masrafını geri bulur."""
    raw = _clean(supplemental_row.aciklama)
    match = re.search(
        r"ana resm[îi] ücret tablosundaki '(.+)' satırından",
        raw,
        flags=re.I,
    )
    if not match:
        return None
    wanted = _norm(match.group(1))
    for row in rows:
        if row.banka != bank:
            continue
        if _norm(row.masraf) == wanted and _has_numeric_fee(row):
            return row
    return None


def _audit_normal_kasa(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Normal Büyük/Orta/Küçük kasa satırında Özel/Süper status'unu seçme."""
    if spec.service != "KASA" or spec.detail not in {"BUYUK", "ORTA", "KUCUK"}:
        return None
    if bank != "AKBANK":
        return None

    token = {"BUYUK": "buyuk boy", "ORTA": "orta boy", "KUCUK": "kucuk boy"}[spec.detail]
    candidates = _audit_rows(
        rows, bank,
        lambda r: (
            token in _norm(r.masraf)
            and "kiralik kasa" in _norm(r.masraf)
            and "ozel" not in _norm(r.masraf)
            and "depozito" not in _norm(r.masraf)
            and "hesaptan odeme" in _norm(r.masraf)
        ),
    )
    if not candidates:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
    row = candidates[0]
    annual = _fee_value_compact(row)
    dep = _extract_deposit_from_description(row)
    value = f"Yıllık: {annual}"
    tax = _bsmv_label(row)
    if tax:
        value += f" ({tax})"
    if dep:
        value += f"\nDepozito: {dep} (BSMV hariç)"
    return (value, row, "NUMERIC")



def _audit_aidat(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "AIDAT":
        return None

    def first(pred) -> Optional[FeeRow]:
        candidates = _audit_rows(rows, bank, pred)
        return candidates[0] if candidates else None

    def range_text(row: Optional[FeeRow]) -> str:
        if row is None:
            return ""
        lo = _display_amount(row.asgari_tutar)
        hi = _display_amount(row.azami_tutar)
        if lo and hi:
            return lo if _norm(lo) == _norm(hi) else f"{lo} – {hi}"
        return hi or lo

    # GARANTİ: Resmî Fatura/Kurum açıklaması aidat/site tahsilatını açıkça kapsıyor.
    if bank == "GARANTİ":
        if wanted_channel == "MOBIL":
            account = first(
                lambda r: (
                    "hesaptan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "aidat" in _norm(r.aciklama)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            card = first(
                lambda r: (
                    "kredi kartindan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if account is None and card is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

            first_row = account or card
            lines: List[str] = []

            if account is not None:
                amount = range_text(account)
                rate = _percent(account.azami_oran) or _percent(account.asgari_oran)
                line = f"Hesaptan: {amount}"
                if rate:
                    line += f" / azami {rate}"
                lines.extend([line, "Kuruma/işleme göre değişir."])

            if card is not None:
                fixed = _display_amount(card.asgari_tutar)
                rate = _percent(card.azami_oran) or _percent(card.asgari_oran)
                if fixed and rate:
                    lines.append(f"Kredi kartından: ≤149,99 TRY: {fixed}")
                    lines.append(f"150 TRY+: {rate}")

            tax = _bsmv_label(first_row)
            if tax:
                lines.append(tax)

            return ("\n".join(lines), first_row, "NUMERIC")

        row = first(
            lambda r: (
                "nakit fatura/kurum odemesi" in _norm(r.masraf)
                and "sube" in _norm(r.masraf)
                and "aidat" in _norm(r.aciklama)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

        lines = [
            f"Nakit/Şube: {range_text(row)}",
            "Kuruma/işleme göre değişir.",
        ]
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), row, "NUMERIC")

    # İŞ BANKASI: hizmet doğrulanıyor, aidata özel sayısal ücret yayımlanmıyor.
    if bank == "İŞBANKASI":
        status = next(
            (
                r for r in rows
                if r.banka == bank
                and "apartman / site aidat odemeleri" in _norm(r.masraf)
            ),
            None,
        )
        if status is None:
            status = (
                _audit_exact_status(rows, bank, "AIDAT", "GENEL")
                or _audit_exact_status(rows, bank, "AIDAT", wanted_channel)
            )
        return (
            "Hizmet var / ayrı tarife yayımlanmıyor.\n"
            "Aidata özel ücret yayımlanmıyor.",
            status,
            "STATUS",
        )

    # AKBANK: Mobil hizmet var. 0,50–55 TRY genel Fatura/Kurum tarifesidir.
    if bank == "AKBANK":
        if wanted_channel == "SUBE":
            return (
                "Bu kanal için ayrı tarife yayımlanmıyor.\n"
                "Aidata özel ücret yayımlanmıyor.",
                None,
                "PUBLICATION_STATUS",
            )

        status = next(
            (
                r for r in rows
                if r.banka == bank
                and "aidat odemeleri - mobil/internet" in _norm(r.masraf)
            ),
            None,
        ) or _audit_exact_status(rows, bank, "AIDAT", "MOBIL")
        generic = first(
            lambda r: (
                "akbank mobil'den fatura / kurum tahsilati" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if generic is None:
            return (
                "Hizmet var / ayrı tarife yayımlanmıyor.\n"
                "Aidata özel ücret yayımlanmıyor.",
                status,
                "STATUS",
            )

        lines = [
            range_text(generic),
            "Genel Fatura/Kurum tarifesinden eşleştirilmiştir.",
            "Aidata özel ücret yayımlanmıyor.",
        ]
        tax = _bsmv_label(generic)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), generic, "GENERIC_TARIFF")

    # YAPI KREDİ: Mobil aidat hizmeti var; 0,63–45 TRY genel kurum tarifesidir.
    if bank == "YAPIKREDI":
        if wanted_channel == "SUBE":
            return (
                "Bu kanal için ayrı tarife yayımlanmıyor.\n"
                "Aidata özel ücret yayımlanmıyor.",
                None,
                "PUBLICATION_STATUS",
            )

        status = next(
            (
                r for r in rows
                if r.banka == bank
                and "aidat odemeleri - mobil/internet" in _norm(r.masraf)
            ),
            None,
        ) or _audit_exact_status(rows, bank, "AIDAT", "MOBIL")
        generic = first(
            lambda r: (
                "fatura ve anlasmali kurum odemeleri" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if generic is None:
            return (
                "Hizmet var / ayrı tarife yayımlanmıyor.\n"
                "Aidata özel ücret yayımlanmıyor.",
                status,
                "STATUS",
            )

        lines = [
            range_text(generic),
            "Genel Fatura/Kurum tarifesinden eşleştirilmiştir.",
            "Aidata özel ücret yayımlanmıyor.",
        ]
        tax = _bsmv_label(generic)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), generic, "GENERIC_TARIFF")

    return None

def _audit_school(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "OZEL_OKUL":
        return None

    allowed = {
        "GARANTİ": {"SUBE"},
        "İŞBANKASI": {"SUBE"},       # Bankamatik de var; karşılaştırmada ATM sütunu yok.
        "AKBANK": {"MOBIL"},
        "YAPIKREDI": {"MOBIL"},
    }

    if wanted_channel not in allowed.get(bank, set()):
        return (
            "Bu kanalda hizmet sunulmuyor.",
            None,
            "NOT_APPLICABLE",
        )

    # Hizmetin ilgili kanalda gerçekten bulunduğunu supplemental resmî kaynakla doğrula.
    def school_status_name_match(r: FeeRow) -> bool:
        mas = _norm(r.masraf)
        if bank == "GARANTİ" and wanted_channel == "SUBE":
            return "ozel okul odemeleri" in mas and "sube" in mas
        if bank == "İŞBANKASI" and wanted_channel == "SUBE":
            return "ozel okul odemeleri" in mas and "sube" in mas
        if bank == "AKBANK" and wanted_channel == "MOBIL":
            return "ozel okul / egitim odemeleri" in mas and "mobil" in mas
        if bank == "YAPIKREDI" and wanted_channel == "MOBIL":
            return "ozel okul / okul taksiti" in mas and "mobil" in mas
        return False

    candidates = _audit_rows(
        rows,
        bank,
        lambda r: (
            (
                "service=ozel_okul" in _norm(r.aciklama)
                and wanted_channel in _channels(r)
                and _status_kind(r) in {"AVAILABLE", "OFFICIAL_FEE"}
            )
            or school_status_name_match(r)
        ),
        numeric_only=False,
    )
    status_row = candidates[0] if candidates else None

    # Garanti Şube: özel okul hizmeti var; 0–50 TRY yalnız genel Nakit
    # Fatura/Kurum tarifesidir ve özel okul komisyonu gibi gösterilemez.
    if bank == "GARANTİ" and wanted_channel == "SUBE":
        if status_row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (
            "Hizmet var.\nÖzel okul ücreti yayımlanmıyor.",
            status_row,
            "STATUS",
        )

    if status_row is None:
        return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

    return (
        "Hizmet var.\nÖzel okul ücreti yayımlanmıyor.",
        status_row,
        "STATUS",
    )

def _audit_phone(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "TELEFON":
        return None

    def first(pred, numeric_only: bool = True) -> Optional[FeeRow]:
        candidates = _audit_rows(rows, bank, pred, numeric_only=numeric_only)
        return candidates[0] if candidates else None

    def range_text(row: Optional[FeeRow]) -> str:
        if row is None:
            return ""
        lo = _display_amount(row.asgari_tutar)
        hi = _display_amount(row.azami_tutar)
        if lo and hi:
            return lo if _norm(lo) == _norm(hi) else f"{lo} – {hi}"
        return hi or lo

    # GARANTİ: telefon/cep telefonu faturaları genel Fatura Ödeme hizmetinin
    # parçasıdır; ilgili kanalın resmî Fatura/Kurum tarifesi gösterilir.
    if bank == "GARANTİ":
        if wanted_channel == "MOBIL":
            row = first(
                lambda r: (
                    "hesaptan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
            lines = [range_text(row)]
            if rate:
                lines[0] += f" / azami {rate}"
            lines.append("Kuruma/işleme göre değişir.")
            tax = _bsmv_label(row)
            if tax:
                lines.append(tax)
            return ("\n".join(lines), row, "NUMERIC")

        row = first(
            lambda r: (
                "nakit fatura/kurum odemesi" in _norm(r.masraf)
                and "sube" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        lines = [f"Nakit/Şube: {range_text(row)}", "Kuruma/işleme göre değişir."]
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), row, "NUMERIC")

    # İŞ BANKASI: telefon faturaları Fatura Ödemeleri kapsamında.
    if bank == "İŞBANKASI":
        if wanted_channel == "MOBIL":
            row = first(
                lambda r: (
                    _norm(r.masraf) == "fatura odemeleri"
                    and "internet sube, iscep, cozum merkezi" in _norm(r.kategori)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
        else:
            row = first(
                lambda r: (
                    _norm(r.masraf) == "fatura odemeleri"
                    and _norm(r.kategori).endswith(" - sube")
                    and "internet sube" not in _norm(r.kategori)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        lines = [range_text(row), "Kuruma/işleme göre değişir."]
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), row, "NUMERIC")

    # AKBANK: Mobil ve Gişe için yayımlanan genel Fatura/Kurum tarifeleri.
    if bank == "AKBANK":
        if wanted_channel == "MOBIL":
            row = first(
                lambda r: (
                    "akbank mobil'den fatura / kurum tahsilati" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
        else:
            row = first(
                lambda r: (
                    "giseden fatura / kurum tahsilati" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        lines = [
            range_text(row),
            "Genel Fatura/Kurum tarifesi uygulanır.",
            "Kuruma/işleme göre değişir.",
        ]
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), row, "GENERIC_TARIFF")

    # YAPI KREDİ:
    # Mobil/İnternet'te vadesiz hesaptan fatura ödemesi ücretsiz.
    # Kredi kartından anlaşmalı kurum fatura tarifesi ayrıca yayımlanıyor.
    if bank == "YAPIKREDI":
        generic = first(
            lambda r: (
                "fatura ve anlasmali kurum odemeleri" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )

        if wanted_channel == "MOBIL":
            free_status = first(
                lambda r: (
                    "hesaptan fatura / kurum odemesi - mobil/internet" in _norm(r.masraf)
                    and "ucretsiz" in _norm(r.aciklama)
                ),
                numeric_only=False,
            )
            card_low = first(
                lambda r: (
                    "anlasmali kurum fatura /sgk prim odemeleri" in _norm(r.masraf)
                    and "0 - 150 tl" in _norm(r.masraf)
                    and "fatura odemeleri" in _norm(r.masraf)
                )
            )
            card_high = first(
                lambda r: (
                    "anlasmali kurum fatura /sgk prim odemeleri" in _norm(r.masraf)
                    and "150,01 tl ve uzeri" in _norm(r.masraf)
                    and "fatura odemeleri" in _norm(r.masraf)
                )
            )

            if free_status is None and card_low is None and card_high is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")

            lines = ["Hesaptan: Ücretsiz / 0 TRY."]
            if card_low is not None:
                low_fee = _display_amount(card_low.azami_tutar) or _display_amount(card_low.asgari_tutar)
                if low_fee:
                    lines.append(f"Kredi kartından: 0–150 TRY: {low_fee}")
            if card_high is not None:
                high_rate = _percent(card_high.azami_oran) or _percent(card_high.asgari_oran)
                if high_rate:
                    lines.append(f"150,01 TRY+: {high_rate}")
            card_tax = _bsmv_label(card_low or card_high)
            if card_tax:
                lines.append(card_tax)
            return ("\n".join(lines), card_low or card_high or free_status, "NUMERIC")

        # Şube için 0,63–45 TRY yalnız genel tarife olarak gösterilir;
        # kaynak kanal/ödeme aracını ayırmıyor.
        if generic is None:
            return (
                "Bu kanal için ayrı tarife yayımlanmıyor.",
                None,
                "PUBLICATION_STATUS",
            )
        lines = [
            range_text(generic),
            "Genel Fatura/Kurum tarifesinden eşleştirilmiştir.",
            "Kanal belirtilmemiş.",
            "Ödeme aracı ayrımı yayımlanmıyor.",
        ]
        tax = _bsmv_label(generic)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), generic, "GENERIC_TARIFF")

    return None

def _audit_vergi(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "VERGI":
        return None

    # "Vergi Tahsilat Komisyonu" diye yeni ürün uydurulmaz. Yalnız açıkça
    # VERGI hizmet statüsü veya doğrudan "Vergi Ödemeleri" numeric satırı kullanılır.
    candidates = _audit_rows(
        rows, bank,
        lambda r: (
            (
                "service=vergi" in _norm(r.aciklama)
                or "vergi odemeleri" in _norm(r.masraf)
            )
            and wanted_channel in _channels(r)
            and "fatura/vergi/sgk" not in _norm(r.masraf)
        ),
    )
    if candidates:
        row = candidates[0]
        return (_audit_institution_range(row), row, "NUMERIC")

    status = _audit_exact_status(rows, bank, "VERGI", wanted_channel)
    if status is not None:
        return (
            _audit_status_text(status, "Hizmet var\nAyrı vergi ücreti yayımlanmıyor"),
            status,
            "STATUS",
        )

    if bank == "GARANTİ" and wanted_channel == "SUBE":
        return (
            "Şube için ayrı vergi ödeme tarifesi yayımlanmıyor",
            None,
            "PUBLICATION_STATUS",
        )

    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")


_DOCUMENT_STATEMENT_SERVICES = {
    "DOC_VADESIZ_TL",
    "DOC_VADESIZ_YP",
    "DOC_KMH_BIREYSEL",
    "DOC_KK_GECMIS",
    "DOC_ARSIV_GENEL",
    "DOC_ESKI_DEKONT_EKSTRE",
    "DOC_AYLIK_POSTA",
    "DOC_ARSIV_SUBE",
    "DOC_ARSIV_BANKA",
    "DOC_KMH_TICARI",
    "DOC_TICARI_KART_GECMIS",
    "DOC_YATIRIM_EKSTRE",
    "DOC_TICARI_AYLIK",
    "DOC_TICARI_KART_BASILI",
    "DOC_TICARI_KART_GONDERIM",
    "DOC_TICARI_HESAP_BASILI",
    "DOC_POS_EKSTRE",
}

    def _audit_document_statement_research(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """
    Belge/ekstre/araştırma ailesini yalnız güvenli ürün adlarıyla çözer.

    Bu bölümde genel ARSIV / HESAP_OZETI etiketleri bilinçli olarak
    kullanılmaz. Bir bankadaki bireysel kart ekstresi başka bankanın
    ticari kart tarifesine veya basılı ekstre ücreti mevduat ekstresine
    taşınamaz.

    Garanti BBVA için önemli not:
    Web sayfasında "TL Hesaplar" ve "YP Hesaplar" birer alt başlıktır;
    MASRAF alanının parçası değildir.

    Örneğin gerçek MASRAF değerleri:
      - KMH Ekstresi
      - Vadesiz TL Hesap Ekstresi- Gerçek Kişiler
      - Vadesiz YP Hesap Ekstresi - Gerçek Kişiler

    Bu nedenle eski sürümde kullanılan:
      "TL Hesaplar - KMH Ekstresi"
    gibi birleşik isimler eşleşmiyordu.
    """

    if spec.service not in _DOCUMENT_STATEMENT_SERVICES:
        return None

    # ------------------------------------------------------------------
    # YARDIMCI FONKSİYONLAR
    # ------------------------------------------------------------------

    def clean_amount(value: str) -> str:
        """
        Yapı Kredi gibi kaynaklardaki:
            45 TL -
            - -
        benzeri gösterimleri temizler.
        """
        parts = [
            part
            for part in _clean(value).split()
            if part != "-"
        ]

        return _display_amount(" ".join(parts)) if parts else ""

    def compact(row: FeeRow) -> str:
        """
        Bir FeeRow içindeki asgari/azami tutarı tek satırlık güvenli
        gösterime çevirir.
        """
        low = clean_amount(row.asgari_tutar)
        high = clean_amount(row.azami_tutar)

        if low and high:
            if _norm(low) == _norm(high):
                return low

            return f"{low} - {high}"

        return high or low

    def exact_rows(
        masraf: str,
        *,
        kategori: str = "",
        numeric_only: bool = True,
    ) -> List[FeeRow]:
        """
        Tam MASRAF adına göre eşleşme.

        Kategori verilmişse kategori içinde de ilgili başlığın geçmesini
        zorunlu tutar.
        """
        wanted_fee = _norm(masraf)
        wanted_category = _norm(kategori)

        return _primary_first(
            _audit_rows(
                rows,
                bank,
                lambda row: (
                    _norm(row.masraf) == wanted_fee
                    and (
                        not wanted_category
                        or wanted_category in _norm(row.kategori)
                    )
                ),
                numeric_only=numeric_only,
            )
        )

    def exact(
        masraf: str,
        *,
        kategori: str = "",
        numeric_only: bool = True,
    ) -> Optional[FeeRow]:
        found_rows = exact_rows(
            masraf,
            kategori=kategori,
            numeric_only=numeric_only,
        )

        return found_rows[0] if found_rows else None

    def safe_candidates(
        predicate,
        *,
        numeric_only: bool = True,
    ) -> List[FeeRow]:
        """
        Exact eşleşmenin kaynak biçimi nedeniyle başarısız olduğu yerlerde
        kontrollü semantic fallback.

        Burada yine:
        - aynı banka,
        - doğru ürün ailesi,
        - gerekiyorsa numeric ücret
        koşulları korunur.
        """
        return _primary_first(
            _audit_rows(
                rows,
                bank,
                predicate,
                numeric_only=numeric_only,
            )
        )

    def first_safe(
        predicate,
        *,
        numeric_only: bool = True,
    ) -> Optional[FeeRow]:
        candidates = safe_candidates(
            predicate,
            numeric_only=numeric_only,
        )

        return candidates[0] if candidates else None

    def published_gap(
        text: str = "Karşılaştırılabilir ayrı tarife yayımlanmıyor",
    ):
        return (
            text,
            None,
            "PUBLICATION_STATUS",
        )

    service = spec.service

    # ==================================================================
    # VADESİZ MEVDUAT EKSTRESİ
    # ==================================================================
    #
    # Garanti'de:
    #
    #     Ekstre Ücreti
    #       TL Hesaplar
    #         KMH Ekstresi
    #         Vadesiz TL Hesap Ekstresi- Gerçek Kişiler
    #
    #       YP Hesaplar
    #         Vadesiz YP Hesap Ekstresi - Gerçek Kişiler
    #
    # "TL Hesaplar" / "YP Hesaplar" MASRAF'ın parçası DEĞİLDİR.
    #
    # Eski hata burada oluşuyordu.
    # ==================================================================

    if service in {
        "DOC_VADESIZ_TL",
        "DOC_VADESIZ_YP",
    }:

        # --------------------------------------------------------------
        # GARANTİ BBVA
        # --------------------------------------------------------------
        if bank == "GARANTİ":

            if service == "DOC_VADESIZ_TL":
                row = exact(
                    "Vadesiz TL Hesap Ekstresi- Gerçek Kişiler",
                    kategori="Ekstre Ücreti",
                )

                # Tire/boşluk veya scraper kategori formatında küçük değişiklik
                # olursa güvenli fallback.
                if row is None:
                    row = first_safe(
                        lambda r: (
                            "vadesiz tl hesap ekstresi"
                            in _norm(r.masraf)
                            and "gercek kisi"
                            in _norm(r.masraf)
                            and "yp hesap"
                            not in _norm(r.masraf)
                            and "ticari"
                            not in _norm(r.text)
                            and "firma kullanim"
                            not in _norm(r.text)
                        )
                    )

            else:
                row = exact(
                    "Vadesiz YP Hesap Ekstresi - Gerçek Kişiler",
                    kategori="Ekstre Ücreti",
                )

                if row is None:
                    row = first_safe(
                        lambda r: (
                            "vadesiz yp hesap ekstresi"
                            in _norm(r.masraf)
                            and "gercek kisi"
                            in _norm(r.masraf)
                            and "ticari"
                            not in _norm(r.text)
                            and "firma kullanim"
                            not in _norm(r.text)
                        )
                    )

            # Garanti kaynağı bu ücreti Mobil/Şube diye ayırmıyor.
            # Karşılaştırma tasarımında ilk hücrede "Genel" açıklaması
            # gösteriliyor; Şube hücresine aynı rakam ikinci kez kopyalanmıyor.
            if wanted_channel == "SUBE":
                return published_gap(
                    "Şube için ayrı tarife yayımlanmıyor"
                )

            if row is None:
                return (
                    _source_gap_text(
                        spec,
                        bank,
                        wanted_channel,
                    ),
                    None,
                    "SOURCE_GAP",
                )

            return (
                "Genel / kanal belirtilmemiş:\n"
                f"{compact(row)}\n"
                "Talep edilen ekstre başına\n"
                "BSMV dahil",
                row,
                "NUMERIC",
            )

        # --------------------------------------------------------------
        # İŞ BANKASI
        # --------------------------------------------------------------
        if bank == "İŞBANKASI":

            if wanted_channel == "MOBIL":
                row = exact(
                    "Ekstre - TP/YP",
                    kategori=(
                        "Mevduat Hesapları - "
                        "Ekstre - Bankamatik"
                    ),
                    numeric_only=False,
                )

                if row is None:
                    return (
                        _source_gap_text(
                            spec,
                            bank,
                            wanted_channel,
                        ),
                        None,
                        "SOURCE_GAP",
                    )

                return (
                    "Bankamatik (Mobil değil):\n"
                    "Ücretsiz / 0 TRY",
                    row,
                    "NUMERIC",
                )

            row = exact(
                "Ekstre - TP/YP",
                kategori=(
                    "Mevduat Hesapları - "
                    "Ekstre - Şube"
                ),
            )

            if row is None:
                return (
                    _source_gap_text(
                        spec,
                        bank,
                        wanted_channel,
                    ),
                    None,
                    "SOURCE_GAP",
                )

            return (
                f"Şube:\n"
                f"{compact(row)}\n"
                "Sayfa başına\n"
                "BSMV hariç",
                row,
                "NUMERIC",
            )

        return published_gap()

    # ==================================================================
    # BİREYSEL KMH HESAP ÖZETİ
    # ==================================================================

    if service == "DOC_KMH_BIREYSEL":

        # --------------------------------------------------------------
        # GARANTİ BBVA
        # --------------------------------------------------------------
        if bank == "GARANTİ":

            # DÜZELTME:
            # Eski:
            #   TL Hesaplar - KMH Ekstresi
            #
            # Gerçek MASRAF:
            #   KMH Ekstresi
            row = exact(
                "KMH Ekstresi",
                kategori="Ekstre Ücreti",
            )

            if row is None:
                row = first_safe(
                    lambda r: (
                        "kmh ekstresi"
                        in _norm(r.masraf)
                        and "ticari"
                        not in _norm(r.text)
                        and "firma kullanim"
                        not in _norm(r.text)
                        and "basili ticari"
                        not in _norm(r.text)
                    )
                )

            if row:
                return (
                    f"{compact(row)}\n"
                    "Kullanılan aylarda aylık\n"
                    "BSMV dahil",
                    row,
                    "NUMERIC",
                )

        # --------------------------------------------------------------
        # YAPI KREDİ
        # --------------------------------------------------------------
        elif bank == "YAPIKREDI":

            row = exact(
                "Ekstre Masrafı - KMH Hesap Özeti "
                "(Esnek Hesap Özeti)"
            )

            if row:
                return (
                    f"{compact(row)} / hesap\n"
                    "Borç devam eden aylarda\n"
                    "BSMV alınmıyor",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # KREDİ KARTI GEÇMİŞ DÖNEM HESAP ÖZETİ
    # ==================================================================

    if service == "DOC_KK_GECMIS":

        # --------------------------------------------------------------
        # GARANTİ BBVA
        # --------------------------------------------------------------
        if bank == "GARANTİ":

            # 1) Önce resmî "Şubeden Hesap Ekstresi" tablosundaki tam isim.
            row = exact(
                "Kredi Kartı Geçmiş Ekstre Ücreti",
                kategori="Şubeden Hesap Ekstresi",
            )

            # 2) Bazı scraper çıktılarında kategori yolu farklı gelebiliyor.
            #    MASRAF tam adı aynıysa yine güvenlidir.
            if row is None:
                row = exact(
                    "Kredi Kartı Geçmiş Ekstre Ücreti",
                )

            # 3) Garanti'nin başka ücret tablosunda aynı hizmet:
            #
            #       Kredi Kartı Geçmiş Ekstre Üc.
            #
            #    olarak kısaltılmış biçimde yayımlanabiliyor.
            if row is None:
                row = exact(
                    "Kredi Kartı Geçmiş Ekstre Üc.",
                )

            # 4) Son güvenli fallback:
            #    yalnız bireysel kredi kartı + geçmiş + ekstre geçen satır.
            if row is None:
                candidates = safe_candidates(
                    lambda r: (
                        "kredi kart"
                        in _norm(r.masraf)
                        and "gecmis"
                        in _norm(r.masraf)
                        and "ekstre"
                        in _norm(r.masraf)
                        and "ticari"
                        not in _norm(r.text)
                        and "firma kullanim"
                        not in _norm(r.text)
                        and "business"
                        not in _norm(r.text)
                    )
                )

                if candidates:
                    # Şubeden Hesap Ekstresi satırını diğer aynı isimli
                    # kayıtlardan öncele.
                    candidates.sort(
                        key=lambda r: (
                            0
                            if (
                                "subeden hesap ekstresi"
                                in _norm(r.kategori)
                            )
                            else 1,
                            0
                            if (
                                _norm(r.masraf)
                                == _norm(
                                    "Kredi Kartı "
                                    "Geçmiş Ekstre Ücreti"
                                )
                            )
                            else 1,
                            0
                            if "bsmv dahil"
                            in _norm(r.aciklama)
                            else 1,
                            1
                            if _is_supplemental_row(r)
                            else 0,
                        )
                    )

                    row = candidates[0]

            if row:
                return (
                    "1 yıldan eski: "
                    f"azami {compact(row)}\n"
                    "BSMV dahil",
                    row,
                    "NUMERIC",
                )

        # --------------------------------------------------------------
        # İŞ BANKASI
        # --------------------------------------------------------------
        elif bank == "İŞBANKASI":

            row = exact(
                "Geçmiş Dönem Kredi Kartı "
                "Hesap Özeti Ücreti-Türkiye Kredi Kartları"
            )

            if row:
                return (
                    "1 yıldan eski: "
                    f"{compact(row)} / hesap özeti\n"
                    "BSMV hariç",
                    row,
                    "NUMERIC",
                )

        # --------------------------------------------------------------
        # YAPI KREDİ
        # --------------------------------------------------------------
        elif bank == "YAPIKREDI":

            row = exact(
                "Diğer Ücretler - Geçmiş Dönem Ekstre "
                "Arşiv Araştırma Ücreti "
                "(kart hamilinin talebinin olması halinde)"
            )

            if row:
                return (
                    f"azami {compact(row)}\n"
                    "BSMV dahil",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # ARŞİV ARAŞTIRMA – GENEL BELGE
    # ==================================================================

    if service == "DOC_ARSIV_GENEL":

        # --------------------------------------------------------------
        # GARANTİ BBVA
        # --------------------------------------------------------------
        if bank == "GARANTİ":

            row = exact(
                "Belge ve Bilgilendirme Ücreti - "
                "Arşiv Araştırma Ücreti",
                kategori="Belge ve Bilgilendirme",
            )

            if row:
                amounts = [
                    _display_amount(
                        f"{amount} TRY"
                    )
                    for amount in re.findall(
                        r"([0-9]+(?:[.,][0-9]+)?)\s*tl",
                        _norm(row.aciklama),
                        flags=re.I,
                    )
                ]

                if len(amounts) < 2:
                    return (
                        _source_gap_text(
                            spec,
                            bank,
                            wanted_channel,
                        ),
                        None,
                        "SOURCE_GAP",
                    )

                older = amounts[0]
                oldest = amounts[1]

                return (
                    "Son 1 yıl: Ücretsiz\n"
                    f"1–3 yıl: {older}\n"
                    f"3 yıldan eski: {oldest}\n"
                    "BSMV hariç",
                    row,
                    "NUMERIC",
                )

        # --------------------------------------------------------------
        # İŞ BANKASI
        # --------------------------------------------------------------
        elif bank == "İŞBANKASI":

            document = exact(
                "Sözleşme, dekont vb. doküman talebi",
                kategori=(
                    "Mevduat Hesapları - "
                    "Özel Rapor Ücretleri"
                ),
            )

            notice = exact(
                "Geçmiş Dönem Bankacılık "
                "İşlemleri Bildirimi"
            )

            if document and notice:
                return (
                    "Belge talebi (>1 yıl): "
                    f"{compact(document)}\n"
                    "İşlem bildirimi (>1 yıl): "
                    f"{compact(notice)}\n"
                    "BSMV hariç",
                    document,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # 1 YILDAN ESKİ DEKONT / EKSTRE
    # ==================================================================

    if service == "DOC_ESKI_DEKONT_EKSTRE":

        # --------------------------------------------------------------
        # AKBANK
        # --------------------------------------------------------------
        if bank == "AKBANK":

            row = exact(
                "1 Yıldan Eski İşlemlere Ait "
                "Geçmişe Yönelik Dekont-Ekstre Verilmesi"
            )

            if row:
                per_page = re.search(
                    r"sayfa basina\s*"
                    r"([0-9]+(?:[.,][0-9]+)?)\s*tl",
                    _norm(row.aciklama),
                    flags=re.I,
                )

                page_text = (
                    _display_amount(
                        f"{per_page.group(1)} TRY"
                    )
                    if per_page
                    else ""
                )

                value = f">1 yıl: {compact(row)}"

                if page_text:
                    value += (
                        f"\n{page_text} / sayfa"
                    )

                value += "\nBSMV hariç"

                return (
                    value,
                    row,
                    "NUMERIC",
                )

        # --------------------------------------------------------------
        # YAPI KREDİ
        # --------------------------------------------------------------
        elif bank == "YAPIKREDI":

            receipt = exact(
                "Dekont Masrafı - "
                "Geçmişe Yönelik Olarak Basılan"
            )

            statements = exact_rows(
                "Ekstre Masrafı"
            )

            if receipt and len(statements) >= 2:

                statement_fees: List[str] = []

                for statement_row in statements:
                    fee = compact(statement_row)

                    if (
                        fee
                        and fee
                        not in statement_fees
                    ):
                        statement_fees.append(fee)

                value = (
                    "Dekont (>1 yıl): "
                    f"{compact(receipt)}"
                )

                if statement_fees:
                    value += (
                        "\nEkstre (>1 yıl), "
                        "yayımlanan tarifeler: "
                        + " | ".join(statement_fees)
                    )

                value += (
                    "\nSayfa başına; "
                    "en fazla 10 sayfa\n"
                    "BSMV hariç"
                )

                return (
                    value,
                    receipt,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # AYLIK HESAP ÖZETİ – POSTA
    # ==================================================================

    if service == "DOC_AYLIK_POSTA":

        if bank == "AKBANK":

            row = exact(
                "Posta İle Aylık Hesap Özeti Gönderimi"
            )

            if row:
                return (
                    f"{compact(row)}\n"
                    "BSMV muaf",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # ARŞİV / ARAŞTIRMA – ŞUBE / BANKA BAZINDA
    # ==================================================================

    if service in {
        "DOC_ARSIV_SUBE",
        "DOC_ARSIV_BANKA",
    }:

        if bank != "İŞBANKASI":
            return published_gap()

        masraf = (
            "Merkezden / Şubeden "
            "Şube Bazında Yapılan - TP/YP"
            if service == "DOC_ARSIV_SUBE"
            else
            "Merkezden Banka Bazında Yapılan - TP/YP"
        )

        row = exact(masraf)

        if row:
            label = (
                "Şube bazında"
                if service == "DOC_ARSIV_SUBE"
                else "Banka bazında"
            )

            return (
                f"{label}: "
                f"{compact(row)}\n"
                "BSMV hariç",
                row,
                "NUMERIC",
            )

        return (
            _source_gap_text(
                spec,
                bank,
                wanted_channel,
            ),
            None,
            "SOURCE_GAP",
        )

    # ==================================================================
    # TİCARİ KMH HESAP ÖZETİ
    # ==================================================================

    if service == "DOC_KMH_TICARI":

        # --------------------------------------------------------------
        # GARANTİ BBVA
        # --------------------------------------------------------------
        if bank == "GARANTİ":

            row = exact(
                "Avans Hesap "
                "(Kredili Mevduat Hesabı) - "
                "Belge ve Bilgilendirme Ücreti "
                "(Ekstre Ücreti)",
                kategori="Firma Kullanım Amaçlı",
                numeric_only=False,
            )

            if row:
                return (
                    "Gönderim maliyeti kadar\n"
                    "Tutar değişken; "
                    "sabit ücret yayımlanmıyor",
                    row,
                    "STATUS",
                )

        # --------------------------------------------------------------
        # YAPI KREDİ
        # --------------------------------------------------------------
        elif bank == "YAPIKREDI":

            row = exact(
                "Esnek Ticari Hesap - Faiz Oranı - "
                "Basılı Ticari KMH Hesap Özeti "
                "(Esnek Ticari Hesap Özeti Ücreti)"
            )

            if row:
                return (
                    f"{compact(row)} / sayfa",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # TİCARİ KART GEÇMİŞ DÖNEM EKSTRESİ
    # ==================================================================

    if service == "DOC_TICARI_KART_GECMIS":

        if bank == "YAPIKREDI":

            row = exact(
                "Diğer Ücretler - "
                "Ticari Kart Geçmiş Dönem Ekstre Ücreti"
            )

            if row:
                return (
                    f"{compact(row)} / sayfa\n"
                    "BSMV hariç",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # YATIRIM HESABI EKSTRESİ
    # ==================================================================

    if service == "DOC_YATIRIM_EKSTRE":

        if bank == "İŞBANKASI":

            row = exact(
                "Ekstre - TP",
                kategori=(
                    "Mevduat Hesapları - "
                    "Ekstre - Yatırım Hesabı Ekstre"
                ),
            )

            if row:
                return (
                    f"{compact(row)} / gönderi\n"
                    "BSMV hariç",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # TİCARİ MÜŞTERİ AYLIK HESAP ÖZETİ
    # ==================================================================

    if service == "DOC_TICARI_AYLIK":

        if bank == "AKBANK":

            row = exact(
                "Aylık Hesap Özeti (Ticari Müşteri)"
            )

            if row:
                return (
                    f"asgari {compact(row)} / hesap\n"
                    "BSMV dahil",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # TİCARİ KART BASILI EKSTRE
    # ==================================================================

    if service == "DOC_TICARI_KART_BASILI":

        if bank == "GARANTİ":

            row = exact(
                "Basılı Ekstre Ücreti",
                kategori="Firma Kullanım Amaçlı",
            )

            if row:
                return (
                    f"azami {compact(row)} / sayfa\n"
                    "BSMV hariç",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # TİCARİ KART EKSTRE GÖNDERİM
    # ==================================================================

    if service == "DOC_TICARI_KART_GONDERIM":

        if bank == "GARANTİ":

            row = exact(
                "Diğer Ücretler - Ekstre Gönderim Ücreti",
                kategori="Firma Kullanım Amaçlı",
            )

            if row:
                return (
                    f"azami {compact(row)} / sayfa\n"
                    "BSMV hariç",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # TİCARİ HESAP BASILI EKSTRE
    # ==================================================================

    if service == "DOC_TICARI_HESAP_BASILI":

        if bank == "AKBANK":

            row = exact(
                "4.6.1 Basılı Ekstre Gönderimi (2)",
                kategori="EK KAYNAK - Akbank Ticari",
            )

            if row:
                return (
                    f"{compact(row)} / sayfa\n"
                    "Posta/kurye maliyeti ilave\n"
                    "BSMV hariç",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ==================================================================
    # ÜYE İŞYERİ / POS – EKSTRE GÖNDERİMİ
    # ==================================================================

    if service == "DOC_POS_EKSTRE":

        if bank == "GARANTİ":

            row = exact(
                "Ekstre Gönderim Ücreti",
                kategori="Üye İşyeri Ücretleri",
            )

            if row:
                amounts = [
                    _display_amount(
                        f"{amount} TRY"
                    )
                    for amount in re.findall(
                        r"([0-9]+(?:[.,][0-9]+)?)\s*tl",
                        _norm(row.aciklama),
                        flags=re.I,
                    )
                ]

                if len(amounts) < 2:
                    return (
                        _source_gap_text(
                            spec,
                            bank,
                            wanted_channel,
                        ),
                        None,
                        "SOURCE_GAP",
                    )

                first = amounts[0]
                second = amounts[1]

                value = (
                    f"0–20 g: {first}"
                    f"\n20–50 g: {second}"
                    "\nBSMV dahil"
                )

                return (
                    value,
                    row,
                    "NUMERIC",
                )

        return published_gap()

    return published_gap()

 
"""
komisyonlar_guncel.xlsx içindeki banka ücretlerini ortak bir masraf sözlüğüne
çevirip KARŞILAŞTIRMA sayfasını otomatik üretir.
 
Temel fark:
- Bankaların MASRAF adlarını birebir eşleştirmez.
- Önce hizmeti (EFT, Havale, FAST, Fatura vb.), kanalı (Mobil/Şube/ATM)
  ve tutar aralığını semantik kurallarla çıkarır.
- Aynı hizmet + aynı kanal + aynı tutar aralığı olan farklı banka satırlarını
  tek bir ORTAK MASRAF ADI altında toplar.
 
Örnek:
  Garanti : "Elektronik Fon Transferi (EFT) Ücreti - EFT Garanti BBVA
             Mobil / İnternet (0-8300 TL)"
  Akbank  : "FAST / Akbank Mobil'den, İnternet'ten ... EFT ...
             (8.300 TL ve Altı)"
  YKB     : "EFT Gönderimi - İnternet/Mobil – 0 - 8.300 TL"
  İşbank  : "EFT Gönderilmesi - 0-8.300 TL"
 
hepsi şu anahtara bağlanır:
  EFT Gönderimi | Mobil | 0 TRY - 8.300 TRY
 
Ana Excel her gün güncellendiğinde KARŞILAŞTIRMA sayfası da aynı veriyle
tekrar üretilir. NOTLAR sütunundaki manuel notlar ortak satır anahtarına göre
korunur.
"""
 
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Set, Tuple
 
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
 
 
COMPARISON_VERSION = "2026-08-26-v30-fast-source-limit-note"
COMPARISON_SHEET = "KARŞILAŞTIRMA"
PREVIEW_LAYOUT_SIGNATURE = "4BANKS|A:I_FILTERED|J:L_EMPTY|M_GLOSSARY_LIST|FAIL_CLOSED_V30|PRIMARY_FIRST|CHANNEL_LABELS|CHEQUE_CURRENT|FAST_SOURCE_LIMITS|RESEARCH_SECURITIES"
 
STATUS_AVAILABLE = "[SUPPLEMENTAL][AVAILABLE_NO_SEPARATE_FEE]"
STATUS_EMPTY = "[SUPPLEMENTAL][PUBLISHED_EMPTY]"
STATUS_NUMERIC = "[SUPPLEMENTAL][OFFICIAL_FEE]"
STATUS_NOT_APPLICABLE = "[SUPPLEMENTAL][NOT_APPLICABLE]"
 
 
# Eski denemelerde oluşmuş karşılaştırma sayfaları kullanıcıyı yanıltmasın.
# v9 çalıştığında KARŞILAŞTIRMA ile başlayan eski sayfaların tamamı silinir
# ve sadece preview ile aynı kompakt sayfa yeniden oluşturulur.
 
# Kullanıcının karşılaştırma şablonundaki banka seti.
BANKS = [
    "GARANTİ",
    "İŞBANKASI",
    "AKBANK",
    "YAPIKREDI",
]
 
DISPLAY_BANKS = {
    "GARANTİ": "Garanti BBVA",
    "İŞBANKASI": "İş Bankası",
    "AKBANK": "Akbank",
    "YAPIKREDI": "Yapı ve Kredi Bankası",
}
 
BANK_COLORS = {
    "GARANTİ": "70AD47",
    "İŞBANKASI": "2E75B6",
    "AKBANK": "E31E24",
    "YAPIKREDI": "17365D",
}
 
 
PRIMARY_SOURCE_URLS = {
    "GARANTİ": "https://www.garantibbva.com.tr/urun-ve-hizmet-ucretleri",
    "İŞBANKASI": "https://www.isbank.com.tr/urun-ve-hizmet-ucretleri",
    "AKBANK": "https://www.akbank.com/urun-ve-hizmet-ucretleri",
    "YAPIKREDI": "https://www.yapikredi.com.tr/bireysel-bankacilik/hesaplama-araclari/bireysel-urun-ve-hizmet-ucretleri",
}
 
HEADER_ALIASES = {
    "banka": "BANKA",
    "kategori": "KATEGORİ",
    "masraf": "MASRAF",
    "asgari tutar": "ASGARİ TUTAR",
    "asgari oran": "ASGARİ ORAN",
    "azami tutar": "AZAMİ TUTAR",
    "azami oran": "AZAMİ ORAN",
    "aciklama": "AÇIKLAMA",
    "site guncelleme tarihi": "GÜNCELLEME TARİHİ",
    "komisyon guncelleme tarihi": "GÜNCELLEME TARİHİ",
    "son kontrol": "SON KONTROL",
    "calistirilma tarihi": "SON KONTROL",
}
 
 
@dataclass(frozen=True)
class FeeRow:
    banka: str
    kategori: str
    masraf: str
    asgari_tutar: str = ""
    asgari_oran: str = ""
    azami_tutar: str = ""
    azami_oran: str = ""
    aciklama: str = ""
    guncelleme_tarihi: str = ""
 
    @property
    def text(self) -> str:
        return " | ".join(
            x for x in (self.kategori, self.masraf, self.aciklama) if x
        )
 
 
@dataclass(frozen=True)
class Band:
    low: Optional[float]
    high: Optional[float]
 
 
@dataclass(frozen=True)
class RowSpec:
    """Karşılaştırmada görünen ORTAK satır tanımı."""
 
    label: str
    service: str
    band_key: Optional[str] = None
    detail: Optional[str] = None
    split_channel: bool = True
 
 
# ---------------------------------------------------------------------------
# ORTAK MASRAF SÖZLÜĞÜ
# ---------------------------------------------------------------------------
# Buradaki label'lar artık banka sitesindeki ham MASRAF adı değildir.
# Tüm bankalar bu ortak isimlere eşleştirilir.
#
# SECTION satırı: ("SECTION", "başlık")
# ROW satırı    : ("ROW", RowSpec(...))
# ---------------------------------------------------------------------------
 
LAYOUT = [
    # Kullanıcı denetimi sonrası güvenlik öncelikli karşılaştırma sözlüğü.
    # Temel ilke: KOMİSYONLAR sayfasında var olan doğru satırı kullan;
    # belirsiz/başka ürün olan veriyi zorla eşleştirme.
 
    ("SECTION", "EFT"),
    ("ROW", RowSpec("0 TRY - 8.300 TRY", "EFT", "TRANSFER_1")),
    ("ROW", RowSpec("8.300,01 TRY - 399.000 TRY", "EFT", "TRANSFER_2")),
    ("ROW", RowSpec("399.000,01 TRY -", "EFT", "TRANSFER_3")),
 
    ("SECTION", "ŞANS OYUNU"),
    ("ROW", RowSpec("Şans Oyunu Ödemeleri", "SANS_OYUNU")),
 
    ("SECTION", "PARA ÇEKME / ATM"),
    ("ROW", RowSpec("Günlük Limit Üzeri Para Çekme", "LIMIT_UZERI_PARA_CEKME", split_channel=False)),
    ("ROW", RowSpec("Ortak ATM Para Çekme", "ORTAK_ATM_PARA_CEKME", split_channel=False)),
    ("ROW", RowSpec("Ortak ATM Bakiye Sorgulama", "BAKIYE_ATM_YURTICI", split_channel=False)),
    ("ROW", RowSpec("Ortak ATM Bakiye Sorgulama – Yurtdışı", "BAKIYE_ATM_YURTDISI", split_channel=False)),
 
    ("SECTION", "FATURA / KURUM ÖDEMELERİ"),
    ("ROW", RowSpec("Hesaptan Fatura / Kurum Ödemesi", "FATURA", detail="HESAPTAN")),
    ("ROW", RowSpec("Kredi Kartından Fatura / Kurum Ödemesi", "FATURA", detail="KREDI_KARTI")),
    ("ROW", RowSpec("Nakit Fatura / Kurum Ödemesi", "FATURA", detail="NAKIT")),
 
    ("SECTION", "HAVALE"),
    ("ROW", RowSpec("0 TRY - 8.300 TRY", "HAVALE", "TRANSFER_1")),
    ("ROW", RowSpec("8.300,01 TRY - 399.000 TRY", "HAVALE", "TRANSFER_2")),
    ("ROW", RowSpec("399.000,01 TRY -", "HAVALE", "TRANSFER_3")),
 
    ("SECTION", "FAST"),
    ("ROW", RowSpec("0 TRY - 8.300 TRY", "FAST", "TRANSFER_1")),
    ("ROW", RowSpec("8.300,01 TRY - 399.000 TRY", "FAST", "TRANSFER_2")),
 
    # İlk sayfada ürün adı olarak bulunmayan sentetik Visa Direct / Global Fast
    # karşılaştırma satırları şimdilik kaldırıldı. Yalnız gerçek SWIFT ailesi kalır.
    ("SECTION", "YURT DIŞI TRANSFERLER"),
    ("ROW", RowSpec("SWIFT - Gelen", "SWIFT_GELEN", split_channel=False)),
    ("ROW", RowSpec("SWIFT - Giden", "SWIFT_GIDEN")),
 
    # Bankaların eşikleri aynı olmadığı için 100 ve 150 TRY kırılımları korunur.
    ("SECTION", "SGK TAHSİLAT"),
    ("ROW", RowSpec("0 TRY - 99,99 TRY", "SGK", "SGK_LOW")),
    ("ROW", RowSpec("100 TRY - 150 TRY", "SGK", "SGK_MID")),
    ("ROW", RowSpec("150,01 TRY -", "SGK", "SGK_HIGH")),
 
    ("SECTION", "KİRALIK KASA"),
    ("ROW", RowSpec("Büyük Kasa", "KASA", detail="BUYUK", split_channel=False)),
    ("ROW", RowSpec("Orta Kasa", "KASA", detail="ORTA", split_channel=False)),
    ("ROW", RowSpec("Küçük Kasa", "KASA", detail="KUCUK", split_channel=False)),
    ("ROW", RowSpec("Özel / Süper Kasa", "KASA", detail="OZEL", split_channel=False)),
 
    # Bankaların çek kırılımları birebir aynı değil. Aynı hizmet ailesindeki
    # gerçek alt tarifeler tek hücrede etiketli gösterilir; zorla tek fiyat seçilmez.
    ("SECTION", "ÇEK"),
    ("ROW", RowSpec("Çek Defteri / Çek Karnesi", "CEK_DEFTERI_GRUP", split_channel=False)),
    ("ROW", RowSpec("Bloke / Keşide Çeki Düzenleme", "CEK_DUZENLEME_STANDART_GRUP", split_channel=False)),
    ("ROW", RowSpec("Dövizli / Özel Nitelikli Çek Düzenleme", "CEK_DUZENLEME_OZEL_GRUP", split_channel=False)),
    ("ROW", RowSpec("Çek İade", "CEK_IADE_GRUP", split_channel=False)),
    ("ROW", RowSpec("Çek Tahsil - Yurtiçi", "CEK_TAHSIL_YURTICI_GRUP", split_channel=False)),
    ("ROW", RowSpec("Çek Tahsil - Dövizli / YP", "CEK_TAHSIL_DOVIZ_GRUP", split_channel=False)),
    ("ROW", RowSpec("Karşılıksız Çek / Düzeltme", "CEK_KARSILIKSIZ_GRUP", split_channel=False)),
 
    ("SECTION", "KKB / RİSK RAPORLARI"),
    ("ROW", RowSpec("KKB Risk Raporu", "KREDI_RISK", split_channel=False)),
    ("ROW", RowSpec("KKB Çek Bilgileri / Çek Risk Raporu", "CEK_RISK", split_channel=False)),
 
    ("SECTION", "HGS"),
    ("ROW", RowSpec("HGS Etiket Bedeli", "HGS_ETIKET", split_channel=False)),
    ("ROW", RowSpec("HGS Kart Bedeli", "HGS_KART", split_channel=False)),
 
    ("SECTION", "VERGİ / DEVLET ÖDEMELERİ"),
    ("ROW", RowSpec("Vergi Ödemeleri", "VERGI")),
 
    ("SECTION", "SENET"),
    ("ROW", RowSpec("Senet İade", "SENET_IADE", split_channel=False)),
    ("ROW", RowSpec("Senet Protesto", "SENET_PROTESTO", split_channel=False)),
    ("ROW", RowSpec("Senet Protesto Kaldırma", "SENET_PROTESTO_KALDIRMA", split_channel=False)),
    ("ROW", RowSpec("Senet Tahsil", "SENET_TAHSIL_GRUP", split_channel=False)),
 
    ("SECTION", "DÜZENLİ TRANSFERLER"),
    ("ROW", RowSpec("Düzenli EFT - 0 TRY - 8.300 TRY", "DUZENLI_EFT", "TRANSFER_1")),
    ("ROW", RowSpec("Düzenli EFT - 8.300,01 TRY - 399.000 TRY", "DUZENLI_EFT", "TRANSFER_2")),
    ("ROW", RowSpec("Düzenli EFT - 399.000,01 TRY -", "DUZENLI_EFT", "TRANSFER_3")),
    ("ROW", RowSpec("Düzenli Havale - 0 TRY - 8.300 TRY", "DUZENLI_HAVALE", "TRANSFER_1")),
    ("ROW", RowSpec("Düzenli Havale - 8.300,01 TRY - 399.000 TRY", "DUZENLI_HAVALE", "TRANSFER_2")),
    ("ROW", RowSpec("Düzenli Havale - 399.000,01 TRY -", "DUZENLI_HAVALE", "TRANSFER_3")),
 
    # Kullanıcı denetimine göre yapay 149,99/150 kırılımı kaldırıldı.
    # Kaynağın yayımladığı gerçek aralık tek hücrede gösterilir.
    ("SECTION", "AİDAT ÖDEMELERİ"),
    ("ROW", RowSpec("Aidat Ödemeleri", "AIDAT")),
 
    ("SECTION", "ÖZEL OKUL ÖDEME"),
    ("ROW", RowSpec("Özel Okul Ödemeleri", "OZEL_OKUL")),
 
    ("SECTION", "TELEFON ÖDEMELERİ"),
    ("ROW", RowSpec("Telefon / Cep Telefonu Faturası Ödemeleri", "TELEFON")),
 
    # Belge/ekstre/araştırma satırları genel HESAP_OZETI veya ARSIV etiketiyle
    # çözülmez. Her satır aşağıdaki ayrı hizmet koduyla banka + tam ürün adına
    # bağlanır; böylece ticari/bireysel, kart/mevduat ve posta/şube tarifeleri
    # birbirine taşınmaz.
    ("SECTION", "BELGE / EKSTRE / ARAŞTIRMA"),
    ("ROW", RowSpec("Vadesiz Mevduat Hesabı Ekstresi – TL", "DOC_VADESIZ_TL")),
    ("ROW", RowSpec("Vadesiz Mevduat Hesabı Ekstresi – YP", "DOC_VADESIZ_YP")),
    ("ROW", RowSpec("Bireysel KMH Hesap Özeti – Basılı", "DOC_KMH_BIREYSEL", split_channel=False)),
    ("ROW", RowSpec("Kredi Kartı Geçmiş Dönem Hesap Özeti", "DOC_KK_GECMIS", split_channel=False)),
    ("ROW", RowSpec("Arşiv Araştırma – Genel Belge", "DOC_ARSIV_GENEL", split_channel=False)),
    ("ROW", RowSpec("Mevduat Araştırma", "DOC_DEPOSIT_RESEARCH", split_channel=False)),
    ("ROW", RowSpec("1 Yıldan Eski Dekont/Ekstre Verilmesi", "DOC_ESKI_DEKONT_EKSTRE", split_channel=False)),
    ("ROW", RowSpec("Aylık Hesap Özeti – Posta Gönderimi", "DOC_AYLIK_POSTA", split_channel=False)),
    ("ROW", RowSpec("Arşiv/Araştırma – Şube Bazında", "DOC_ARSIV_SUBE", split_channel=False)),
    ("ROW", RowSpec("Arşiv/Araştırma – Banka Bazında", "DOC_ARSIV_BANKA", split_channel=False)),
 
    ("SECTION", "BANKAYA ÖZEL EKSTRE TARİFELERİ"),
    ("ROW", RowSpec("Ticari KMH Hesap Özeti – Basılı", "DOC_KMH_TICARI", split_channel=False)),
    ("ROW", RowSpec("Ticari Kart Geçmiş Dönem Ekstresi", "DOC_TICARI_KART_GECMIS", split_channel=False)),
    ("ROW", RowSpec("Yatırım Hesabı Ekstresi – Gönderim", "DOC_YATIRIM_EKSTRE", split_channel=False)),
    ("ROW", RowSpec("Ticari Müşteri Aylık Hesap Özeti", "DOC_TICARI_AYLIK", split_channel=False)),
    ("ROW", RowSpec("Ticari Kart – Basılı Ekstre Ücreti", "DOC_TICARI_KART_BASILI", split_channel=False)),
    ("ROW", RowSpec("Ticari Kart – Ekstre Gönderim Ücreti", "DOC_TICARI_KART_GONDERIM", split_channel=False)),
    ("ROW", RowSpec("Ticari Hesap – Basılı Ekstre Gönderimi", "DOC_TICARI_HESAP_BASILI", split_channel=False)),
    ("ROW", RowSpec("Üye İşyeri/POS – Ekstre Gönderimi", "DOC_POS_EKSTRE", split_channel=False)),
 
    # Menkul kıymet tarifeleri tek bir genel rakama indirgenemez. Hisse,
    # virman ve saklama aileleri ayrı gösterilir; banka yalnız açıklama/formül
    # yayımlıyorsa bu durum rakam uydurulmadan hücrede açıkça belirtilir.
    ("SECTION", "YURTİÇİ MENKUL KIYMET İŞLEMLERİ"),
    ("ROW", RowSpec("Hisse Senedi Alım-Satım Komisyonu", "SEC_DOMESTIC_STOCK_TRADE", split_channel=False)),
    ("ROW", RowSpec("Hisse Senedi Virmanı", "SEC_DOMESTIC_STOCK_TRANSFER", split_channel=False)),
    ("ROW", RowSpec("Bono/Tahvil Virmanı", "SEC_DOMESTIC_BOND_TRANSFER", split_channel=False)),
    ("ROW", RowSpec("Yatırım Hesabı / Saklama Ücreti", "SEC_DOMESTIC_CUSTODY", split_channel=False)),
 
    ("SECTION", "YURTDIŞI MENKUL KIYMET İŞLEMLERİ"),
    ("ROW", RowSpec("Yurtdışı Pay Alım-Satım Komisyonu", "SEC_FOREIGN_STOCK_TRADE", split_channel=False)),
    ("ROW", RowSpec("Yurtdışı Pay Saklama Ücreti", "SEC_FOREIGN_STOCK_CUSTODY", split_channel=False)),
    ("ROW", RowSpec("Eurobond Virmanı", "SEC_EUROBOND_TRANSFER", split_channel=False)),
    ("ROW", RowSpec("Eurobond Saklama Ücreti", "SEC_EUROBOND_CUSTODY", split_channel=False)),
]
 
# ---------------------------------------------------------------------------
# NORMALİZASYON
# ---------------------------------------------------------------------------
 
@lru_cache(maxsize=65536)
def _norm(value) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\u0307", "").replace("\xa0", " ")
    text = " ".join(text.split())
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.translate(str.maketrans({
        "ı": "i", "İ": "i", "ğ": "g", "Ğ": "g",
        "ü": "u", "Ü": "u", "ş": "s", "Ş": "s",
        "ö": "o", "Ö": "o", "ç": "c", "Ç": "c",
    }))
    return text.lower().strip()
 
 
@lru_cache(maxsize=65536)
def _has_word(text: str, word: str) -> bool:
    """Kısa hizmet kodlarını başka kelimelerin içinden yanlış yakalamaz."""
    t = _norm(text)
    w = _norm(word)
    return re.search(rf"(?<![a-z0-9]){re.escape(w)}(?![a-z0-9])", t) is not None
 
 
def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
 
 
def _parse_number(raw: str) -> Optional[float]:
    s = re.sub(r"[^0-9.,]", "", raw or "")
    if not s:
        return None
 
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        parts = s.split(".")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]):
            s = "".join(parts)
 
    try:
        return float(s)
    except ValueError:
        return None
 
 
def _parse_band(text: str) -> Optional[Band]:
    """MASRAF adından işlem TUTARI aralığını çıkarır; ücret kolonlarına bakmaz."""
    t = _norm(text)
 
    # 8.300 TL ve altı
    m = re.search(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*(?:ve\s*)?(?:alti|altinda)",
        t,
    )
    if m:
        high = _parse_number(m.group(1))
        if high is not None:
            return Band(0.0, high)
 
    # 399.000,01 TL ve üzeri / üstü
    m = re.search(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*(?:ve\s*)?(?:uzeri|ustu)",
        t,
    )
    if m:
        low = _parse_number(m.group(1))
        if low is not None:
            return Band(low, None)
 
    # 8.300,01 TL - 399.000 TL
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*[-–—]\s*"
        r"(\d[\d.,]*)\s*(?:tl|try)?",
        t,
    ):
        local = t[max(0, m.start() - 6): min(len(t), m.end() + 6)]
        if ":" in local:  # saat aralığı (16:00-17:15) olmasın
            continue
 
        low = _parse_number(m.group(1))
        high = _parse_number(m.group(2))
 
        if low is None or high is None or high < low:
            continue
 
        return Band(low, high)
 
    return None
 
 
 
@lru_cache(maxsize=32768)
def _all_band_keys(text: str) -> Set[str]:
    """
    MASRAF + AÇIKLAMA içinde geçen bütün standart tutar bantlarını bulur.
 
    Böylece:
      - "149,99 TL'ye kadar ... 150 TL ve üzeri ..."
      - "0-150 TL ... 150,01 TL ve üzeri"
      - "100 TL altı ... 100 TL ve üzeri"
    gibi tek satır içinde iki farklı kademe yayınlayan bankalar da
    iki karşılaştırma satırına bağlanabilir.
    """
    t = _norm(text)
    bands: List[Band] = []
 
    # "8.300 TL ve altı", "100 TL altında"
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*(?:ve\s*)?(?:alti|altinda)",
        t,
    ):
        high = _parse_number(m.group(1))
        if high is not None:
            bands.append(Band(0.0, high))
 
    # "149,99 TL'ye kadar"
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?(?:'?[yea]*)?\s*kadar",
        t,
    ):
        high = _parse_number(m.group(1))
        if high is not None:
            bands.append(Band(0.0, high))
 
    # "399.000,01 TL ve üzeri / üstü"
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*(?:ve\s*)?(?:uzeri|ustu)",
        t,
    ):
        low = _parse_number(m.group(1))
        if low is not None:
            bands.append(Band(low, None))
 
    # Normal aralıklar: 0-8.300 / 8.300,01-399.000
    for m in re.finditer(
        r"(\d[\d.,]*)\s*(?:tl|try)?\s*[-–—]\s*"
        r"(\d[\d.,]*)\s*(?:tl|try)?(?:\s*arasi)?",
        t,
    ):
        local = t[max(0, m.start() - 6): min(len(t), m.end() + 6)]
 
        if ":" in local:
            continue
 
        low = _parse_number(m.group(1))
        high = _parse_number(m.group(2))
 
        if low is None or high is None or high < low:
            continue
 
        bands.append(Band(low, high))
 
    keys: Set[str] = set()
 
    for band in bands:
        key = _band_key(band)
 
        if key and not key.startswith("RAW:"):
            keys.add(key)
 
    return keys
 
 
def _close(a: Optional[float], b: Optional[float], tolerance: float = 1.01) -> bool:
    if a is None or b is None:
        return a is b
    return abs(a - b) <= tolerance
 
 
def _band_key(band: Optional[Band]) -> Optional[str]:
    if band is None:
        return None
 
    low, high = band.low, band.high
 
    # EFT / Havale / FAST mevzuat dilimleri.
    if low is not None and low <= 1.0 and _close(high, 8300.0):
        return "TRANSFER_1"
    if _close(low, 8300.01) and _close(high, 399000.0):
        return "TRANSFER_2"
    if low is not None and abs(low - 399000.01) <= 1.01 and high is None:
        return "TRANSFER_3"
 
    # Fatura / Aidat / Okul örnek şablon dilimleri.
    if low is not None and low <= 1.0 and _close(high, 149.99):
        return "FATURA_1"
    if low is not None and abs(low - 150.0) <= 1.01 and high is None:
        return "FATURA_2"
 
    # SGK örnek şablon dilimleri.
    if low is not None and low <= 1.0 and _close(high, 99.99):
        return "SGK_1"
    if low is not None and abs(low - 100.0) <= 1.01 and high is None:
        return "SGK_2"
 
    return f"RAW:{low}:{high}"
 
 
# ---------------------------------------------------------------------------
# HİZMET / KANAL SINIFLANDIRMA
# ---------------------------------------------------------------------------
 
@lru_cache(maxsize=16384)
def _service_tags(row: FeeRow) -> Set[str]:
    """Farklı bankaların adlandırmalarını ortak hizmet kimliklerine çevirir."""
    cat = _norm(row.kategori)
    mas = _norm(row.masraf)
    short = f"{cat} | {mas}"
    full = _norm(row.text)
 
    tags: Set[str] = set()
 
    # Supplemental satırları kendi kanonik hizmetini açıkça taşır.
    marker = re.search(r"service\s*=\s*([a-z0-9_]+)", full, flags=re.I)
    if marker:
        tags.add(marker.group(1).upper())
 
    # ---------------- PARA TRANSFERİ ----------------
    international = any(x in full for x in (
        "swift", "uluslararasi fon transfer", "yurt disi fast",
        "global fast", "western union", "fast uluslararasi",
    ))
    package = any(x in short for x in ("paket", "kota"))
    card_cash = any(x in short for x in ("nakit avans", "faiz orani"))
 
    if ((_has_word(full, "eft") or "elektronik fon transfer" in full)
            and not international and not package and "altin eft" not in full):
        if not card_cash or any(x in short for x in (
            "eft ucreti", "eft gonder", "elektronik fon transfer",
        )):
            tags.add("EFT")
 
    if ((_has_word(full, "fast") or "fonlarin anlik ve surekli transferi" in full)
            and not international and not package):
        tags.add("FAST")
 
    if (any(x in full for x in ("havale", "hesaptan hesaba"))
            and not international and not package):
        if not card_cash or any(x in short for x in (
            "havale ucreti", "havale gonder", "hesaptan hesaba",
        )):
            tags.add("HAVALE")
 
    # Referans karşılaştırma tablosundaki ek transfer aileleri.
    if any(x in full for x in ("swift", "uluslararasi fon transfer")):
        if any(x in full for x in ("gelen swift", "gelen doviz", "gelen uluslararasi", "gelen fon transfer")):
            tags.add("SWIFT_GELEN")
        if any(x in full for x in ("giden swift", "giden doviz", "doviz havale", "hesaptan giden", "gonderim")):
            tags.add("SWIFT_GIDEN")
 
    # Yurt dışı hızlı transferleri ürün markasına göre değil hedefe göre ayır:
    # 1) hesaba transfer, 2) karta transfer. Böylece Visa Direct / MoneySend /
    # Global Fast Karta gibi aynı işi farklı isimle sunan ürünler eşleşebilir.
    is_global_fast_card = (
        "global fast" in full
        and any(x in full for x in ("karta para gonder", "karta gonderim"))
    )
 
    is_fast_account = (
        any(x in full for x in (
            "yurt disi fast", "yurtdisi fast",
            "fast uluslararasi", "fast uluslararasi turkiye disina",
            "akbank fast uluslararasi",
        ))
        or ("global fast" in full and not is_global_fast_card)
        # Visa ile Yurt Dışı Para Transferi ayrı bir ürün ailesidir.
        # "Hızlı hesaba transfer" satırında yalnız bankaların FAST/Global FAST
        # benzeri kanonik ürünleri karşılaştırılır; aksi halde aynı hücreye iki
        # farklı ürün tarifesi yığılabiliyor.
    )
    if is_fast_account:
        tags.add("YURT_DISI_FAST")
 
    is_card_transfer = (
        # MasterCard kart ağı üzerinden yurt dışı karta gönderim.
        (
            "moneysend" in full
            and "alici" not in full
            and (
                any(x in full for x in (
                    "gonderici",
                    "yurtdisi banka kartina",
                    "yurt disi banka kartina",
                    "yurtdisi kart",
                    "yurt disi kart",
                ))
                # İş Bankası ücret tablosunda Bankamatik Kartından satırları
                # kategori altında "Moneysend" diye yayınlanır; MASRAF adında
                # "yurtdışı" kelimesi her zaman tekrar edilmez.
                or "bankamatik kartindan" in full
            )
        )
        # Yapı Kredi aynı işi Global Fast - Karta adıyla yayımlıyor.
        or is_global_fast_card
        # İş Bankası sözleşmelerinde aynı aile Moneysend / VISA Direct olarak geçiyor.
        or (
            "visa direct" in full
            and "alici" not in full
            and "hesaba" not in full
        )
    )
    if is_card_transfer:
        tags.add("KART_YURTDISI_TRANSFER")
 
    if "duzenli" in full and (_has_word(full, "eft") or "elektronik fon transfer" in full):
        tags.add("DUZENLI_EFT")
    if "duzenli" in full and "havale" in full:
        tags.add("DUZENLI_HAVALE")
 
    # ---------------- ATM / NAKİT ----------------
    if any(x in full for x in ("gunluk limit uzeri para cekme", "limit uzeri para cekme", "gunluk limit ustu para cekme", "limit ustu para cekme")):
        tags.add("LIMIT_UZERI_PARA_CEKME")
 
    if (
        "para cekme" in full
        and any(x in full for x in ("ortak atm", "diger banka atm", "baska kurulus", "baska banka atm"))
        and "limit uzeri" not in full
        and "nakit avans" not in full
    ):
        tags.add("ORTAK_ATM_PARA_CEKME")
 
    if (
        "para yatirma" in full
        and any(x in full for x in ("ortak atm", "diger banka atm", "baska kurulus", "baska banka atm", "atm"))
        and "nakit avans" not in mas
    ):
        tags.add("PARA_YATIRMA")
 
    if (
        any(x in mas for x in (
            "altin transfer",
            "ats ile altin gonderimi",
            "kiymetli maden transferi ucreti - altin",
            "kiymetli maden transfer - altin",
        ))
        and not any(x in mas for x in (
            "fiziki", "teslim", "kulce altin cekme",
            "western union", "eft", "havale", "fast",
        ))
    ):
        tags.add("ALTIN_TRANSFER")
 
    # ---------------- KASA / RAPOR ----------------
    if any(x in full for x in ("kiralik kasa", "kasa kiralama", "kasa ucreti")):
        tags.add("KASA")
 
    if (
        any(x in full for x in (
            "kiymetli maden teslim", "fiziki altin teslim", "altin teslim",
            "kulce altin cekme", "kulce altin teslim", "fiziki altin cekme",
            "musteriye fiziki altin teslim",
        ))
        and not any(x in full for x in (
            "musteriden fiziki altin kabul",
            "fiziki altin kabul",
        ))
    ):
        tags.add("KIYMETLI_MADEN_TESLIM")
 
    # Garanti gibi bazı bankalar "KKB Çek / Risk Raporu"nu tek satırda
    # yayımlar. Böyle bir satır hem çek risk hem genel risk raporunu temsil eder.
    combined_risk = any(x in full for x in (
        "kkb cek / risk raporu", "kkb cek/risk raporu", "cek / risk raporu",
    ))
    if combined_risk:
        tags.update({"CEK_RISK", "KREDI_RISK"})
    else:
        if any(x in full for x in (
            "cek risk raporu", "cek bilgileri raporu", "kkb cek", "findeks cek raporu",
            "cek sorgu raporu",
        )):
            tags.add("CEK_RISK")
        if any(x in full for x in (
            "risk raporu", "kkb risk", "kredi risk", "findeks risk", "risk merkezi raporu",
        )):
            tags.add("KREDI_RISK")
 
    # ---------------- TAHSİLAT / ÖDEME ----------------
    if (("fatura" in short or "fatura / kurum" in short or "fatura/kurum" in short
         or "fatura tahsil" in short or "kurum tahsil" in short)
            and "e-fatura" not in short):
        tags.add("FATURA")
 
    if "sgk" in short or "sosyal guvenlik" in full:
        tags.add("SGK")
 
    # HGS Etiket ve HGS Kart farklı ürünlerdir.
    # "HGS" kelimesi tek başına yeterli değildir; ürün tipi MASRAF/KATEGORİ
    # alanında açıkça görünmelidir.
    hgs_structural = short
    if "hgs" in hgs_structural:
        has_etiket = "etiket" in hgs_structural
        has_kart = bool(re.search(r"\bkart(?:i|ı|lari|ları|ucreti|ücreti)?\b", hgs_structural))
        if has_etiket and not has_kart:
            tags.add("HGS_ETIKET")
        if has_kart and not has_etiket:
            tags.add("HGS_KART")
 
    if "sans oyun" in full or any(x in full for x in (
        "bilyoner", "nesine", "tuttur", "oley", "misli", "sisal sans", "tjk",
    )):
        tags.add("SANS_OYUNU")
 
    # Aidat ödeme hizmetini kredi kartı "Aidatsız Kart" / kart aidatı kavramından ayır.
    has_aidat_word = bool(re.search(r"\baidat\b", short)) or (
        bool(re.search(r"\baidat\b", full))
        and any(x in full for x in ("fatura", "tahsilat", "odeme", "site", "apartman"))
    )
    if has_aidat_word and not any(x in full for x in (
        "aidatsiz kart", "kart aidati", "yillik kart ucreti", "uyelik ucreti - kart",
    )):
        tags.add("AIDAT")
 
    # Özel okul ÖDEME hizmetini "Vize ve Özel Okullar İçin Düzenlenen Mektup"
    # gibi belge hizmetlerinden kesin olarak ayır.
    if (any(x in full for x in (
        "ozel okul", "okul odeme", "okul taksiti", "egitim odeme", "egitim kurumu odeme",
    )) and not any(x in full for x in (
        "mektup", "vize", "konsolosluk", "referans yazisi", "referans mektubu",
    ))):
        tags.add("OZEL_OKUL")
 
    telefon_candidate = (
        any(x in short for x in (
            "telefon odeme", "telefon fatur", "cep telefonu fatur",
            "telefon operatorleri odemelerine aracilik",
            "gsm odeme", "telekom odeme", "turkcell", "vodafone",
        ))
        or (
            any(x in full for x in ("turkcell", "vodafone", "superonline", "tellcom", "turk telekom"))
            and any(x in short for x in ("fatura", "kurum odeme"))
        )
    )
    if (
        telefon_candidate
        and "tl/paket yukleme" not in mas
        and "paket yukleme" not in mas
        and "otomatik fatura odeme" not in mas
        and "alisveris faiz" not in mas
    ):
        tags.add("TELEFON")
 
    if ("vergi" in short and any(x in short for x in (
        "vergi tahsil", "vergi odeme", "fatura/vergi/sgk", "fatura / vergi / sgk", "mtv",
    )) and not any(x in short for x in ("vergi numarasi", "vergi yazisi", "kredi"))):
        tags.add("VERGI")
 
    # ---------------- BELGE / HESAP ----------------
    if (any(x in full for x in (
        "arsiv arastirma", "gecmis donem ekstre arsiv", "gecmis donem bankacilik islemleri bildirimi",
        "arsivden belge", "dokuman arastirma", "dokuman talebi",
        "1 yildan eski islemlere ait gecmise yonelik dekont", "gecmise yonelik dekont",
        "gecmis donem ekstre", "dekont masrafi - gecmise yonelik",
    )) or ("arsiv" in full and "arastirma" in full)):
        tags.add("ARSIV")
 
    if (any(x in full for x in ("mevduat arastirma", "mevduat hesap arastirma"))
            or ("arsiv" in cat and "arastirma" in cat and mas.startswith("merkezden"))):
        tags.add("MEVDUAT_ARASTIRMA")
 
    if any(x in short for x in (
        "referans mektubu", "referans yazisi", "banka referans",
        "itibar mektubu", "itibar /niyet/referans", "itibar/niyet/referans",
    )):
        tags.add("REFERANS_MEKTUBU")
 
    if any(x in full for x in (
        "vize icin", "konsolosluk icin mektup", "konsolosluk icin",
        "ozel okullar icin duzenlenen mektup", "ozel okul icin duzenlenen mektup",
    )):
        tags.add("VIZE_MEKTUBU")
 
    postal_statement = any(x in full for x in (
        "posta ile aylik hesap ozeti", "basili ekstre", "basili hesap ozeti",
        "ekstre gonderim", "hesap ozeti gonderimi", "hesap ozeti posta", "ekstre posta",
        "gecmis donem kredi karti hesap ozeti gonderimi",
    ))
 
    if postal_statement:
        tags.add("HESAP_OZETI_POSTA")
    elif (
        "hesap ozeti" in short
        or "ekstre masraf" in short
        or "ekstre ucret" in short
        or "ekstre veril" in short
        or "ekstre - " in short
    ):
        tags.add("HESAP_OZETI")
 
    if any(x in full for x in ("hesap arastirma", "hesap tespit", "hesap bulma")):
        tags.add("HESAP_ARASTIRMA")
    if any(x in full for x in ("borcu yoktur", "borc yoktur", "borcsuzluk")):
        tags.add("BORCU_YOKTUR")
 
    # ---------------- ATM BAKİYE ----------------
    if (any(x in full for x in ("bakiye sorma", "bakiye sorgu", "bakiye goruntule"))
            and (any(x in full for x in ("atm", "bankamatik"))
                 or "baska kurulus araciligiyla yapilan islemler" in full)):
        if any(x in full for x in ("yurt disi", "yurtdisi")):
            tags.add("BAKIYE_ATM_YURTDISI")
        else:
            tags.add("BAKIYE_ATM_YURTICI")
 
    # ---------------- ÇEK ----------------
    # Çek tarafında yanlış eşleştirme riski yüksek olduğu için mümkün olduğunca
    # MASRAF + KATEGORİ gibi yapısal alanlar kullanılır; açıklama tek başına
    # ürün kimliği belirlemez.
    cek_struct = short
 
    # Çek defteri: yaprak başı ve 10 yapraklı karne ayrı karşılaştırılır.
    if (
        any(x in cek_struct for x in ("cek defteri", "cek karnesi", "cek yaprak", "cek kitabi"))
        and any(x in cek_struct for x in ("yaprak basi", "yaprak başi", "yaprak başı", "50-350 yaprakli", "351 yaprak"))
    ):
        tags.add("CEK_DEFTERI_YAPRAK")
 
    if (
        any(x in cek_struct for x in ("cek defteri", "cek karnesi"))
        and any(x in cek_struct for x in ("10 yaprakli", "10'luk", "10luk"))
    ):
        tags.add("CEK_KARNESI_10")
 
    # Standart bloke/keşide çeki düzenleme. Hediye/armağan, seyahat,
    # dövizli/DTH ve karşılıklı çekler bu aileye giremez.
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "bloke cek duzenleme",
            "bloke/ keside ceki duzenleme",
            "bloke/keside ceki duzenleme",
            "keside ceki / bloke cek duzenleme",
            "keside cek duzenleme",
            "duzenleme - tp/yp",
        ))
        and not any(x in cek_struct for x in (
            "hediye", "armagan", "seyahat", "doviz", "dovizi natik",
            "dth", "karsilikli", "odeme", "durdurma",
        ))
    ):
        tags.add("CEK_DUZENLEME_STANDART")
 
    # Dövizli / DTH çek düzenleme.
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "dovizli cek duzenleme",
            "dovizi natik cek duzenleme",
            "dth'dan cek duzenlenmesi",
            "dth dan cek duzenlenmesi",
        ))
        and not any(x in cek_struct for x in ("odeme", "durdurma", "tahsil"))
    ):
        tags.add("CEK_DUZENLEME_DOVIZ")
 
    # İşlemsiz/muamelesiz çek iadesi. Senet iadesi veya başka iade türleri alınmaz.
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "cek muamelesiz iade",
            "cekin islemsiz iades",
            "ceklerin islemsiz iades",
            "tahsile verilen cekin islemsiz iades",
            "cek iade ucreti",
        ))
        and "senet" not in cek_struct
    ):
        tags.add("CEK_IADE")
 
    # Çek tahsilatı yalnız açık tahsil/tahsile alma ifadelerinden üretilir.
    # Gişeden çek ödeme / bloke çek ödeme ayrı işlemdir ve tahsilata girmez.
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "cek tahsil",
            "tahsile alinan cek",
            "tahsile alinan bankamiz ceki",
            "tahsile alinan diger banka ceki",
            "bankamiz ceki (tl-yp) tahsile alma",
            "baska banka ceki (tl-yp) tahsile alma",
            "yurtici banka ceki",
            "ykb ceki",
            "yp cek tahsilati",
            "dovizli cek - tahsile alinan",
            "tahsile alinan dovizli cek",
        ))
        and not any(x in cek_struct for x in (
            "gişeden cek odeme", "giseden cek odeme", "bloke cek odeme",
            "seyahat ceki odeme", "karsiliksiz cek elden odeme",
            "cek odeme -", "odemeyi durdurma",
        ))
    ):
        tags.add("CEK_TAHSIL")
 
    if (
        "karsiliksiz cek" in cek_struct
        and any(x in cek_struct for x in ("belgelendirme", "elden odeme"))
    ):
        tags.add("CEK_KARSILIKSIZ")
 
    if (
        "cek" in cek_struct
        and any(x in cek_struct for x in (
            "duzeltme hakki",
            "duzeltme ucreti",
        ))
    ):
        tags.add("CEK_DUZELTME_HAKKI")
 
    # ---------------- SENET ----------------
    # "Çekler ve Senetler" kategori başlığı yüzünden çek satırlarının senet
    # olarak etiketlenmesini engelle. Öncelik MASRAF adıdır; MASRAF kısa ise
    # yalnız açık "Senet Tahsil" kategorisi fallback olarak kullanılabilir.
    if "senet" in mas and "iade" in mas:
        tags.add("SENET_IADE")
 
    protesto_name = (
        "senet" in mas
        and "protesto" in mas
        and "protestosuz" not in mas
        and "iskonto" not in mas
        and "istira" not in mas
    )
    if protesto_name and "kaldir" not in mas:
        tags.add("SENET_PROTESTO")
 
    if protesto_name and "kaldir" in mas:
        tags.add("SENET_PROTESTO_KALDIRMA")
 
    senet_tahsil_context = (
        ("senet" in mas and any(x in mas for x in ("tahsil", "tahsile alma")))
        or (
            "senet tahsil" in cat
            and "cek" not in mas
            and any(x in mas for x in ("ayni sube", "diger sube", "tp/yp"))
        )
    )
    if senet_tahsil_context:
        tags.add("SENET_TAHSIL")
 
    return tags
 
 
@lru_cache(maxsize=16384)
def _channels(row: FeeRow) -> Set[str]:
    """
    Bir satırın geçerli olduğu kanal kümesini döndürür.
 
    Kanal önce KATEGORİ + MASRAF gibi yapısal alanlardan çıkarılır.
    "İnternet Şube" fiziksel şube değildir. Özellikle İş Bankası'nın
    "İnternet Şube, İşCep, Çözüm Merkezi" birleşik dijital tarifesi yalnız
    MOBİL/DİJİTAL olarak sınıflandırılır; fiziksel "... - Şube" ayrı kalır.
    """
    cat = _norm(row.kategori)
    mas = _norm(row.masraf)
    desc = _norm(row.aciklama)
    full = _norm(row.text)
 
    explicit = re.search(r"channel\s*=\s*([a-z0-9_]+)", full, flags=re.I)
    if explicit:
        channel = explicit.group(1).upper()
        if channel == "GENEL":
            return {"GENEL"}
        return {channel}
 
    structural = f"{cat} | {mas}"
 
    # İş Bankası'nın dijital birleşik adı: "İnternet Şube, İşCep, Çözüm Merkezi".
    # Buradaki "Şube" ve "Çözüm Merkezi" fiziksel şube tarifesi anlamına gelmez.
    is_isbank_digital_bundle = (
        row.banka == "İŞBANKASI"
        and (
            "internet sube" in structural
            or "iscep" in structural
        )
        and "fatura odemeleri" in structural
    )
    if is_isbank_digital_bundle:
        return {"MOBIL"}
 
    structural2 = (
        structural
        .replace("internet subesi", "internet")
        .replace("internet sube", "internet")
    )
 
    if "tum kanal" in structural2:
        return {"MOBIL", "SUBE", "ATM"}
 
    channels: Set[str] = set()
 
    if any(x in structural2 for x in (
        "mobil", "internet", "dijital", "cepteteb", "iscep",
        "asistan", "sgk.gov.tr", "web",
    )):
        channels.add("MOBIL")
 
    if any(x in structural2 for x in (
        "sube", "subeden", "musteri iletisim merkezi",
        "cozum merkezi", "telefon subesi", "kasadan",
    )):
        channels.add("SUBE")
 
    if any(x in structural2 for x in (
        "atm", "btm", "kiosk", "bankamatik",
    )):
        channels.add("ATM")
 
    if channels:
        return channels
 
    # Açıklama fallback'i yalnız açık işlem-kanalı ifadelerinde kullanılır.
    desc2 = (
        desc
        .replace("internet subesi", "internet")
        .replace("internet sube", "internet")
    )
 
    if any(x in desc2 for x in (
        "mobil uzerinden", "mobil'den", "mobilden",
        "internet uzerinden", "dijital kanaldan",
        "yalniz mobil", "sadece mobil",
    )):
        channels.add("MOBIL")
 
    if any(x in desc2 for x in (
        "subeden yapilan", "subeden gerceklestir",
        "sube kanalindan", "yalniz sube", "sadece sube",
        "musteri iletisim merkezi'nden", "musteri iletisim merkezinden",
    )):
        channels.add("SUBE")
 
    if any(x in desc2 for x in (
        "atm'den", "atmden", "atm uzerinden",
        "bankamatikten", "bankamatik uzerinden",
    )):
        channels.add("ATM")
 
    return channels or {"GENEL"}
 
def _detail_match(row: FeeRow, detail: Optional[str]) -> bool:
    if not detail:
        return True
 
    text = _norm(row.text)
 
    tests = {
        "BUYUK": lambda: (
            any(x in text for x in ("buyuk", "large"))
            and not any(x in text for x in ("ozel", "super", "extra buyuk", "xl"))
        ),
        "ORTA": lambda: (
            any(x in text for x in ("orta", "medium"))
            and not any(x in text for x in ("ozel", "super", "extra buyuk", "xl"))
        ),
        "KUCUK": lambda: (
            any(x in text for x in ("kucuk", "small"))
            and not any(x in text for x in ("ozel", "super", "extra buyuk", "xl"))
        ),
        "OZEL": lambda: any(
            x in text
            for x in (
                "ozel",
                "super",
                "extra buyuk",
                "xl",
            )
        ),
        "AYNI": lambda: any(
            x in text
            for x in (
                "ayni banka",
                "ayni sube",
                "bankamiza ait",
                "bankamiz cek",
                "bankamiz ceki",
                "ykb ceki",
                "gb subesinden",
                "kendi bankasi",
            )
        ),
        "DIGER": lambda: any(
            x in text
            for x in (
                "diger banka",
                "diger sube",
                "farkli sube",
                "muhabir banka",
                "muhabirden",
                "baska banka",
                "yurtici banka ceki",
            )
        ),
        "DOVIZ": lambda: any(
            x in text
            for x in (
                "doviz cek",
                "dovizli cek",
                "dovizi natik",
                "yp cek",
                "yabanci banka",
                "yurt disi",
            )
        ),
    }
 
    fn = tests.get(detail)
    return fn() if fn else True
 
 
 
# ---------------------------------------------------------------------------
# EŞLEŞME PUANI
# ---------------------------------------------------------------------------
 
def _candidate_score(row: FeeRow, spec: RowSpec, wanted_channel: str) -> int:
    tags = _service_tags(row)
 
    if spec.service not in tags:
        return -10_000
 
    score = 100
    mas = _norm(row.masraf)
    cat = _norm(row.kategori)
    full = _norm(row.text)
 
    # Tekil transfer karşılaştırmasına KOBİ/paket/kota fiyatı asla girmez.
    # Eski bir çıktıda "Kobi Giden Swift Paketi 250" tek SWIFT ücreti gibi
    # seçilip 685.714,29 TL gösterilmişti. Bu artık hard-reject.
    transfer_services = {
        "EFT", "HAVALE", "FAST", "SWIFT_GELEN", "SWIFT_GIDEN",
        "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER",
        "DUZENLI_EFT", "DUZENLI_HAVALE",
    }
    if spec.service in transfer_services:
        if (
            any(x in mas for x in ("paket", "kobi", "kota"))
            or "urun ve hizmet paket" in cat
            or "ek kaynak - akbank ticari" in cat
        ):
            return -10_000
 
    # Kaynak önceliği:
    # - Bireysel ana Ürün/Hizmet Ücretleri satırı temel referanstır.
    # - Ek resmî kaynak yalnız ana kaynakta eksik kalan kalemi tamamlar.
    # - Ticari paket/ücret, bireysel karşılaştırmaya taşınmaz.
    is_supplemental = cat.startswith("ek kaynak")
    if not is_supplemental:
        score += 38
    if "ek kaynak - akbank ticari" in cat:
        score -= 55
 
    if STATUS_NUMERIC.lower() in row.aciklama.lower():
        score += 18
    if STATUS_AVAILABLE.lower() in row.aciklama.lower():
        score -= 55
    if STATUS_EMPTY.lower() in row.aciklama.lower():
        score -= 45
    if STATUS_NOT_APPLICABLE.lower() in row.aciklama.lower():
        score -= 80
 
    # ---------------- BAND ----------------
    if spec.band_key:
        masraf_keys = _all_band_keys(row.masraf)
        all_keys = _all_band_keys(
            f"{row.masraf} | {row.aciklama}"
        )
 
        if spec.band_key in all_keys:
            score += 40
 
            # Band doğrudan MASRAF adındaysa en güvenilir eşleşme.
            if spec.band_key in masraf_keys:
                score += 22
 
        else:
            # Bazı bankalar Fatura / SGK gibi hizmetleri tutar bandı
            # belirtmeden "tüm tutarlara %X" olarak yayınlıyor.
            # Bu durumda iki ortak banda da aynı tarife uygulanabilir.
            if spec.service in {
                "FATURA",
                "SGK",
                "AIDAT",
                "OZEL_OKUL",
            }:
                score -= 12
            else:
                return -10_000
 
    # ---------------- DETAY ----------------
    # Çek tahsilatında "aynı şube çek ödeme" ile "bankanın kendi çekini
    # tahsile alma" aynı işlem değildir. Yanlış rakam riskini azaltmak için
    # çek tahsil detayları fail-closed / açık ürün adı ile eşleşir.
    if spec.service == "CEK_TAHSIL":
        structural = f"{cat} | {mas}"
 
        if spec.detail == "AYNI":
            same_bank = any(x in structural for x in (
                "bankamiz ceki",
                "bankamiza ait",
                "tahsile alinan bankamiz ceki",
                "ykb ceki",
                "garanti bankasi ceki",
                "kendi bankasi ceki",
            ))
            if not same_bank:
                return -10_000
 
        elif spec.detail == "DIGER":
            other_bank = any(x in structural for x in (
                "diger banka ceki",
                "baska banka ceki",
                "tahsile alinan diger banka ceki",
                "yurtici banka ceki",
            ))
            if not other_bank:
                return -10_000
 
        elif spec.detail == "DOVIZ":
            fx_cheque = any(x in structural for x in (
                "yp cek tahsil",
                "dovizli cek - tahsile alinan",
                "tahsile alinan dovizli cek",
                "tahsile alinan yp cekler",
                "diger banka yp - tahsile alinan",
                "yurt disi yabanci banka doviz ceki tahsile alinmasi",
            ))
            if not fx_cheque:
                return -10_000
 
        # Açık aynı/diğer/döviz tahsil ifadesi bulundu: yüksek güven.
        score += 80
        if "tahsile alinan" in structural or "tahsile alma" in structural:
            score += 25
        if "teminata alinan" in structural:
            score -= 30
 
    elif not _detail_match(row, spec.detail):
        generic_detail_fallback = False
 
        # Senette bazı bankalar aynı/diğer ayrımı yapmadan tek genel tarife
        # yayımlar. Yalnız senet tarafında düşük öncelikli fallback korunur.
        if spec.service == "SENET_TAHSIL":
            generic_detail_fallback = (
                "senet tahsil" in full
                and not any(
                    x in full
                    for x in (
                        "ayni sube",
                        "diger sube",
                        "muhabir",
                        "baska banka",
                    )
                )
            )
 
        if generic_detail_fallback:
            score -= 18
        else:
            return -10_000
 
    # ---------------- KANAL ----------------
    if spec.split_channel:
        channels = _channels(row)
 
        if wanted_channel in channels:
            score += 60
        elif "GENEL" in channels:
            score += 15
        else:
            return -10_000
 
    # Hizmet MASRAF adında açıkça geçiyorsa açıklama fallback'ından üstündür.
    if spec.service == "EFT" and _has_word(mas, "eft"):
        score += 30
    elif spec.service == "FAST" and _has_word(mas, "fast"):
        score += 35
    elif spec.service == "HAVALE" and "havale" in mas:
        score += 30
    elif spec.service in _service_tags(
        FeeRow(
            banka=row.banka,
            kategori=row.kategori,
            masraf=row.masraf,
        )
    ):
        score += 20
 
    if spec.service in {"SWIFT_GELEN", "SWIFT_GIDEN"}:
        if "swift" in mas:
            score += 35
        if spec.service == "SWIFT_GELEN" and "gelen" in full:
            score += 45
        if spec.service == "SWIFT_GIDEN" and any(x in full for x in ("giden", "gonder")):
            score += 45
 
    if spec.service == "YURT_DISI_FAST" and (
        any(x in full for x in (
            "yurt disi fast", "yurtdisi fast", "fast uluslararasi",
            "akbank fast uluslararasi",
        ))
        or ("global fast" in full and "karta para gonder" not in full)
        or "visa ile yurt disi para transferi" in full
        or "visa ile yurtdisi para transferi" in full
    ):
        score += 65
    if spec.service == "KART_YURTDISI_TRANSFER" and any(x in full for x in (
        "moneysend", "visa direct", "global fast - karta", "karta para gonderim",
    )):
        score += 65
    if spec.service == "LIMIT_UZERI_PARA_CEKME" and any(x in full for x in ("limit uzeri", "limit ustu")):
        score += 60
    if spec.service == "ORTAK_ATM_PARA_CEKME" and "para cekme" in full:
        score += 45
    if spec.service == "PARA_YATIRMA" and "para yatirma" in full:
        score += 45
    if spec.service == "DUZENLI_EFT" and "duzenli" in full and _has_word(full, "eft"):
        score += 55
    if spec.service == "DUZENLI_HAVALE" and "duzenli" in full and "havale" in full:
        score += 55
    if spec.service == "ALTIN_TRANSFER" and "altin" in full:
        score += 45
 
    # Arşiv kategorisi birçok farklı özel raporu içeriyor.
    # "Borcu Yoktur", Risk Raporu, Vize Mektubu gibi alt hizmetlerin
    # Arşiv Araştırma satırına yanlış düşmesini engelle.
    if spec.service == "ARSIV":
        specialized = {
            "BORCU_YOKTUR",
            "KREDI_RISK",
            "CEK_RISK",
            "VIZE_MEKTUBU",
            "REFERANS_MEKTUBU",
            "MEVDUAT_ARASTIRMA",
        }
 
        if tags & specialized:
            score -= 90
 
        if any(
            x in mas
            for x in (
                "gecmis donem bankacilik islemleri bildirimi",
                "sozlesme, dekont",
                "dokuman talebi",
                "1 yildan eski",
                "gecmise yonelik",
                "arsiv arastirma",
            )
        ):
            score += 65
 
    if spec.service == "HESAP_OZETI":
        if any(x in full for x in (
            "posta ile", "basili ekstre", "basili hesap ozeti",
            "ekstre gonderim", "hesap ozeti gonderimi",
        )):
            return -10_000
        if any(x in mas for x in ("ekstre", "hesap ozeti")):
            score += 25
 
    if spec.service == "SENET_TAHSIL":
        if any(x in mas for x in ("iskonto", "istira", "teminat")):
            return -10_000
        if "cek" in mas:
            return -10_000
 
    if spec.service == "HESAP_OZETI_POSTA":
        if "kktc" in full:
            score -= 40
 
        if any(
            x in full
            for x in (
                "posta ile",
                "basili ekstre",
                "ekstre gonderim",
                "hesap ozeti gonderimi",
            )
        ):
            score += 55
 
    if spec.service == "BAKIYE_ATM_YURTDISI":
        if "kktc" in full:
            score -= 90
        if any(x in full for x in ("turkiye kart", "turkiye bireysel", "t.c. kart")):
            score += 70
 
    # Standart işlem ücretini özel varyantlardan öne al.
    if any(
        x in mas
        for x in (
            "gonderimi",
            "gonderilmesi",
            "gonderme",
        )
    ):
        score += 18
 
    if any(
        x in mas
        for x in (
            "gec eft",
            "gec havale",
            "gec fast",
        )
    ):
        score -= 35
 
    if (
        spec.service in {"EFT", "HAVALE", "FAST", "SWIFT_GELEN", "SWIFT_GIDEN", "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER", "DUZENLI_EFT", "DUZENLI_HAVALE"}
        and any(
            x in mas
            for x in (
                "duzenli",
                "talimat",
                "supurme",
            )
        )
    ):
        score -= 12
 
    if any(
        x in mas
        for x in (
            "cebe",
            "kartsiz",
            "kartli para yatirma",
        )
    ):
        score -= 24
 
    if any(
        x in cat
        for x in (
            "kampanyali",
            "urun ve hizmet paket",
        )
    ):
        score -= 45
 
    # Template'teki Fatura satırları kart ödemesi odaklı.
    if spec.service == "FATURA":
        if (
            "kredi kart" in full
            or "kartindan" in full
        ):
            score += 25
 
        if "nakit" in mas:
            score -= 20
 
    if (
        spec.service == "SGK"
        and "kredi kart" in full
    ):
        score += 20
 
    if spec.service == "CEK_RISK":
        if any(x in mas for x in ("cek risk", "cek bilgileri", "kkb cek", "cek sorgu")):
            score += 70
 
    if spec.service == "KREDI_RISK":
        if "gercek kisiler" in full:
            score += 120
        if "ek kaynak - akbank ticari" in cat or "urun ve hizmet paketleri" in cat:
            score -= 110
 
    if spec.service == "PARA_YATIRMA":
        if "banka kartiyla" in mas and "cari hesaba para yatirma" in mas:
            score += 85
        if "t.c. subelerinden" in mas:
            score += 35
        if "standart atm" in mas:
            score += 28
        if "k.k.t.c." in mas:
            score -= 55
        if "kredi kartiyla" in mas or "maxipara" in mas:
            score -= 45
        if "ek kaynak - is bankasi banka karti sozlesmesi" in cat:
            score += 125
 
    if spec.service == "TELEFON":
        if any(x in mas for x in ("tl/paket", "paket yukleme", "otomatik fatura odeme faizi")):
            return -10_000
 
    if spec.service == "SENET_IADE":
        if (
            "senet" not in mas
            or "iade" not in mas
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            return -10_000
        score += 70
 
    if spec.service == "SENET_PROTESTO":
        if (
            "senet" not in mas
            or "protesto" not in mas
            or "protestosuz" in mas
            or "kaldir" in mas
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            return -10_000
        score += 80
 
    if spec.service == "SENET_PROTESTO_KALDIRMA":
        if (
            not ("senet" in mas and "protesto" in mas and "kaldir" in mas)
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            return -10_000
        score += 80
 
    if spec.service == "HESAP_OZETI_POSTA":
        if any(x in full for x in ("kredi kart", "ticari kart", "business kart")):
            return -10_000
 
    if spec.service == "KASA":
        if "depozito" in full:
            score -= 80
 
        if "aylik" in full:
            score -= 35
 
        if "yillik" in full:
            score += 45
 
    # ---------------- YÜKSEK RİSKLİ HİZMETLER ----------------
    # Bu grupta "yakın isim" yeterli değildir. Ürünün yapısal adında açık
    # eşleşme yoksa rakam göstermek yerine kaynak boşluğu bırakılır.
    structural = f"{cat} | {mas}"
 
    if spec.service == "HGS_ETIKET":
        if not ("hgs" in structural and "etiket" in structural and "kart" not in mas):
            return -10_000
        score += 120
 
    if spec.service == "HGS_KART":
        if not ("hgs" in structural and "kart" in structural and "etiket" not in mas):
            return -10_000
        score += 120
 
    if spec.service == "CEK_DEFTERI_YAPRAK":
        if not (
            any(x in structural for x in ("cek defteri", "cek karnesi"))
            and any(x in structural for x in (
                "yaprak basi", "50-350 yaprakli", "351 yaprak",
            ))
        ):
            return -10_000
        if any(x in mas for x in ("10 yaprakli", "10'luk", "25 yaprakli", "25'lik")):
            return -10_000
        score += 110
 
    if spec.service == "CEK_KARNESI_10":
        if not (
            any(x in structural for x in ("cek defteri", "cek karnesi"))
            and any(x in structural for x in ("10 yaprakli", "10'luk", "10luk"))
        ):
            return -10_000
        score += 110
 
    if spec.service == "CEK_DUZENLEME_STANDART":
        if not any(x in structural for x in (
            "bloke cek duzenleme",
            "bloke/ keside ceki duzenleme",
            "bloke/keside ceki duzenleme",
            "keside ceki / bloke cek duzenleme",
            "duzenleme - tp/yp",
        )):
            return -10_000
        if any(x in structural for x in (
            "hediye", "armagan", "seyahat", "doviz", "dovizi natik",
            "dth", "karsilikli", "odeme", "durdurma",
        )):
            return -10_000
        score += 120
 
    if spec.service == "CEK_DUZENLEME_DOVIZ":
        if not any(x in structural for x in (
            "dovizli cek duzenleme",
            "dovizi natik cek duzenleme",
            "dth'dan cek duzenlenmesi",
            "dth dan cek duzenlenmesi",
        )):
            return -10_000
        if any(x in structural for x in ("odeme", "durdurma", "tahsil")):
            return -10_000
        score += 120
 
    if spec.service == "CEK_IADE":
        if not (
            "cek" in structural
            and any(x in structural for x in (
                "muamelesiz iade",
                "islemsiz iade",
                "cek iade ucreti",
            ))
        ):
            return -10_000
        score += 110
 
    if spec.service == "CEK_KARSILIKSIZ":
        if not (
            "karsiliksiz cek" in structural
            and any(x in structural for x in ("belgelendirme", "elden odeme"))
        ):
            return -10_000
        score += 100
 
    if spec.service == "CEK_DUZELTME_HAKKI":
        if not (
            "cek" in structural
            and any(x in structural for x in ("duzeltme hakki", "duzeltme ucreti"))
        ):
            return -10_000
        score += 100
 
    return score
 
 
 
_AMBIGUITY_LOGGED: Set[Tuple[str, str, str, str]] = set()
 
 
def _row_fee_signature(row: FeeRow) -> Tuple[str, str, str, str]:
    """Adayların gerçekten aynı tarifeyi taşıyıp taşımadığını karşılaştırır."""
    return tuple(
        _clean(getattr(row, attr, ""))
        for attr in ("asgari_tutar", "asgari_oran", "azami_tutar", "azami_oran")
    )
 
 
def _best_match(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
) -> Optional[FeeRow]:
    """
    En iyi adayı seçer; belirsiz durumda FAIL-CLOSED davranır.
 
    Önceki sürüm her zaman en yüksek puanlı ilk satırı seçiyordu. İki farklı
    tarife birbirine çok yakın puan alırsa bu yaklaşım yanlış rakam
    üretebiliyordu. Artık yakın puanlı ve farklı ücretli aday varsa hiçbirini
    seçmiyoruz; hücre "Kontrol gerekli" / kaynak boşluğu olarak kalıyor.
    """
    candidates: List[Tuple[int, FeeRow]] = []
 
    for row in rows:
        if row.banka != bank:
            continue
 
        score = _candidate_score(row, spec, wanted_channel)
        if score <= -10_000:
            continue
 
        candidates.append((score, row))
 
    if not candidates:
        return None
 
    candidates.sort(
        key=lambda item: (
            item[0],
            -len(_norm(item[1].masraf)),
        ),
        reverse=True,
    )
 
    best_score, best = candidates[0]
 
    # Çok düşük güvenli eşleşmeyi yayımlama.
    min_score = 150
    exact_high_risk = {
        "HGS_ETIKET", "HGS_KART",
        "CEK_DEFTERI_YAPRAK", "CEK_KARNESI_10",
        "CEK_DUZENLEME_STANDART", "CEK_DUZENLEME_DOVIZ",
        "CEK_IADE", "CEK_TAHSIL", "CEK_KARSILIKSIZ",
        "CEK_DUZELTME_HAKKI",
    }
    transfer_high_risk = {
        "SWIFT_GELEN", "SWIFT_GIDEN",
        "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER",
    }
    high_risk = exact_high_risk | transfer_high_risk
    if spec.service in exact_high_risk:
        min_score = 170
    elif spec.service in transfer_high_risk:
        min_score = 185
 
    if best_score < min_score:
        return None
 
    best_fee = _row_fee_signature(best)
 
    # Aynı puanlı farklı ücret: kesin belirsizlik.
    conflicting_same_score = [
        row for score, row in candidates[1:]
        if score == best_score and _row_fee_signature(row) != best_fee
    ]
 
    # Yüksek riskli hizmetlerde puanı çok yakın farklı ücret de belirsizliktir.
    close_margin = 12 if spec.service in high_risk else 6
    conflicting_close = [
        row for score, row in candidates[1:]
        if best_score - score <= close_margin
        and _row_fee_signature(row) != best_fee
    ]
 
    if conflicting_same_score or conflicting_close:
        key = (bank, spec.service, spec.label, wanted_channel)
        if key not in _AMBIGUITY_LOGGED:
            _AMBIGUITY_LOGGED.add(key)
            conflict = (conflicting_same_score or conflicting_close)[0]
            print(
                "[comparison][AMBIGUOUS] "
                f"{spec.service} | {spec.label} | {bank} | {wanted_channel}: "
                f"'{best.masraf}' ({best_score}) ile "
                f"'{conflict.masraf}' arasında güvenli seçim yapılamadı. "
                "Yanlış tutar yazmak yerine hücre boşluğu bırakıldı."
            )
        return None
 
    return best
 
 
# ---------------------------------------------------------------------------
# EXCEL OKUMA / GÖSTERİM
# ---------------------------------------------------------------------------
 
def _find_source_sheet(wb):
    for ws in wb.worksheets:
        if ws.title == COMPARISON_SHEET:
            continue
 
        for row_idx in range(1, min(ws.max_row, 20) + 1):
            vals = [
                _norm(ws.cell(row=row_idx, column=col).value)
                for col in range(1, min(ws.max_column, 30) + 1)
            ]
            if "banka" in vals and "masraf" in vals:
                return ws, row_idx
 
    raise RuntimeError("BANKA ve MASRAF kolonlarını içeren ana veri sayfası bulunamadı.")
 
 
def _header_map(ws, header_row: int) -> Dict[str, int]:
    result: Dict[str, int] = {}
 
    for col in range(1, ws.max_column + 1):
        key = _norm(ws.cell(row=header_row, column=col).value)
        if key in HEADER_ALIASES:
            result[HEADER_ALIASES[key]] = col
 
    required = {"BANKA", "KATEGORİ", "MASRAF"}
    missing = required - set(result)
    if missing:
        raise RuntimeError("Eksik kolon: " + ", ".join(sorted(missing)))
 
    return result
 
 
def _read_rows(ws, header_row: int, cols: Mapping[str, int]) -> List[FeeRow]:
    def get(row_idx: int, name: str) -> str:
        col = cols.get(name)
        return _clean(ws.cell(row=row_idx, column=col).value) if col else ""
 
    result: List[FeeRow] = []
 
    for row_idx in range(header_row + 1, ws.max_row + 1):
        bank = get(row_idx, "BANKA")
        masraf = get(row_idx, "MASRAF")
        if not bank or not masraf:
            continue
 
        if bank not in BANKS:
            continue
 
        result.append(FeeRow(
            banka=bank,
            kategori=get(row_idx, "KATEGORİ"),
            masraf=masraf,
            asgari_tutar=get(row_idx, "ASGARİ TUTAR"),
            asgari_oran=get(row_idx, "ASGARİ ORAN"),
            azami_tutar=get(row_idx, "AZAMİ TUTAR"),
            azami_oran=get(row_idx, "AZAMİ ORAN"),
            aciklama=get(row_idx, "AÇIKLAMA"),
            guncelleme_tarihi=get(row_idx, "GÜNCELLEME TARİHİ"),
        ))
 
    return result
 
 
def _percent(value: str) -> str:
    value = _clean(value)
    if not value or value in ("-", "0", "0.0", "0.00"):
        return ""
 
    if "%" in value:
        return value
 
    try:
        number = float(value.replace(",", "."))
        # API bazı bankalarda 0,7'yi "0.7", bazılarında 0.007 gibi verebilir.
        if 0 < abs(number) < 0.1:
            number *= 100
        return (f"{number:g}%").replace(".", ",")
    except ValueError:
        return value
 
 
def _display_amount(value: str) -> str:
    value = _clean(value)
    if not value or value == "-":
        return ""
 
    # Birimi standartlaştır. Sayısal nokta ondalıksa Türkçe virgüle çevir.
    value = re.sub(r"\s*TL\b", " TRY", value, flags=re.I)
    value = re.sub(r"\s*TRY\b", " TRY", value, flags=re.I)
    value = re.sub(r"\s*(USD|EUR)\b", r" \1", value, flags=re.I)
 
    m = re.fullmatch(r"\s*([0-9.,]+)\s*(TRY|USD|EUR)?\s*", value, flags=re.I)
    if not m:
        return value
 
    raw, currency = m.group(1), (m.group(2) or "").upper()
 
    # 1234.56 -> 1.234,56 ; 1.234,56 zaten Türkçe formattadır.
    number = _parse_number(raw)
    if number is None:
        return value
 
    if abs(number - round(number)) < 1e-9:
        formatted = f"{int(round(number)):,}".replace(",", ".")
    else:
        formatted = f"{number:,.2f}"
        formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
        formatted = formatted.rstrip("0").rstrip(",")
 
    return f"{formatted} {currency}".strip()
 
 
 
def _fee_value_compact(row: Optional[FeeRow]) -> str:
    """Tutar/oranı BSMV notu olmadan tek satırda özetler."""
    if row is None:
        return ""
 
    min_amount = _display_amount(row.asgari_tutar)
    max_amount = _display_amount(row.azami_tutar)
    min_rate = _percent(row.asgari_oran)
    max_rate = _percent(row.azami_oran)
 
    if min_amount and max_amount:
        amount = (
            min_amount
            if _norm(min_amount) == _norm(max_amount)
            else f"{min_amount} - {max_amount}"
        )
    else:
        amount = max_amount or min_amount
 
    if min_rate and max_rate:
        rate = (
            min_rate
            if _norm(min_rate) == _norm(max_rate)
            else f"{min_rate} - {max_rate}"
        )
    else:
        rate = max_rate or min_rate
 
    if amount and rate:
        return f"{amount} / {rate}"
 
    return amount or rate
 
 
def _extract_deposit_from_description(row: Optional[FeeRow]) -> str:
    """
    Akbank gibi depozitoyu ayrı satır yerine yıllık kira açıklamasında
    yayımlayan bankalar için açıklamadan depozito tutarını çıkarır.
    """
    if row is None:
        return ""
 
    raw = _clean(row.aciklama)
 
    patterns = (
        r"depozito\s+bedeli\s*[:\-]?\s*([0-9][0-9.\s]*(?:,[0-9]+)?)\s*(TL|TRY)",
        r"depozito\s+ucreti\s*[:\-]?\s*([0-9][0-9.\s]*(?:,[0-9]+)?)\s*(TL|TRY)",
    )
 
    norm_raw = _norm(raw)
 
    # _norm Türkçe karakterleri sadeleştirir fakat rakam biçimini korur.
    for pattern in patterns:
        m = re.search(pattern, norm_raw, flags=re.I)
        if not m:
            continue
 
        amount = _display_amount(f"{m.group(1).strip()} {m.group(2)}")
        if amount:
            return amount
 
    return ""
 
 
def _deposit_match_score(row: FeeRow, spec: RowSpec) -> int:
    full = _norm(row.text)
    short = _norm(
        f"{row.kategori} | {row.masraf}"
    )
 
    # Açıklamasında "depozito bedeli ..." geçen yıllık kira satırını
    # ayrı depozito kaydı sanma. Ayrı satır fallback'i yalnız kategori
    # veya MASRAF adında depozito açıkça yazıyorsa çalışır.
    if "depozito" not in short:
        return -10_000
 
    if "KASA" not in _service_tags(row):
        return -10_000
 
    score = 100
 
    if spec.detail == "BUYUK":
        if "buyuk" in full:
            score += 50
        elif "standart kasa depozito" in full:
            score += 15
        elif any(x in full for x in ("orta", "kucuk", "ozel")):
            return -10_000
 
    elif spec.detail == "ORTA":
        if "orta" in full:
            score += 50
        elif "standart kasa depozito" in full:
            score += 15
        elif any(x in full for x in ("buyuk", "kucuk", "ozel")):
            return -10_000
 
    elif spec.detail == "KUCUK":
        if "kucuk" in full:
            score += 50
        elif "standart kasa depozito" in full:
            score += 15
        elif any(x in full for x in ("buyuk", "orta", "ozel")):
            return -10_000
 
    elif spec.detail == "OZEL":
        if "ozel" in full:
            score += 60
        else:
            return -10_000
 
    # Kasa24 tipleri boy adı olmayan ayrı sınıflar; yanlış boyla
    # eşleştirmemek için standart büyük/orta/küçük satırlarda geriye at.
    if "kasa24" in full and spec.detail in {"BUYUK", "ORTA", "KUCUK"}:
        score -= 40
 
    return score
 
 
def _best_deposit_match(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    annual_row: Optional[FeeRow],
) -> Tuple[str, str]:
    """(depozito_değeri, BSMV_durumu) döndürür."""
    candidates = []
    for row in rows:
        if row.banka != bank:
            continue
        score = _deposit_match_score(row, spec)
        if score <= -10_000:
            continue
        candidates.append((score, row))
 
    if candidates:
        candidates.sort(key=lambda item: (item[0], -len(_norm(item[1].masraf))), reverse=True)
        dep_row = candidates[0][1]
        value = _fee_value_compact(dep_row)
        return value, _bsmv_label(dep_row)
 
    # Akbank gibi depozitoyu yıllık satır açıklamasında veren kaynaklar.
    value = _extract_deposit_from_description(annual_row)
    if not value or annual_row is None:
        return "", ""
 
    desc = _norm(annual_row.aciklama)
    # Depozito cümlesinin yakınında BSMV bilgisi varsa onu kullan; yoksa tahmin etme.
    m = re.search(r"depozito.{0,160}?(bsmv\s+(?:dahil|haric))", desc, flags=re.I)
    if not m:
        m = re.search(r"(bsmv\s+(?:dahil|haric)).{0,160}?depozito", desc, flags=re.I)
    tax = ""
    if m:
        tax = "BSMV dahil" if "dahil" in _norm(m.group(1)) else "BSMV hariç"
    return value, tax
 
 
def _bsmv_label(row: Optional[FeeRow]) -> str:
    if row is None:
        return ""
    desc = _norm(row.aciklama)
    if "bsmv dahil" in desc:
        return "BSMV dahil"
    if "bsmv haric" in desc or "bsmv'den haric" in desc:
        return "BSMV hariç"
    if "bsiv dahil" in desc:
        return "BSİV dahil"
    if "bsiv haric" in desc:
        return "BSİV hariç"
    return ""
 
 
def _status_meta(row: Optional[FeeRow]) -> Dict[str, str]:
    if row is None:
        return {}
    desc = row.aciklama or ""
    meta: Dict[str, str] = {}
    for key in ("SERVICE", "CHANNEL", "BAND", "DISPLAY_TEXT"):
        m = re.search(rf"(?:^|[;|]\s*){key}\s*=\s*([^;|]+)", desc, flags=re.I)
        if m:
            meta[key] = _clean(m.group(1)).replace("\\n", "\n")
    return meta
 
 
def _status_kind(row: Optional[FeeRow]) -> str:
    if row is None:
        return ""
    desc = row.aciklama or ""
    if STATUS_NOT_APPLICABLE in desc:
        return "NOT_APPLICABLE"
    if STATUS_EMPTY in desc:
        return "PUBLISHED_EMPTY"
    if STATUS_AVAILABLE in desc:
        return "AVAILABLE"
    if STATUS_NUMERIC in desc:
        return "OFFICIAL_FEE"
    return ""
 
 
def _has_numeric_fee(row: Optional[FeeRow]) -> bool:
    if row is None:
        return False
    return any(_clean(getattr(row, attr, "")) not in ("", "-", "0", "0.0", "0.00")
               for attr in ("asgari_tutar", "asgari_oran", "azami_tutar", "azami_oran"))
 
 
def _fee_text(row: Optional[FeeRow], spec: Optional[RowSpec] = None) -> str:
    if row is None:
        return "N/A"
 
    status = _status_kind(row)
    meta = _status_meta(row)
    display_text = meta.get("DISPLAY_TEXT", "").strip()
    if display_text and status in {"NOT_APPLICABLE", "PUBLISHED_EMPTY", "AVAILABLE"}:
        return display_text
 
    if status == "NOT_APPLICABLE" and not _has_numeric_fee(row):
        return "Uygulanmıyor"
    if status == "PUBLISHED_EMPTY" and not _has_numeric_fee(row):
        return "Ücret ilan edilmemiş"
    if status == "AVAILABLE" and not _has_numeric_fee(row):
        return "Hizmet var\nAyrı ücret ilan edilmemiş"
 
    min_amount = _display_amount(row.asgari_tutar)
    max_amount = _display_amount(row.azami_tutar)
    min_rate = _percent(row.asgari_oran)
    max_rate = _percent(row.azami_oran)
 
    amount = ""
    if min_amount and max_amount:
        amount = min_amount if _norm(min_amount) == _norm(max_amount) else f"min {min_amount}\nmax {max_amount}"
    else:
        amount = max_amount or min_amount
 
    rate = ""
    if min_rate and max_rate:
        rate = min_rate if _norm(min_rate) == _norm(max_rate) else f"min {min_rate}\nmax {max_rate}"
    else:
        rate = max_rate or min_rate
 
    if spec and spec.band_key in {"FATURA_1", "SGK_1"}:
        keys = _all_band_keys(f"{row.masraf} | {row.aciklama}")
        upper_key = "FATURA_2" if spec.band_key == "FATURA_1" else "SGK_2"
        if spec.band_key in keys and upper_key in keys and amount:
            result = amount
        elif amount and rate:
            result = f"{amount}\n{rate}"
        else:
            result = amount or rate
    elif spec and spec.band_key in {"FATURA_2", "SGK_2"}:
        keys = _all_band_keys(f"{row.masraf} | {row.aciklama}")
        lower_key = "FATURA_1" if spec.band_key == "FATURA_2" else "SGK_1"
        if spec.band_key in keys and lower_key in keys and rate:
            result = rate
        else:
            result = rate or amount
    else:
        parts = [x for x in (amount, rate) if x]
        result = "\n".join(parts)
 
    tax = _bsmv_label(row)
    if tax:
        result = f"{result}\n{tax}" if result else tax
    return result or "Ücret bilgisi açıklamada"
 
 
 
def _service_status_row(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
    *,
    kinds: Optional[Set[str]] = None,
) -> Optional[FeeRow]:
    candidates = []
    allowed = kinds or {"AVAILABLE", "PUBLISHED_EMPTY", "NOT_APPLICABLE"}
 
    for row in rows:
        if row.banka != bank or spec.service not in _service_tags(row):
            continue
 
        status = _status_kind(row)
        if status not in allowed:
            continue
 
        meta = _status_meta(row)
        row_band = meta.get("BAND", "").upper()
 
        # Status belirli bir banda özel ise yalnız o bandı çözsün.
        if row_band and spec.band_key and row_band != spec.band_key:
            continue
        if row_band and not spec.band_key:
            continue
 
        channels = _channels(row)
        if wanted_channel in channels:
            score = 120
        elif "GENEL" in channels:
            score = 75
        else:
            continue
 
        if row_band and spec.band_key and row_band == spec.band_key:
            score += 80
        if status == "NOT_APPLICABLE":
            score += 30
 
        candidates.append((score, row))
 
    return max(candidates, key=lambda x: x[0])[1] if candidates else None
 
 
def _fast_limit_note(
    rows: Sequence[FeeRow], bank: str, wanted_channel: str,
) -> str:
    """FAST limit notunu bankanın resmî supplemental statüsünden üretir."""
    upper_band = RowSpec(
        label="FAST üst limit durumu",
        service="FAST",
        band_key="TRANSFER_3",
    )
    status_row = _service_status_row(
        rows,
        bank,
        upper_band,
        wanted_channel,
        kinds={"NOT_APPLICABLE"},
    )
    if status_row is None:
        return ""
 
    display_text = _status_meta(status_row).get("DISPLAY_TEXT", "")
    match = re.search(r"FAST\s+limiti\s+([^\n]+)", display_text, flags=re.I)
    if not match:
        return ""
    return f"FAST işlemleri {match.group(1).strip()}'ye kadar yapılmaktadır."
 
 
 
def _generic_institution_fee(
    rows: Sequence[FeeRow], bank: str, wanted_channel: str,
) -> Optional[FeeRow]:
    """Aidat/okul/telefon için yalnız gerçek Fatura/Kurum tahsilat tarifesini seçer."""
    candidates = []
 
    for row in rows:
        if row.banka != bank or not _has_numeric_fee(row):
            continue
 
        full = _norm(row.text)
        mas = _norm(row.masraf)
        cat = _norm(row.kategori)
 
        if not any(x in full for x in (
            "fatura/kurum", "fatura / kurum", "fatura ve anlasmali kurum",
            "fatura odemeleri", "kurum tahsilat", "kurum odeme",
        )):
            continue
 
        # Faiz / kredi ürünü / SGK / şans / vergi / telefon yükleme gibi
        # komşu fakat farklı ücretleri genel kurum tarifesine sokma.
        if any(x in mas for x in (
            "faiz", "otomatik fatura odeme", "talimatli fatura odeme islem faizi",
            "alisveris faiz", "sgk", "sans oyun", "vergi", "tl/paket yukleme",
            "paket yukleme", "nakit avans", "konsolosluk", "vize randevu",
        )):
            continue
 
        channels = _channels(row)
        score = 100
 
        if wanted_channel in channels:
            score += 70
        elif "GENEL" in channels:
            score += 28
        else:
            continue
 
        # Bankaların ana, karşılaştırılabilir kurum tarifeleri.
        if bank == "YAPIKREDI" and "fatura ve anlasmali kurum odemeleri" in mas:
            score += 220
        if bank == "AKBANK" and "fatura / kurum tahsil" in mas:
            score += 190
        if bank == "GARANTİ" and "fatura/kurum odemesi" in mas:
            score += 190
        if bank == "İŞBANKASI" and mas == "fatura odemeleri":
            score += 190
 
        # Kart faiz başlıklarını ve ticari paketleri geriye at.
        if "kredi karti faiz" in cat:
            score -= 200
        if "paket" in mas or "ticari" in cat:
            score -= 90
 
        candidates.append((score, row))
 
    if not candidates:
        return None
 
    candidates.sort(
        key=lambda item: (item[0], -len(_norm(item[1].masraf))),
        reverse=True,
    )
    return candidates[0][1]
 
 
 
def _fast_branch_publication_status(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[str]:
    if spec.service != "FAST" or wanted_channel != "SUBE":
        return None
    mobile = _best_match(rows, bank, spec, "MOBIL")
    if mobile is None:
        return None
    # Bankanın kaynak satırlarında bu band için şube FAST kaydı yoksa bunu eşleşme
    # hatası gibi N/A göstermek yerine kaynakta ayrı tarife bulunmadığını belirt.
    for row in rows:
        if row.banka != bank or "FAST" not in _service_tags(row):
            continue
        if spec.band_key and spec.band_key not in _all_band_keys(f"{row.masraf} | {row.aciklama}"):
            continue
        if "SUBE" in _channels(row):
            return None
    return "Şube FAST tarifesi\nyayımlanmıyor"
 
 
 
def _compact_hint(row: FeeRow) -> str:
    """Ücret hücresinde kısa bant/varyant etiketi üretir."""
    raw = _clean(row.masraf)
    parens = re.findall(r"\(([^()]*(?:TL|TRY|gr|USD|EUR)[^()]*)\)", raw, flags=re.I)
    if parens:
        hint = parens[-1].strip()
        if len(hint) <= 48:
            return hint
    parts = [p.strip() for p in re.split(r"\s+-\s+", raw) if p.strip()]
    if parts:
        tail = parts[-1]
        if len(tail) <= 48:
            return tail
    return ""
 
 
def _card_transfer_description_fee(row: FeeRow) -> str:
    """
    Numeric kolonları boş olup ücretini açıklamada yayımlayan kart-transfer
    satırlarını güvenli biçimde özetler. Şimdilik bunun gerekli olduğu resmî
    primary örnek Garanti MoneySend (Gönderici) satırıdır.
    """
    full = _norm(row.text)
    if "moneysend" not in full or "gonderici" not in full:
        return ""
 
    text = _norm(row.aciklama)
 
    p1 = re.search(
        r"([0-9][0-9.,]*)\s*tl(?:'?[a-z]+)?\s*kadar\s*%\s*([0-9.,]+)\s*\+\s*([0-9][0-9.,]*)\s*tl",
        text,
        flags=re.I,
    )
    p2 = re.search(
        r"([0-9][0-9.,]*)\s*[-–—]\s*([0-9][0-9.,]*)\s*tl\s*arasinda\s*%\s*([0-9.,]+)\s*\+\s*([0-9][0-9.,]*)\s*tl",
        text,
        flags=re.I,
    )
    p3 = re.search(
        r"([0-9][0-9.,]*)\s*tl\s*(?:uzerinde|ve uzeri)\s*%\s*([0-9.,]+)\s*\+\s*([0-9][0-9.,]*)\s*tl",
        text,
        flags=re.I,
    )
 
    if not (p1 and p2 and p3):
        return ""
 
    hi1, rate1, fixed1 = p1.groups()
    low2, hi2, rate2, fixed2 = p2.groups()
    low3, rate3, fixed3 = p3.groups()
 
    def amount(raw: str) -> str:
        return _display_amount(f"{raw} TRY")
 
    def rate(raw: str) -> str:
        return _percent(raw)
 
    lines = [
        f"0-{amount(hi1)}: {amount(fixed1)} + {rate(rate1)}",
        f"{amount(low2)}-{amount(hi2)}: {amount(fixed2)} + {rate(rate2)}",
        f"{amount(low3)}+: {amount(fixed3)} + {rate(rate3)}",
    ]
 
    if "bsmv" in text:
        lines[-1] += " (BSMV eklenir)"
 
    return "\n".join(lines)
 
 
def _transfer_band(row: FeeRow) -> Optional[Band]:
    """Uluslararası transfer satırının işlem tutarı bandını çıkarır."""
    return _parse_band(f"{row.masraf} | {row.aciklama}")
 
 
def _transfer_band_label(row: FeeRow) -> str:
    """0–12.000 TRY / 12.000 TRY+ gibi kısa ve okunabilir bant etiketi."""
    band = _transfer_band(row)
    if band is None:
        return ""
 
    def fmt(value: Optional[float]) -> str:
        if value is None:
            return ""
        return _display_amount(f"{value} TRY")
 
    low = band.low
    high = band.high
 
    if low is not None and low <= 1.0 and high is not None:
        return f"0–{fmt(high)}"
    if low is not None and high is not None:
        return f"{fmt(low)}–{fmt(high)}"
    if low is not None and high is None:
        return f"{fmt(low)}+"
    if low is None and high is not None:
        return f"≤{fmt(high)}"
    return ""
 
 
def _canonical_transfer_candidate(
    row: FeeRow,
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
) -> bool:
    """
    Yurt dışı transferlerde yalnız gerçekten aynı ürünü karşılaştır.
 
    Bu kurallar bilinçli olarak banka bazında dar tutulur. Amaç hücreyi
    doldurmak değil, İsme/Kasaya, araştırma/iade, farklı valör, kart/hesap,
    paket/toplu ödeme gibi başka ürünleri yanlışlıkla aynı hücreye sokmamaktır.
    """
    mas = _norm(row.masraf)
    cat = _norm(row.kategori)
    full = _norm(row.text)
 
    if any(x in mas for x in ("paket", "kobi", "kota")):
        return False
    if "ek kaynak - akbank ticari" in cat:
        return False
 
    if spec.service == "SWIFT_GELEN":
        # Araştırma/iade ve özel banka/isme-kasaya varyantları standart
        # "hesaba gelen" karşılaştırmasına dahil edilmez.
        if any(x in mas for x in (
            "arastirma", "iade", "isme", "kasaya",
            "garanti bbva international", "gb international",
            "isbank ag",
        )):
            return False
 
        if bank == "GARANTİ":
            return (
                "diger bankadan gelen doviz havale" in mas
                or "diger bankadan gelen doviz havalesi" in mas
            )
 
        if bank == "İŞBANKASI":
            return (
                "diger bankadan gelen doviz havalesi - odeme" in mas
                or "diger bankadan gelen doviz havale - odeme" in mas
            )
 
        if bank == "AKBANK":
            return (
                "gelen doviz havalesi - hesaba" in mas
                or "gelen doviz havale - hesaba" in mas
            )
 
        if bank == "YAPIKREDI":
            return (
                "gelen doviz havale" in mas
                and "arastirma" not in mas
                and "iade" not in mas
            )
 
        return False
 
    if spec.service == "SWIFT_GIDEN":
        if any(x in full for x in (
            "gelen swift",
            "gelen doviz",
            "yurtdisindan",
            "yurt disindan",
            "arastirma/iade",
            "arastirma / iade",
            "nydo toplu",
            "toplu odeme",
        )):
            return False
 
        if bank == "GARANTİ":
            if "diger bankaya giden doviz havale" not in mas:
                return False
            if wanted_channel == "MOBIL":
                return "mobil/internet" in mas or "mobil / internet" in mas
            if wanted_channel == "SUBE":
                return "sube" in mas
            return False
 
        if bank == "İŞBANKASI":
            if mas != "hesaptan giden swift":
                return False
            # Kanal İş Bankası'nda kategori başlığında taşınıyor.
            if wanted_channel == "MOBIL":
                return any(x in cat for x in ("internet/iscep", "internet / iscep"))
            if wanted_channel == "SUBE":
                return "sube kanali" in cat
            return False
 
        if bank == "AKBANK":
            if wanted_channel == "MOBIL":
                # Mobil ve İnternet aynı tarifeyi iki ayrı satırla yayınlıyor.
                # Tek bir kanonik kaynak olarak Mobil satırını kullan.
                return (
                    "akbank mobil baska bankaya doviz transferi" in mas
                    and "internet" not in mas
                )
            if wanted_channel == "SUBE":
                # Kanal yazmayan "Giden - 30.000..." satırı bankanın standart
                # şube/genel tarifesidir. Mobil/İnternet satırlarını alma.
                return (
                    "swift - giden -" in mas
                    and "akbank mobil" not in mas
                    and "akbank internet" not in mas
                )
            return False
 
        if bank == "YAPIKREDI":
            if "doviz havale gonderimi" not in mas:
                return False
            # Yapı Kredi aynı hizmet için İleri Gün / 1 Gün / Aynı Gün
            # valörlü üç gerçek tarife yayımlıyor. Üçü de aynı hücrede
            # etiketli olarak gösterilir; hiçbiri kaybedilmez.
            valor_ok = any(x in mas for x in (
                "ileri gun valorlu",
                "1 gun valorlu",
                "ayni gun valorlu",
            ))
            if not valor_ok:
                return False
            if wanted_channel == "MOBIL":
                return any(x in mas for x in (
                    "internet / mobil / telefon",
                    "internet/mobil/telefon",
                ))
            if wanted_channel == "SUBE":
                return "subeden" in mas
            return False
 
        return False
 
    if spec.service == "YURT_DISI_FAST":
        # Aynı hücreye Visa transferi, Western Union vb. eklenmez.
        if any(x in full for x in (
            "moneysend",
            "western union",
            "karta para gonder",
            "visa ile yurt disi para transferi",
            "visa ile yurtdisi para transferi",
        )):
            return False
 
        if bank == "GARANTİ":
            return "yurt disi fast" in mas or "yurtdisi fast" in mas
 
        if bank == "İŞBANKASI":
            # Güncel ücret tablosunda diğer üç bankadaki FAST/Global FAST
            # hesaba gönderimle birebir karşılaştırılabilir ayrı bir ürün yok.
            return False
 
        if bank == "AKBANK":
            return (
                "akbank fast uluslararasi" in full
                and "turkiye disina para gonderim" in full
            )
 
        if bank == "YAPIKREDI":
            return (
                "global fast" in full
                and "hesaba para gonderim" in full
                and "karta" not in full
            )
 
        return False
 
    if spec.service == "KART_YURTDISI_TRANSFER":
        if "alici" in full or "kktc" in full:
            return False
 
        if bank == "GARANTİ":
            return "moneysend" in full and "gonderici" in full
 
        if bank == "İŞBANKASI":
            # Aynı ücret kredi kartı/ön ödemeli/banka kartı varyantlarında
            # tekrarlanabiliyor. Karşılaştırma için debit kart (Bankamatik)
            # + Türkiye Kartları + TL tarifesi kanonik kaynak seçilir.
            return (
                "moneysend" in full
                and "bankamatik kartindan" in mas
                and "turkiye kartlari" in mas
                and " tl" in mas
            )
 
        if bank == "AKBANK":
            return False
 
        if bank == "YAPIKREDI":
            return (
                "global fast" in full
                and "karta para gonderim" in full
            )
 
        return False
 
    return True
 
 
def _canonical_transfer_rows(
    rows: Sequence[FeeRow],
    bank: str,
    spec: RowSpec,
    wanted_channel: str,
) -> List[Tuple[int, FeeRow]]:
    """Kanonik transfer adaylarını puanlayıp tekrarsız sırala."""
    lookup_channel = wanted_channel if spec.split_channel else "GENEL"
    candidates: List[Tuple[int, FeeRow]] = []
 
    for row in rows:
        if row.banka != bank:
            continue
 
        score = _candidate_score(row, spec, lookup_channel)
        special_card_fee = (
            _card_transfer_description_fee(row)
            if spec.service == "KART_YURTDISI_TRANSFER"
            else ""
        )
 
        if score <= -10_000:
            continue
        if not _has_numeric_fee(row) and not special_card_fee:
            continue
        if not _canonical_transfer_candidate(row, bank, spec, wanted_channel):
            continue
 
        candidates.append((score, row))
 
    # Aynı bant + aynı ücret farklı bir satırda tekrar yayınlanmışsa tekilleştir.
    unique: Dict[Tuple[str, Tuple[str, str, str, str]], Tuple[int, FeeRow]] = {}
 
    for score, row in candidates:
        band_label = _transfer_band_label(row)
        fee_signature = _row_fee_signature(row)
        key = (_norm(band_label), fee_signature)
 
        existing = unique.get(key)
        if existing is None or score > existing[0]:
            unique[key] = (score, row)
 
    result = list(unique.values())
 
    def sort_key(item: Tuple[int, FeeRow]):
        score, row = item
        band = _transfer_band(row)
        low = band.low if band and band.low is not None else -1.0
        high = band.high if band and band.high is not None else float("inf")
        return (low, high, -score, _norm(row.masraf))
 
    result.sort(key=sort_key)
    return result
 
 
def _common_tax_label(rows: Sequence[FeeRow]) -> str:
    labels = [_bsmv_label(row) for row in rows]
    nonempty = [x for x in labels if x]
    if not nonempty:
        return ""
    normalized = {_norm(x) for x in nonempty}
    if len(normalized) == 1 and len(nonempty) == len(rows):
        return nonempty[0]
    return ""
 
 
def _fixed_transfer_surcharge(row: FeeRow) -> str:
    """
    Bazı hızlı yurt dışı transfer tarifelerinde yüzde/min-max ücretine ek
    sabit komisyon açıklama alanında yayınlanır (Akbank/Yapı Kredi örneği).
    Bu ek tutarı kaybetmeden hücre özetine taşır.
    """
    text = _norm(row.aciklama)
 
    patterns = (
        r"sabit komisyon\s*([0-9][0-9.,]*)\s*tl",
        r"islem ucreti\s*([0-9][0-9.,]*)\s*tl\s*ek",
        r"([0-9][0-9.,]*)\s*tl\s*ek olarak tahsil",
    )
 
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return _display_amount(f"{match.group(1)} TRY")
 
    return ""
 
 
def _compact_transfer_blocks(
    rows: Sequence[FeeRow],
    spec: RowSpec,
) -> str:
    """
    Aynı kanonik SWIFT ürününün gerçek tutar bantlarını kaybetmeden gösterir.
 
    Önceki sürüm 4+ bandı "x-y (3 kademe)" diye sıkıştırıyordu; bu görünüm
    gerçek eşikleri gizlediği için kaldırıldı. Artık her resmî bant ayrı satırdır.
    Yapı Kredi'de İleri Gün / 1 Gün / Aynı Gün valörleri de ayrı etiketlenir.
    """
    if not rows:
        return ""
 
    common_tax = _common_tax_label(rows)
    lines: List[str] = []
    seen: Set[str] = set()
 
    def variant_label(row: FeeRow) -> str:
        text = _norm(row.masraf)
        if "ileri gun valorlu" in text:
            return "İleri Gün"
        if "ayni gun valorlu" in text:
            return "Aynı Gün"
        if "1 gun valorlu" in text:
            return "1 Gün"
        return ""
 
    for row in rows:
        fee = _fee_value_compact(row)
        if not fee:
            continue
 
        band_label = _transfer_band_label(row)
        variant = variant_label(row)
 
        labels = [x for x in (variant, band_label) if x]
        line = f"{' | '.join(labels)}: {fee}" if labels else fee
 
        # Vergi bilgisi tüm satırlarda aynı değilse satır bazında koru.
        if not common_tax:
            tax = _bsmv_label(row)
            if tax:
                line += f" ({tax})"
 
        key = _norm(line)
        if key in seen:
            continue
        seen.add(key)
        lines.append(line)
 
    if common_tax:
        lines.append(common_tax)
 
    return "\n".join(lines)
 
 
# ---------------------------------------------------------------------------
# KULLANICI DENETİMİ SONRASI YÜKSEK RİSKLİ ÖZEL ÇÖZÜCÜLER (V20)
# ---------------------------------------------------------------------------
 
def _audit_rows(
    rows: Sequence[FeeRow],
    bank: str,
    predicate,
    *,
    numeric_only: bool = True,
) -> List[FeeRow]:
    result: List[FeeRow] = []
    seen: Set[Tuple[str, str, str, str, str]] = set()
    for row in rows:
        if row.banka != bank:
            continue
        if numeric_only and not _has_numeric_fee(row):
            continue
        try:
            ok = predicate(row)
        except Exception:
            ok = False
        if not ok:
            continue
        key = (
            _norm(row.kategori), _norm(row.masraf),
            _norm(row.asgari_tutar), _norm(row.asgari_oran),
            _norm(row.azami_tutar + "|" + row.azami_oran),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result
 
 
def _is_supplemental_row(row: FeeRow) -> bool:
    """Ek kaynağı dataclass özelliği varsaymadan kategori adından belirler."""
    return _norm(row.kategori).startswith("ek kaynak")
 
 
def _primary_first(rows: Sequence[FeeRow]) -> List[FeeRow]:
    """Aynı hizmet için ana Ürün/Hizmet tablosunu ek kaynağın önüne alır."""
    return sorted(rows, key=lambda row: (1 if _is_supplemental_row(row) else 0))
 
 
def _audit_fee(row: FeeRow, *, prefix: str = "", suffix: str = "") -> str:
    fee = _fee_value_compact(row)
    if not fee:
        return ""
    value = f"{prefix}: {fee}" if prefix else fee
    if suffix:
        value += f" {suffix}"
    tax = _bsmv_label(row)
    if tax:
        value += f"\n{tax}"
    return value
 
 
def _audit_common_tax(rows: Sequence[FeeRow]) -> str:
    labels = {_bsmv_label(row) for row in rows if _bsmv_label(row)}
    return next(iter(labels)) if len(labels) == 1 else ""
 
 
def _audit_lines(rows: Sequence[FeeRow], labeler) -> Tuple[str, Optional[FeeRow]]:
    lines: List[str] = []
    seen: Set[str] = set()
    first: Optional[FeeRow] = None
    common_tax = _audit_common_tax(rows)
 
    for row in rows:
        fee = _fee_value_compact(row)
        if not fee:
            continue
        label = labeler(row)
        line = f"{label}: {fee}" if label else fee
        if not common_tax:
            tax = _bsmv_label(row)
            if tax:
                line += f" ({tax})"
        key = _norm(line)
        if key in seen:
            continue
        seen.add(key)
        first = first or row
        lines.append(line)
 
    if common_tax and lines:
        lines.append(common_tax)
    return "\n".join(lines), first
 
 
def _audit_status_text(row: Optional[FeeRow], fallback: str) -> str:
    if row is None:
        return fallback
    meta = _status_meta(row)
    display = meta.get("DISPLAY_TEXT", "").replace("\\n", "\n").strip()
    return display or fallback
 
 
def _audit_institution_range(row: FeeRow, *, label: str = "") -> str:
    value = _fee_value_compact(row)
    if not value:
        return ""
    desc = _norm(row.aciklama)
    if any(x in desc for x in ("kurum", "anlasma", "fatura turune gore", "islem bazinda")):
        value += "\n(kuruma/işleme göre)"
    tax = _bsmv_label(row)
    if tax:
        value += f"\n{tax}"
    if label:
        value = f"{label}: {value}"
    return value
 
 
def _audit_exact_status(
    rows: Sequence[FeeRow], bank: str, service: str, channel: str,
) -> Optional[FeeRow]:
    spec = RowSpec("audit", service)
    return _service_status_row(rows, bank, spec, channel)
 
 
 
 
def _audit_fatura_methods(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """
    Fatura/Kurum ödemelerini ödeme yöntemine göre ayırır.
 
    Kritik kural:
    - "Hesaptan", "kredi kartından" ve "nakit" tarifeleri birbirine taşınmaz.
    - Banka yalnız genel/gişe Fatura-Kurum tarifesi yayımlıyorsa hücre bunu
      açıkça "Genel tarife" diye belirtir; yöntem-spesifik ücretmiş gibi sunmaz.
    - İş Bankası kredi kartı kuralı resmî SSS sayfasındaki doğrulanmış
      supplemental status satırından alınır.
    - Yapı Kredi Mobil/İnternet vadesiz hesaptan fatura ödemesinin ücretsiz
      olduğu resmî fatura sayfasındaki supplemental status ile doğrulanır.
    """
    if spec.service != "FATURA":
        return None
 
    method = (spec.detail or "").upper()
    if method not in {"HESAPTAN", "KREDI_KARTI", "NAKIT"}:
        return None
 
    def first(pred) -> Optional[FeeRow]:
        candidates = _audit_rows(rows, bank, pred)
        return candidates[0] if candidates else None
 
    def amount_range(row: Optional[FeeRow]) -> str:
        if row is None:
            return ""
        lo = _display_amount(row.asgari_tutar)
        hi = _display_amount(row.azami_tutar)
        if lo and hi:
            return lo if _norm(lo) == _norm(hi) else f"{lo} - {hi}"
        return hi or lo
 
    def with_tax(lines: List[str], row: Optional[FeeRow]) -> str:
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return "\n".join(x for x in lines if x)
 
    # ------------------------------------------------------------------
    # GARANTİ BBVA
    # ------------------------------------------------------------------
    if bank == "GARANTİ":
        if method == "HESAPTAN":
            if wanted_channel == "SUBE":
                return (
                    "Ayrı hesaptan-Şube tarifesi yayımlanmıyor",
                    None,
                    "PUBLICATION_STATUS",
                )
 
            row = first(
                lambda r: (
                    "hesaptan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
            lines = [amount_range(row)]
            rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
            if rate:
                lines.append(f"azami {rate}")
            lines.append("Kuruma/işleme göre")
            return (with_tax(lines, row), row, "NUMERIC")
 
        if method == "KREDI_KARTI":
            if wanted_channel == "SUBE":
                return (
                    "Ayrı kredi kartı-Şube tarifesi yayımlanmıyor",
                    None,
                    "PUBLICATION_STATUS",
                )
 
            row = first(
                lambda r: (
                    "kredi kartindan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
            fixed = _display_amount(row.asgari_tutar)
            rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
            if not fixed or not rate:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
            return (
                with_tax(
                    [
                        f"149,99 TRY'ye kadar: {fixed}",
                        f"150 TRY ve üzeri: {rate}",
                    ],
                    row,
                ),
                row,
                "NUMERIC",
            )
 
        # NAKİT
        if wanted_channel == "MOBIL":
            return (
                "Mobil/İnternet kanalında nakit ödeme uygulanmıyor",
                None,
                "NOT_APPLICABLE",
            )
 
        row = first(
            lambda r: (
                "nakit fatura/kurum odemesi" in _norm(r.masraf)
                and "sube" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
        return (
            with_tax(
                [amount_range(row), "Kuruma/işleme göre"],
                row,
            ),
            row,
            "NUMERIC",
        )
 
    # ------------------------------------------------------------------
    # İŞ BANKASI
    # ------------------------------------------------------------------
    if bank == "İŞBANKASI":
        if method == "HESAPTAN":
            if wanted_channel == "MOBIL":
                row = first(
                    lambda r: (
                        _norm(r.masraf) == "fatura odemeleri"
                        and (
                            "internet sube" in _norm(r.kategori)
                            or "iscep" in _norm(r.kategori)
                        )
                        and "bankamatik" not in _norm(r.kategori)
                    )
                )
                if row is None:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                return (
                    with_tax(
                        [amount_range(row), "Kuruma/işleme göre"],
                        row,
                    ),
                    row,
                    "NUMERIC",
                )
 
            row = first(
                lambda r: (
                    _norm(r.masraf) == "fatura odemeleri"
                    and "fatura odemeleri - sube" in _norm(r.kategori)
                    and "internet sube" not in _norm(r.kategori)
                    and "bankamatik" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            return (
                with_tax(
                    [
                        "Genel gişe tarifesi:",
                        amount_range(row),
                        "Hesap/nakit ayrımı yayımlanmıyor",
                    ],
                    row,
                ),
                row,
                "GENERIC_TARIFF",
            )
 
        if method == "KREDI_KARTI":
            faq_rows = _audit_rows(
                rows,
                bank,
                lambda r: (
                    "service=fatura_kredi_karti" in _norm(r.aciklama)
                    and _status_kind(r) in {"AVAILABLE", "OFFICIAL_FEE"}
                ),
                numeric_only=False,
            )
            if not faq_rows:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            row = faq_rows[0]
            value = _audit_status_text(
                row,
                "Genel kredi kartı tarifesi:\n"
                "0-150 TRY: 5 TRY\n"
                "150 TRY üzeri: %3,50 + BSMV\n"
                "Kanal ayrımı yayımlanmıyor",
            )
            return (value, row, "NUMERIC")
 
        # NAKİT
        if wanted_channel == "MOBIL":
            return (
                "Mobil/İnternet kanalında nakit ödeme uygulanmıyor",
                None,
                "NOT_APPLICABLE",
            )
 
        row = first(
            lambda r: (
                _norm(r.masraf) == "fatura odemeleri"
                and "fatura odemeleri - sube" in _norm(r.kategori)
                and "internet sube" not in _norm(r.kategori)
                and "bankamatik" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (
            with_tax(
                [
                    "Genel gişe tarifesi:",
                    amount_range(row),
                    "Nakit/hesap ayrımı yayımlanmıyor",
                ],
                row,
            ),
            row,
            "GENERIC_TARIFF",
        )
 
    # ------------------------------------------------------------------
    # AKBANK
    # ------------------------------------------------------------------
    if bank == "AKBANK":
        if method == "HESAPTAN":
            if wanted_channel == "MOBIL":
                row = first(
                    lambda r: (
                        "akbank mobil'den fatura / kurum tahsilati" in _norm(r.masraf)
                        and "ek kaynak" not in _norm(r.kategori)
                    )
                )
                label = "Genel Mobil tarifesi:"
            else:
                row = first(
                    lambda r: (
                        "giseden fatura / kurum tahsilati" in _norm(r.masraf)
                        and "ek kaynak" not in _norm(r.kategori)
                    )
                )
                label = "Genel gişe tarifesi:"
 
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
            return (
                with_tax(
                    [
                        label,
                        amount_range(row),
                        "Ödeme aracı ayrılmıyor",
                    ],
                    row,
                ),
                row,
                "GENERIC_TARIFF",
            )
 
        if method == "KREDI_KARTI":
            if wanted_channel == "MOBIL":
                row = first(
                    lambda r: (
                        "kredi kartindan anlik fatura odeme ucreti (akbank internet)" in _norm(r.masraf)
                        and "ek kaynak" not in _norm(r.kategori)
                    )
                )
                if row is None:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                max_fee = _display_amount(row.azami_tutar) or _display_amount(row.asgari_tutar)
                return (
                    with_tax(
                        [
                            "Mobil için ayrı tarife yayımlanmıyor",
                            f"Akbank İnternet: azami {max_fee}",
                            "Firmaya göre",
                        ],
                        row,
                    ),
                    row,
                    "GENERIC_TARIFF",
                )
 
            row = first(
                lambda r: (
                    "kredi kartindan anlik fatura odeme ucreti (sube)" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            max_fee = _display_amount(row.azami_tutar) or _display_amount(row.asgari_tutar)
            return (
                with_tax(
                    [f"Azami {max_fee}", "Firmaya göre"],
                    row,
                ),
                row,
                "NUMERIC",
            )
 
        # NAKİT
        if wanted_channel == "MOBIL":
            return (
                "Mobil/İnternet kanalında nakit ödeme uygulanmıyor",
                None,
                "NOT_APPLICABLE",
            )
 
        row = first(
            lambda r: (
                "giseden fatura / kurum tahsilati" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (
            with_tax(
                [
                    "Genel gişe tarifesi:",
                    amount_range(row),
                    "Nakit/hesap ayrımı yayımlanmıyor",
                ],
                row,
            ),
            row,
            "GENERIC_TARIFF",
        )
 
    # ------------------------------------------------------------------
    # YAPI KREDİ
    # ------------------------------------------------------------------
    if bank == "YAPIKREDI":
        general = first(
            lambda r: (
                "fatura ve anlasmali kurum odemeleri" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
 
        if method == "HESAPTAN":
            if wanted_channel == "MOBIL":
                status_rows = _audit_rows(
                    rows,
                    bank,
                    lambda r: (
                        "service=fatura_hesaptan" in _norm(r.aciklama)
                        and "channel=mobil" in _norm(r.aciklama)
                        and _status_kind(r) in {"AVAILABLE", "OFFICIAL_FEE"}
                    ),
                    numeric_only=False,
                )
                if not status_rows:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                row = status_rows[0]
                return (
                    _audit_status_text(
                        row,
                        "Ücretsiz / 0 TRY\nVadesiz hesaptan",
                    ),
                    row,
                    "NUMERIC",
                )
 
            if general is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            return (
                with_tax(
                    [
                        "Genel tarife:",
                        amount_range(general),
                        "Kanal/ödeme aracı ayrılmıyor",
                    ],
                    general,
                ),
                general,
                "GENERIC_TARIFF",
            )
 
        if method == "KREDI_KARTI":
            card_rows = _audit_rows(
                rows,
                bank,
                lambda r: (
                    "anlasmali kurum fatura /sgk prim odemeleri" in _norm(r.masraf)
                    and "sube veya sgk.gov.tr" not in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                ),
            )
            low = next(
                (r for r in card_rows if "0 - 150" in _norm(r.masraf) or "0-150" in _norm(r.masraf)),
                None,
            )
            high = next(
                (r for r in card_rows if "150,01" in _norm(r.masraf) or "150.01" in _norm(r.masraf)),
                None,
            )
            if low is None or high is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
            fixed = _display_amount(low.azami_tutar) or _display_amount(low.asgari_tutar)
            rate = _percent(high.azami_oran) or _percent(high.asgari_oran)
            if not fixed or not rate:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
            return (
                with_tax(
                    [
                        "Genel kredi kartı tarifesi:",
                        f"0-150 TRY: {fixed}",
                        f"150,01 TRY ve üzeri: {rate}",
                        "Kanal ayrımı yayımlanmıyor",
                    ],
                    low,
                ),
                low,
                "GENERIC_TARIFF",
            )
 
        # NAKİT
        if wanted_channel == "MOBIL":
            return (
                "Mobil/İnternet kanalında nakit ödeme uygulanmıyor",
                None,
                "NOT_APPLICABLE",
            )
 
        if general is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (
            with_tax(
                [
                    "Genel tarife:",
                    amount_range(general),
                    "Nakit/hesap ayrımı yayımlanmıyor",
                ],
                general,
            ),
            general,
            "GENERIC_TARIFF",
        )
 
    return None
 
 
def _audit_atm(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service not in {
        "LIMIT_UZERI_PARA_CEKME",
        "ORTAK_ATM_PARA_CEKME",
        "BAKIYE_ATM_YURTICI",
        "BAKIYE_ATM_YURTDISI",
    }:
        return None
 
    # -----------------------------------------------------
    # İŞ BANKASI - ilk sayfadaki gerçek bireysel Türkiye tarifeleri
    # -----------------------------------------------------
    if bank == "İŞBANKASI":
        if spec.service == "BAKIYE_ATM_YURTDISI":
            candidates = _audit_rows(
                rows,
                bank,
                lambda r: (
                    "bankamatik karti - yurt disi atm bakiye sorgulama" in _norm(r.kategori)
                    and "turkiye bireysel kartlari" in _norm(r.masraf)
                    and "kktc" not in _norm(r.text)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
            eur = next((r for r in candidates if "eur" in _norm(r.asgari_tutar)), None)
            usd = next((r for r in candidates if "usd" in _norm(r.asgari_tutar)), None)
            if eur is None or usd is None:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            return (
                f"{_display_amount(eur.asgari_tutar)} / {_display_amount(usd.asgari_tutar)}\n"
                "Türkiye bireysel Bankamatik kartı\nBSMV hariç",
                eur,
                "NUMERIC",
            )
 
        if spec.service == "LIMIT_UZERI_PARA_CEKME":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    (
                        _norm(r.masraf) == "limit ustu para cekme"
                        or "mevduat hesabindan gunluk" in _norm(r.masraf)
                    )
                    and "turkiye" in _norm(r.text)
                    and "kktc" not in _norm(r.text)
                ),
            )
            # Aynı ücret Bankamatik/Kredi Kartı ana başlıklarında tekrar
            # yayımlanabiliyor. Türkiye bireysel mevduat/Bankamatik satırını
            # öncele.
            if not candidates:
                candidates = _audit_rows(
                    rows, bank,
                    lambda r: (
                        "limit ustu para cekme" in _norm(r.text)
                        and "kktc" not in _norm(r.text)
                        and "23,50" in _clean(r.asgari_tutar)
                    ),
                )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
            candidates.sort(
                key=lambda r: (
                    "bankamatik karti" not in _norm(r.kategori),
                    "mevduat hesaplari" not in _norm(r.kategori),
                    len(_norm(r.masraf)),
                )
            )
            row = candidates[0]
            amount = _display_amount(row.asgari_tutar) or _display_amount(row.azami_tutar)
            rate = _percent(row.asgari_oran) or _percent(row.azami_oran)
            pieces = []
            if amount:
                pieces.append(f"min {amount}")
            if rate:
                pieces.append(rate)
            value = " + ".join(pieces) if pieces else _fee_value_compact(row)
            tax = _bsmv_label(row)
            if tax:
                value += f"\n{tax}"
            return (value, row, "NUMERIC")
 
        if spec.service == "ORTAK_ATM_PARA_CEKME":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    "bankamatik karti - ortak atm" in _norm(r.kategori)
                    and "turkiye subelerinden verilmis kartlar" in _norm(r.masraf)
                    and "cari hesaptan para cekme" in _norm(r.masraf)
                    and "kibris" not in _norm(r.text)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
            def labeler(r: FeeRow) -> str:
                t = _norm(r.masraf)
                if "tek atm" in t:
                    return "Tek ATM"
                return "Standart ATM"
 
            candidates.sort(key=lambda r: 1 if "tek atm" in _norm(r.masraf) else 0)
            value, first = _audit_lines(candidates, labeler)
            return (value, first, "NUMERIC")
 
        if spec.service == "BAKIYE_ATM_YURTICI":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    " - bankamatik karti - " in _norm(r.kategori)
                    and "ortak atm" in _norm(r.kategori)
                    and "bakiye sorgulama" in _norm(r.text)
                    and "yurt ici diger banka atm" in _norm(r.masraf)
                    and "maxipara" not in _norm(r.kategori)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
            def labeler(r: FeeRow) -> str:
                t = _norm(r.masraf)
                if "tek atm" in t:
                    return "Tek ATM"
                return "Standart ATM"
 
            candidates.sort(key=lambda r: 1 if "tek atm" in _norm(r.masraf) else 0)
            value, first = _audit_lines(candidates, labeler)
            return (value, first, "NUMERIC")
 
    if spec.service == "BAKIYE_ATM_YURTDISI":
        return (
            "Karşılaştırılabilir ayrı yurtdışı ATM bakiye sorgulama tarifesi yayımlanmıyor",
            None,
            "PUBLICATION_STATUS",
        )
 
    # -----------------------------------------------------
    # YAPI KREDİ - Ortak ATM oranına ek yayımlanan maktu tutar
    # -----------------------------------------------------
    if bank == "YAPIKREDI" and spec.service == "ORTAK_ATM_PARA_CEKME":
        candidates = _audit_rows(
            rows,
            bank,
            lambda r: (
                "ortak atm" in _norm(r.masraf)
                and "para cekme" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
        row = candidates[0]
        rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
        fixed_match = re.search(
            r"ayrica\s+([0-9][0-9.,]*)\s*tl\s+maktu",
            _norm(row.aciklama),
        )
        fixed = (
            _display_amount(f"{fixed_match.group(1)} TL")
            if fixed_match
            else ""
        )
        if not fixed or not rate:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
        value = f"{fixed} + {rate}"
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")
 
    # -----------------------------------------------------
    # GARANTİ - yalnız yurtiçi TRY satırları
    # -----------------------------------------------------
    if bank != "GARANTİ":
        return None
 
    def pred(row: FeeRow) -> bool:
        if spec.service not in _service_tags(row):
            return False
        t = _norm(row.text)
        # Yalnız yurtiçi / TRY tarife. KKTÇ, yurtdışı ve dövizli alternatifleri reddet.
        if any(x in t for x in ("kktc", "yurt disi", "yurtdisi", "usd", "eur", "gbp", "sterlin")):
            return False
        if spec.service == "LIMIT_UZERI_PARA_CEKME":
            return "(tl)" in _norm(row.masraf) or "(try)" in _norm(row.masraf)
        return True
 
    candidates = _audit_rows(rows, bank, pred)
    if not candidates:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
    row = candidates[0]
    if spec.service == "LIMIT_UZERI_PARA_CEKME":
        amount = _display_amount(row.asgari_tutar) or _display_amount(row.azami_tutar)
        rate = _percent(row.asgari_oran) or _percent(row.azami_oran)
        lines = []
        if amount:
            lines.append(f"min {amount} (BSMV dahil)")
        if rate:
            lines.append(f"{rate} (BSMV hariç)")
        return ("\n".join(lines), row, "NUMERIC")
    return (_fee_text(row, spec), row, "NUMERIC")
 
 
def _audit_sans_oyunu(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Kanalsız genel şans oyunu tarifesini iki sütuna kopyalamaz."""
    if spec.service != "SANS_OYUNU":
        return None
 
    candidates = _primary_first(
        _audit_rows(
            rows,
            bank,
            lambda r: "SANS_OYUNU" in _service_tags(r),
        )
    )
    if not candidates:
        return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
    exact = [row for row in candidates if wanted_channel in _channels(row)]
    if exact:
        row = exact[0]
        return (_fee_text(row, None), row, "NUMERIC")
 
    general = [row for row in candidates if "GENEL" in _channels(row)]
    if general:
        row = general[0]
        if wanted_channel == "MOBIL":
            return (
                f"Genel / kanal belirtilmemiş:\n{_fee_text(row, None)}",
                row,
                "GENERIC_TARIFF",
            )
        return (
            "Şube için ayrı tarife\nyayımlanmıyor",
            row,
            "PUBLICATION_STATUS",
        )
 
    return (
        f"{wanted_channel.title()} için ayrı tarife yayımlanmıyor",
        None,
        "PUBLICATION_STATUS",
    )
 
 
def _audit_hgs(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """HGS ürün tipini ve kaynakta yayımlanan kanal bilgisini açık gösterir."""
    if spec.service not in {"HGS_ETIKET", "HGS_KART"}:
        return None
 
    candidates = _primary_first(
        _audit_rows(
            rows,
            bank,
            lambda r: spec.service in _service_tags(r),
        )
    )
    if not candidates:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
    row = candidates[0]
    if bank == "İŞBANKASI":
        channel_label = "Mobil / İnternet / Şube"
    else:
        channel_label = "Genel / kanal belirtilmemiş"
    return (
        f"{channel_label}:\n{_fee_text(row, None)}",
        row,
        "NUMERIC",
    )
 
def _audit_sgk(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "SGK":
        return None
 
    band = spec.band_key or ""
 
    if bank == "GARANTİ":
        candidates = _audit_rows(
            rows, bank,
            lambda r: _norm(r.masraf) == "sgk" and "ek kaynak" not in _norm(r.kategori),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        row = candidates[0]
        if band == "SGK_LOW":
            fee = _display_amount(row.asgari_tutar) or _display_amount(row.azami_tutar)
        else:
            fee = _percent(row.asgari_oran) or _percent(row.azami_oran)
        if not fee:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        value = (
            "Genel SGK tarifesi:\n"
            f"{fee}\n"
            "Kanal ayrımı yayımlanmıyor"
        )
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")
 
    if bank == "İŞBANKASI":
        if wanted_channel == "SUBE":
            status = _audit_exact_status(rows, bank, "SGK", "SUBE")
            return (
                _audit_status_text(status, "Şube için ayrı SGK kart ödeme tarifesi yayımlanmıyor"),
                status,
                "STATUS",
            )
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                "sgk prim odemesi" in _norm(r.text)
                and ("internet sube" in _norm(r.text) or "iscep" in _norm(r.text))
                and "ek kaynak" not in _norm(r.kategori)
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        row = candidates[0]
        # 0 TL minimum + %3,10 oran; sabit ücret aralığı gibi göstermiyoruz.
        rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
        value = rate or _fee_value_compact(row)
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")
 
    if bank == "AKBANK":
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                "anlik sgk prim odeme" in _norm(r.text)
                and "sgk.gov.tr" not in _norm(r.text)
                and "talimat" not in _norm(r.text)
                and "ticari" not in _norm(r.kategori)
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        # Kanal ayrımı yayımlanmayan genel Anlık SGK tarifesini kullan.
        # Min tutar + oran birlikte yayımlanan satır varsa onu tercih et.
        candidates.sort(
            key=lambda r: (
                bool(_display_amount(r.asgari_tutar) or _display_amount(r.azami_tutar)),
                bool(_percent(r.asgari_oran) or _percent(r.azami_oran)),
            ),
            reverse=True,
        )
        row = candidates[0]
        fee = _fee_value_compact(row)
        if not fee:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        value = (
            "Genel SGK tarifesi:\n"
            f"{fee}\n"
            "Kanal ayrımı yayımlanmıyor"
        )
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")
 
    if bank == "YAPIKREDI":
        # Fatura satırındaki 5 TRY / %3,5 SGK'ya taşınamaz. SGK için yayımlanan
        # tek doğru kredi kartı tarifesi Şube veya sgk.gov.tr satırıdır:
        # 0-100 TRY 3 TRY, 100,01 TRY ve üzeri %3, BSMV hariç.
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                "sube veya sgk.gov.tr" in _norm(r.text)
                and "sgk" in _norm(r.text)
                and "ek kaynak" not in _norm(r.kategori)
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
        row = candidates[0]
        fixed = _display_amount(row.asgari_tutar) or "3 TRY"
        rate = _percent(row.asgari_oran) or _percent(row.azami_oran) or "3%"
        channel_label = "Şube" if wanted_channel == "SUBE" else "sgk.gov.tr (referans)"
 
        if band == "SGK_LOW":
            value = f"{channel_label}:\n{fixed}"
        elif band == "SGK_MID":
            value = (
                f"{channel_label}:\n"
                f"100 TRY: {fixed}\n"
                f"100,01-150 TRY: {rate}"
            )
        else:
            value = f"{channel_label}:\n{rate}"
        tax = _bsmv_label(row)
        if tax:
            value += f"\n{tax}"
        return (value, row, "NUMERIC")
 
    return None
 
 
def _audit_isbank_special_kasa(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if bank != "İŞBANKASI" or spec.service != "KASA" or spec.detail != "OZEL":
        return None
 
    annual = _audit_rows(
        rows, bank,
        lambda r: (
            "ozel kasa" in _norm(r.masraf)
            and "depozito" not in _norm(r.text)
            and "kasa(yillik)" in _norm(r.kategori)
        ),
    )
    deposit_rows = _audit_rows(
        rows, bank,
        lambda r: "ozel kasa" in _norm(r.text) and "depozito" in _norm(r.text),
    )
    if not annual:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
    order = {"kucuk": 0, "orta": 1, "buyuk": 2}
    annual.sort(key=lambda r: next((v for k, v in order.items() if k in _norm(r.masraf)), 9))
    deposit = _fee_value_compact(deposit_rows[0]) if deposit_rows else ""
 
    lines: List[str] = []
    for row in annual:
        t = _norm(row.masraf)
        if "kucuk" in t:
            label = "Küçük Özel"
        elif "orta" in t:
            label = "Orta Özel"
        elif "buyuk" in t:
            label = "Büyük Özel"
        else:
            continue
        line = f"{label}: Yıllık {_fee_value_compact(row)}"
        if deposit:
            line += f" | Depozito {deposit}"
        lines.append(line)
 
    tax = _audit_common_tax(annual)
    if tax:
        lines.append(tax)
    return ("\n".join(lines), annual[0], "NUMERIC")
 
 
def _audit_cheque(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    service = spec.service
    cheque_services = {
        "CEK_DEFTERI_GRUP", "CEK_DUZENLEME_STANDART_GRUP",
        "CEK_DUZENLEME_OZEL_GRUP", "CEK_IADE_GRUP",
        "CEK_TAHSIL_YURTICI_GRUP", "CEK_TAHSIL_DOVIZ_GRUP",
        "CEK_KARSILIKSIZ_GRUP",
    }
    if service not in cheque_services:
        return None
 
    def safe_cheque(row: FeeRow) -> bool:
        t = _norm(row.text)
        masraf = _norm(row.masraf)
        kategori = _norm(row.kategori)
        if any(x in t for x in (
            "hediye ceki", "armagan ceki", "seyahat ceki", "cek paketi",
            "kobi cek", "kota", "istihbarat", "kayip cek", "sahte cek",
            "odemeyi durdurma", "odeme durdurma",
        )):
            return False
        # Ana sayfadaki satırlarda "çek" masraf adında bulunur. İş Bankası BHS
        # ek kaynağında ise "Bankamız TP/YP - Tahsile Alınan" gibi satır adlarında
        # "çek" yazmayabilir; bu durumda kategori açıkça BHS-Çek olmalıdır.
        return (
            "cek" in masraf
            or "is bankasi bhs - cek" in kategori
            or "cek defteri" in kategori
        )
 
    # -----------------------------------------------------
    # ÇEK DEFTERİ / KARNESİ
    # -----------------------------------------------------
    if service == "CEK_DEFTERI_GRUP":
        if bank == "GARANTİ":
            candidates = _audit_rows(
                rows, bank,
                lambda r: safe_cheque(r) and "cek defteri teslimi" in _norm(r.masraf) and "yaprak" in _norm(r.masraf),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            row = candidates[0]
            lines = [f"Şube / yaprak: {_fee_value_compact(row)}"]
            desc = _clean(row.aciklama)
            m = re.search(r"dijital[^0-9]{0,120}([0-9][0-9.,]*)\s*TL", desc, flags=re.I | re.S)
            if m:
                lines.append(f"Dijital başvuru / yaprak: {_display_amount(m.group(1) + ' TL')}")
            tax = _bsmv_label(row)
            if tax:
                lines.append(tax)
            if "degerli kagit" in _norm(desc):
                lines.append("Değerli kâğıt bedeli ayrıca")
            return ("\n".join(lines), row, "NUMERIC")
 
        if bank == "İŞBANKASI":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    safe_cheque(r)
                    and "cek defteri" in _norm(r.text)
                    and any(x in _norm(r.text) for x in (
                        "10 yaprak", "25 yaprak", "50-350", "50 - 350",
                        "50 – 350", "351 yaprak",
                    ))
                    and "k.k.t.c" not in _norm(r.text)
                    and "kktc" not in _norm(r.text)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
            # Ana tarife ile BHS aynı kalemi tekrar edebiliyor. Aynı grup/kademe
            # için primary satır önde tutulur; BHS yalnız eksik kademeyi tamamlar.
            slots: Dict[Tuple[str, str], FeeRow] = {}
            for row in _primary_first(candidates):
                t = _norm(row.text)
                group = "Karekodlu + Logolu" if "logolu" in t else "Karekodlu"
                if "10 yaprak" in t:
                    tier = "10 yaprak"
                elif "25 yaprak" in t:
                    tier = "25 yaprak"
                elif "50" in t and "350" in t:
                    tier = "50–350"
                elif "351" in t:
                    tier = "351+"
                else:
                    continue
                slots.setdefault((group, tier), row)
 
            lines: List[str] = []
            first: Optional[FeeRow] = None
            tier_order = ("10 yaprak", "25 yaprak", "50–350", "351+")
            used_rows: List[FeeRow] = []
            for group_name in ("Karekodlu", "Karekodlu + Logolu"):
                bits: List[str] = []
                for tier in tier_order:
                    row = slots.get((group_name, tier))
                    if row is None:
                        continue
                    fee = _fee_value_compact(row)
                    if tier in {"50–350", "351+"}:
                        fee += " / yaprak"
                    bits.append(f"{tier}: {fee}")
                    first = first or row
                    used_rows.append(row)
                if not bits:
                    continue
                lines.append(group_name + ":\n" + "\n".join(bits))
            tax = _audit_common_tax(used_rows)
            if tax:
                lines.append(tax)
            lines.append("Değerli kâğıt bedeli ayrıca")
            return ("\n".join(lines), first, "NUMERIC")
 
        if bank == "AKBANK":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    safe_cheque(r)
                    and (
                        "cek defteri" in _norm(r.masraf)
                        or "cek karnesi" in _norm(r.masraf)
                    )
                    and "paket" not in _norm(r.kategori)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
            candidates = _primary_first(candidates)
            lines: List[str] = []
            first: Optional[FeeRow] = None
            for label, token in (
                ("10 yaprak", "10 yaprak"),
                ("25 yaprak", "25 yaprak"),
                ("Yaprak başı", "yaprak basi"),
            ):
                row = next((r for r in candidates if token in _norm(r.masraf)), None)
                if row is None:
                    continue
                lines.append(f"{label}: {_fee_value_compact(row)}")
                first = first or row
 
            if not lines or first is None:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            lines.append("Değerli kâğıt bedeli ayrıca")
            tax = _audit_common_tax(candidates)
            if tax:
                lines.append(tax)
            return ("\n".join(lines), first, "NUMERIC")
 
        if bank == "YAPIKREDI":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    safe_cheque(r)
                    and ("cek karnesi" in _norm(r.text) or "cek defteri" in _norm(r.text))
                    and any(x in _norm(r.text) for x in ("10 yaprak", "10'luk", "25 yaprak", "25'lik", "100", "500"))
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            def labeler(r: FeeRow) -> str:
                t = _norm(r.masraf)
                if "10 yaprak" in t or "10'luk" in t: return "10 yaprak"
                if "25 yaprak" in t or "25'lik" in t: return "25 yaprak"
                if "100" in t: return "100'lük sürekli form"
                if "500" in t: return "500'lük sürekli form"
                return _clean(r.masraf)
            value, first = _audit_lines(candidates, labeler)
            return (value or _source_gap_text(spec, bank, "GENEL"), first, "NUMERIC" if value else "SOURCE_GAP")
 
    # -----------------------------------------------------
    # STANDART BLOKE / KEŞİDE DÜZENLEME
    # -----------------------------------------------------
    if service == "CEK_DUZENLEME_STANDART_GRUP":
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                safe_cheque(r)
                and any(x in _norm(r.masraf) for x in ("bloke cek duzenleme", "keside ceki duzenleme", "bloke/keside", "bloke / keside"))
                and not any(x in _norm(r.masraf) for x in ("dovizli", "dovizi natik", "dth"))
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
        row = candidates[0]
        return (_fee_text(row, None), row, "NUMERIC")
 
    # -----------------------------------------------------
    # DÖVİZLİ / ÖZEL NİTELİKLİ DÜZENLEME
    # -----------------------------------------------------
    if service == "CEK_DUZENLEME_OZEL_GRUP":
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                safe_cheque(r)
                and any(x in _norm(r.masraf) for x in ("duzenleme", "duzenlen"))
                and any(x in _norm(r.masraf) for x in ("dovizli", "dovizi natik", "dth", "ozel nitelikli"))
            ),
        )
        if candidates:
            candidates = _primary_first(candidates)
            value, first = _audit_lines(candidates, lambda r: _clean(r.masraf))
 
            if bank == "AKBANK":
                published = _audit_rows(
                    rows,
                    bank,
                    lambda r: (
                        safe_cheque(r)
                        and "ozel nitelikli cek duzenleme" in _norm(r.masraf)
                    ),
                    numeric_only=False,
                )
                if published and not _has_numeric_fee(published[0]):
                    value += (
                        "\nÖzel Nitelikli Çek Düzenleme: "
                        "Ücret tutarı belirtilmemiş"
                    )
            return (value, first, "NUMERIC")
 
        # Akbank'ın resmî ticari tablosunda "Özel Nitelikli Çek Düzenleme"
        # kalemi mevcut ancak ücret alanları boş. Bunu kaynak boşluğu gibi
        # göstermeyip yayımlanmış-boş tarife olarak açıklıyoruz.
        if bank == "AKBANK":
            published = _audit_rows(
                rows, bank,
                lambda r: (
                    safe_cheque(r)
                    and "ozel nitelikli cek duzenleme" in _norm(r.masraf)
                ),
                numeric_only=False,
            )
            if published:
                return (
                    "Kalem yayımlanıyor\nAyrı ücret tutarı belirtilmemiş",
                    published[0],
                    "STATUS",
                )
 
        # Garanti'de hediye/seyahat çeki gibi farklı ürünleri bu alana
        # taşımıyoruz. Doğrudan karşılaştırılabilir ayrı tarife yok.
        if bank == "GARANTİ":
            return (
                "Ayrı dövizli / özel nitelikli çek düzenleme tarifesi yayımlanmıyor",
                None,
                "PUBLICATION_STATUS",
            )
 
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
    # -----------------------------------------------------
    # ÇEK İADE
    # -----------------------------------------------------
    if service == "CEK_IADE_GRUP":
        candidates = _audit_rows(
            rows, bank,
            lambda r: safe_cheque(r) and "iade" in _norm(r.masraf),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
        # Aynı resmî tarife primary + supplemental olarak iki kez gelmişse tek göster.
        # Gerçekten farklı iade varyantları ücret imzası farklı olduğu için korunur.
        unique: List[FeeRow] = []
        seen_fee: Set[Tuple[str, str, str, str, str]] = set()
        for row in candidates:
            key = (*_row_fee_signature(row), _bsmv_label(row))
            if key in seen_fee:
                continue
            seen_fee.add(key)
            unique.append(row)
 
        def iade_label(r: FeeRow) -> str:
            t = _norm(r.masraf)
            if bank == "GARANTİ":
                if "takas gunu" in t:
                    return "Takas günü iade"
                return "Normal / işlemsiz iade"
            if bank == "İŞBANKASI":
                return "Muamelesiz / işlemsiz iade"
            if bank == "AKBANK":
                return "Çek iade"
            if bank == "YAPIKREDI":
                return "İşlemsiz iade"
            return _clean(r.masraf)
 
        value, first = _audit_lines(unique, iade_label)
 
        # Bazı bankalar Takas Günü iade ücretini ayrı satır yerine açıklamada yayımlar.
        desc = " ".join(_clean(r.aciklama) for r in candidates)
        m = re.search(
            r"takas\s+gunu.{0,120}?([0-9][0-9.,]*)\s*TL",
            _norm(desc),
            flags=re.I | re.S,
        )
        if m:
            extra = f"Takas günü: {_display_amount(m.group(1) + ' TL')}"
            if _norm(extra) not in _norm(value):
                value += f"\n{extra}"
 
        return (value, first, "NUMERIC")
 
    # -----------------------------------------------------
    # ÇEK TAHSİL YURTİÇİ
    # -----------------------------------------------------
    if service == "CEK_TAHSIL_YURTICI_GRUP":
        # Akbank'ta 4.7.3 üst satırı 380,95-4.000 TRY aralığı şeklinde
        # yayımlanıyor; ancak ilk sayfada bunun gerçek alt işlem tarifeleri
        # ayrıca bulunuyor. Üst satırı tek aralık diye göstermek yanıltıcıdır.
        # Bu nedenle doğrudan üç gerçek işlem satırını etiketli gösteriyoruz.
        if bank == "AKBANK":
            candidates = _audit_rows(
                rows, bank,
                lambda r: (
                    _norm(r.kategori) == "cekler ve senetler"
                    and "tahsile alinan cek" in _norm(r.masraf)
                    and "dovizli" not in _norm(r.masraf)
                    and "doviz" not in _norm(r.masraf)
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
            order = {
                "hesaba yatan": 0,
                "nakit odenen": 1,
                "farkli subeden": 2,
            }
            def akbank_label(r: FeeRow) -> str:
                t = _norm(r.masraf)
                if "farkli subeden" in t:
                    return "Farklı şubeden nakit ödenen"
                if "hesaba yatan" in t:
                    return "Hesaba yatan"
                if "nakit odenen" in t:
                    return "Nakit ödenen"
                return _clean(r.masraf)
 
            def akbank_sort(r: FeeRow) -> int:
                t = _norm(r.masraf)
                if "hesaba yatan" in t:
                    return 0
                if "farkli subeden" in t:
                    return 2
                if "nakit odenen" in t:
                    return 1
                return 9
 
            candidates.sort(key=akbank_sort)
            value, first = _audit_lines(candidates, akbank_label)
            return (value, first, "NUMERIC")
 
        def yurtiçi_pred(r: FeeRow) -> bool:
            if not safe_cheque(r):
                return False
            m = _norm(r.masraf)
            if bank == "GARANTİ":
                return (
                    "cek tahsil" in m
                    and (
                        "bankamiz ceki" in m
                        or "baska banka ceki (tl-yp)" in m
                        or "baska banka ceki (tl / yp)" in m
                    )
                    and "takasa" not in m
                )
            if bank == "İŞBANKASI":
                return (
                    "tahsile alinan" in m
                    and (
                        "bankamiz tp/yp" in m
                        or "diger banka tp" in m
                    )
                    and "diger banka yp" not in m
                )
            if bank == "YAPIKREDI":
                return (
                    "cek tahsilati" in m
                    and "(tl)" in m
                )
            return False
 
        candidates = _audit_rows(rows, bank, yurtiçi_pred)
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
        def labeler(r: FeeRow) -> str:
            t = _norm(r.masraf)
            if any(x in t for x in ("bankamiz", "ykb", "ayni banka", "ayni sube")):
                return "Bankanın kendi çeki"
            if any(x in t for x in ("diger banka", "yurtici banka", "farkli sube", "baska banka")):
                return "Diğer banka çeki"
            return "Genel çek tahsil"
 
        value, first = _audit_lines(candidates, labeler)
        return (value, first, "NUMERIC")
 
    # -----------------------------------------------------
    # ÇEK TAHSİL DÖVİZLİ / YP
    # -----------------------------------------------------
    if service == "CEK_TAHSIL_DOVIZ_GRUP":
        if bank == "İŞBANKASI":
            # BHS'deki 1.170 TRY satırı, ana Ürün/Hizmet tablosunda daha sonra
            # 1.543 TRY olarak güncellenmiştir. Ana kaynak varken eski ek kaynak
            # hiçbir koşulda karşılaştırmaya giremez.
            candidates = _audit_rows(
                rows,
                bank,
                lambda r: (
                    not _is_supplemental_row(r)
                    and _norm(r.masraf) == "diger banka - tahsile alinan - yp"
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            row = candidates[0]
            return (
                _audit_fee(row, prefix="Diğer Banka YP - Tahsile Alınan"),
                row,
                "NUMERIC",
            )
 
        if bank == "AKBANK":
            # Güncel ana bireysel tabloda iki ayrı resmî kalem yayımlanıyor.
            # Eski ticari 571,43-12.952,38 TRY / %1 satırı bunların yerine geçmez.
            candidates = _audit_rows(
                rows,
                bank,
                lambda r: (
                    not _is_supplemental_row(r)
                    and (
                        _norm(r.masraf).startswith("baska banka yp cek")
                        or _norm(r.masraf) == "tahsile alinan dovizli cekler"
                    )
                ),
            )
            if not candidates:
                return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
            candidates.sort(
                key=lambda r: 0 if _norm(r.masraf).startswith("baska banka yp cek") else 1
            )
            value, first = _audit_lines(candidates, lambda r: _clean(r.masraf))
            return (value, first, "NUMERIC")
 
        def doviz_pred(r: FeeRow) -> bool:
            if not safe_cheque(r):
                return False
            m = _norm(r.masraf)
            if bank == "GARANTİ":
                return (
                    "yp cek tahsil" in m
                    or ("baska banka" in m and "yp cek" in m and "takasa" in m)
                )
            if bank == "YAPIKREDI":
                return (
                    ("cek tahsilati" in m and "(yp)" in m)
                    or "tahsile alinan yp cekler" in m
                )
            return False
 
        candidates = _audit_rows(rows, bank, doviz_pred)
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
        value, first = _audit_lines(candidates, lambda r: _clean(r.masraf))
        return (value, first, "NUMERIC")
 
    # -----------------------------------------------------
    # KARŞILIKSIZ ÇEK / DÜZELTME
    # -----------------------------------------------------
    if service == "CEK_KARSILIKSIZ_GRUP":
        numeric = _audit_rows(
            rows, bank,
            lambda r: (
                safe_cheque(r)
                and (
                    "karsiliksiz" in _norm(r.masraf)
                    or "duzeltme" in _norm(r.masraf)
                    or "elden odeme" in _norm(r.masraf)
                )
            ),
        )
 
        # Yapı Kredi'de düzeltme hakkı ile karşılıksız çek elden ödeme
        # birbirinden farklı ücretlerdir. TL/YP aynı ücretle tekrar eden satırları
        # tekilleştirip işlem tiplerini açıkça etiketliyoruz.
        if bank == "YAPIKREDI" and numeric:
            picked: List[Tuple[str, FeeRow]] = []
            seen: Set[Tuple[str, Tuple[str, str, str, str]]] = set()
 
            for row in numeric:
                t = _norm(row.masraf)
                if "duzeltme hakki" in t:
                    label = "Düzeltme Hakkı"
                    order = 0
                elif "ayni sube" in t and "karsiliksiz" in t and "elden odeme" in t:
                    label = "Aynı Şube Karşılıksız Çek Elden Ödeme"
                    order = 1
                elif "baska sube" in t and "karsiliksiz" in t and "elden odeme" in t:
                    label = "Başka Şube Karşılıksız Çek Elden Ödeme"
                    order = 2
                else:
                    continue
 
                key = (label, _row_fee_signature(row))
                if key in seen:
                    continue
                seen.add(key)
                picked.append((f"{order}|{label}", row))
 
            picked.sort(key=lambda item: int(item[0].split("|", 1)[0]))
            lines: List[str] = []
            first: Optional[FeeRow] = None
            common_tax = _audit_common_tax([row for _, row in picked])
 
            for encoded_label, row in picked:
                label = encoded_label.split("|", 1)[1]
                fee = _fee_value_compact(row)
                if not fee:
                    continue
                line = f"{label}: {fee}"
                if not common_tax:
                    tax = _bsmv_label(row)
                    if tax:
                        line += f" ({tax})"
                lines.append(line)
                first = first or row
 
            if common_tax and lines:
                lines.append(common_tax)
 
            if lines:
                return ("\n".join(lines), first, "NUMERIC")
 
        # İş Bankası BHS: belgelendirme ayrı ücret yayımlamıyor; düzeltme numeric.
        status_candidates = [
            r for r in rows
            if r.banka == bank
            and "karsiliksiz cek belgelendirme" in _norm(r.masraf)
            and _status_kind(r)
        ]
        lines: List[str] = []
        first: Optional[FeeRow] = None
        seen: Set[str] = set()
        for row in numeric:
            t = _norm(row.masraf)
            label = "Karşılıksız / elden ödeme" if "karsiliksiz" in t else "Düzeltme"
            line = f"{label}: {_fee_value_compact(row)}"
            tax = _bsmv_label(row)
            if tax:
                line += f" ({tax})"
            if _norm(line) not in seen:
                seen.add(_norm(line))
                lines.append(line)
                first = first or row
        if status_candidates:
            st = status_candidates[0]
            lines.insert(0, "Belgelendirme: Ayrı ücret yayımlanmıyor")
            first = first or st
        if not lines:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
        return ("\n".join(lines), first, "NUMERIC" if numeric else "STATUS")
 
    return None
 
 
def _audit_kkb(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service not in {"KREDI_RISK", "CEK_RISK"}:
        return None
 
    if bank == "İŞBANKASI" and spec.service == "KREDI_RISK":
        candidates = _audit_rows(
            rows, bank,
            lambda r: (
                "kredi risk raporu" in _norm(r.masraf)
                and any(x in _norm(r.text) for x in ("iscep", "internet", "bankamatik"))
            ),
        )
        if not candidates:
            return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
        # Aynı 95 TL tarife farklı kanallarda tekrar eder; bir kez göster.
        row = candidates[0]
        return (_fee_text(row, None), row, "NUMERIC")
 
    if bank == "İŞBANKASI" and spec.service == "CEK_RISK":
        status = _audit_exact_status(rows, bank, "CEK_RISK", "MOBIL") or _audit_exact_status(rows, bank, "CEK_RISK", "GENEL")
        if status is not None:
            return (_audit_status_text(status, "Çek Raporu paket tarifeleri yayımlanıyor"), status, "STATUS")
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
    return None
 
 
def _audit_senet_tahsil(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "SENET_TAHSIL_GRUP":
        return None
 
    candidates = _audit_rows(
        rows, bank,
        lambda r: (
            "senet" in _norm(r.masraf)
            and "tahsil" in _norm(r.masraf)
            and not any(x in _norm(r.masraf) for x in ("iskonto", "istira", "teminat"))
        ),
    )
    if not candidates:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
 
    # Yapı Kredi'de iki eş numeric satırı tekilleştir; açıklamadaki muhabir
    # masrafını ayrıca görünür kıl.
    unique_by_fee: List[FeeRow] = []
    seen_sig: Set[Tuple[str, str, str, str]] = set()
    for row in candidates:
        sig = _row_fee_signature(row)
        if sig in seen_sig and bank == "YAPIKREDI":
            continue
        seen_sig.add(sig)
        unique_by_fee.append(row)
 
    def labeler(r: FeeRow) -> str:
        t = _norm(r.masraf)
        if "muhabir" in t:
            return "Muhabir banka"
        if "farkli sube" in t:
            return "Farklı şube"
        if "ayni sube" in t or "ayni banka" in t:
            return "Aynı banka/şube"
        return "Senet tahsil"
 
    value, first = _audit_lines(unique_by_fee, labeler)
    if bank == "YAPIKREDI":
        desc = " ".join(_clean(r.aciklama) for r in candidates)
        m = re.search(r"binde\s*4.{0,80}?asgari\s*([0-9][0-9.,]*)\s*TL", desc, flags=re.I | re.S)
        if m:
            value += f"\nMuhabir masrafı: min {_display_amount(m.group(1) + ' TL')} + binde 4"
    return (value, first, "NUMERIC")
 
 
def _audit_normal_transfer_row(
    rows: Sequence[FeeRow], bank: str, service: str, band_key: str, channel: str,
) -> Optional[FeeRow]:
    """Normal EFT/Havale tarifesini ürün adına göre kesin seçer."""
 
    def band_ok(row: FeeRow) -> bool:
        return _band_key(_parse_band(row.masraf)) == band_key
 
    candidates: List[FeeRow] = []
    for row in rows:
        if row.banka != bank or not _has_numeric_fee(row) or not band_ok(row):
            continue
        mas = _norm(row.masraf)
 
        if bank == "YAPIKREDI":
            if service == "EFT":
                if "eft gonderimi" not in mas or "gec" in mas or "fast" in mas:
                    continue
                if channel == "MOBIL" and "internet/mobil" not in mas:
                    continue
                if channel == "SUBE" and "sube/diger" not in mas:
                    continue
            elif service == "HAVALE":
                if "havale gonderimi" not in mas:
                    continue
                if channel == "MOBIL" and "internet/mobil" not in mas:
                    continue
                if channel == "SUBE" and "sube/diger" not in mas:
                    continue
            else:
                continue
            candidates.append(row)
            continue
 
        if bank == "AKBANK" and service == "EFT" and channel == "MOBIL":
            if (
                "akbank mobil'den" in mas
                and "internet'ten" in mas
                and "eft" in mas
                and "gec eft" not in mas
            ):
                candidates.append(row)
            continue
 
        # Diğer durumlarda genel yüksek-güvenli skoru kullan; geç/paket ürünleri alma.
        if any(x in mas for x in ("gec eft", "paket", "kota")):
            continue
        base_spec = RowSpec("audit", service, band_key)
        score = _candidate_score(row, base_spec, channel)
        if score > -10_000:
            candidates.append(row)
 
    if not candidates:
        return None
 
    # Aynı ücret tekrarları varsa ilkini al; farklı ücretli birden fazla kesin aday
    # kalırsa otomatik seçim yapma.
    sigs = {_row_fee_signature(r) for r in candidates}
    if len(sigs) > 1 and bank not in {"YAPIKREDI", "AKBANK"}:
        return None
    return candidates[0]
 
def _audit_regular_transfer(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service not in {"DUZENLI_EFT", "DUZENLI_HAVALE"}:
        return None
 
    # Yapı Kredi: Mobil/İnternet düzenli talimatta normal dijital tarife
    # uygulanıyor. Şube için ayrı düzenli tarife yayımlanmadığından normal şube
    # ücretini düzenli ödeme ücreti gibi değil, yalnız açık bir referans olarak
    # gösteriyoruz.
    if bank == "YAPIKREDI":
        base_service = "EFT" if spec.service == "DUZENLI_EFT" else "HAVALE"
        row = _audit_normal_transfer_row(rows, bank, base_service, spec.band_key or "", wanted_channel)
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        fee = _fee_text(row, None)
        if wanted_channel == "MOBIL":
            value = f"Normal Mobil/İnternet tarifesi:\n{fee}"
            return (value, row, "NUMERIC")
 
        transfer_name = "EFT" if spec.service == "DUZENLI_EFT" else "Havale"
        value = (
            f"Normal Şube tarifesi (referans):\n{fee}\n"
            f"Düzenli {transfer_name} için Şube tarifesi ayrıca yayımlanmıyor"
        )
        return (value, row, "REFERENCE_TARIFF")
 
    # Akbank şubeden düzenli EFT talimatında kendi açıklamasına göre Mobil/İnternet
    # ücretleri uygulanır. Normal şube EFT tarifesini kullanmak yanlıştı.
    if bank == "AKBANK" and spec.service == "DUZENLI_EFT" and wanted_channel == "SUBE":
        row = _audit_normal_transfer_row(rows, bank, "EFT", spec.band_key or "", "MOBIL")
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (_fee_text(row, None), row, "NUMERIC")
 
    return None
 
 
def _audit_akbank_havale(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Akbank normal Havale'de cep telefonu/ATM transferini ana havale diye seçme."""
    if bank != "AKBANK" or spec.service != "HAVALE" or not spec.band_key:
        return None
 
    def pred(row: FeeRow) -> bool:
        if _band_key(_parse_band(row.masraf)) != spec.band_key:
            return False
        mas = _norm(row.masraf)
        if wanted_channel == "MOBIL":
            return (
                "akbank mobil'den" in mas
                and "hesaptan hesaba / isme tl ve yp havale" in mas
                and "cep telefonu numarasi" not in mas
            )
        if wanted_channel == "SUBE":
            return (
                "hesaptan hesaba / isme tl ve yp havale-subeden" in mas
                and "ileri vadeli" not in mas
            )
        return False
 
    candidates = _audit_rows(rows, bank, pred)
    if not candidates:
        return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
    row = candidates[0]
    return (_fee_text(row, None), row, "NUMERIC")
 
 
def _audit_backfill_primary_source(
    rows: Sequence[FeeRow], bank: str, supplemental_row: FeeRow,
) -> Optional[FeeRow]:
    """Supplemental satırın açıkça belirttiği primary kaynak masrafını geri bulur."""
    raw = _clean(supplemental_row.aciklama)
    match = re.search(
        r"ana resm[îi] ücret tablosundaki '(.+)' satırından",
        raw,
        flags=re.I,
    )
    if not match:
        return None
    wanted = _norm(match.group(1))
    for row in rows:
        if row.banka != bank:
            continue
        if _norm(row.masraf) == wanted and _has_numeric_fee(row):
            return row
    return None
 
 
def _audit_normal_kasa(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Normal Büyük/Orta/Küçük kasa satırında Özel/Süper status'unu seçme."""
    if spec.service != "KASA" or spec.detail not in {"BUYUK", "ORTA", "KUCUK"}:
        return None
    if bank != "AKBANK":
        return None
 
    token = {"BUYUK": "buyuk boy", "ORTA": "orta boy", "KUCUK": "kucuk boy"}[spec.detail]
    candidates = _audit_rows(
        rows, bank,
        lambda r: (
            token in _norm(r.masraf)
            and "kiralik kasa" in _norm(r.masraf)
            and "ozel" not in _norm(r.masraf)
            and "depozito" not in _norm(r.masraf)
            and "hesaptan odeme" in _norm(r.masraf)
        ),
    )
    if not candidates:
        return (_source_gap_text(spec, bank, "GENEL"), None, "SOURCE_GAP")
    row = candidates[0]
    annual = _fee_value_compact(row)
    dep = _extract_deposit_from_description(row)
    value = f"Yıllık: {annual}"
    tax = _bsmv_label(row)
    if tax:
        value += f" ({tax})"
    if dep:
        value += f"\nDepozito: {dep} (BSMV hariç)"
    return (value, row, "NUMERIC")
 
 
 
def _audit_aidat(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "AIDAT":
        return None
 
    def first(pred) -> Optional[FeeRow]:
        candidates = _audit_rows(rows, bank, pred)
        return candidates[0] if candidates else None
 
    def range_text(row: Optional[FeeRow]) -> str:
        if row is None:
            return ""
        lo = _display_amount(row.asgari_tutar)
        hi = _display_amount(row.azami_tutar)
        if lo and hi:
            return lo if _norm(lo) == _norm(hi) else f"{lo} – {hi}"
        return hi or lo
 
    # GARANTİ: Resmî Fatura/Kurum açıklaması aidat/site tahsilatını açıkça kapsıyor.
    if bank == "GARANTİ":
        if wanted_channel == "MOBIL":
            account = first(
                lambda r: (
                    "hesaptan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "aidat" in _norm(r.aciklama)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            card = first(
                lambda r: (
                    "kredi kartindan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if account is None and card is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
            first_row = account or card
            lines: List[str] = []
 
            if account is not None:
                amount = range_text(account)
                rate = _percent(account.azami_oran) or _percent(account.asgari_oran)
                line = f"Hesaptan: {amount}"
                if rate:
                    line += f" / azami {rate}"
                lines.extend([line, "Kuruma/işleme göre değişir."])
 
            if card is not None:
                fixed = _display_amount(card.asgari_tutar)
                rate = _percent(card.azami_oran) or _percent(card.asgari_oran)
                if fixed and rate:
                    lines.append(f"Kredi kartından: ≤149,99 TRY: {fixed}")
                    lines.append(f"150 TRY+: {rate}")
 
            tax = _bsmv_label(first_row)
            if tax:
                lines.append(tax)
 
            return ("\n".join(lines), first_row, "NUMERIC")
 
        row = first(
            lambda r: (
                "nakit fatura/kurum odemesi" in _norm(r.masraf)
                and "sube" in _norm(r.masraf)
                and "aidat" in _norm(r.aciklama)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
        lines = [
            f"Nakit/Şube: {range_text(row)}",
            "Kuruma/işleme göre değişir.",
        ]
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), row, "NUMERIC")
 
    # İŞ BANKASI: hizmet doğrulanıyor, aidata özel sayısal ücret yayımlanmıyor.
    if bank == "İŞBANKASI":
        status = next(
            (
                r for r in rows
                if r.banka == bank
                and "apartman / site aidat odemeleri" in _norm(r.masraf)
            ),
            None,
        )
        if status is None:
            status = (
                _audit_exact_status(rows, bank, "AIDAT", "GENEL")
                or _audit_exact_status(rows, bank, "AIDAT", wanted_channel)
            )
        return (
            "Hizmet var / ayrı tarife yayımlanmıyor.\n"
            "Aidata özel ücret yayımlanmıyor.",
            status,
            "STATUS",
        )
 
    # AKBANK: Mobil hizmet var. 0,50–55 TRY genel Fatura/Kurum tarifesidir.
    if bank == "AKBANK":
        if wanted_channel == "SUBE":
            return (
                "Bu kanal için ayrı tarife yayımlanmıyor.\n"
                "Aidata özel ücret yayımlanmıyor.",
                None,
                "PUBLICATION_STATUS",
            )
 
        status = next(
            (
                r for r in rows
                if r.banka == bank
                and "aidat odemeleri - mobil/internet" in _norm(r.masraf)
            ),
            None,
        ) or _audit_exact_status(rows, bank, "AIDAT", "MOBIL")
        generic = first(
            lambda r: (
                "akbank mobil'den fatura / kurum tahsilati" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if generic is None:
            return (
                "Hizmet var / ayrı tarife yayımlanmıyor.\n"
                "Aidata özel ücret yayımlanmıyor.",
                status,
                "STATUS",
            )
 
        lines = [
            range_text(generic),
            "Genel Fatura/Kurum tarifesinden eşleştirilmiştir.",
            "Aidata özel ücret yayımlanmıyor.",
        ]
        tax = _bsmv_label(generic)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), generic, "GENERIC_TARIFF")
 
    # YAPI KREDİ: Mobil aidat hizmeti var; 0,63–45 TRY genel kurum tarifesidir.
    if bank == "YAPIKREDI":
        if wanted_channel == "SUBE":
            return (
                "Bu kanal için ayrı tarife yayımlanmıyor.\n"
                "Aidata özel ücret yayımlanmıyor.",
                None,
                "PUBLICATION_STATUS",
            )
 
        status = next(
            (
                r for r in rows
                if r.banka == bank
                and "aidat odemeleri - mobil/internet" in _norm(r.masraf)
            ),
            None,
        ) or _audit_exact_status(rows, bank, "AIDAT", "MOBIL")
        generic = first(
            lambda r: (
                "fatura ve anlasmali kurum odemeleri" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if generic is None:
            return (
                "Hizmet var / ayrı tarife yayımlanmıyor.\n"
                "Aidata özel ücret yayımlanmıyor.",
                status,
                "STATUS",
            )
 
        lines = [
            range_text(generic),
            "Genel Fatura/Kurum tarifesinden eşleştirilmiştir.",
            "Aidata özel ücret yayımlanmıyor.",
        ]
        tax = _bsmv_label(generic)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), generic, "GENERIC_TARIFF")
 
    return None
 
def _audit_school(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "OZEL_OKUL":
        return None
 
    allowed = {
        "GARANTİ": {"SUBE"},
        "İŞBANKASI": {"SUBE"},       # Bankamatik de var; karşılaştırmada ATM sütunu yok.
        "AKBANK": {"MOBIL"},
        "YAPIKREDI": {"MOBIL"},
    }
 
    if wanted_channel not in allowed.get(bank, set()):
        return (
            "Bu kanalda hizmet sunulmuyor.",
            None,
            "NOT_APPLICABLE",
        )
 
    # Hizmetin ilgili kanalda gerçekten bulunduğunu supplemental resmî kaynakla doğrula.
    def school_status_name_match(r: FeeRow) -> bool:
        mas = _norm(r.masraf)
        if bank == "GARANTİ" and wanted_channel == "SUBE":
            return "ozel okul odemeleri" in mas and "sube" in mas
        if bank == "İŞBANKASI" and wanted_channel == "SUBE":
            return "ozel okul odemeleri" in mas and "sube" in mas
        if bank == "AKBANK" and wanted_channel == "MOBIL":
            return "ozel okul / egitim odemeleri" in mas and "mobil" in mas
        if bank == "YAPIKREDI" and wanted_channel == "MOBIL":
            return "ozel okul / okul taksiti" in mas and "mobil" in mas
        return False
 
    candidates = _audit_rows(
        rows,
        bank,
        lambda r: (
            (
                "service=ozel_okul" in _norm(r.aciklama)
                and wanted_channel in _channels(r)
                and _status_kind(r) in {"AVAILABLE", "OFFICIAL_FEE"}
            )
            or school_status_name_match(r)
        ),
        numeric_only=False,
    )
    status_row = candidates[0] if candidates else None
 
    # Garanti Şube: özel okul hizmeti var; 0–50 TRY yalnız genel Nakit
    # Fatura/Kurum tarifesidir ve özel okul komisyonu gibi gösterilemez.
    if bank == "GARANTİ" and wanted_channel == "SUBE":
        if status_row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        return (
            "Hizmet var.\nÖzel okul ücreti yayımlanmıyor.",
            status_row,
            "STATUS",
        )
 
    if status_row is None:
        return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
    return (
        "Hizmet var.\nÖzel okul ücreti yayımlanmıyor.",
        status_row,
        "STATUS",
    )
 
def _audit_phone(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "TELEFON":
        return None
 
    def first(pred, numeric_only: bool = True) -> Optional[FeeRow]:
        candidates = _audit_rows(rows, bank, pred, numeric_only=numeric_only)
        return candidates[0] if candidates else None
 
    def range_text(row: Optional[FeeRow]) -> str:
        if row is None:
            return ""
        lo = _display_amount(row.asgari_tutar)
        hi = _display_amount(row.azami_tutar)
        if lo and hi:
            return lo if _norm(lo) == _norm(hi) else f"{lo} – {hi}"
        return hi or lo
 
    # GARANTİ: telefon/cep telefonu faturaları genel Fatura Ödeme hizmetinin
    # parçasıdır; ilgili kanalın resmî Fatura/Kurum tarifesi gösterilir.
    if bank == "GARANTİ":
        if wanted_channel == "MOBIL":
            row = first(
                lambda r: (
                    "hesaptan fatura/kurum odemesi" in _norm(r.masraf)
                    and "mobil" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            rate = _percent(row.azami_oran) or _percent(row.asgari_oran)
            lines = [range_text(row)]
            if rate:
                lines[0] += f" / azami {rate}"
            lines.append("Kuruma/işleme göre değişir.")
            tax = _bsmv_label(row)
            if tax:
                lines.append(tax)
            return ("\n".join(lines), row, "NUMERIC")
 
        row = first(
            lambda r: (
                "nakit fatura/kurum odemesi" in _norm(r.masraf)
                and "sube" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        lines = [f"Nakit/Şube: {range_text(row)}", "Kuruma/işleme göre değişir."]
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), row, "NUMERIC")
 
    # İŞ BANKASI: telefon faturaları Fatura Ödemeleri kapsamında.
    if bank == "İŞBANKASI":
        if wanted_channel == "MOBIL":
            row = first(
                lambda r: (
                    _norm(r.masraf) == "fatura odemeleri"
                    and "internet sube, iscep, cozum merkezi" in _norm(r.kategori)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
        else:
            row = first(
                lambda r: (
                    _norm(r.masraf) == "fatura odemeleri"
                    and _norm(r.kategori).endswith(" - sube")
                    and "internet sube" not in _norm(r.kategori)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        lines = [range_text(row), "Kuruma/işleme göre değişir."]
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), row, "NUMERIC")
 
    # AKBANK: Mobil ve Gişe için yayımlanan genel Fatura/Kurum tarifeleri.
    if bank == "AKBANK":
        if wanted_channel == "MOBIL":
            row = first(
                lambda r: (
                    "akbank mobil'den fatura / kurum tahsilati" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
        else:
            row = first(
                lambda r: (
                    "giseden fatura / kurum tahsilati" in _norm(r.masraf)
                    and "ek kaynak" not in _norm(r.kategori)
                )
            )
        if row is None:
            return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
        lines = [
            range_text(row),
            "Genel Fatura/Kurum tarifesi uygulanır.",
            "Kuruma/işleme göre değişir.",
        ]
        tax = _bsmv_label(row)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), row, "GENERIC_TARIFF")
 
    # YAPI KREDİ:
    # Mobil/İnternet'te vadesiz hesaptan fatura ödemesi ücretsiz.
    # Kredi kartından anlaşmalı kurum fatura tarifesi ayrıca yayımlanıyor.
    if bank == "YAPIKREDI":
        generic = first(
            lambda r: (
                "fatura ve anlasmali kurum odemeleri" in _norm(r.masraf)
                and "ek kaynak" not in _norm(r.kategori)
            )
        )
 
        if wanted_channel == "MOBIL":
            free_status = first(
                lambda r: (
                    "hesaptan fatura / kurum odemesi - mobil/internet" in _norm(r.masraf)
                    and "ucretsiz" in _norm(r.aciklama)
                ),
                numeric_only=False,
            )
            card_low = first(
                lambda r: (
                    "anlasmali kurum fatura /sgk prim odemeleri" in _norm(r.masraf)
                    and "0 - 150 tl" in _norm(r.masraf)
                    and "fatura odemeleri" in _norm(r.masraf)
                )
            )
            card_high = first(
                lambda r: (
                    "anlasmali kurum fatura /sgk prim odemeleri" in _norm(r.masraf)
                    and "150,01 tl ve uzeri" in _norm(r.masraf)
                    and "fatura odemeleri" in _norm(r.masraf)
                )
            )
 
            if free_status is None and card_low is None and card_high is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
            lines = ["Hesaptan: Ücretsiz / 0 TRY."]
            if card_low is not None:
                low_fee = _display_amount(card_low.azami_tutar) or _display_amount(card_low.asgari_tutar)
                if low_fee:
                    lines.append(f"Kredi kartından: 0–150 TRY: {low_fee}")
            if card_high is not None:
                high_rate = _percent(card_high.azami_oran) or _percent(card_high.asgari_oran)
                if high_rate:
                    lines.append(f"150,01 TRY+: {high_rate}")
            card_tax = _bsmv_label(card_low or card_high)
            if card_tax:
                lines.append(card_tax)
            return ("\n".join(lines), card_low or card_high or free_status, "NUMERIC")
 
        # Şube için 0,63–45 TRY yalnız genel tarife olarak gösterilir;
        # kaynak kanal/ödeme aracını ayırmıyor.
        if generic is None:
            return (
                "Bu kanal için ayrı tarife yayımlanmıyor.",
                None,
                "PUBLICATION_STATUS",
            )
        lines = [
            range_text(generic),
            "Genel Fatura/Kurum tarifesinden eşleştirilmiştir.",
            "Kanal belirtilmemiş.",
            "Ödeme aracı ayrımı yayımlanmıyor.",
        ]
        tax = _bsmv_label(generic)
        if tax:
            lines.append(tax)
        return ("\n".join(lines), generic, "GENERIC_TARIFF")
 
    return None
 
def _audit_vergi(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    if spec.service != "VERGI":
        return None
 
    # "Vergi Tahsilat Komisyonu" diye yeni ürün uydurulmaz. Yalnız açıkça
    # VERGI hizmet statüsü veya doğrudan "Vergi Ödemeleri" numeric satırı kullanılır.
    candidates = _audit_rows(
        rows, bank,
        lambda r: (
            (
                "service=vergi" in _norm(r.aciklama)
                or "vergi odemeleri" in _norm(r.masraf)
            )
            and wanted_channel in _channels(r)
            and "fatura/vergi/sgk" not in _norm(r.masraf)
        ),
    )
    if candidates:
        row = candidates[0]
        return (_audit_institution_range(row), row, "NUMERIC")
 
    status = _audit_exact_status(rows, bank, "VERGI", wanted_channel)
    if status is not None:
        return (
            _audit_status_text(status, "Hizmet var\nAyrı vergi ücreti yayımlanmıyor"),
            status,
            "STATUS",
        )
 
    if bank == "GARANTİ" and wanted_channel == "SUBE":
        return (
            "Şube için ayrı vergi ödeme tarifesi yayımlanmıyor",
            None,
            "PUBLICATION_STATUS",
        )
 
    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
 
_DOCUMENT_STATEMENT_SERVICES = {
    "DOC_VADESIZ_TL",
    "DOC_VADESIZ_YP",
    "DOC_KMH_BIREYSEL",
    "DOC_KK_GECMIS",
    "DOC_ARSIV_GENEL",
    "DOC_ESKI_DEKONT_EKSTRE",
    "DOC_AYLIK_POSTA",
    "DOC_ARSIV_SUBE",
    "DOC_ARSIV_BANKA",
    "DOC_KMH_TICARI",
    "DOC_TICARI_KART_GECMIS",
    "DOC_YATIRIM_EKSTRE",
    "DOC_TICARI_AYLIK",
    "DOC_TICARI_KART_BASILI",
    "DOC_TICARI_KART_GONDERIM",
    "DOC_TICARI_HESAP_BASILI",
    "DOC_POS_EKSTRE",
}
 
 
def _audit_document_statement_research(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Belge/ekstre/araştırma ailesini yalnız tam ürün adlarıyla çözer.
 
    Bu bölümde genel ``ARSIV`` / ``HESAP_OZETI`` tag'leri bilinçli olarak
    kullanılmaz. Bir bankadaki bireysel kart ekstresi başka bankanın ticari
    kart tarifesine veya basılı ekstre ücreti mevduat ekstresine taşınamaz.
    """
    if spec.service not in _DOCUMENT_STATEMENT_SERVICES:
        return None
 
    def clean_amount(value: str) -> str:
        # Yapı Kredi kaynağındaki ``45 TL -`` / ``- -`` gösterimini güvenli
        # biçimde temizle; para birimini veya sayıyı tahmin etme.
        parts = [part for part in _clean(value).split() if part != "-"]
        return _display_amount(" ".join(parts)) if parts else ""
 
    def compact(row: FeeRow) -> str:
        low = clean_amount(row.asgari_tutar)
        high = clean_amount(row.azami_tutar)
        if low and high:
            return low if _norm(low) == _norm(high) else f"{low} - {high}"
        return high or low
 
    def exact_rows(
        masraf: str,
        *,
        kategori: str = "",
        numeric_only: bool = True,
    ) -> List[FeeRow]:
        wanted_fee = _norm(masraf)
        wanted_category = _norm(kategori)
        return _primary_first(_audit_rows(
            rows,
            bank,
            lambda row: (
                _norm(row.masraf) == wanted_fee
                and (not wanted_category or wanted_category in _norm(row.kategori))
            ),
            numeric_only=numeric_only,
        ))
 
    def exact(
        masraf: str,
        *,
        kategori: str = "",
        numeric_only: bool = True,
    ) -> Optional[FeeRow]:
        found = exact_rows(
            masraf,
            kategori=kategori,
            numeric_only=numeric_only,
        )
        return found[0] if found else None
 
    def published_gap(text: str = "Karşılaştırılabilir ayrı tarife yayımlanmıyor"):
        return (text, None, "PUBLICATION_STATUS")
 
    service = spec.service
 
    # ------------------------------------------------------------------
    # VADESİZ MEVDUAT EKSTRESİ: İş Bankası Bankamatik ve Şube tarifeleri
    # açıkça ayrılır. İlk banka alt sütunu ``Mobil`` olsa da hücre içinde
    # ``Bankamatik`` etiketi yazılarak kanal uydurulmaz.
    # ------------------------------------------------------------------
    if service in {"DOC_VADESIZ_TL", "DOC_VADESIZ_YP"}:
        if bank == "GARANTİ":
            masraf = (
                "TL Hesaplar - Vadesiz TL Hesap Ekstresi- Gerçek Kişiler"
                if service == "DOC_VADESIZ_TL"
                else "YP Hesaplar - Vadesiz YP Hesap Ekstresi - Gerçek Kişiler"
            )
            row = exact(masraf, kategori="Ekstre Ücreti")
            if wanted_channel == "SUBE":
                return published_gap("Şube için ayrı tarife yayımlanmıyor")
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            return (
                "Genel / kanal belirtilmemiş:\n"
                f"{compact(row)}\nTalep edilen ekstre başına\nBSMV dahil",
                row,
                "NUMERIC",
            )
 
        if bank == "İŞBANKASI":
            if wanted_channel == "MOBIL":
                row = exact(
                    "Ekstre - TP/YP",
                    kategori="Mevduat Hesapları - Ekstre - Bankamatik",
                    numeric_only=False,
                )
                if row is None:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                return ("Bankamatik (Mobil değil):\nÜcretsiz / 0 TRY", row, "NUMERIC")
 
            row = exact(
                "Ekstre - TP/YP",
                kategori="Mevduat Hesapları - Ekstre - Şube",
            )
            if row is None:
                return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
            return (
                f"Şube:\n{compact(row)}\nSayfa başına\nBSMV hariç",
                row,
                "NUMERIC",
            )
 
        return published_gap()
 
    if service == "DOC_KMH_BIREYSEL":
        if bank == "GARANTİ":
            row = exact("TL Hesaplar - KMH Ekstresi", kategori="Ekstre Ücreti")
            if row:
                return (
                    f"{compact(row)}\nKullanılan aylarda aylık\nBSMV dahil",
                    row,
                    "NUMERIC",
                )
        elif bank == "YAPIKREDI":
            row = exact("Ekstre Masrafı - KMH Hesap Özeti (Esnek Hesap Özeti)")
            if row:
                return (
                    f"{compact(row)} / hesap\nBorç devam eden aylarda\nBSMV alınmıyor",
                    row,
                    "NUMERIC",
                )
        return published_gap()
 
    if service == "DOC_KK_GECMIS":
        if bank == "GARANTİ":
            row = exact(
                "Kredi Kartı Geçmiş Ekstre Ücreti",
                kategori="Şubeden Hesap Ekstresi",
            )
            if row:
                return (
                    f"1 yıldan eski: azami {compact(row)}\nBSMV dahil",
                    row,
                    "NUMERIC",
                )
        elif bank == "İŞBANKASI":
            row = exact("Geçmiş Dönem Kredi Kartı Hesap Özeti Ücreti-Türkiye Kredi Kartları")
            if row:
                return (
                    f"1 yıldan eski: {compact(row)} / hesap özeti\nBSMV hariç",
                    row,
                    "NUMERIC",
                )
        elif bank == "YAPIKREDI":
            row = exact(
                "Diğer Ücretler - Geçmiş Dönem Ekstre Arşiv Araştırma Ücreti (kart hamilinin talebinin olması halinde)"
            )
            if row:
                return (f"azami {compact(row)}\nBSMV dahil", row, "NUMERIC")
        return published_gap()
 
    if service == "DOC_ARSIV_GENEL":
        if bank == "GARANTİ":
            row = exact(
                "Belge ve Bilgilendirme Ücreti - Arşiv Araştırma Ücreti",
                kategori="Belge ve Bilgilendirme",
            )
            if row:
                amounts = [
                    _display_amount(f"{amount} TRY")
                    for amount in re.findall(
                        r"([0-9]+(?:[.,][0-9]+)?)\s*tl",
                        _norm(row.aciklama),
                        flags=re.I,
                    )
                ]
                if len(amounts) < 2:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                older = amounts[0]
                oldest = amounts[1]
                return (
                    "Son 1 yıl: Ücretsiz\n"
                    f"1–3 yıl: {older}\n"
                    f"3 yıldan eski: {oldest}\n"
                    "BSMV hariç",
                    row,
                    "NUMERIC",
                )
 
        elif bank == "İŞBANKASI":
            document = exact(
                "Sözleşme, dekont vb. doküman talebi",
                kategori="Mevduat Hesapları - Özel Rapor Ücretleri",
            )
            notice = exact("Geçmiş Dönem Bankacılık İşlemleri Bildirimi")
            if document and notice:
                return (
                    f"Belge talebi (>1 yıl): {compact(document)}\n"
                    f"İşlem bildirimi (>1 yıl): {compact(notice)}\n"
                    "BSMV hariç",
                    document,
                    "NUMERIC",
                )
        return published_gap()
 
    if service == "DOC_ESKI_DEKONT_EKSTRE":
        if bank == "AKBANK":
            row = exact("1 Yıldan Eski İşlemlere Ait Geçmişe Yönelik Dekont-Ekstre Verilmesi")
            if row:
                per_page = re.search(
                    r"sayfa basina\s*([0-9]+(?:[.,][0-9]+)?)\s*tl",
                    _norm(row.aciklama),
                    flags=re.I,
                )
                page_text = (
                    _display_amount(f"{per_page.group(1)} TRY")
                    if per_page else ""
                )
                value = f">1 yıl: {compact(row)}"
                if page_text:
                    value += f"\n{page_text} / sayfa"
                value += "\nBSMV hariç"
                return (value, row, "NUMERIC")
 
        elif bank == "YAPIKREDI":
            receipt = exact("Dekont Masrafı - Geçmişe Yönelik Olarak Basılan")
            statements = exact_rows("Ekstre Masrafı")
            if receipt and len(statements) >= 2:
                statement_fees: List[str] = []
                for row in statements:
                    fee = compact(row)
                    if fee and fee not in statement_fees:
                        statement_fees.append(fee)
                value = f"Dekont (>1 yıl): {compact(receipt)}"
                if statement_fees:
                    value += "\nEkstre (>1 yıl), yayımlanan tarifeler: " + " | ".join(statement_fees)
                value += "\nSayfa başına; en fazla 10 sayfa\nBSMV hariç"
                return (value, receipt, "NUMERIC")
        return published_gap()
 
    if service == "DOC_AYLIK_POSTA":
        if bank == "AKBANK":
            row = exact("Posta İle Aylık Hesap Özeti Gönderimi")
            if row:
                return (f"{compact(row)}\nBSMV muaf", row, "NUMERIC")
        return published_gap()
 
    if service in {"DOC_ARSIV_SUBE", "DOC_ARSIV_BANKA"}:
        if bank != "İŞBANKASI":
            return published_gap()
        masraf = (
            "Merkezden / Şubeden Şube Bazında Yapılan - TP/YP"
            if service == "DOC_ARSIV_SUBE"
            else "Merkezden Banka Bazında Yapılan - TP/YP"
        )
        row = exact(masraf)
        if row:
            label = "Şube bazında" if service == "DOC_ARSIV_SUBE" else "Banka bazında"
            return (f"{label}: {compact(row)}\nBSMV hariç", row, "NUMERIC")
        return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
 
    if service == "DOC_KMH_TICARI":
        if bank == "GARANTİ":
            row = exact(
                "Avans Hesap (Kredili Mevduat Hesabı) - Belge ve Bilgilendirme Ücreti (Ekstre Ücreti)",
                kategori="Firma Kullanım Amaçlı",
                numeric_only=False,
            )
            if row:
                return (
                    "Gönderim maliyeti kadar\nTutar değişken; sabit ücret yayımlanmıyor",
                    row,
                    "STATUS",
                )
        elif bank == "YAPIKREDI":
            row = exact(
                "Esnek Ticari Hesap - Faiz Oranı - Basılı Ticari KMH Hesap Özeti (Esnek Ticari Hesap Özeti Ücreti)"
            )
            if row:
                return (f"{compact(row)} / sayfa", row, "NUMERIC")
        return published_gap()
 
    if service == "DOC_TICARI_KART_GECMIS":
        if bank == "YAPIKREDI":
            row = exact("Diğer Ücretler - Ticari Kart Geçmiş Dönem Ekstre Ücreti")
            if row:
                return (f"{compact(row)} / sayfa\nBSMV hariç", row, "NUMERIC")
        return published_gap()
 
    if service == "DOC_YATIRIM_EKSTRE":
        if bank == "İŞBANKASI":
            row = exact(
                "Ekstre - TP",
                kategori="Mevduat Hesapları - Ekstre - Yatırım Hesabı Ekstre",
            )
            if row:
                return (f"{compact(row)} / gönderi\nBSMV hariç", row, "NUMERIC")
        return published_gap()
 
    if service == "DOC_TICARI_AYLIK":
        if bank == "AKBANK":
            row = exact("Aylık Hesap Özeti (Ticari Müşteri)")
            if row:
                return (f"asgari {compact(row)} / hesap\nBSMV dahil", row, "NUMERIC")
        return published_gap()
 
    if service == "DOC_TICARI_KART_BASILI":
        if bank == "GARANTİ":
            row = exact("Basılı Ekstre Ücreti", kategori="Firma Kullanım Amaçlı")
            if row:
                return (f"azami {compact(row)} / sayfa\nBSMV hariç", row, "NUMERIC")
        return published_gap()
 
    if service == "DOC_TICARI_KART_GONDERIM":
        if bank == "GARANTİ":
            row = exact(
                "Diğer Ücretler - Ekstre Gönderim Ücreti",
                kategori="Firma Kullanım Amaçlı",
            )
            if row:
                return (f"azami {compact(row)} / sayfa\nBSMV hariç", row, "NUMERIC")
        return published_gap()
 
    if service == "DOC_TICARI_HESAP_BASILI":
        if bank == "AKBANK":
            row = exact("4.6.1 Basılı Ekstre Gönderimi (2)", kategori="EK KAYNAK - Akbank Ticari")
            if row:
                return (
                    f"{compact(row)} / sayfa\nPosta/kurye maliyeti ilave\nBSMV hariç",
                    row,
                    "NUMERIC",
                )
        return published_gap()
 
    if service == "DOC_POS_EKSTRE":
        if bank == "GARANTİ":
            row = exact("Ekstre Gönderim Ücreti", kategori="Üye İşyeri Ücretleri")
            if row:
                amounts = [
                    _display_amount(f"{amount} TRY")
                    for amount in re.findall(
                        r"([0-9]+(?:[.,][0-9]+)?)\s*tl",
                        _norm(row.aciklama),
                        flags=re.I,
                    )
                ]
                if len(amounts) < 2:
                    return (_source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP")
                first = amounts[0]
                second = amounts[1]
                value = f"0–20 g: {first}"
                value += f"\n20–50 g: {second}"
                value += "\nBSMV dahil"
                return (value, row, "NUMERIC")
        return published_gap()
 
    return published_gap()
 
 
_RESEARCH_SECURITIES_SERVICES = {
    "DOC_DEPOSIT_RESEARCH",
    "SEC_DOMESTIC_STOCK_TRADE",
    "SEC_DOMESTIC_STOCK_TRANSFER",
    "SEC_DOMESTIC_BOND_TRANSFER",
    "SEC_DOMESTIC_CUSTODY",
    "SEC_FOREIGN_STOCK_TRADE",
    "SEC_FOREIGN_STOCK_CUSTODY",
    "SEC_EUROBOND_TRANSFER",
    "SEC_EUROBOND_CUSTODY",
}
 
 
def _audit_research_and_securities(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Mevduat araştırma ve menkul kıymet ailelerini tam adla çözer.
 
    Menkul kıymetlerde bankaların ürün ve dönem kırılımları aynı değildir.
    Bu nedenle tek bir genel oran seçilmez; karşılaştırılabilir alt hizmetler
    ayrı satırlarda, kaynağın kendi kapsam etiketiyle gösterilir.
    """
    service = spec.service
    if service not in _RESEARCH_SECURITIES_SERVICES:
        return None
 
    def found(predicate, *, numeric_only: bool = False) -> List[FeeRow]:
        return _primary_first(_audit_rows(
            rows,
            bank,
            predicate,
            numeric_only=numeric_only,
        ))
 
    def exact(masraf: str, *, kategori: str = "", numeric_only: bool = False) -> Optional[FeeRow]:
        target_fee = _norm(masraf)
        target_cat = _norm(kategori)
        matches = found(
            lambda row: (
                _norm(row.masraf) == target_fee
                and (not target_cat or target_cat in _norm(row.kategori))
            ),
            numeric_only=numeric_only,
        )
        return matches[0] if matches else None
 
    def published_gap() -> Tuple[str, Optional[FeeRow], str]:
        return (
            "Karşılaştırılabilir ayrı tarife yayımlanmıyor",
            None,
            "PUBLICATION_STATUS",
        )
 
    def number_text(value: float, decimals: int = 6) -> str:
        raw = f"{value:.{decimals}f}".rstrip("0").rstrip(".")
        return raw.replace(".", ",")
 
    def numeric_rate(value: str) -> Optional[float]:
        raw = _clean(value)
        if not raw or raw in {"-", "0", "0.0", "0.00", "0%", "0 %"}:
            return None
        match = re.search(r"-?[0-9]+(?:[.,][0-9]+)?", raw)
        if not match:
            return None
        try:
            # Oran alanındaki nokta binlik ayırıcı değil ondalıktır
            # (örn. ``0.185 %`` = yüzde 0,185).
            number = float(match.group(0).replace(",", "."))
        except ValueError:
            return None
        if "%" not in raw and 0 < abs(number) < 0.1:
            number *= 100
        return number
 
    def rate_text(value: str) -> str:
        number = numeric_rate(value)
        return f"{number_text(number)}%" if number is not None else ""
 
    def rate_range(items: Sequence[FeeRow]) -> str:
        rates: List[float] = []
        for row in items:
            for attr in ("asgari_oran", "azami_oran"):
                value = numeric_rate(getattr(row, attr, ""))
                if value is not None:
                    rates.append(value)
        if not rates:
            return ""
        low, high = min(rates), max(rates)
        if abs(low - high) < 1e-12:
            return f"{number_text(low)}%"
        return f"{number_text(low)}% - {number_text(high)}%"
 
    # ------------------------------------------------------------------
    # MEVDUAT ARAŞTIRMA
    # ------------------------------------------------------------------
    if service == "DOC_DEPOSIT_RESEARCH":
        matches = found(
            lambda row: "mevduat arastirma" in _norm(row.text),
            numeric_only=False,
        )
        if not matches:
            return published_gap()
        row = matches[0]
        value = _fee_text(row, spec) if _has_numeric_fee(row) else "Ücret tutarı belirtilmemiş"
        return (value, row, "NUMERIC" if _has_numeric_fee(row) else "STATUS")
 
    # ------------------------------------------------------------------
    # YURTİÇİ HİSSE SENEDİ ALIM-SATIM
    # ------------------------------------------------------------------
    if service == "SEC_DOMESTIC_STOCK_TRADE":
        if bank == "GARANTİ":
            row = exact("Hisse Senedi Sabit Komisyon Oranı - Garanti BBVA Mobil")
            if row:
                rates = rate_range([row])
                return (f"Mobil/İnternet: {rates}\nİşlem başına\nBSMV hariç", row, "NUMERIC")
 
        elif bank == "İŞBANKASI":
            matches = found(lambda row: (
                "iscep sen. alis-satis" in _norm(row.kategori)
                and "islem hacmine gore alim-satim komisyonu" in _norm(row.masraf)
            ))
            if matches:
                return (
                    f"İşCep/İnternet: {rate_range(matches)}\n"
                    "Asgari 1 TRY\n3 aylık hacme göre\nBSMV hariç",
                    matches[0],
                    "NUMERIC",
                )
 
        elif bank == "AKBANK":
            matches = found(lambda row: _norm(row.masraf).startswith(
                "hisse senedi komisyonu aylik islem hacmi"
            ))
            if matches:
                return (
                    f"Dijital kanallar: {rate_range(matches)}\n"
                    "Aylık işlem hacmine göre\nBSMV hariç",
                    matches[0],
                    "NUMERIC",
                )
 
        elif bank == "YAPIKREDI":
            row = exact(
                "Hisse Senedi İşlemleri - İşlem Komisyonları - İnternet Bankacılığı / Mobil Bankacılık"
            )
            if row:
                match = re.search(r"%\s*([0-9]+(?:[.,][0-9]+)?)", _clean(row.aciklama))
                if match:
                    return (
                        f"Mobil/İnternet: {number_text(_parse_number(match.group(1)) or 0)}%\n"
                        "Alım-satım işlem tutarı üzerinden",
                        row,
                        "NUMERIC",
                    )
 
        return published_gap()
 
    # ------------------------------------------------------------------
    # YURTİÇİ HİSSE SENEDİ VİRMANI
    # ------------------------------------------------------------------
    if service == "SEC_DOMESTIC_STOCK_TRANSFER":
        if bank == "İŞBANKASI":
            outgoing = exact("Virman İşlemi", kategori="Hisse Senedi - Varant Virman İşlemleri")
            incoming = exact("Virman İşlemi", kategori="Yatırım Kuruluşlarından Bankamıza Hisse Senedi")
            if outgoing and incoming:
                return (
                    f"Dışarı/hesaplar arası: {_display_amount(outgoing.asgari_tutar)}\n"
                    f"Kuruma gelen: {_display_amount(incoming.asgari_tutar) or '0 TRY'}\n"
                    "BSMV hariç",
                    outgoing,
                    "NUMERIC",
                )
 
        elif bank == "AKBANK":
            outside = exact("Kurum dışı hisse senedi virmanı")
            inside = exact("Kurum içi hisse senedi virmanı")
            if outside and inside:
                return (
                    f"Kurum dışı: {_fee_value_compact(outside)} (BSMV hariç)\n"
                    f"Kurum içi: {_fee_value_compact(inside)} (BSMV dahil)",
                    outside,
                    "NUMERIC",
                )
 
        elif bank == "YAPIKREDI":
            row = exact("Hisse Senedi İşlemleri - Kıymet Transferi - Virman")
            if row:
                return (
                    "Transfer tutarı üzerinden\nMKK/Takasbank oranının 5 katı",
                    row,
                    "STATUS",
                )
 
        return published_gap()
 
    # ------------------------------------------------------------------
    # YURTİÇİ BONO/TAHVİL VİRMANI
    # ------------------------------------------------------------------
    if service == "SEC_DOMESTIC_BOND_TRANSFER":
        if bank == "İŞBANKASI":
            matches = found(lambda row: (
                "sb.get.men.kiym.oz.sektor tah.virmani" in _norm(row.kategori)
                and _norm(row.masraf) == "virman islemi"
            ))
            if matches:
                row = matches[0]
                amount = _display_amount(row.asgari_tutar) or _display_amount(row.azami_tutar)
                rate = rate_text(row.asgari_oran or row.azami_oran)
                value = " / ".join(part for part in (amount, rate) if part)
                return (f"{value}\nBSMV hariç", row, "NUMERIC")
 
        elif bank == "AKBANK":
            row = exact("Hazine Bonosu / Devlet Tahvili Virmanı")
            if row:
                return (f"{_fee_value_compact(row)}\nBSMV hariç", row, "NUMERIC")
 
        elif bank == "YAPIKREDI":
            row = exact("Sabit Getirili Menkul Kıymet İşlem Ücretleri - Başka Kuruma Bono/Tahvil Virmanı")
            if row:
                return (f"{_fee_value_compact(row)}\nKıymet başına\nBSMV hariç", row, "NUMERIC")
 
        return published_gap()
 
    # ------------------------------------------------------------------
    # YURTİÇİ YATIRIM HESABI / SAKLAMA
    # ------------------------------------------------------------------
    if service == "SEC_DOMESTIC_CUSTODY":
        if bank == "GARANTİ":
            row = exact("Hisse Senedi Sabit Komisyon Oranı - Hisse Senedi Hesap Bakım Ücreti")
            if row:
                return (
                    f"{_fee_value_compact(row)} / 6 ay\n500 TRY / yıl\nBSMV dahil",
                    row,
                    "NUMERIC",
                )
 
        elif bank == "İŞBANKASI":
            matches = found(lambda row: (
                "yatirim hesabi saklama" in _norm(row.kategori)
                and "altin" not in _norm(row.kategori)
            ), numeric_only=False)
            amounts = [
                _parse_number(_clean(row.asgari_tutar).replace("TL", ""))
                for row in matches
            ]
            amounts = [value for value in amounts if value is not None]
            if matches and amounts:
                low = _display_amount(f"{min(amounts)} TL")
                high = _display_amount(f"{max(amounts)} TL")
                return (
                    f"{low} - {high} / 3 ay\n"
                    "Bakiye kademesine göre\nBSMV hariç",
                    matches[0],
                    "NUMERIC",
                )
 
        elif bank == "AKBANK":
            row = exact("MKK Hesap Bakım Ücreti (günlük)")
            if row:
                match = re.search(r"gunluk\s*([0-9]+(?:[.,][0-9]+)?)\s*tl", _norm(row.aciklama))
                if match:
                    exact_amount = match.group(1).replace(".", ",") + " TRY"
                    return (
                        f"{exact_amount} / gün\n"
                        "Saklama bakiyesi 1.000 TRY ve üzeri\nBSMV hariç",
                        row,
                        "NUMERIC",
                    )
 
        elif bank == "YAPIKREDI":
            row = exact(
                "Merkezi Kayıt Kuruluşu Ücretleri - Yatırımcı Hesabı Açma ve Bakım Hizmetleri - Hesap Bakım"
            )
            if row:
                return (
                    f"{_display_amount(row.asgari_tutar)} / gün\n"
                    "Saklama bakiyesi 1.000 TRY ve üzeri",
                    row,
                    "NUMERIC",
                )
 
        return published_gap()
 
    # ------------------------------------------------------------------
    # YURTDIŞI PAY ALIM-SATIM / SAKLAMA
    # ------------------------------------------------------------------
    if service == "SEC_FOREIGN_STOCK_TRADE":
        if bank == "İŞBANKASI":
            row = exact("Alım Satım Aracılık Hizmeti", kategori="Yurt Dışı Piyasalar - Pay İşlemleri")
            if row:
                return (
                    f"Azami {rate_text(row.azami_oran)}\n"
                    "Asgari komisyon ülke/borsaya göre\nÖrn. NYSE/NASDAQ: 1 USD",
                    row,
                    "NUMERIC",
                )
        return published_gap()
 
    if service == "SEC_FOREIGN_STOCK_CUSTODY":
        if bank == "İŞBANKASI":
            common = exact(
                "Pay (Saklama - Yurt Dışı) (ABD, İngiltere, İtalya, Fransa, Almanya, Avusturya)"
            )
            other = exact("Pay (Saklama - Yurt Dışı) (Diğer Ülkeler)")
            adr = exact("Pay (ADR-GDR Depo Saklama Masrafı)")
            if common and other and adr:
                return (
                    f"Belirtilen ülkeler: {rate_text(common.azami_oran)} / ay\n"
                    f"Diğer ülkeler: {rate_text(other.azami_oran)} / ay\n"
                    "ADR/GDR: 0,03 USD/pay, 6 ayda bir",
                    common,
                    "NUMERIC",
                )
        return published_gap()
 
    # ------------------------------------------------------------------
    # EUROBOND VİRMAN / SAKLAMA
    # ------------------------------------------------------------------
    if service == "SEC_EUROBOND_TRANSFER":
        if bank == "İŞBANKASI":
            row = exact("SWIFT - Eurobond virmanı")
            if row:
                return (
                    "20 USD/EUR\nEk SWIFT: 10 USD/EUR\nBSMV hariç",
                    row,
                    "NUMERIC",
                )
        elif bank == "AKBANK":
            matches = found(lambda row: _norm(row.masraf) == "eurobond virmani")
            if matches:
                return ("20 USD veya 20 EUR\nBSMV hariç", matches[0], "NUMERIC")
        elif bank == "YAPIKREDI":
            row = exact("Sabit Getirili Menkul Kıymet İşlem Ücretleri - Başka Kuruma Eurobond Virmanı")
            if row:
                return (f"{_fee_value_compact(row)}\nKıymet başına\nBSMV hariç", row, "NUMERIC")
        return published_gap()
 
    if service == "SEC_EUROBOND_CUSTODY":
        if bank == "İŞBANKASI":
            row = exact("Eurobond saklama masrafı")
            if row:
                return ("Ücretsiz / 0 USD-EUR\nBSMV hariç", row, "NUMERIC")
        return published_gap()
 
    return None
 
 
def _user_audit_override(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Kullanıcı tarafından tek tek denetlenen yüksek riskli aileler."""
    for resolver in (
        _audit_fatura_methods,
        _audit_akbank_havale,
        _audit_sgk,
        _audit_sans_oyunu,
        _audit_regular_transfer,
        _audit_aidat,
        _audit_school,
        _audit_phone,
        _audit_vergi,
        _audit_document_statement_research,
        _audit_research_and_securities,
    ):
        result = resolver(rows, bank, spec, wanted_channel)
        if result is not None:
            return result
 
    for resolver in (
        _audit_atm,
        _audit_hgs,
        _audit_normal_kasa,
        _audit_isbank_special_kasa,
        _audit_cheque,
        _audit_kkb,
        _audit_senet_tahsil,
    ):
        result = resolver(rows, bank, spec)
        if result is not None:
            return result
 
    return None
 
 
def _aggregate_service_fee(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, FeeRow]]:
    """
    Çok satırlı tarifeleri tek hücrede özetler.
 
    Yurt dışı transferlerde eski "ilk 5 adayı bas" mantığı kullanılmaz.
    Önce banka bazında tek kanonik ürün ailesi seçilir; yalnız o ürünün
    gerçek tutar bantları gösterilir.
    """
    transfer_services = {
        "SWIFT_GELEN",
        "SWIFT_GIDEN",
        "YURT_DISI_FAST",
        "KART_YURTDISI_TRANSFER",
    }
 
    if spec.service in transfer_services:
        candidates = _canonical_transfer_rows(
            rows,
            bank,
            spec,
            wanted_channel,
        )
 
        if not candidates:
            return None
 
        canonical_rows = [row for _, row in candidates]
        value = _compact_transfer_blocks(
            canonical_rows,
            spec,
        )
 
        if not value:
            return None
 
        if spec.service == "SWIFT_GELEN":
            channel_label = (
                "Şube"
                if bank == "GARANTİ"
                else "Genel / kanal belirtilmemiş"
            )
            value = f"{channel_label}:\n{value}"
 
        return value, canonical_rows[0]
 
    # Altın transferinde eski aggregate davranışı korunur; burada aynı
    # ürünün gram/tutar kademeleri gerçekten aynı hizmetin parçasıdır.
    if spec.service != "ALTIN_TRANSFER":
        return None
 
    candidates = []
 
    for row in rows:
        if row.banka != bank:
            continue
 
        score = _candidate_score(row, spec, "GENEL")
        if score <= -10_000 or not _has_numeric_fee(row):
            continue
 
        mas = _norm(row.masraf)
 
        if not any(x in mas for x in (
            "altin transfer",
            "ats ile altin gonderimi",
            "kiymetli maden transferi ucreti - altin",
        )):
            continue
 
        if any(x in mas for x in (
            "western union", "eft", "havale", "fast", "fiziki", "teslim",
        )):
            continue
 
        candidates.append((score, row))
 
    if not candidates:
        return None
 
    candidates.sort(
        key=lambda item: (
            item[0],
            -len(_norm(item[1].masraf)),
        ),
        reverse=True,
    )
 
    blocks: List[str] = []
    seen: Set[str] = set()
    first_row = candidates[0][1]
 
    for _, row in candidates:
        fee = _fee_value_compact(row)
        if not fee:
            continue
 
        hint = _compact_hint(row)
        block = f"{hint}: {fee}" if hint else fee
 
        tax = _bsmv_label(row)
        if tax:
            block += f" ({tax})"
 
        key = _norm(block)
        if key in seen:
            continue
 
        seen.add(key)
        blocks.append(block)
 
        if len(blocks) >= 5:
            break
 
    if not blocks:
        return None
 
    return "\n".join(blocks), first_row
 
 
def _kasa24_special_fee(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec
) -> Optional[Tuple[str, FeeRow]]:
    """Yapı Kredi Özel/Süper satırında Kasa24 A-E tarifesini dürüstçe özetler."""
    if bank != "YAPIKREDI" or spec.service != "KASA" or spec.detail != "OZEL":
        return None
 
    blocks: List[str] = []
    used_rows: List[FeeRow] = []
    first: Optional[FeeRow] = None
 
    for kasa_type in ("A", "B", "C", "D", "E"):
        token = _norm(f"{kasa_type} Tipi Kasa24")
        annual = None
        deposit = None
        for row in rows:
            if row.banka != bank:
                continue
            full = _norm(row.text)
            if "kasa24" not in full or token not in full:
                continue
            if "depozito" in full:
                deposit = row
            elif "yillik" in full or "kasa24 ucreti" in full:
                annual = row
 
        if annual is None:
            continue
        first = first or annual
        used_rows.append(annual)
        annual_fee = _fee_value_compact(annual)
        dep_fee = _fee_value_compact(deposit) if deposit else ""
        if deposit is not None:
            used_rows.append(deposit)
        line = f"Kasa24 {kasa_type}: Yıllık {annual_fee}"
        if dep_fee:
            line += f" | Depozito {dep_fee}"
        blocks.append(line)
 
    if not blocks or first is None:
        return None
    tax = _audit_common_tax(used_rows)
    if tax:
        blocks.append(tax)
    return "\n".join(blocks), first
 
 
def _source_url(row: Optional[FeeRow]) -> str:
    if row is None:
        return ""
    m = re.search(r"Resmî ek kaynak:\s*(https?://[^\s|]+)", row.aciklama or "", flags=re.I)
    if m:
        return m.group(1).rstrip(".,;")
    return PRIMARY_SOURCE_URLS.get(row.banka, "")
 
 
def _cell_comment(row: Optional[FeeRow], resolution: str, value: str) -> Optional[Comment]:
    if row is None:
        return None
    parts = [
        f"Çözüm türü: {resolution}",
        f"Banka: {row.banka}",
        f"Kaynak masraf: {row.masraf}",
    ]
    if row.kategori:
        parts.append(f"Kategori: {row.kategori}")
    if row.guncelleme_tarihi:
        parts.append(f"Kaynak güncelleme: {row.guncelleme_tarihi}")
    url = _source_url(row)
    if url:
        parts.append(f"Kaynak: {url}")
    return Comment("\n".join(parts), "Otomatik Eşleştirme")
 
 
def _source_gap_text(spec: RowSpec, bank: str, wanted_channel: str) -> str:
    # "Kontrol gerekli" = hizmet yok / ücretsiz demek değildir.
    # Yalnız güvenli bir karşılaştırılabilir tarife seçilemediğini gösterir.
    return "Kontrol gerekli"
 
 
def _assert_safe_source(row: Optional[FeeRow], spec: RowSpec) -> None:
    """
    Son güvenlik ağı.
 
    Eşleştirme kurallarında ileride yapılacak bir değişiklik yanlış ürünü
    seçerse final Excel yayınlanmadan pipeline'ı durdurur.
    """
    if row is None:
        return
 
    structural = _norm(f"{row.kategori} | {row.masraf}")
    mas = _norm(row.masraf)
 
    if spec.service == "HGS_ETIKET":
        if not ("hgs" in structural and "etiket" in structural and "kart" not in mas):
            raise RuntimeError(
                f"Güvenlik: HGS Etiket için yanlış kaynak seçildi: {row.masraf}"
            )
 
    elif spec.service == "HGS_KART":
        if not ("hgs" in structural and "kart" in structural and "etiket" not in mas):
            raise RuntimeError(
                f"Güvenlik: HGS Kart için yanlış kaynak seçildi: {row.masraf}"
            )
 
    elif spec.service == "CEK_DEFTERI_YAPRAK":
        if not (
            any(x in structural for x in ("cek defteri", "cek karnesi"))
            and any(x in structural for x in ("yaprak basi", "50-350 yaprakli", "351 yaprak"))
        ):
            raise RuntimeError(
                f"Güvenlik: Çek defteri yaprak-başı için yanlış kaynak: {row.masraf}"
            )
 
    elif spec.service == "CEK_KARNESI_10":
        if not (
            any(x in structural for x in ("cek defteri", "cek karnesi"))
            and any(x in structural for x in ("10 yaprakli", "10'luk", "10luk"))
        ):
            raise RuntimeError(
                f"Güvenlik: 10 yapraklı çek karnesi için yanlış kaynak: {row.masraf}"
            )
 
    elif spec.service == "CEK_DUZENLEME_STANDART":
        forbidden = (
            "hediye", "armagan", "seyahat", "doviz", "dovizi natik",
            "dth", "karsilikli", "odeme", "durdurma",
        )
        if (
            not any(x in structural for x in (
                "bloke cek duzenleme",
                "bloke/ keside ceki duzenleme",
                "bloke/keside ceki duzenleme",
                "keside ceki / bloke cek duzenleme",
                "duzenleme - tp/yp",
            ))
            or any(x in structural for x in forbidden)
        ):
            raise RuntimeError(
                f"Güvenlik: Standart çek düzenleme için yanlış kaynak: {row.masraf}"
            )
 
    elif spec.service == "CEK_DUZENLEME_DOVIZ":
        if not any(x in structural for x in (
            "dovizli cek duzenleme",
            "dovizi natik cek duzenleme",
            "dth'dan cek duzenlenmesi",
            "dth dan cek duzenlenmesi",
        )):
            raise RuntimeError(
                f"Güvenlik: Dövizli çek düzenleme için yanlış kaynak: {row.masraf}"
            )
 
    elif spec.service == "CEK_TAHSIL":
        if any(x in structural for x in (
            "gişeden cek odeme", "giseden cek odeme", "bloke cek odeme",
            "karsiliksiz cek elden odeme", "seyahat ceki odeme",
        )):
            raise RuntimeError(
                f"Güvenlik: Çek tahsilatına çek ödeme satırı karıştı: {row.masraf}"
            )
 
    elif spec.service == "SENET_IADE":
        if (
            "senet" not in mas
            or "iade" not in mas
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            raise RuntimeError(
                f"Güvenlik: Senet iade için yanlış kaynak seçildi: {row.masraf}"
            )
 
    elif spec.service == "SENET_PROTESTO":
        if (
            "senet" not in mas
            or "protesto" not in mas
            or "protestosuz" in mas
            or "kaldir" in mas
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            raise RuntimeError(
                f"Güvenlik: Senet protesto için yanlış kaynak seçildi: {row.masraf}"
            )
 
    elif spec.service == "SENET_PROTESTO_KALDIRMA":
        if not ("senet" in mas and "protesto" in mas and "kaldir" in mas):
            raise RuntimeError(
                f"Güvenlik: Senet protesto kaldırma için yanlış kaynak: {row.masraf}"
            )
 
    elif spec.service == "HESAP_OZETI_POSTA":
        if any(x in structural for x in ("kredi kart", "ticari kart", "business kart")):
            raise RuntimeError(
                f"Güvenlik: Hesap özeti posta satırına kart ekstresi karıştı: {row.masraf}"
            )
 
    elif spec.service in {
        "SWIFT_GELEN", "SWIFT_GIDEN", "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER"
    }:
        if any(x in mas for x in ("paket", "kobi", "kota")):
            raise RuntimeError(
                f"Güvenlik: Tekil transfer tarifesine paket/kota karıştı: {row.masraf}"
            )
 
        # Yön/ürün güvenliği. Aggregate tarafı daha da dar banka-bazlı
        # filtre uygular; burası son savunma katmanıdır.
        if spec.service == "SWIFT_GELEN" and any(x in mas for x in (
            "arastirma", "iade", "isme", "kasaya", "giden",
        )):
            raise RuntimeError(
                f"Güvenlik: Gelen SWIFT'e farklı varyant karıştı: {row.masraf}"
            )
 
        if spec.service == "YURT_DISI_FAST" and any(x in _norm(row.text) for x in (
            "moneysend", "western union", "karta para gonder",
            "visa ile yurt disi para transferi",
            "visa ile yurtdisi para transferi",
        )):
            raise RuntimeError(
                f"Güvenlik: Hızlı hesaba transfere farklı ürün karıştı: {row.masraf}"
            )
 
        if spec.service == "KART_YURTDISI_TRANSFER" and any(x in _norm(row.text) for x in (
            "alici", "kktc",
        )):
            raise RuntimeError(
                f"Güvenlik: Karta transfere alıcı/KKTC varyantı karıştı: {row.masraf}"
            )
 
 
def _resolve_cell(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Tuple[str, Optional[FeeRow], str]:
    """Hücre metni, dayanak satır ve çözüm türünü döndürür."""
 
    lookup_channel = wanted_channel if spec.split_channel else "GENEL"
 
    # Kullanıcı tarafından doğrulanmış yüksek riskli aileler önce çözülür.
    # Böylece genel/fuzzy fallback doğru ilk-sayfa verisinin önüne geçemez.
    audited = _user_audit_override(rows, bank, spec, wanted_channel)
    if audited is not None:
        return audited
 
    # Resmî olarak uygulanmadığı/kanalda yayımlanmadığı doğrulanan durum,
    # aynı isimli EFT tarife satırının FAST'e yanlış taşınmasını engellemek için
    # sayısal eşleşmeden ÖNCE değerlendirilir.
    pre_status = _service_status_row(
        rows, bank, spec, lookup_channel, kinds={"NOT_APPLICABLE"}
    )
    if pre_status is not None:
        return _fee_text(pre_status, spec), pre_status, "NOT_APPLICABLE"
 
    # Yapı Kredi Kasa24, klasik Büyük/Orta/Küçük kasa ailesinden farklı
    # yayımlandığı için Özel/Süper satırında A-E birlikte gösterilir.
    kasa24 = _kasa24_special_fee(rows, bank, spec)
    if kasa24 is not None:
        value, source = kasa24
        return value, source, "NUMERIC"
 
    aggregated = _aggregate_service_fee(rows, bank, spec, wanted_channel)
    if aggregated is not None:
        value, agg_row = aggregated
        _assert_safe_source(agg_row, spec)
        return value, agg_row, "NUMERIC"
 
    row = _best_match(rows, bank, spec, lookup_channel)
 
    if row is None and not spec.split_channel:
        found = [_best_match(rows, bank, spec, possible) for possible in ("GENEL", "MOBIL", "SUBE")]
        row = next((x for x in found if x is not None), None)
 
    # Status satırı sayısal ücretin önüne geçmesin; NOT_APPLICABLE yukarıda
    # zaten özel olarak ele alındı.
    if row is not None and _has_numeric_fee(row):
        _assert_safe_source(row, spec)
        value = _fee_text(row, spec)
        if spec.service == "FAST" and spec.band_key == "TRANSFER_2":
            limit_note = _fast_limit_note(rows, bank, wanted_channel)
            if limit_note:
                value += f"\n{limit_note}"
        if spec.service == "KASA":
            annual_tax = _bsmv_label(row)
            annual_fee = _fee_value_compact(row)
            annual = f"Yıllık: {annual_fee}" if annual_fee else "Yıllık: ücret bilgisi açıklamada"
            if annual_tax:
                annual += f" ({annual_tax})"
            dep_value, dep_tax = _best_deposit_match(rows, bank, spec, row)
            value = annual
            if dep_value:
                value += f"\nDepozito: {dep_value}"
                if dep_tax:
                    value += f" ({dep_tax})"
        return value, row, "NUMERIC"
 
    # Aidat/okul: hizmetin resmî varlık kanıtı varsa, bankanın genel
    # Fatura/Kurum tarifesi açıkça "genel tarife" etiketiyle kullanılabilir.
    if spec.service in {"AIDAT", "OZEL_OKUL", "TELEFON"}:
        status_row = _service_status_row(rows, bank, spec, lookup_channel)
        if status_row is not None:
            generic = _generic_institution_fee(rows, bank, lookup_channel)
            if generic is not None:
                fee = _fee_text(generic, None)
                return f"Genel kurum tarifesi:\n{fee}", generic, "GENERIC_TARIFF"
            return _fee_text(status_row, spec), status_row, "STATUS"
 
    if row is not None and _status_kind(row):
        return _fee_text(row, spec), row, "STATUS"
 
    status_row = _service_status_row(rows, bank, spec, lookup_channel)
    if status_row is not None:
        kind = "NOT_APPLICABLE" if _status_kind(status_row) == "NOT_APPLICABLE" else "STATUS"
        return _fee_text(status_row, spec), status_row, kind
 
    fast_status = _fast_branch_publication_status(rows, bank, spec, wanted_channel)
    if fast_status:
        return fast_status, None, "PUBLICATION_STATUS"
 
    # Son çare olarak veri uydurmak yerine açıkça kaynak boşluğu belirtilir.
    # Böylece N/A'nın "0 TL" veya "hizmet yok" sanılması engellenir.
    return _source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP"
 
 
 
def _preserve_notes(old_ws) -> Dict[Tuple[str, str], str]:
    """Notları satır numarasına değil ORTAK satır adına göre korur."""
    notes: Dict[Tuple[str, str], str] = {}
    if old_ws is None:
        return notes
 
    section = ""
    for row in range(3, old_ws.max_row + 1):
        label = _clean(old_ws.cell(row=row, column=1).value)
        note = _clean(old_ws.cell(row=row, column=13).value)
        if not label:
            continue
 
        # Section satırları bankalarda veri taşımadığı için kalın başlık gibi davranır.
        if label in {payload for kind, payload in LAYOUT if kind == "SECTION"}:
            section = label
            continue
 
        if note and not _looks_like_generated_note(note):
            notes[(section, label)] = note
 
    return notes
 
 
 
# ---------------------------------------------------------------------------
# NOTLAR SÖZLÜĞÜ
# ---------------------------------------------------------------------------
# NOTLAR hücresinde yalnız satırda gerçekten geçen kavramların kısa anlamı
# gösterilir. Metinler özellikle kısa ve düz tutulur; kalın/rich-text yoktur.
_AUTO_NOTE_LABELS = {
    "BSMV dahil",
    "BSMV hariç",
    "Genel Fatura/Kurum tarifesi",
    "Genel kurum tarifesi",
    "Genel tarife",
    "Genel gişe tarifesi",
    "Kanal ayrımı yayımlanmıyor",
    "Ödeme aracı ayrımı yayımlanmıyor",
    "Hesap/nakit ayrımı yayımlanmıyor",
    "Nakit/hesap ayrımı yayımlanmıyor",
    "Ayrı ücret yayımlanmıyor",
    "Bu kanalda tarife yayımlanmıyor",
    "Hizmet var",
    "Ücretsiz / 0 TRY",
    "Kuruma/işleme göre",
    "Firmaya göre",
    "Fatura türüne göre",
    "Asgari",
    "Azami",
    "Valör",
    "YP",
    "Depozito",
    "Değerli kağıt bedeli hariç",
}
 
 
def _append_note(notes: List[str], label: str, explanation: str) -> None:
    line = f"• {label}: {explanation}"
    if line not in notes:
        notes.append(line)
 
 
def _automatic_notes(values: Sequence[str]) -> str:
    """
    B:I hücrelerinde kullanılan terimleri algılar ve M sütununda
    tek tek kısa açıklama listesine dönüştürür.
    """
    raw = "\n".join(_clean(v) for v in values if _clean(v))
    n = _norm(raw)
    notes: List[str] = []
 
    if "bsmv dahil" in n:
        _append_note(
            notes,
            "BSMV dahil",
            "BSMV gösterilen tutarın içindedir.",
        )
    if "bsmv haric" in n:
        _append_note(
            notes,
            "BSMV hariç",
            "BSMV gösterilen tutara dahil değildir; ayrıca eklenebilir.",
        )
 
    # En spesifik genel-tarife ifadesini kullan; aynı satırda gereksiz tekrar yaratma.
    if "genel fatura/kurum tarifesi" in n:
        _append_note(
            notes,
            "Genel Fatura/Kurum tarifesi",
            "Gösterilen ücret bu hizmete özel değil, bankanın genel fatura/kurum tarifesidir.",
        )
    elif "genel kurum tarifesi" in n:
        _append_note(
            notes,
            "Genel kurum tarifesi",
            "Gösterilen ücret bu alt hizmete özel değil, genel kurum ödeme tarifesidir.",
        )
    elif "genel tarife" in n:
        _append_note(
            notes,
            "Genel tarife",
            "Banka işlemi daha alt türlere ayırmadan ortak bir ücret yayımlamıştır.",
        )
 
    if "genel gise tarifesi" in n:
        _append_note(
            notes,
            "Genel gişe tarifesi",
            "Şubede/gişede yapılan işlem için yayımlanan genel ücrettir.",
        )
 
    if "kanal/odeme araci ayrimi yayimlanmiyor" in n:
        _append_note(
            notes,
            "Kanal ayrımı yayımlanmıyor",
            "Banka Mobil ve Şube için ayrı ücret açıklamamıştır.",
        )
        _append_note(
            notes,
            "Ödeme aracı ayrımı yayımlanmıyor",
            "Hesap, nakit veya kredi kartı için ayrı fiyat açıklanmamıştır.",
        )
    else:
        if "kanal ayrimi yayimlanmiyor" in n:
            _append_note(
                notes,
                "Kanal ayrımı yayımlanmıyor",
                "Banka Mobil ve Şube için ayrı ücret açıklamamıştır.",
            )
        if "odeme araci ayrimi yayimlanmiyor" in n:
            _append_note(
                notes,
                "Ödeme aracı ayrımı yayımlanmıyor",
                "Hesap, nakit veya kredi kartı için ayrı fiyat açıklanmamıştır.",
            )
 
    if "hesap/nakit ayrimi yayimlanmiyor" in n:
        _append_note(
            notes,
            "Hesap/nakit ayrımı yayımlanmıyor",
            "Banka hesaptan ve nakit ödemeyi ayrı fiyatlandırmamıştır.",
        )
    if "nakit/hesap ayrimi yayimlanmiyor" in n:
        _append_note(
            notes,
            "Nakit/hesap ayrımı yayımlanmıyor",
            "Banka nakit ve hesaptan ödemeyi ayrı fiyatlandırmamıştır.",
        )
 
    # "Ayrı ... ücret/tarife yayımlanmıyor" türlerinin tümü tek açıklamada birleşir.
    if (
        "ayri aidat ucreti yayimlanmiyor" in n
        or "aidata ozel ucret yayimlanmiyor" in n
        or "ayri ozel okul ucreti yayimlanmiyor" in n
        or "ayri ozel okul/egitim ucreti yayimlanmiyor" in n
        or "ayri ucret tutari belirtilmemis" in n
        or "ayri ucret ilan edilmemis" in n
        or "ayri vergi ucreti yayimlanmiyor" in n
    ):
        _append_note(
            notes,
            "Ayrı ücret yayımlanmıyor",
            "Hizmet veya kalem kaynakta var, ancak ona özel sayısal ücret açıklanmamıştır.",
        )
 
    if (
        "bu kanalda tarife yayimlanmiyor" in n
        or "bu kanalda ozel okul odeme tarifesi yayimlanmiyor" in n
        or "sube icin ayri" in n
        or "mobil icin ayri" in n
    ):
        _append_note(
            notes,
            "Bu kanalda tarife yayımlanmıyor",
            "İlgili kanal için ayrı ve güvenilir bir ücret tarifesi yayımlanmamıştır.",
        )
 
    if "hizmet var" in n:
        _append_note(
            notes,
            "Hizmet var",
            "Bankanın bu işlemi sunduğu resmî kaynakta doğrulanmıştır.",
        )
 
    if "ucretsiz / 0 try" in n or "ucretsiz/0 try" in n:
        _append_note(
            notes,
            "Ücretsiz / 0 TRY",
            "Banka bu işlemden ücret alınmadığını belirtmektedir.",
        )
 
    if "kuruma/isleme gore" in n:
        _append_note(
            notes,
            "Kuruma/işleme göre",
            "Ücret, ödeme yapılan kurum veya işlem türüne göre değişebilir.",
        )
    if "firmaya gore" in n:
        _append_note(
            notes,
            "Firmaya göre",
            "Ücret, ödeme yapılan firmaya göre değişebilir.",
        )
    if "fatura turune gore" in n:
        _append_note(
            notes,
            "Fatura türüne göre",
            "Ücret, ödenen fatura türüne göre değişebilir.",
        )
 
    # Sık kullanılan finansal terimler.
    if re.search(r"\basgari\b|\bmin\b", n):
        _append_note(
            notes,
            "Asgari",
            "Uygulanabilecek en düşük ücret veya tutardır.",
        )
    if re.search(r"\bazami\b|\bmax\b", n):
        _append_note(
            notes,
            "Azami",
            "Uygulanabilecek en yüksek ücret veya tutardır.",
        )
    if "valor" in n:
        _append_note(
            notes,
            "Valör",
            "Transferin hesaba değer kazanacağı/işleme alınacağı günü ifade eder.",
        )
    if re.search(r"(^|[^a-z])yp([^a-z]|$)", n):
        _append_note(
            notes,
            "YP",
            "Yabancı para anlamına gelir.",
        )
    if "depozito" in n:
        _append_note(
            notes,
            "Depozito",
            "Kiralık kasa için ayrıca alınabilen güvence bedelidir.",
        )
    if "degerli kagit" in n and "haric" in n:
        _append_note(
            notes,
            "Değerli kağıt bedeli hariç",
            "Gösterilen çek ücretine resmî değerli kağıt bedeli dahil değildir.",
        )
 
    return "\n".join(notes)
 
 
def _looks_like_generated_note(note: str) -> bool:
    """Önceki V25 çalışmasının otomatik sözlük notlarını manuel not sanma."""
    lines = [x.strip() for x in str(note or "").splitlines() if x.strip()]
    if not lines:
        return False
    if not all(line.startswith("• ") and ":" in line for line in lines):
        return False
 
    heads = {line[2:].split(":", 1)[0].strip() for line in lines}
    return bool(heads) and heads.issubset(_AUTO_NOTE_LABELS)
 
 
def _compose_notes(auto_note: str, manual_note: str) -> str:
    """
    Otomatik terim açıklamalarını ve varsa gerçek manuel notu aynı hücrede,
    yine madde madde gösterir.
    """
    lines = [x.strip() for x in str(auto_note or "").splitlines() if x.strip()]
 
    manual_note = _clean(manual_note)
    if manual_note and not _looks_like_generated_note(manual_note):
        for raw_line in str(manual_note).splitlines():
            clean_line = raw_line.strip()
            if not clean_line:
                continue
            if clean_line.startswith("• "):
                line = clean_line
            else:
                line = f"• Not: {clean_line}"
            if line not in lines:
                lines.append(line)
 
    return "\n".join(lines)
 
 
 
# ---------------------------------------------------------------------------
# NOTLAR / TERİM AÇIKLAMALARI
# ---------------------------------------------------------------------------
# Kullanıcının istediği sabit sözlük. NOTLAR sütununda her açıklama ayrı satır
# halinde, düz (bold olmayan) fontla gösterilir.
NOTES_GLOSSARY: Sequence[Tuple[str, str]] = (
    ("Ayrı ücret tutarı yayımlanmıyor.", "Banka bu hizmeti sunuyor ancak bu işlem için özel bir ücret rakamı açıklamıyor."),
    ("Ücret tutarı belirtilmemiş.", "İşlem bankanın tarifesinde yer alıyor fakat ilgili satırda sayısal ücret bulunmuyor."),
    ("Hizmet var / ayrı tarife yayımlanmıyor.", "İşlemin yapılabildiği doğrulandı ancak yalnız bu hizmete ait ayrı bir fiyat tarifesi yok."),
    ("Genel Fatura/Kurum tarifesi uygulanır.", "Bu işlem için özel fiyat yerine bankanın genel fatura/kurum ödeme tarifesi kullanılıyor."),
    ("Genel Fatura/Kurum tarifesinden eşleştirilmiştir.", "Gösterilen ücret özel olarak bu hizmetin değil, onu kapsayan genel tarifenin ücretidir."),
    ("Genel tarife.", "Banka işlemi daha alt türlere ayırmadan tek ortak ücret tarifesi yayımlıyor."),
    ("Genel gişe tarifesi.", "Ücret şubede/gişede yapılan işlem için verilmiş, ödeme şekli ayrıca ayrılmamış."),
    ("Kanal ayrımı yayımlanmıyor.", "Banka Mobil, İnternet veya Şube için ayrı ayrı ücret belirtmiyor."),
    ("Hesap/Nakit ayrımı yayımlanmıyor.", "Banka ödemenin hesaptan mı nakit mi yapıldığına göre ayrı ücret açıklamıyor."),
    ("Ödeme aracı ayrımı yayımlanmıyor.", "Banka hesap, nakit veya kredi kartı kullanımına göre tarifeyi ayırmıyor."),
    ("Aidata özel ücret yayımlanmıyor.", "Aidat ödemesi yapılabiliyor ancak yalnız aidata özel bir komisyon tarifesi yok."),
    ("Özel okul ücreti yayımlanmıyor.", "Özel okul ödemesi yapılabiliyor fakat bu işleme özel komisyon rakamı açıklanmıyor."),
    ("Genel SGK tarifesi – kanal ayrımı yayımlanmıyor.", "SGK için ücret var fakat banka Mobil ve Şube ücretlerini ayrı göstermiyor."),
    ("Ücretsiz / 0 TRY.", "Bankanın resmî açıklamasına göre bu işlemden ücret alınmıyor."),
    ("Bu kanalda hizmet sunulmuyor.", "İşlem ilgili Mobil veya Şube kanalından yapılamıyor."),
    ("Bu kanal için ayrı tarife yayımlanmıyor.", "İşlem kanalda olabilir ancak o kanala özel ücret açıklanmamış."),
    ("Kaynakta sayısal tarife bulunamadı.", "Resmî kaynakta hizmete ilişkin güvenilir bir ücret rakamı tespit edilemedi."),
    ("Kaynakta tutar görünmüyor.", "Bankanın ilgili satırı mevcut fakat ücret alanı boş."),
    ("Kanal belirtilmemiş.", "Ücret yayımlanmış ancak hangi işlem kanalında geçerli olduğu belirtilmemiş."),
    ("Kuruma/işleme göre değişir.", "Ücret sabit değil; ödeme yapılan kurum veya işlem türüne göre belirtilen sınırlar içinde değişebilir."),
    ("Fatura türüne göre değişir.", "Telefon, elektrik, su vb. fatura türüne göre uygulanacak ücret farklı olabilir."),
    ("BSMV dahil.", "Gösterilen tutarın içinde BSMV zaten vardır."),
    ("BSMV hariç.", "Gösterilen tutara işlem sırasında ayrıca BSMV eklenebilir."),
    ("KDV dahil.", "Gösterilen ücretin içine KDV dahildir."),
    ("Değerli kağıt bedeli hariç.", "Gösterilen çek ücreti dışında ayrıca değerli kağıt bedeli alınabilir."),
    ("Bkz. FATURA bölümü.", "Aynı tarife tekrar yazılmadı; detaylı ücret Fatura bölümünde gösteriliyor."),
)
 
 
def _write_notes_glossary(ws, thin: Side) -> None:
    """NOTLAR sütununu sabit terim sözlüğü olarak yazar."""
    start_row = 3
 
    # Eski otomatik/manüel not kalıntılarını temizle.
    for r in range(start_row, max(ws.max_row, start_row + len(NOTES_GLOSSARY)) + 1):
        c = ws.cell(row=r, column=13)
        c.value = None
        c.comment = None
        c.font = Font(bold=False, italic=False, color="595959", size=8.5)
        c.fill = PatternFill("solid", fgColor="FFF9E6")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    for offset, (term, meaning) in enumerate(NOTES_GLOSSARY):
        row = start_row + offset
        c = ws.cell(row=row, column=13)
        c.value = f'• “{term}” → {meaning}'
        c.font = Font(bold=False, italic=False, color="595959", size=8.5)
        c.fill = PatternFill("solid", fgColor="FFF9E6")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
 
        # Açıklamalar tek tek rahat okunsun; mevcut daha yüksek satırları küçültme.
        current_height = ws.row_dimensions[row].height or 15
        ws.row_dimensions[row].height = max(current_height, 34)
 
 
def _sheet_row_key(section: str, label: str) -> Tuple[str, str]:
    return (section, label)
 
 
def _write_comparison(ws, rows: Sequence[FeeRow], notes: Mapping[Tuple[str, str], str]) -> int:
    thin = Side(style="thin", color="B7B7B7")
    medium = Side(style="medium", color="7F7F7F")
 
    ws["M1"] = "NOTLAR"
    ws.merge_cells("M1:M2")
    ws["M1"].font = Font(bold=True, color="1F1F1F")
    ws["M1"].fill = PatternFill("solid", fgColor="D9EAD3")
    ws["M1"].alignment = Alignment(horizontal="center", vertical="center")
 
    ws["A2"] = "ORTAK MASRAF ADI"
    ws["A2"].font = Font(bold=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
 
    col = 2
    for bank in BANKS:
        color = BANK_COLORS[bank]
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        cell = ws.cell(row=1, column=col)
        cell.value = DISPLAY_BANKS[bank]
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for offset, sub in enumerate(("Mobil", "Şube")):
            c = ws.cell(row=2, column=col + offset)
            c.value = sub
            c.fill = PatternFill("solid", fgColor=color)
            c.font = Font(bold=True, italic=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
        col += 2
 
    current_row = 3
    section = ""
 
    for kind, payload in LAYOUT:
        if kind == "SECTION":
            section = payload
            ws.cell(row=current_row, column=1).value = payload
            for col_idx in range(1, 10):
                c = ws.cell(row=current_row, column=col_idx)
                c.fill = PatternFill("solid", fgColor="E7E6E6")
                c.font = Font(bold=True, color="595959")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border(top=medium, bottom=thin)
 
            ws.row_dimensions[current_row].height = 24
            current_row += 1
            continue
 
        spec: RowSpec = payload
        ws.cell(row=current_row, column=1).value = spec.label
        ws.cell(row=current_row, column=1).font = Font(color="595959")
        ws.cell(row=current_row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
 
        for bank_index, bank in enumerate(BANKS):
            start_col = 2 + bank_index * 2
 
            if spec.split_channel:
                for offset, wanted_channel in enumerate(("MOBIL", "SUBE")):
                    value, row, resolution = _resolve_cell(rows, bank, spec, wanted_channel)
                    c = ws.cell(row=current_row, column=start_col + offset)
                    c.value = value
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.font = Font(
                        bold=resolution not in {"SOURCE_GAP"},
                        color=BANK_COLORS[bank] if resolution != "SOURCE_GAP" else "A6A6A6",
                        size=9,
                    )
                    comment = _cell_comment(row, resolution, value)
                    if comment:
                        c.comment = comment
            else:
                # Kanal ayrımı anlamsız olan kasa, çek, senet, rapor vb. satırlarda
                # Mobil/Şube hücrelerini tek mantıksal banka hücresine birleştir.
                ws.merge_cells(
                    start_row=current_row, start_column=start_col,
                    end_row=current_row, end_column=start_col + 1,
                )
                value, row, resolution = _resolve_cell(rows, bank, spec, "GENEL")
                c = ws.cell(row=current_row, column=start_col)
                c.value = value
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.font = Font(
                    bold=resolution not in {"SOURCE_GAP"},
                    color=BANK_COLORS[bank] if resolution != "SOURCE_GAP" else "A6A6A6",
                    size=9,
                )
                comment = _cell_comment(row, resolution, value)
                if comment:
                    c.comment = comment
 
        # NOTLAR sütunu satır-bazlı otomatik not yerine sabit sözlük/sidebar
        # olarak kullanılır. Buradaki hücreler döngü sonunda topluca yazılır.
        note_value = ""
 
        if spec.service == "KASA" and spec.detail == "OZEL":
            ws.row_dimensions[current_row].height = 105
        elif spec.service == "KASA":
            ws.row_dimensions[current_row].height = 72
        elif spec.service == "CEK_DEFTERI_GRUP":
            ws.row_dimensions[current_row].height = 150
        elif spec.service in {"CEK_DUZENLEME_OZEL_GRUP", "CEK_TAHSIL_DOVIZ_GRUP"}:
            ws.row_dimensions[current_row].height = 100
        elif spec.service in {"DUZENLI_EFT", "DUZENLI_HAVALE"}:
            ws.row_dimensions[current_row].height = 90
        elif spec.service == "DOC_ESKI_DEKONT_EKSTRE":
            ws.row_dimensions[current_row].height = 92
        elif spec.service in _DOCUMENT_STATEMENT_SERVICES:
            ws.row_dimensions[current_row].height = 72
        elif spec.service in _RESEARCH_SECURITIES_SERVICES:
            ws.row_dimensions[current_row].height = 88
        elif spec.service in {
            "SWIFT_GELEN",
            "SWIFT_GIDEN",
            "YURT_DISI_FAST",
            "KART_YURTDISI_TRANSFER",
            "HGS_ETIKET",
            "HGS_KART",
            "SANS_OYUNU",
            "BAKIYE_ATM_YURTDISI",
        }:
            ws.row_dimensions[current_row].height = 66
        else:
            ws.row_dimensions[current_row].height = 52
 
        # Not listesinin alt satırları görünür kalsın; gereksiz dev satır oluşturma.
        if note_value:
            note_lines = len(note_value.splitlines())
            ws.row_dimensions[current_row].height = max(
                ws.row_dimensions[current_row].height or 52,
                min(108, 18 + note_lines * 13),
            )
 
        current_row += 1
 
    # Sağdaki NOTLAR alanı: kullanıcının istediği sabit terim açıklamaları.
    _write_notes_glossary(ws, thin)
 
    for row_cells in ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=13):
        for cell in row_cells:
            if cell.column <= 9 and cell.row > 2:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    ws.column_dimensions["A"].width = 44
    for col_letter in "BCDEFGHI":
        ws.column_dimensions[col_letter].width = 18
    for col_letter in "JKL":
        ws.column_dimensions[col_letter].width = 3
    ws.column_dimensions["M"].width = 58
 
    # Birleşik banka başlıklarını bozmadan gerçek Excel filtresi.
    # A2:I... aralığı kullanılır: A=masraf adı, B:I=4 bankanın Mobil/Şube sütunları.
    ws.auto_filter.ref = f"A2:I{current_row - 1}"
 
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "B3"
    ws.sheet_view.zoomScale = 80
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
 
    # Teknik eşleşme özeti artık görünür Excel alanına yazılmaz.
    # Ayrıntılı kalite kontrolü yalnız GitHub Actions logunda basılır.
    return current_row - 1
 
 
 
def _print_transfer_audit(rows: Sequence[FeeRow]) -> None:
    """Kritik transfer eşleşmelerini GitHub logunda görünür yapar."""
    print("[comparison] ===== ORTAK MASRAF EŞLEŞME KONTROLÜ =====")
 
    audit_specs = [
        spec
        for kind, spec in LAYOUT
        if kind == "ROW" and (
            spec.service in {
                "EFT", "HAVALE", "FAST", "SWIFT_GELEN", "SWIFT_GIDEN",
                "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER", "DUZENLI_EFT",
                "DUZENLI_HAVALE", "ALTIN_TRANSFER",
                "HGS_ETIKET", "HGS_KART", "VERGI",
                "CEK_DEFTERI_YAPRAK", "CEK_KARNESI_10",
                "CEK_DUZENLEME_STANDART", "CEK_DUZENLEME_DOVIZ",
                "CEK_IADE", "CEK_TAHSIL", "CEK_KARSILIKSIZ",
                "CEK_DUZELTME_HAKKI",
            }
            or spec.service in _DOCUMENT_STATEMENT_SERVICES
            or spec.service in _RESEARCH_SECURITIES_SERVICES
        )
    ]
 
    for spec in audit_specs:
        for bank in BANKS:
            channels = ("MOBIL", "SUBE") if spec.split_channel else ("GENEL",)
            for channel in channels:
                value, row, resolution = _resolve_cell(rows, bank, spec, channel)
                if row is None:
                    raw = value.replace("\n", " / ")
                else:
                    raw = row.masraf
                    if len(raw) > 120:
                        raw = raw[:117] + "..."
                print(
                    f"[comparison][match] {spec.service} | {spec.label} | "
                    f"{bank} | {channel} <- {raw} [{resolution}]"
                )
 
    print("[comparison] ==========================================")
 
 
 
def _is_comparison_sheet_name(title: str) -> bool:
    key = _norm(title).replace(" ", "")
    return (
        key.startswith("karsilastirma")
        or key.startswith("comparison")
    )
 
 
def _remove_old_comparison_sheets(wb) -> List[str]:
    """
    Önceki 10-bankalı / 5x5 / test karşılaştırma sayfalarını temizler.
    Böylece kullanıcı yanlışlıkla eski sheet'e bakmaz.
    """
    removed: List[str] = []
 
    for ws in list(wb.worksheets):
        if _is_comparison_sheet_name(ws.title):
            removed.append(ws.title)
            wb.remove(ws)
 
    return removed
 
 
def _assert_preview_layout(ws) -> None:
    """
    Preview'den sapmayı fatal hata yapar.
    Eski 10 bankalı formatın sessizce geri gelmesini engeller.
    """
    expected_headers = {
        "B1": "Garanti BBVA",
        "D1": "İş Bankası",
        "F1": "Akbank",
        "H1": "Yapı ve Kredi Bankası",
        "A2": "ORTAK MASRAF ADI",
        "M1": "NOTLAR",
    }
 
    for cell_ref, expected in expected_headers.items():
        actual = _clean(ws[cell_ref].value)
        if actual != expected:
            raise RuntimeError(
                f"Preview layout doğrulaması başarısız: "
                f"{cell_ref} beklenen={expected!r}, gelen={actual!r}"
            )
 
    # Preview'de banka alanı B:I, J:L boş ve M notlardır.
    for col in ("J", "K", "L"):
        if any(
            _clean(ws[f"{col}{row}"].value)
            for row in range(1, min(ws.max_row, 12) + 1)
        ):
            raise RuntimeError(
                f"Preview layout doğrulaması başarısız: {col} kolonu boş olmalı."
            )
 
    # Eski formatta kullanılan banka isimleri görünür başlıklarda olmamalı.
    forbidden = {
        "QNB",
        "DenizBank",
        "Halkbank",
        "VakıfBank",
        "TEB",
        "Ziraat",
    }
    visible_headers = {
        _clean(ws.cell(row=row, column=col).value)
        for row in (1, 2)
        for col in range(1, 14)
    }
 
    bad = sorted(forbidden & visible_headers)
    if bad:
        raise RuntimeError(
            "Preview layout yerine eski çok-bankalı format oluştu: "
            + ", ".join(bad)
        )
 
    expected_filter = f"A2:I{ws.max_row}"
    if ws.auto_filter.ref != expected_filter:
        raise RuntimeError(
            "Karşılaştırma filtresi oluşmadı: "
            f"beklenen={expected_filter!r}, gelen={ws.auto_filter.ref!r}"
        )
 
 
def _assert_document_statement_sources(rows: Sequence[FeeRow]) -> None:
    """Yeni 17 satırın kritik resmî dayanaklarını yazımdan önce doğrular."""
    document_specs = {
        spec.service: spec
        for kind, spec in LAYOUT
        if kind == "ROW" and spec.service in _DOCUMENT_STATEMENT_SERVICES
    }
    if len(document_specs) != 17:
        raise RuntimeError(
            "Belge/ekstre/araştırma sözlüğü eksik veya mükerrer: "
            f"beklenen=17, gelen={len(document_specs)}"
        )
 
    required = (
        ("DOC_VADESIZ_TL", "GARANTİ", "MOBIL"),
        ("DOC_VADESIZ_TL", "İŞBANKASI", "MOBIL"),
        ("DOC_VADESIZ_TL", "İŞBANKASI", "SUBE"),
        ("DOC_VADESIZ_YP", "GARANTİ", "MOBIL"),
        ("DOC_VADESIZ_YP", "İŞBANKASI", "MOBIL"),
        ("DOC_VADESIZ_YP", "İŞBANKASI", "SUBE"),
        ("DOC_KMH_BIREYSEL", "GARANTİ", "GENEL"),
        ("DOC_KMH_BIREYSEL", "YAPIKREDI", "GENEL"),
        ("DOC_KK_GECMIS", "GARANTİ", "GENEL"),
        ("DOC_KK_GECMIS", "İŞBANKASI", "GENEL"),
        ("DOC_KK_GECMIS", "YAPIKREDI", "GENEL"),
        ("DOC_ARSIV_GENEL", "GARANTİ", "GENEL"),
        ("DOC_ARSIV_GENEL", "İŞBANKASI", "GENEL"),
        ("DOC_ESKI_DEKONT_EKSTRE", "AKBANK", "GENEL"),
        ("DOC_ESKI_DEKONT_EKSTRE", "YAPIKREDI", "GENEL"),
        ("DOC_AYLIK_POSTA", "AKBANK", "GENEL"),
        ("DOC_ARSIV_SUBE", "İŞBANKASI", "GENEL"),
        ("DOC_ARSIV_BANKA", "İŞBANKASI", "GENEL"),
        ("DOC_KMH_TICARI", "GARANTİ", "GENEL"),
        ("DOC_KMH_TICARI", "YAPIKREDI", "GENEL"),
        ("DOC_TICARI_KART_GECMIS", "YAPIKREDI", "GENEL"),
        ("DOC_YATIRIM_EKSTRE", "İŞBANKASI", "GENEL"),
        ("DOC_TICARI_AYLIK", "AKBANK", "GENEL"),
        ("DOC_TICARI_KART_BASILI", "GARANTİ", "GENEL"),
        ("DOC_TICARI_KART_GONDERIM", "GARANTİ", "GENEL"),
        ("DOC_TICARI_HESAP_BASILI", "AKBANK", "GENEL"),
        ("DOC_POS_EKSTRE", "GARANTİ", "GENEL"),
    )
 
    failures: List[str] = []
    for service, bank, channel in required:
        spec = document_specs[service]
        value, row, resolution = _resolve_cell(rows, bank, spec, channel)
        if row is None or resolution not in {"NUMERIC", "STATUS"}:
            failures.append(
                f"{service}/{bank}/{channel}: {resolution} ({value!r})"
            )
 
    if failures:
        raise RuntimeError(
            "Belge/ekstre/araştırma kritik kaynak kontrolü başarısız: "
            + "; ".join(failures)
        )
 
    print(
        "[comparison] BELGE/EKSTRE/ARAŞTIRMA: "
        f"17 satır ve {len(required)} kritik kaynak eşleşmesi doğrulandı."
    )
 
 
def _assert_research_securities_sources(rows: Sequence[FeeRow]) -> None:
    """Yeni araştırma, yurtdışı ATM ve menkul kıymet satırlarını fail-closed doğrular."""
    specs = {
        spec.service: spec
        for kind, spec in LAYOUT
        if kind == "ROW" and (
            spec.service in _RESEARCH_SECURITIES_SERVICES
            or spec.service == "BAKIYE_ATM_YURTDISI"
        )
    }
    expected_services = _RESEARCH_SECURITIES_SERVICES | {"BAKIYE_ATM_YURTDISI"}
    if set(specs) != expected_services:
        raise RuntimeError(
            "Araştırma/menkul kıymet sözlüğü eksik veya mükerrer: "
            f"beklenen={sorted(expected_services)}, gelen={sorted(specs)}"
        )
 
    numeric_required = (
        ("BAKIYE_ATM_YURTDISI", "İŞBANKASI"),
        ("SEC_DOMESTIC_STOCK_TRADE", "GARANTİ"),
        ("SEC_DOMESTIC_STOCK_TRADE", "İŞBANKASI"),
        ("SEC_DOMESTIC_STOCK_TRADE", "AKBANK"),
        ("SEC_DOMESTIC_STOCK_TRADE", "YAPIKREDI"),
        ("SEC_DOMESTIC_STOCK_TRANSFER", "İŞBANKASI"),
        ("SEC_DOMESTIC_STOCK_TRANSFER", "AKBANK"),
        ("SEC_DOMESTIC_STOCK_TRANSFER", "YAPIKREDI"),
        ("SEC_DOMESTIC_BOND_TRANSFER", "İŞBANKASI"),
        ("SEC_DOMESTIC_BOND_TRANSFER", "AKBANK"),
        ("SEC_DOMESTIC_BOND_TRANSFER", "YAPIKREDI"),
        ("SEC_DOMESTIC_CUSTODY", "GARANTİ"),
        ("SEC_DOMESTIC_CUSTODY", "İŞBANKASI"),
        ("SEC_DOMESTIC_CUSTODY", "AKBANK"),
        ("SEC_DOMESTIC_CUSTODY", "YAPIKREDI"),
        ("SEC_FOREIGN_STOCK_TRADE", "İŞBANKASI"),
        ("SEC_FOREIGN_STOCK_CUSTODY", "İŞBANKASI"),
        ("SEC_EUROBOND_TRANSFER", "İŞBANKASI"),
        ("SEC_EUROBOND_TRANSFER", "AKBANK"),
        ("SEC_EUROBOND_TRANSFER", "YAPIKREDI"),
        ("SEC_EUROBOND_CUSTODY", "İŞBANKASI"),
    )
 
    failures: List[str] = []
    for service, bank in numeric_required:
        value, row, resolution = _resolve_cell(rows, bank, specs[service], "GENEL")
        if row is None or resolution not in {"NUMERIC", "STATUS"}:
            failures.append(f"{service}/{bank}: {resolution} ({value!r})")
 
    deposit = specs["DOC_DEPOSIT_RESEARCH"]
    for bank in BANKS:
        value, _, resolution = _resolve_cell(rows, bank, deposit, "GENEL")
        if resolution not in {"NUMERIC", "STATUS", "PUBLICATION_STATUS"}:
            failures.append(f"DOC_DEPOSIT_RESEARCH/{bank}: {resolution} ({value!r})")
 
    if failures:
        raise RuntimeError(
            "Araştırma/menkul kıymet kritik kaynak kontrolü başarısız: "
            + "; ".join(failures)
        )
 
    print(
        "[comparison] ARAŞTIRMA/MENKUL KIYMET: "
        f"{len(expected_services)} satır ailesi ve {len(numeric_required)} "
        "kritik kaynak eşleşmesi doğrulandı."
    )
 
 
def update_comparison_sheet(excel_path: str = "komisyonlar_guncel.xlsx") -> Dict[str, int]:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Ana Excel bulunamadı: {path}")
 
    print(f"[comparison] SÜRÜM: {COMPARISON_VERSION}")
    print(f"[comparison] MODÜL: {Path(__file__).resolve()}")
    print(f"[comparison] LAYOUT: {PREVIEW_LAYOUT_SIGNATURE}")
    print(f"[comparison] Ana Excel okunuyor: {path}")
 
    wb = load_workbook(path)
    source_ws, header_row = _find_source_sheet(wb)
    cols = _header_map(source_ws, header_row)
    rows = _read_rows(source_ws, header_row, cols)
 
    if not rows:
        raise RuntimeError("Karşılaştırma için uygun banka verisi bulunamadı.")
 
    _assert_document_statement_sources(rows)
    _assert_research_securities_sources(rows)
 
    old_ws = wb[COMPARISON_SHEET] if COMPARISON_SHEET in wb.sheetnames else None
    notes = _preserve_notes(old_ws)
 
    removed_sheets = _remove_old_comparison_sheets(wb)
 
    ws = wb.create_sheet(COMPARISON_SHEET)
    comparison_rows = _write_comparison(ws, rows, notes)
    # Dosya açıldığında kullanıcı doğrudan karşılaştırma sekmesini görsün.
    wb.active = wb.sheetnames.index(COMPARISON_SHEET)
    ws.sheet_state = "visible"
 
    # Oluşan sheet preview ile aynı yapıdan saparsa Excel'i kaydetme.
    _assert_preview_layout(ws)
 
    _print_transfer_audit(rows)
 
    # Ana Excel'i yarım/bozuk kaydetmemek için atomik karşılaştırma güncellemesi.
    tmp_path = path.with_name(path.stem + ".comparison.tmp" + path.suffix)
    wb.save(tmp_path)
    tmp_path.replace(path)
 
    print(
        f"[comparison] {COMPARISON_SHEET} referans-sıralı kanonik formatta güncellendi. "
        f"Kaynak={source_ws.title}, 4 banka kaynak satırı={len(rows)}, "
        f"karşılaştırma satırı={comparison_rows}, korunan not={len(notes)}"
    )
    print(
        f"[comparison] Silinen eski karşılaştırma sayfaları: "
        f"{removed_sheets or 'yok'}"
    )
    print(
        "[comparison] Görünüm doğrulandı: "
        "A=ortak masraf, B:I=4 banka Mobil/Şube (filtreli), J:L=boş, M=NOTLAR (sabit sözlük, düz font, tek tek açıklamalar)."
    )
 
    resolution_counts = {
        "NUMERIC": 0,
        "GENERIC_TARIFF": 0,
        "STATUS": 0,
        "PUBLICATION_STATUS": 0,
        "NOT_APPLICABLE": 0,
        "SOURCE_GAP": 0,
        "N/A": 0,
    }
    possible_cells = 0
 
    for kind, payload in LAYOUT:
        if kind != "ROW":
            continue
        spec = payload
        for bank in BANKS:
            channels = ("MOBIL", "SUBE") if spec.split_channel else ("GENEL",)
            for channel in channels:
                possible_cells += 1
                _, _, resolution = _resolve_cell(rows, bank, spec, channel)
                resolution_counts[resolution] = resolution_counts.get(resolution, 0) + 1
 
    source_gaps = resolution_counts.get("SOURCE_GAP", 0)
    true_na = resolution_counts.get("N/A", 0)
    verified_cells = possible_cells - source_gaps - true_na
    numeric_like = resolution_counts.get("NUMERIC", 0) + resolution_counts.get("GENERIC_TARIFF", 0)
    status_like = (
        resolution_counts.get("STATUS", 0)
        + resolution_counts.get("PUBLICATION_STATUS", 0)
        + resolution_counts.get("NOT_APPLICABLE", 0)
    )
 
    print(
        f"[comparison] KALİTE: doğrulanmış={verified_cells}/{possible_cells} "
        f"(%{(verified_cells / possible_cells * 100):.1f}) | "
        f"sayısal={resolution_counts.get('NUMERIC', 0)} | "
        f"genel_tarife={resolution_counts.get('GENERIC_TARIFF', 0)} | "
        f"resmî_durum={status_like} | "
        f"kaynak_boşluğu={source_gaps} | "
        f"belirsiz_eşleşme={len(_AMBIGUITY_LOGGED)} | N/A={true_na}"
    )
 
 
    return {
        "source_rows": len(rows),
        "comparison_rows": comparison_rows,
        "notes_preserved": len(notes),
        "matched_cells": verified_cells,
        "numeric_cells": numeric_like,
        "status_cells": status_like,
        "source_gap_cells": source_gaps,
        "ambiguous_cells": len(_AMBIGUITY_LOGGED),
        "missing_cells": true_na,
        "possible_cells": possible_cells,
    }
 
 
if __name__ == "__main__":
    print(update_comparison_sheet())
 


_RESEARCH_SECURITIES_SERVICES = {
    "DOC_DEPOSIT_RESEARCH",
    "SEC_DOMESTIC_STOCK_TRADE",
    "SEC_DOMESTIC_STOCK_TRANSFER",
    "SEC_DOMESTIC_BOND_TRANSFER",
    "SEC_DOMESTIC_CUSTODY",
    "SEC_FOREIGN_STOCK_TRADE",
    "SEC_FOREIGN_STOCK_CUSTODY",
    "SEC_EUROBOND_TRANSFER",
    "SEC_EUROBOND_CUSTODY",
}


def _audit_research_and_securities(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Mevduat araştırma ve menkul kıymet ailelerini tam adla çözer.

    Menkul kıymetlerde bankaların ürün ve dönem kırılımları aynı değildir.
    Bu nedenle tek bir genel oran seçilmez; karşılaştırılabilir alt hizmetler
    ayrı satırlarda, kaynağın kendi kapsam etiketiyle gösterilir.
    """
    service = spec.service
    if service not in _RESEARCH_SECURITIES_SERVICES:
        return None

    def found(predicate, *, numeric_only: bool = False) -> List[FeeRow]:
        return _primary_first(_audit_rows(
            rows,
            bank,
            predicate,
            numeric_only=numeric_only,
        ))

    def exact(masraf: str, *, kategori: str = "", numeric_only: bool = False) -> Optional[FeeRow]:
        target_fee = _norm(masraf)
        target_cat = _norm(kategori)
        matches = found(
            lambda row: (
                _norm(row.masraf) == target_fee
                and (not target_cat or target_cat in _norm(row.kategori))
            ),
            numeric_only=numeric_only,
        )
        return matches[0] if matches else None

    def published_gap() -> Tuple[str, Optional[FeeRow], str]:
        return (
            "Karşılaştırılabilir ayrı tarife yayımlanmıyor",
            None,
            "PUBLICATION_STATUS",
        )

    def number_text(value: float, decimals: int = 6) -> str:
        raw = f"{value:.{decimals}f}".rstrip("0").rstrip(".")
        return raw.replace(".", ",")

    def numeric_rate(value: str) -> Optional[float]:
        raw = _clean(value)
        if not raw or raw in {"-", "0", "0.0", "0.00", "0%", "0 %"}:
            return None
        match = re.search(r"-?[0-9]+(?:[.,][0-9]+)?", raw)
        if not match:
            return None
        try:
            # Oran alanındaki nokta binlik ayırıcı değil ondalıktır
            # (örn. ``0.185 %`` = yüzde 0,185).
            number = float(match.group(0).replace(",", "."))
        except ValueError:
            return None
        if "%" not in raw and 0 < abs(number) < 0.1:
            number *= 100
        return number

    def rate_text(value: str) -> str:
        number = numeric_rate(value)
        return f"{number_text(number)}%" if number is not None else ""

    def rate_range(items: Sequence[FeeRow]) -> str:
        rates: List[float] = []
        for row in items:
            for attr in ("asgari_oran", "azami_oran"):
                value = numeric_rate(getattr(row, attr, ""))
                if value is not None:
                    rates.append(value)
        if not rates:
            return ""
        low, high = min(rates), max(rates)
        if abs(low - high) < 1e-12:
            return f"{number_text(low)}%"
        return f"{number_text(low)}% - {number_text(high)}%"

    # ------------------------------------------------------------------
    # MEVDUAT ARAŞTIRMA
    # ------------------------------------------------------------------
    if service == "DOC_DEPOSIT_RESEARCH":
        matches = found(
            lambda row: "mevduat arastirma" in _norm(row.text),
            numeric_only=False,
        )
        if not matches:
            return published_gap()
        row = matches[0]
        value = _fee_text(row, spec) if _has_numeric_fee(row) else "Ücret tutarı belirtilmemiş"
        return (value, row, "NUMERIC" if _has_numeric_fee(row) else "STATUS")

    # ------------------------------------------------------------------
    # YURTİÇİ HİSSE SENEDİ ALIM-SATIM
    # ------------------------------------------------------------------
    if service == "SEC_DOMESTIC_STOCK_TRADE":
        if bank == "GARANTİ":
            row = exact("Hisse Senedi Sabit Komisyon Oranı - Garanti BBVA Mobil")
            if row:
                rates = rate_range([row])
                return (f"Mobil/İnternet: {rates}\nİşlem başına\nBSMV hariç", row, "NUMERIC")

        elif bank == "İŞBANKASI":
            matches = found(lambda row: (
                "iscep sen. alis-satis" in _norm(row.kategori)
                and "islem hacmine gore alim-satim komisyonu" in _norm(row.masraf)
            ))
            if matches:
                return (
                    f"İşCep/İnternet: {rate_range(matches)}\n"
                    "Asgari 1 TRY\n3 aylık hacme göre\nBSMV hariç",
                    matches[0],
                    "NUMERIC",
                )

        elif bank == "AKBANK":
            matches = found(lambda row: _norm(row.masraf).startswith(
                "hisse senedi komisyonu aylik islem hacmi"
            ))
            if matches:
                return (
                    f"Dijital kanallar: {rate_range(matches)}\n"
                    "Aylık işlem hacmine göre\nBSMV hariç",
                    matches[0],
                    "NUMERIC",
                )

        elif bank == "YAPIKREDI":
            row = exact(
                "Hisse Senedi İşlemleri - İşlem Komisyonları - İnternet Bankacılığı / Mobil Bankacılık"
            )
            if row:
                match = re.search(r"%\s*([0-9]+(?:[.,][0-9]+)?)", _clean(row.aciklama))
                if match:
                    return (
                        f"Mobil/İnternet: {number_text(_parse_number(match.group(1)) or 0)}%\n"
                        "Alım-satım işlem tutarı üzerinden",
                        row,
                        "NUMERIC",
                    )

        return published_gap()

    # ------------------------------------------------------------------
    # YURTİÇİ HİSSE SENEDİ VİRMANI
    # ------------------------------------------------------------------
    if service == "SEC_DOMESTIC_STOCK_TRANSFER":
        if bank == "İŞBANKASI":
            outgoing = exact("Virman İşlemi", kategori="Hisse Senedi - Varant Virman İşlemleri")
            incoming = exact("Virman İşlemi", kategori="Yatırım Kuruluşlarından Bankamıza Hisse Senedi")
            if outgoing and incoming:
                return (
                    f"Dışarı/hesaplar arası: {_display_amount(outgoing.asgari_tutar)}\n"
                    f"Kuruma gelen: {_display_amount(incoming.asgari_tutar) or '0 TRY'}\n"
                    "BSMV hariç",
                    outgoing,
                    "NUMERIC",
                )

        elif bank == "AKBANK":
            outside = exact("Kurum dışı hisse senedi virmanı")
            inside = exact("Kurum içi hisse senedi virmanı")
            if outside and inside:
                return (
                    f"Kurum dışı: {_fee_value_compact(outside)} (BSMV hariç)\n"
                    f"Kurum içi: {_fee_value_compact(inside)} (BSMV dahil)",
                    outside,
                    "NUMERIC",
                )

        elif bank == "YAPIKREDI":
            row = exact("Hisse Senedi İşlemleri - Kıymet Transferi - Virman")
            if row:
                return (
                    "Transfer tutarı üzerinden\nMKK/Takasbank oranının 5 katı",
                    row,
                    "STATUS",
                )

        return published_gap()

    # ------------------------------------------------------------------
    # YURTİÇİ BONO/TAHVİL VİRMANI
    # ------------------------------------------------------------------
    if service == "SEC_DOMESTIC_BOND_TRANSFER":
        if bank == "İŞBANKASI":
            matches = found(lambda row: (
                "sb.get.men.kiym.oz.sektor tah.virmani" in _norm(row.kategori)
                and _norm(row.masraf) == "virman islemi"
            ))
            if matches:
                row = matches[0]
                amount = _display_amount(row.asgari_tutar) or _display_amount(row.azami_tutar)
                rate = rate_text(row.asgari_oran or row.azami_oran)
                value = " / ".join(part for part in (amount, rate) if part)
                return (f"{value}\nBSMV hariç", row, "NUMERIC")

        elif bank == "AKBANK":
            row = exact("Hazine Bonosu / Devlet Tahvili Virmanı")
            if row:
                return (f"{_fee_value_compact(row)}\nBSMV hariç", row, "NUMERIC")

        elif bank == "YAPIKREDI":
            row = exact("Sabit Getirili Menkul Kıymet İşlem Ücretleri - Başka Kuruma Bono/Tahvil Virmanı")
            if row:
                return (f"{_fee_value_compact(row)}\nKıymet başına\nBSMV hariç", row, "NUMERIC")

        return published_gap()

    # ------------------------------------------------------------------
    # YURTİÇİ YATIRIM HESABI / SAKLAMA
    # ------------------------------------------------------------------
    if service == "SEC_DOMESTIC_CUSTODY":
        if bank == "GARANTİ":
            row = exact("Hisse Senedi Sabit Komisyon Oranı - Hisse Senedi Hesap Bakım Ücreti")
            if row:
                return (
                    f"{_fee_value_compact(row)} / 6 ay\n500 TRY / yıl\nBSMV dahil",
                    row,
                    "NUMERIC",
                )

        elif bank == "İŞBANKASI":
            matches = found(lambda row: (
                "yatirim hesabi saklama" in _norm(row.kategori)
                and "altin" not in _norm(row.kategori)
            ), numeric_only=False)
            amounts = [
                _parse_number(_clean(row.asgari_tutar).replace("TL", ""))
                for row in matches
            ]
            amounts = [value for value in amounts if value is not None]
            if matches and amounts:
                low = _display_amount(f"{min(amounts)} TL")
                high = _display_amount(f"{max(amounts)} TL")
                return (
                    f"{low} - {high} / 3 ay\n"
                    "Bakiye kademesine göre\nBSMV hariç",
                    matches[0],
                    "NUMERIC",
                )

        elif bank == "AKBANK":
            row = exact("MKK Hesap Bakım Ücreti (günlük)")
            if row:
                match = re.search(r"gunluk\s*([0-9]+(?:[.,][0-9]+)?)\s*tl", _norm(row.aciklama))
                if match:
                    exact_amount = match.group(1).replace(".", ",") + " TRY"
                    return (
                        f"{exact_amount} / gün\n"
                        "Saklama bakiyesi 1.000 TRY ve üzeri\nBSMV hariç",
                        row,
                        "NUMERIC",
                    )

        elif bank == "YAPIKREDI":
            row = exact(
                "Merkezi Kayıt Kuruluşu Ücretleri - Yatırımcı Hesabı Açma ve Bakım Hizmetleri - Hesap Bakım"
            )
            if row:
                return (
                    f"{_display_amount(row.asgari_tutar)} / gün\n"
                    "Saklama bakiyesi 1.000 TRY ve üzeri",
                    row,
                    "NUMERIC",
                )

        return published_gap()

    # ------------------------------------------------------------------
    # YURTDIŞI PAY ALIM-SATIM / SAKLAMA
    # ------------------------------------------------------------------
    if service == "SEC_FOREIGN_STOCK_TRADE":
        if bank == "İŞBANKASI":
            row = exact("Alım Satım Aracılık Hizmeti", kategori="Yurt Dışı Piyasalar - Pay İşlemleri")
            if row:
                return (
                    f"Azami {rate_text(row.azami_oran)}\n"
                    "Asgari komisyon ülke/borsaya göre\nÖrn. NYSE/NASDAQ: 1 USD",
                    row,
                    "NUMERIC",
                )
        return published_gap()

    if service == "SEC_FOREIGN_STOCK_CUSTODY":
        if bank == "İŞBANKASI":
            common = exact(
                "Pay (Saklama - Yurt Dışı) (ABD, İngiltere, İtalya, Fransa, Almanya, Avusturya)"
            )
            other = exact("Pay (Saklama - Yurt Dışı) (Diğer Ülkeler)")
            adr = exact("Pay (ADR-GDR Depo Saklama Masrafı)")
            if common and other and adr:
                return (
                    f"Belirtilen ülkeler: {rate_text(common.azami_oran)} / ay\n"
                    f"Diğer ülkeler: {rate_text(other.azami_oran)} / ay\n"
                    "ADR/GDR: 0,03 USD/pay, 6 ayda bir",
                    common,
                    "NUMERIC",
                )
        return published_gap()

    # ------------------------------------------------------------------
    # EUROBOND VİRMAN / SAKLAMA
    # ------------------------------------------------------------------
    if service == "SEC_EUROBOND_TRANSFER":
        if bank == "İŞBANKASI":
            row = exact("SWIFT - Eurobond virmanı")
            if row:
                return (
                    "20 USD/EUR\nEk SWIFT: 10 USD/EUR\nBSMV hariç",
                    row,
                    "NUMERIC",
                )
        elif bank == "AKBANK":
            matches = found(lambda row: _norm(row.masraf) == "eurobond virmani")
            if matches:
                return ("20 USD veya 20 EUR\nBSMV hariç", matches[0], "NUMERIC")
        elif bank == "YAPIKREDI":
            row = exact("Sabit Getirili Menkul Kıymet İşlem Ücretleri - Başka Kuruma Eurobond Virmanı")
            if row:
                return (f"{_fee_value_compact(row)}\nKıymet başına\nBSMV hariç", row, "NUMERIC")
        return published_gap()

    if service == "SEC_EUROBOND_CUSTODY":
        if bank == "İŞBANKASI":
            row = exact("Eurobond saklama masrafı")
            if row:
                return ("Ücretsiz / 0 USD-EUR\nBSMV hariç", row, "NUMERIC")
        return published_gap()

    return None


def _user_audit_override(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, Optional[FeeRow], str]]:
    """Kullanıcı tarafından tek tek denetlenen yüksek riskli aileler."""
    for resolver in (
        _audit_fatura_methods,
        _audit_akbank_havale,
        _audit_sgk,
        _audit_sans_oyunu,
        _audit_regular_transfer,
        _audit_aidat,
        _audit_school,
        _audit_phone,
        _audit_vergi,
        _audit_document_statement_research,
        _audit_research_and_securities,
    ):
        result = resolver(rows, bank, spec, wanted_channel)
        if result is not None:
            return result

    for resolver in (
        _audit_atm,
        _audit_hgs,
        _audit_normal_kasa,
        _audit_isbank_special_kasa,
        _audit_cheque,
        _audit_kkb,
        _audit_senet_tahsil,
    ):
        result = resolver(rows, bank, spec)
        if result is not None:
            return result

    return None


def _aggregate_service_fee(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Optional[Tuple[str, FeeRow]]:
    """
    Çok satırlı tarifeleri tek hücrede özetler.

    Yurt dışı transferlerde eski "ilk 5 adayı bas" mantığı kullanılmaz.
    Önce banka bazında tek kanonik ürün ailesi seçilir; yalnız o ürünün
    gerçek tutar bantları gösterilir.
    """
    transfer_services = {
        "SWIFT_GELEN",
        "SWIFT_GIDEN",
        "YURT_DISI_FAST",
        "KART_YURTDISI_TRANSFER",
    }

    if spec.service in transfer_services:
        candidates = _canonical_transfer_rows(
            rows,
            bank,
            spec,
            wanted_channel,
        )

        if not candidates:
            return None

        canonical_rows = [row for _, row in candidates]
        value = _compact_transfer_blocks(
            canonical_rows,
            spec,
        )

        if not value:
            return None

        if spec.service == "SWIFT_GELEN":
            channel_label = (
                "Şube"
                if bank == "GARANTİ"
                else "Genel / kanal belirtilmemiş"
            )
            value = f"{channel_label}:\n{value}"

        return value, canonical_rows[0]

    # Altın transferinde eski aggregate davranışı korunur; burada aynı
    # ürünün gram/tutar kademeleri gerçekten aynı hizmetin parçasıdır.
    if spec.service != "ALTIN_TRANSFER":
        return None

    candidates = []

    for row in rows:
        if row.banka != bank:
            continue

        score = _candidate_score(row, spec, "GENEL")
        if score <= -10_000 or not _has_numeric_fee(row):
            continue

        mas = _norm(row.masraf)

        if not any(x in mas for x in (
            "altin transfer",
            "ats ile altin gonderimi",
            "kiymetli maden transferi ucreti - altin",
        )):
            continue

        if any(x in mas for x in (
            "western union", "eft", "havale", "fast", "fiziki", "teslim",
        )):
            continue

        candidates.append((score, row))

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: (
            item[0],
            -len(_norm(item[1].masraf)),
        ),
        reverse=True,
    )

    blocks: List[str] = []
    seen: Set[str] = set()
    first_row = candidates[0][1]

    for _, row in candidates:
        fee = _fee_value_compact(row)
        if not fee:
            continue

        hint = _compact_hint(row)
        block = f"{hint}: {fee}" if hint else fee

        tax = _bsmv_label(row)
        if tax:
            block += f" ({tax})"

        key = _norm(block)
        if key in seen:
            continue

        seen.add(key)
        blocks.append(block)

        if len(blocks) >= 5:
            break

    if not blocks:
        return None

    return "\n".join(blocks), first_row


def _kasa24_special_fee(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec
) -> Optional[Tuple[str, FeeRow]]:
    """Yapı Kredi Özel/Süper satırında Kasa24 A-E tarifesini dürüstçe özetler."""
    if bank != "YAPIKREDI" or spec.service != "KASA" or spec.detail != "OZEL":
        return None

    blocks: List[str] = []
    used_rows: List[FeeRow] = []
    first: Optional[FeeRow] = None

    for kasa_type in ("A", "B", "C", "D", "E"):
        token = _norm(f"{kasa_type} Tipi Kasa24")
        annual = None
        deposit = None
        for row in rows:
            if row.banka != bank:
                continue
            full = _norm(row.text)
            if "kasa24" not in full or token not in full:
                continue
            if "depozito" in full:
                deposit = row
            elif "yillik" in full or "kasa24 ucreti" in full:
                annual = row

        if annual is None:
            continue
        first = first or annual
        used_rows.append(annual)
        annual_fee = _fee_value_compact(annual)
        dep_fee = _fee_value_compact(deposit) if deposit else ""
        if deposit is not None:
            used_rows.append(deposit)
        line = f"Kasa24 {kasa_type}: Yıllık {annual_fee}"
        if dep_fee:
            line += f" | Depozito {dep_fee}"
        blocks.append(line)

    if not blocks or first is None:
        return None
    tax = _audit_common_tax(used_rows)
    if tax:
        blocks.append(tax)
    return "\n".join(blocks), first


def _source_url(row: Optional[FeeRow]) -> str:
    if row is None:
        return ""
    m = re.search(r"Resmî ek kaynak:\s*(https?://[^\s|]+)", row.aciklama or "", flags=re.I)
    if m:
        return m.group(1).rstrip(".,;")
    return PRIMARY_SOURCE_URLS.get(row.banka, "")


def _cell_comment(row: Optional[FeeRow], resolution: str, value: str) -> Optional[Comment]:
    if row is None:
        return None
    parts = [
        f"Çözüm türü: {resolution}",
        f"Banka: {row.banka}",
        f"Kaynak masraf: {row.masraf}",
    ]
    if row.kategori:
        parts.append(f"Kategori: {row.kategori}")
    if row.guncelleme_tarihi:
        parts.append(f"Kaynak güncelleme: {row.guncelleme_tarihi}")
    url = _source_url(row)
    if url:
        parts.append(f"Kaynak: {url}")
    return Comment("\n".join(parts), "Otomatik Eşleştirme")


def _source_gap_text(spec: RowSpec, bank: str, wanted_channel: str) -> str:
    # "Kontrol gerekli" = hizmet yok / ücretsiz demek değildir.
    # Yalnız güvenli bir karşılaştırılabilir tarife seçilemediğini gösterir.
    return "Kontrol gerekli"


def _assert_safe_source(row: Optional[FeeRow], spec: RowSpec) -> None:
    """
    Son güvenlik ağı.

    Eşleştirme kurallarında ileride yapılacak bir değişiklik yanlış ürünü
    seçerse final Excel yayınlanmadan pipeline'ı durdurur.
    """
    if row is None:
        return

    structural = _norm(f"{row.kategori} | {row.masraf}")
    mas = _norm(row.masraf)

    if spec.service == "HGS_ETIKET":
        if not ("hgs" in structural and "etiket" in structural and "kart" not in mas):
            raise RuntimeError(
                f"Güvenlik: HGS Etiket için yanlış kaynak seçildi: {row.masraf}"
            )

    elif spec.service == "HGS_KART":
        if not ("hgs" in structural and "kart" in structural and "etiket" not in mas):
            raise RuntimeError(
                f"Güvenlik: HGS Kart için yanlış kaynak seçildi: {row.masraf}"
            )

    elif spec.service == "CEK_DEFTERI_YAPRAK":
        if not (
            any(x in structural for x in ("cek defteri", "cek karnesi"))
            and any(x in structural for x in ("yaprak basi", "50-350 yaprakli", "351 yaprak"))
        ):
            raise RuntimeError(
                f"Güvenlik: Çek defteri yaprak-başı için yanlış kaynak: {row.masraf}"
            )

    elif spec.service == "CEK_KARNESI_10":
        if not (
            any(x in structural for x in ("cek defteri", "cek karnesi"))
            and any(x in structural for x in ("10 yaprakli", "10'luk", "10luk"))
        ):
            raise RuntimeError(
                f"Güvenlik: 10 yapraklı çek karnesi için yanlış kaynak: {row.masraf}"
            )

    elif spec.service == "CEK_DUZENLEME_STANDART":
        forbidden = (
            "hediye", "armagan", "seyahat", "doviz", "dovizi natik",
            "dth", "karsilikli", "odeme", "durdurma",
        )
        if (
            not any(x in structural for x in (
                "bloke cek duzenleme",
                "bloke/ keside ceki duzenleme",
                "bloke/keside ceki duzenleme",
                "keside ceki / bloke cek duzenleme",
                "duzenleme - tp/yp",
            ))
            or any(x in structural for x in forbidden)
        ):
            raise RuntimeError(
                f"Güvenlik: Standart çek düzenleme için yanlış kaynak: {row.masraf}"
            )

    elif spec.service == "CEK_DUZENLEME_DOVIZ":
        if not any(x in structural for x in (
            "dovizli cek duzenleme",
            "dovizi natik cek duzenleme",
            "dth'dan cek duzenlenmesi",
            "dth dan cek duzenlenmesi",
        )):
            raise RuntimeError(
                f"Güvenlik: Dövizli çek düzenleme için yanlış kaynak: {row.masraf}"
            )

    elif spec.service == "CEK_TAHSIL":
        if any(x in structural for x in (
            "gişeden cek odeme", "giseden cek odeme", "bloke cek odeme",
            "karsiliksiz cek elden odeme", "seyahat ceki odeme",
        )):
            raise RuntimeError(
                f"Güvenlik: Çek tahsilatına çek ödeme satırı karıştı: {row.masraf}"
            )

    elif spec.service == "SENET_IADE":
        if (
            "senet" not in mas
            or "iade" not in mas
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            raise RuntimeError(
                f"Güvenlik: Senet iade için yanlış kaynak seçildi: {row.masraf}"
            )

    elif spec.service == "SENET_PROTESTO":
        if (
            "senet" not in mas
            or "protesto" not in mas
            or "protestosuz" in mas
            or "kaldir" in mas
            or any(x in mas for x in ("iskonto", "istira", "teminat"))
        ):
            raise RuntimeError(
                f"Güvenlik: Senet protesto için yanlış kaynak seçildi: {row.masraf}"
            )

    elif spec.service == "SENET_PROTESTO_KALDIRMA":
        if not ("senet" in mas and "protesto" in mas and "kaldir" in mas):
            raise RuntimeError(
                f"Güvenlik: Senet protesto kaldırma için yanlış kaynak: {row.masraf}"
            )

    elif spec.service == "HESAP_OZETI_POSTA":
        if any(x in structural for x in ("kredi kart", "ticari kart", "business kart")):
            raise RuntimeError(
                f"Güvenlik: Hesap özeti posta satırına kart ekstresi karıştı: {row.masraf}"
            )

    elif spec.service in {
        "SWIFT_GELEN", "SWIFT_GIDEN", "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER"
    }:
        if any(x in mas for x in ("paket", "kobi", "kota")):
            raise RuntimeError(
                f"Güvenlik: Tekil transfer tarifesine paket/kota karıştı: {row.masraf}"
            )

        # Yön/ürün güvenliği. Aggregate tarafı daha da dar banka-bazlı
        # filtre uygular; burası son savunma katmanıdır.
        if spec.service == "SWIFT_GELEN" and any(x in mas for x in (
            "arastirma", "iade", "isme", "kasaya", "giden",
        )):
            raise RuntimeError(
                f"Güvenlik: Gelen SWIFT'e farklı varyant karıştı: {row.masraf}"
            )

        if spec.service == "YURT_DISI_FAST" and any(x in _norm(row.text) for x in (
            "moneysend", "western union", "karta para gonder",
            "visa ile yurt disi para transferi",
            "visa ile yurtdisi para transferi",
        )):
            raise RuntimeError(
                f"Güvenlik: Hızlı hesaba transfere farklı ürün karıştı: {row.masraf}"
            )

        if spec.service == "KART_YURTDISI_TRANSFER" and any(x in _norm(row.text) for x in (
            "alici", "kktc",
        )):
            raise RuntimeError(
                f"Güvenlik: Karta transfere alıcı/KKTC varyantı karıştı: {row.masraf}"
            )


def _resolve_cell(
    rows: Sequence[FeeRow], bank: str, spec: RowSpec, wanted_channel: str,
) -> Tuple[str, Optional[FeeRow], str]:
    """Hücre metni, dayanak satır ve çözüm türünü döndürür."""

    lookup_channel = wanted_channel if spec.split_channel else "GENEL"

    # Kullanıcı tarafından doğrulanmış yüksek riskli aileler önce çözülür.
    # Böylece genel/fuzzy fallback doğru ilk-sayfa verisinin önüne geçemez.
    audited = _user_audit_override(rows, bank, spec, wanted_channel)
    if audited is not None:
        return audited

    # Resmî olarak uygulanmadığı/kanalda yayımlanmadığı doğrulanan durum,
    # aynı isimli EFT tarife satırının FAST'e yanlış taşınmasını engellemek için
    # sayısal eşleşmeden ÖNCE değerlendirilir.
    pre_status = _service_status_row(
        rows, bank, spec, lookup_channel, kinds={"NOT_APPLICABLE"}
    )
    if pre_status is not None:
        return _fee_text(pre_status, spec), pre_status, "NOT_APPLICABLE"

    # Yapı Kredi Kasa24, klasik Büyük/Orta/Küçük kasa ailesinden farklı
    # yayımlandığı için Özel/Süper satırında A-E birlikte gösterilir.
    kasa24 = _kasa24_special_fee(rows, bank, spec)
    if kasa24 is not None:
        value, source = kasa24
        return value, source, "NUMERIC"

    aggregated = _aggregate_service_fee(rows, bank, spec, wanted_channel)
    if aggregated is not None:
        value, agg_row = aggregated
        _assert_safe_source(agg_row, spec)
        return value, agg_row, "NUMERIC"

    row = _best_match(rows, bank, spec, lookup_channel)

    if row is None and not spec.split_channel:
        found = [_best_match(rows, bank, spec, possible) for possible in ("GENEL", "MOBIL", "SUBE")]
        row = next((x for x in found if x is not None), None)

    # Status satırı sayısal ücretin önüne geçmesin; NOT_APPLICABLE yukarıda
    # zaten özel olarak ele alındı.
    if row is not None and _has_numeric_fee(row):
        _assert_safe_source(row, spec)
        value = _fee_text(row, spec)
        if spec.service == "FAST" and spec.band_key == "TRANSFER_2":
            limit_note = _fast_limit_note(rows, bank, wanted_channel)
            if limit_note:
                value += f"\n{limit_note}"
        if spec.service == "KASA":
            annual_tax = _bsmv_label(row)
            annual_fee = _fee_value_compact(row)
            annual = f"Yıllık: {annual_fee}" if annual_fee else "Yıllık: ücret bilgisi açıklamada"
            if annual_tax:
                annual += f" ({annual_tax})"
            dep_value, dep_tax = _best_deposit_match(rows, bank, spec, row)
            value = annual
            if dep_value:
                value += f"\nDepozito: {dep_value}"
                if dep_tax:
                    value += f" ({dep_tax})"
        return value, row, "NUMERIC"

    # Aidat/okul: hizmetin resmî varlık kanıtı varsa, bankanın genel
    # Fatura/Kurum tarifesi açıkça "genel tarife" etiketiyle kullanılabilir.
    if spec.service in {"AIDAT", "OZEL_OKUL", "TELEFON"}:
        status_row = _service_status_row(rows, bank, spec, lookup_channel)
        if status_row is not None:
            generic = _generic_institution_fee(rows, bank, lookup_channel)
            if generic is not None:
                fee = _fee_text(generic, None)
                return f"Genel kurum tarifesi:\n{fee}", generic, "GENERIC_TARIFF"
            return _fee_text(status_row, spec), status_row, "STATUS"

    if row is not None and _status_kind(row):
        return _fee_text(row, spec), row, "STATUS"

    status_row = _service_status_row(rows, bank, spec, lookup_channel)
    if status_row is not None:
        kind = "NOT_APPLICABLE" if _status_kind(status_row) == "NOT_APPLICABLE" else "STATUS"
        return _fee_text(status_row, spec), status_row, kind

    fast_status = _fast_branch_publication_status(rows, bank, spec, wanted_channel)
    if fast_status:
        return fast_status, None, "PUBLICATION_STATUS"

    # Son çare olarak veri uydurmak yerine açıkça kaynak boşluğu belirtilir.
    # Böylece N/A'nın "0 TL" veya "hizmet yok" sanılması engellenir.
    return _source_gap_text(spec, bank, wanted_channel), None, "SOURCE_GAP"



def _preserve_notes(old_ws) -> Dict[Tuple[str, str], str]:
    """Notları satır numarasına değil ORTAK satır adına göre korur."""
    notes: Dict[Tuple[str, str], str] = {}
    if old_ws is None:
        return notes

    section = ""
    for row in range(3, old_ws.max_row + 1):
        label = _clean(old_ws.cell(row=row, column=1).value)
        note = _clean(old_ws.cell(row=row, column=13).value)
        if not label:
            continue

        # Section satırları bankalarda veri taşımadığı için kalın başlık gibi davranır.
        if label in {payload for kind, payload in LAYOUT if kind == "SECTION"}:
            section = label
            continue

        if note and not _looks_like_generated_note(note):
            notes[(section, label)] = note

    return notes



# ---------------------------------------------------------------------------
# NOTLAR SÖZLÜĞÜ
# ---------------------------------------------------------------------------
# NOTLAR hücresinde yalnız satırda gerçekten geçen kavramların kısa anlamı
# gösterilir. Metinler özellikle kısa ve düz tutulur; kalın/rich-text yoktur.
_AUTO_NOTE_LABELS = {
    "BSMV dahil",
    "BSMV hariç",
    "Genel Fatura/Kurum tarifesi",
    "Genel kurum tarifesi",
    "Genel tarife",
    "Genel gişe tarifesi",
    "Kanal ayrımı yayımlanmıyor",
    "Ödeme aracı ayrımı yayımlanmıyor",
    "Hesap/nakit ayrımı yayımlanmıyor",
    "Nakit/hesap ayrımı yayımlanmıyor",
    "Ayrı ücret yayımlanmıyor",
    "Bu kanalda tarife yayımlanmıyor",
    "Hizmet var",
    "Ücretsiz / 0 TRY",
    "Kuruma/işleme göre",
    "Firmaya göre",
    "Fatura türüne göre",
    "Asgari",
    "Azami",
    "Valör",
    "YP",
    "Depozito",
    "Değerli kağıt bedeli hariç",
}


def _append_note(notes: List[str], label: str, explanation: str) -> None:
    line = f"• {label}: {explanation}"
    if line not in notes:
        notes.append(line)


def _automatic_notes(values: Sequence[str]) -> str:
    """
    B:I hücrelerinde kullanılan terimleri algılar ve M sütununda
    tek tek kısa açıklama listesine dönüştürür.
    """
    raw = "\n".join(_clean(v) for v in values if _clean(v))
    n = _norm(raw)
    notes: List[str] = []

    if "bsmv dahil" in n:
        _append_note(
            notes,
            "BSMV dahil",
            "BSMV gösterilen tutarın içindedir.",
        )
    if "bsmv haric" in n:
        _append_note(
            notes,
            "BSMV hariç",
            "BSMV gösterilen tutara dahil değildir; ayrıca eklenebilir.",
        )

    # En spesifik genel-tarife ifadesini kullan; aynı satırda gereksiz tekrar yaratma.
    if "genel fatura/kurum tarifesi" in n:
        _append_note(
            notes,
            "Genel Fatura/Kurum tarifesi",
            "Gösterilen ücret bu hizmete özel değil, bankanın genel fatura/kurum tarifesidir.",
        )
    elif "genel kurum tarifesi" in n:
        _append_note(
            notes,
            "Genel kurum tarifesi",
            "Gösterilen ücret bu alt hizmete özel değil, genel kurum ödeme tarifesidir.",
        )
    elif "genel tarife" in n:
        _append_note(
            notes,
            "Genel tarife",
            "Banka işlemi daha alt türlere ayırmadan ortak bir ücret yayımlamıştır.",
        )

    if "genel gise tarifesi" in n:
        _append_note(
            notes,
            "Genel gişe tarifesi",
            "Şubede/gişede yapılan işlem için yayımlanan genel ücrettir.",
        )

    if "kanal/odeme araci ayrimi yayimlanmiyor" in n:
        _append_note(
            notes,
            "Kanal ayrımı yayımlanmıyor",
            "Banka Mobil ve Şube için ayrı ücret açıklamamıştır.",
        )
        _append_note(
            notes,
            "Ödeme aracı ayrımı yayımlanmıyor",
            "Hesap, nakit veya kredi kartı için ayrı fiyat açıklanmamıştır.",
        )
    else:
        if "kanal ayrimi yayimlanmiyor" in n:
            _append_note(
                notes,
                "Kanal ayrımı yayımlanmıyor",
                "Banka Mobil ve Şube için ayrı ücret açıklamamıştır.",
            )
        if "odeme araci ayrimi yayimlanmiyor" in n:
            _append_note(
                notes,
                "Ödeme aracı ayrımı yayımlanmıyor",
                "Hesap, nakit veya kredi kartı için ayrı fiyat açıklanmamıştır.",
            )

    if "hesap/nakit ayrimi yayimlanmiyor" in n:
        _append_note(
            notes,
            "Hesap/nakit ayrımı yayımlanmıyor",
            "Banka hesaptan ve nakit ödemeyi ayrı fiyatlandırmamıştır.",
        )
    if "nakit/hesap ayrimi yayimlanmiyor" in n:
        _append_note(
            notes,
            "Nakit/hesap ayrımı yayımlanmıyor",
            "Banka nakit ve hesaptan ödemeyi ayrı fiyatlandırmamıştır.",
        )

    # "Ayrı ... ücret/tarife yayımlanmıyor" türlerinin tümü tek açıklamada birleşir.
    if (
        "ayri aidat ucreti yayimlanmiyor" in n
        or "aidata ozel ucret yayimlanmiyor" in n
        or "ayri ozel okul ucreti yayimlanmiyor" in n
        or "ayri ozel okul/egitim ucreti yayimlanmiyor" in n
        or "ayri ucret tutari belirtilmemis" in n
        or "ayri ucret ilan edilmemis" in n
        or "ayri vergi ucreti yayimlanmiyor" in n
    ):
        _append_note(
            notes,
            "Ayrı ücret yayımlanmıyor",
            "Hizmet veya kalem kaynakta var, ancak ona özel sayısal ücret açıklanmamıştır.",
        )

    if (
        "bu kanalda tarife yayimlanmiyor" in n
        or "bu kanalda ozel okul odeme tarifesi yayimlanmiyor" in n
        or "sube icin ayri" in n
        or "mobil icin ayri" in n
    ):
        _append_note(
            notes,
            "Bu kanalda tarife yayımlanmıyor",
            "İlgili kanal için ayrı ve güvenilir bir ücret tarifesi yayımlanmamıştır.",
        )

    if "hizmet var" in n:
        _append_note(
            notes,
            "Hizmet var",
            "Bankanın bu işlemi sunduğu resmî kaynakta doğrulanmıştır.",
        )

    if "ucretsiz / 0 try" in n or "ucretsiz/0 try" in n:
        _append_note(
            notes,
            "Ücretsiz / 0 TRY",
            "Banka bu işlemden ücret alınmadığını belirtmektedir.",
        )

    if "kuruma/isleme gore" in n:
        _append_note(
            notes,
            "Kuruma/işleme göre",
            "Ücret, ödeme yapılan kurum veya işlem türüne göre değişebilir.",
        )
    if "firmaya gore" in n:
        _append_note(
            notes,
            "Firmaya göre",
            "Ücret, ödeme yapılan firmaya göre değişebilir.",
        )
    if "fatura turune gore" in n:
        _append_note(
            notes,
            "Fatura türüne göre",
            "Ücret, ödenen fatura türüne göre değişebilir.",
        )

    # Sık kullanılan finansal terimler.
    if re.search(r"\basgari\b|\bmin\b", n):
        _append_note(
            notes,
            "Asgari",
            "Uygulanabilecek en düşük ücret veya tutardır.",
        )
    if re.search(r"\bazami\b|\bmax\b", n):
        _append_note(
            notes,
            "Azami",
            "Uygulanabilecek en yüksek ücret veya tutardır.",
        )
    if "valor" in n:
        _append_note(
            notes,
            "Valör",
            "Transferin hesaba değer kazanacağı/işleme alınacağı günü ifade eder.",
        )
    if re.search(r"(^|[^a-z])yp([^a-z]|$)", n):
        _append_note(
            notes,
            "YP",
            "Yabancı para anlamına gelir.",
        )
    if "depozito" in n:
        _append_note(
            notes,
            "Depozito",
            "Kiralık kasa için ayrıca alınabilen güvence bedelidir.",
        )
    if "degerli kagit" in n and "haric" in n:
        _append_note(
            notes,
            "Değerli kağıt bedeli hariç",
            "Gösterilen çek ücretine resmî değerli kağıt bedeli dahil değildir.",
        )

    return "\n".join(notes)


def _looks_like_generated_note(note: str) -> bool:
    """Önceki V25 çalışmasının otomatik sözlük notlarını manuel not sanma."""
    lines = [x.strip() for x in str(note or "").splitlines() if x.strip()]
    if not lines:
        return False
    if not all(line.startswith("• ") and ":" in line for line in lines):
        return False

    heads = {line[2:].split(":", 1)[0].strip() for line in lines}
    return bool(heads) and heads.issubset(_AUTO_NOTE_LABELS)


def _compose_notes(auto_note: str, manual_note: str) -> str:
    """
    Otomatik terim açıklamalarını ve varsa gerçek manuel notu aynı hücrede,
    yine madde madde gösterir.
    """
    lines = [x.strip() for x in str(auto_note or "").splitlines() if x.strip()]

    manual_note = _clean(manual_note)
    if manual_note and not _looks_like_generated_note(manual_note):
        for raw_line in str(manual_note).splitlines():
            clean_line = raw_line.strip()
            if not clean_line:
                continue
            if clean_line.startswith("• "):
                line = clean_line
            else:
                line = f"• Not: {clean_line}"
            if line not in lines:
                lines.append(line)

    return "\n".join(lines)



# ---------------------------------------------------------------------------
# NOTLAR / TERİM AÇIKLAMALARI
# ---------------------------------------------------------------------------
# Kullanıcının istediği sabit sözlük. NOTLAR sütununda her açıklama ayrı satır
# halinde, düz (bold olmayan) fontla gösterilir.
NOTES_GLOSSARY: Sequence[Tuple[str, str]] = (
    ("Ayrı ücret tutarı yayımlanmıyor.", "Banka bu hizmeti sunuyor ancak bu işlem için özel bir ücret rakamı açıklamıyor."),
    ("Ücret tutarı belirtilmemiş.", "İşlem bankanın tarifesinde yer alıyor fakat ilgili satırda sayısal ücret bulunmuyor."),
    ("Hizmet var / ayrı tarife yayımlanmıyor.", "İşlemin yapılabildiği doğrulandı ancak yalnız bu hizmete ait ayrı bir fiyat tarifesi yok."),
    ("Genel Fatura/Kurum tarifesi uygulanır.", "Bu işlem için özel fiyat yerine bankanın genel fatura/kurum ödeme tarifesi kullanılıyor."),
    ("Genel Fatura/Kurum tarifesinden eşleştirilmiştir.", "Gösterilen ücret özel olarak bu hizmetin değil, onu kapsayan genel tarifenin ücretidir."),
    ("Genel tarife.", "Banka işlemi daha alt türlere ayırmadan tek ortak ücret tarifesi yayımlıyor."),
    ("Genel gişe tarifesi.", "Ücret şubede/gişede yapılan işlem için verilmiş, ödeme şekli ayrıca ayrılmamış."),
    ("Kanal ayrımı yayımlanmıyor.", "Banka Mobil, İnternet veya Şube için ayrı ayrı ücret belirtmiyor."),
    ("Hesap/Nakit ayrımı yayımlanmıyor.", "Banka ödemenin hesaptan mı nakit mi yapıldığına göre ayrı ücret açıklamıyor."),
    ("Ödeme aracı ayrımı yayımlanmıyor.", "Banka hesap, nakit veya kredi kartı kullanımına göre tarifeyi ayırmıyor."),
    ("Aidata özel ücret yayımlanmıyor.", "Aidat ödemesi yapılabiliyor ancak yalnız aidata özel bir komisyon tarifesi yok."),
    ("Özel okul ücreti yayımlanmıyor.", "Özel okul ödemesi yapılabiliyor fakat bu işleme özel komisyon rakamı açıklanmıyor."),
    ("Genel SGK tarifesi – kanal ayrımı yayımlanmıyor.", "SGK için ücret var fakat banka Mobil ve Şube ücretlerini ayrı göstermiyor."),
    ("Ücretsiz / 0 TRY.", "Bankanın resmî açıklamasına göre bu işlemden ücret alınmıyor."),
    ("Bu kanalda hizmet sunulmuyor.", "İşlem ilgili Mobil veya Şube kanalından yapılamıyor."),
    ("Bu kanal için ayrı tarife yayımlanmıyor.", "İşlem kanalda olabilir ancak o kanala özel ücret açıklanmamış."),
    ("Kaynakta sayısal tarife bulunamadı.", "Resmî kaynakta hizmete ilişkin güvenilir bir ücret rakamı tespit edilemedi."),
    ("Kaynakta tutar görünmüyor.", "Bankanın ilgili satırı mevcut fakat ücret alanı boş."),
    ("Kanal belirtilmemiş.", "Ücret yayımlanmış ancak hangi işlem kanalında geçerli olduğu belirtilmemiş."),
    ("Kuruma/işleme göre değişir.", "Ücret sabit değil; ödeme yapılan kurum veya işlem türüne göre belirtilen sınırlar içinde değişebilir."),
    ("Fatura türüne göre değişir.", "Telefon, elektrik, su vb. fatura türüne göre uygulanacak ücret farklı olabilir."),
    ("BSMV dahil.", "Gösterilen tutarın içinde BSMV zaten vardır."),
    ("BSMV hariç.", "Gösterilen tutara işlem sırasında ayrıca BSMV eklenebilir."),
    ("KDV dahil.", "Gösterilen ücretin içine KDV dahildir."),
    ("Değerli kağıt bedeli hariç.", "Gösterilen çek ücreti dışında ayrıca değerli kağıt bedeli alınabilir."),
    ("Bkz. FATURA bölümü.", "Aynı tarife tekrar yazılmadı; detaylı ücret Fatura bölümünde gösteriliyor."),
)


def _write_notes_glossary(ws, thin: Side) -> None:
    """NOTLAR sütununu sabit terim sözlüğü olarak yazar."""
    start_row = 3

    # Eski otomatik/manüel not kalıntılarını temizle.
    for r in range(start_row, max(ws.max_row, start_row + len(NOTES_GLOSSARY)) + 1):
        c = ws.cell(row=r, column=13)
        c.value = None
        c.comment = None
        c.font = Font(bold=False, italic=False, color="595959", size=8.5)
        c.fill = PatternFill("solid", fgColor="FFF9E6")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for offset, (term, meaning) in enumerate(NOTES_GLOSSARY):
        row = start_row + offset
        c = ws.cell(row=row, column=13)
        c.value = f'• “{term}” → {meaning}'
        c.font = Font(bold=False, italic=False, color="595959", size=8.5)
        c.fill = PatternFill("solid", fgColor="FFF9E6")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Açıklamalar tek tek rahat okunsun; mevcut daha yüksek satırları küçültme.
        current_height = ws.row_dimensions[row].height or 15
        ws.row_dimensions[row].height = max(current_height, 34)


def _sheet_row_key(section: str, label: str) -> Tuple[str, str]:
    return (section, label)


def _write_comparison(ws, rows: Sequence[FeeRow], notes: Mapping[Tuple[str, str], str]) -> int:
    thin = Side(style="thin", color="B7B7B7")
    medium = Side(style="medium", color="7F7F7F")

    ws["M1"] = "NOTLAR"
    ws.merge_cells("M1:M2")
    ws["M1"].font = Font(bold=True, color="1F1F1F")
    ws["M1"].fill = PatternFill("solid", fgColor="D9EAD3")
    ws["M1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = "ORTAK MASRAF ADI"
    ws["A2"].font = Font(bold=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    col = 2
    for bank in BANKS:
        color = BANK_COLORS[bank]
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        cell = ws.cell(row=1, column=col)
        cell.value = DISPLAY_BANKS[bank]
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for offset, sub in enumerate(("Mobil", "Şube")):
            c = ws.cell(row=2, column=col + offset)
            c.value = sub
            c.fill = PatternFill("solid", fgColor=color)
            c.font = Font(bold=True, italic=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
        col += 2

    current_row = 3
    section = ""

    for kind, payload in LAYOUT:
        if kind == "SECTION":
            section = payload
            ws.cell(row=current_row, column=1).value = payload
            for col_idx in range(1, 10):
                c = ws.cell(row=current_row, column=col_idx)
                c.fill = PatternFill("solid", fgColor="E7E6E6")
                c.font = Font(bold=True, color="595959")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border(top=medium, bottom=thin)

            ws.row_dimensions[current_row].height = 24
            current_row += 1
            continue

        spec: RowSpec = payload
        ws.cell(row=current_row, column=1).value = spec.label
        ws.cell(row=current_row, column=1).font = Font(color="595959")
        ws.cell(row=current_row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

        for bank_index, bank in enumerate(BANKS):
            start_col = 2 + bank_index * 2

            if spec.split_channel:
                for offset, wanted_channel in enumerate(("MOBIL", "SUBE")):
                    value, row, resolution = _resolve_cell(rows, bank, spec, wanted_channel)
                    c = ws.cell(row=current_row, column=start_col + offset)
                    c.value = value
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.font = Font(
                        bold=resolution not in {"SOURCE_GAP"},
                        color=BANK_COLORS[bank] if resolution != "SOURCE_GAP" else "A6A6A6",
                        size=9,
                    )
                    comment = _cell_comment(row, resolution, value)
                    if comment:
                        c.comment = comment
            else:
                # Kanal ayrımı anlamsız olan kasa, çek, senet, rapor vb. satırlarda
                # Mobil/Şube hücrelerini tek mantıksal banka hücresine birleştir.
                ws.merge_cells(
                    start_row=current_row, start_column=start_col,
                    end_row=current_row, end_column=start_col + 1,
                )
                value, row, resolution = _resolve_cell(rows, bank, spec, "GENEL")
                c = ws.cell(row=current_row, column=start_col)
                c.value = value
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.font = Font(
                    bold=resolution not in {"SOURCE_GAP"},
                    color=BANK_COLORS[bank] if resolution != "SOURCE_GAP" else "A6A6A6",
                    size=9,
                )
                comment = _cell_comment(row, resolution, value)
                if comment:
                    c.comment = comment

        # NOTLAR sütunu satır-bazlı otomatik not yerine sabit sözlük/sidebar
        # olarak kullanılır. Buradaki hücreler döngü sonunda topluca yazılır.
        note_value = ""

        if spec.service == "KASA" and spec.detail == "OZEL":
            ws.row_dimensions[current_row].height = 105
        elif spec.service == "KASA":
            ws.row_dimensions[current_row].height = 72
        elif spec.service == "CEK_DEFTERI_GRUP":
            ws.row_dimensions[current_row].height = 150
        elif spec.service in {"CEK_DUZENLEME_OZEL_GRUP", "CEK_TAHSIL_DOVIZ_GRUP"}:
            ws.row_dimensions[current_row].height = 100
        elif spec.service in {"DUZENLI_EFT", "DUZENLI_HAVALE"}:
            ws.row_dimensions[current_row].height = 90
        elif spec.service == "DOC_ESKI_DEKONT_EKSTRE":
            ws.row_dimensions[current_row].height = 92
        elif spec.service in _DOCUMENT_STATEMENT_SERVICES:
            ws.row_dimensions[current_row].height = 72
        elif spec.service in _RESEARCH_SECURITIES_SERVICES:
            ws.row_dimensions[current_row].height = 88
        elif spec.service in {
            "SWIFT_GELEN",
            "SWIFT_GIDEN",
            "YURT_DISI_FAST",
            "KART_YURTDISI_TRANSFER",
            "HGS_ETIKET",
            "HGS_KART",
            "SANS_OYUNU",
            "BAKIYE_ATM_YURTDISI",
        }:
            ws.row_dimensions[current_row].height = 66
        else:
            ws.row_dimensions[current_row].height = 52

        # Not listesinin alt satırları görünür kalsın; gereksiz dev satır oluşturma.
        if note_value:
            note_lines = len(note_value.splitlines())
            ws.row_dimensions[current_row].height = max(
                ws.row_dimensions[current_row].height or 52,
                min(108, 18 + note_lines * 13),
            )

        current_row += 1

    # Sağdaki NOTLAR alanı: kullanıcının istediği sabit terim açıklamaları.
    _write_notes_glossary(ws, thin)

    for row_cells in ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=13):
        for cell in row_cells:
            if cell.column <= 9 and cell.row > 2:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 44
    for col_letter in "BCDEFGHI":
        ws.column_dimensions[col_letter].width = 18
    for col_letter in "JKL":
        ws.column_dimensions[col_letter].width = 3
    ws.column_dimensions["M"].width = 58

    # Birleşik banka başlıklarını bozmadan gerçek Excel filtresi.
    # A2:I... aralığı kullanılır: A=masraf adı, B:I=4 bankanın Mobil/Şube sütunları.
    ws.auto_filter.ref = f"A2:I{current_row - 1}"

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "B3"
    ws.sheet_view.zoomScale = 80
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Teknik eşleşme özeti artık görünür Excel alanına yazılmaz.
    # Ayrıntılı kalite kontrolü yalnız GitHub Actions logunda basılır.
    return current_row - 1



def _print_transfer_audit(rows: Sequence[FeeRow]) -> None:
    """Kritik transfer eşleşmelerini GitHub logunda görünür yapar."""
    print("[comparison] ===== ORTAK MASRAF EŞLEŞME KONTROLÜ =====")

    audit_specs = [
        spec
        for kind, spec in LAYOUT
        if kind == "ROW" and (
            spec.service in {
                "EFT", "HAVALE", "FAST", "SWIFT_GELEN", "SWIFT_GIDEN",
                "YURT_DISI_FAST", "KART_YURTDISI_TRANSFER", "DUZENLI_EFT",
                "DUZENLI_HAVALE", "ALTIN_TRANSFER",
                "HGS_ETIKET", "HGS_KART", "VERGI",
                "CEK_DEFTERI_YAPRAK", "CEK_KARNESI_10",
                "CEK_DUZENLEME_STANDART", "CEK_DUZENLEME_DOVIZ",
                "CEK_IADE", "CEK_TAHSIL", "CEK_KARSILIKSIZ",
                "CEK_DUZELTME_HAKKI",
            }
            or spec.service in _DOCUMENT_STATEMENT_SERVICES
            or spec.service in _RESEARCH_SECURITIES_SERVICES
        )
    ]

    for spec in audit_specs:
        for bank in BANKS:
            channels = ("MOBIL", "SUBE") if spec.split_channel else ("GENEL",)
            for channel in channels:
                value, row, resolution = _resolve_cell(rows, bank, spec, channel)
                if row is None:
                    raw = value.replace("\n", " / ")
                else:
                    raw = row.masraf
                    if len(raw) > 120:
                        raw = raw[:117] + "..."
                print(
                    f"[comparison][match] {spec.service} | {spec.label} | "
                    f"{bank} | {channel} <- {raw} [{resolution}]"
                )

    print("[comparison] ==========================================")



def _is_comparison_sheet_name(title: str) -> bool:
    key = _norm(title).replace(" ", "")
    return (
        key.startswith("karsilastirma")
        or key.startswith("comparison")
    )


def _remove_old_comparison_sheets(wb) -> List[str]:
    """
    Önceki 10-bankalı / 5x5 / test karşılaştırma sayfalarını temizler.
    Böylece kullanıcı yanlışlıkla eski sheet'e bakmaz.
    """
    removed: List[str] = []

    for ws in list(wb.worksheets):
        if _is_comparison_sheet_name(ws.title):
            removed.append(ws.title)
            wb.remove(ws)

    return removed


def _assert_preview_layout(ws) -> None:
    """
    Preview'den sapmayı fatal hata yapar.
    Eski 10 bankalı formatın sessizce geri gelmesini engeller.
    """
    expected_headers = {
        "B1": "Garanti BBVA",
        "D1": "İş Bankası",
        "F1": "Akbank",
        "H1": "Yapı ve Kredi Bankası",
        "A2": "ORTAK MASRAF ADI",
        "M1": "NOTLAR",
    }

    for cell_ref, expected in expected_headers.items():
        actual = _clean(ws[cell_ref].value)
        if actual != expected:
            raise RuntimeError(
                f"Preview layout doğrulaması başarısız: "
                f"{cell_ref} beklenen={expected!r}, gelen={actual!r}"
            )

    # Preview'de banka alanı B:I, J:L boş ve M notlardır.
    for col in ("J", "K", "L"):
        if any(
            _clean(ws[f"{col}{row}"].value)
            for row in range(1, min(ws.max_row, 12) + 1)
        ):
            raise RuntimeError(
                f"Preview layout doğrulaması başarısız: {col} kolonu boş olmalı."
            )

    # Eski formatta kullanılan banka isimleri görünür başlıklarda olmamalı.
    forbidden = {
        "QNB",
        "DenizBank",
        "Halkbank",
        "VakıfBank",
        "TEB",
        "Ziraat",
    }
    visible_headers = {
        _clean(ws.cell(row=row, column=col).value)
        for row in (1, 2)
        for col in range(1, 14)
    }

    bad = sorted(forbidden & visible_headers)
    if bad:
        raise RuntimeError(
            "Preview layout yerine eski çok-bankalı format oluştu: "
            + ", ".join(bad)
        )

    expected_filter = f"A2:I{ws.max_row}"
    if ws.auto_filter.ref != expected_filter:
        raise RuntimeError(
            "Karşılaştırma filtresi oluşmadı: "
            f"beklenen={expected_filter!r}, gelen={ws.auto_filter.ref!r}"
        )


def _assert_document_statement_sources(rows: Sequence[FeeRow]) -> None:
    """Yeni 17 satırın kritik resmî dayanaklarını yazımdan önce doğrular."""
    document_specs = {
        spec.service: spec
        for kind, spec in LAYOUT
        if kind == "ROW" and spec.service in _DOCUMENT_STATEMENT_SERVICES
    }
    if len(document_specs) != 17:
        raise RuntimeError(
            "Belge/ekstre/araştırma sözlüğü eksik veya mükerrer: "
            f"beklenen=17, gelen={len(document_specs)}"
        )

    required = (
        ("DOC_VADESIZ_TL", "GARANTİ", "MOBIL"),
        ("DOC_VADESIZ_TL", "İŞBANKASI", "MOBIL"),
        ("DOC_VADESIZ_TL", "İŞBANKASI", "SUBE"),
        ("DOC_VADESIZ_YP", "GARANTİ", "MOBIL"),
        ("DOC_VADESIZ_YP", "İŞBANKASI", "MOBIL"),
        ("DOC_VADESIZ_YP", "İŞBANKASI", "SUBE"),
        ("DOC_KMH_BIREYSEL", "GARANTİ", "GENEL"),
        ("DOC_KMH_BIREYSEL", "YAPIKREDI", "GENEL"),
        ("DOC_KK_GECMIS", "GARANTİ", "GENEL"),
        ("DOC_KK_GECMIS", "İŞBANKASI", "GENEL"),
        ("DOC_KK_GECMIS", "YAPIKREDI", "GENEL"),
        ("DOC_ARSIV_GENEL", "GARANTİ", "GENEL"),
        ("DOC_ARSIV_GENEL", "İŞBANKASI", "GENEL"),
        ("DOC_ESKI_DEKONT_EKSTRE", "AKBANK", "GENEL"),
        ("DOC_ESKI_DEKONT_EKSTRE", "YAPIKREDI", "GENEL"),
        ("DOC_AYLIK_POSTA", "AKBANK", "GENEL"),
        ("DOC_ARSIV_SUBE", "İŞBANKASI", "GENEL"),
        ("DOC_ARSIV_BANKA", "İŞBANKASI", "GENEL"),
        ("DOC_KMH_TICARI", "GARANTİ", "GENEL"),
        ("DOC_KMH_TICARI", "YAPIKREDI", "GENEL"),
        ("DOC_TICARI_KART_GECMIS", "YAPIKREDI", "GENEL"),
        ("DOC_YATIRIM_EKSTRE", "İŞBANKASI", "GENEL"),
        ("DOC_TICARI_AYLIK", "AKBANK", "GENEL"),
        ("DOC_TICARI_KART_BASILI", "GARANTİ", "GENEL"),
        ("DOC_TICARI_KART_GONDERIM", "GARANTİ", "GENEL"),
        ("DOC_TICARI_HESAP_BASILI", "AKBANK", "GENEL"),
        ("DOC_POS_EKSTRE", "GARANTİ", "GENEL"),
    )

    failures: List[str] = []
    for service, bank, channel in required:
        spec = document_specs[service]
        value, row, resolution = _resolve_cell(rows, bank, spec, channel)
        if row is None or resolution not in {"NUMERIC", "STATUS"}:
            failures.append(
                f"{service}/{bank}/{channel}: {resolution} ({value!r})"
            )

    if failures:
        raise RuntimeError(
            "Belge/ekstre/araştırma kritik kaynak kontrolü başarısız: "
            + "; ".join(failures)
        )

    print(
        "[comparison] BELGE/EKSTRE/ARAŞTIRMA: "
        f"17 satır ve {len(required)} kritik kaynak eşleşmesi doğrulandı."
    )


def _assert_research_securities_sources(rows: Sequence[FeeRow]) -> None:
    """Yeni araştırma, yurtdışı ATM ve menkul kıymet satırlarını fail-closed doğrular."""
    specs = {
        spec.service: spec
        for kind, spec in LAYOUT
        if kind == "ROW" and (
            spec.service in _RESEARCH_SECURITIES_SERVICES
            or spec.service == "BAKIYE_ATM_YURTDISI"
        )
    }
    expected_services = _RESEARCH_SECURITIES_SERVICES | {"BAKIYE_ATM_YURTDISI"}
    if set(specs) != expected_services:
        raise RuntimeError(
            "Araştırma/menkul kıymet sözlüğü eksik veya mükerrer: "
            f"beklenen={sorted(expected_services)}, gelen={sorted(specs)}"
        )

    numeric_required = (
        ("BAKIYE_ATM_YURTDISI", "İŞBANKASI"),
        ("SEC_DOMESTIC_STOCK_TRADE", "GARANTİ"),
        ("SEC_DOMESTIC_STOCK_TRADE", "İŞBANKASI"),
        ("SEC_DOMESTIC_STOCK_TRADE", "AKBANK"),
        ("SEC_DOMESTIC_STOCK_TRADE", "YAPIKREDI"),
        ("SEC_DOMESTIC_STOCK_TRANSFER", "İŞBANKASI"),
        ("SEC_DOMESTIC_STOCK_TRANSFER", "AKBANK"),
        ("SEC_DOMESTIC_STOCK_TRANSFER", "YAPIKREDI"),
        ("SEC_DOMESTIC_BOND_TRANSFER", "İŞBANKASI"),
        ("SEC_DOMESTIC_BOND_TRANSFER", "AKBANK"),
        ("SEC_DOMESTIC_BOND_TRANSFER", "YAPIKREDI"),
        ("SEC_DOMESTIC_CUSTODY", "GARANTİ"),
        ("SEC_DOMESTIC_CUSTODY", "İŞBANKASI"),
        ("SEC_DOMESTIC_CUSTODY", "AKBANK"),
        ("SEC_DOMESTIC_CUSTODY", "YAPIKREDI"),
        ("SEC_FOREIGN_STOCK_TRADE", "İŞBANKASI"),
        ("SEC_FOREIGN_STOCK_CUSTODY", "İŞBANKASI"),
        ("SEC_EUROBOND_TRANSFER", "İŞBANKASI"),
        ("SEC_EUROBOND_TRANSFER", "AKBANK"),
        ("SEC_EUROBOND_TRANSFER", "YAPIKREDI"),
        ("SEC_EUROBOND_CUSTODY", "İŞBANKASI"),
    )

    failures: List[str] = []
    for service, bank in numeric_required:
        value, row, resolution = _resolve_cell(rows, bank, specs[service], "GENEL")
        if row is None or resolution not in {"NUMERIC", "STATUS"}:
            failures.append(f"{service}/{bank}: {resolution} ({value!r})")

    deposit = specs["DOC_DEPOSIT_RESEARCH"]
    for bank in BANKS:
        value, _, resolution = _resolve_cell(rows, bank, deposit, "GENEL")
        if resolution not in {"NUMERIC", "STATUS", "PUBLICATION_STATUS"}:
            failures.append(f"DOC_DEPOSIT_RESEARCH/{bank}: {resolution} ({value!r})")

    if failures:
        raise RuntimeError(
            "Araştırma/menkul kıymet kritik kaynak kontrolü başarısız: "
            + "; ".join(failures)
        )

    print(
        "[comparison] ARAŞTIRMA/MENKUL KIYMET: "
        f"{len(expected_services)} satır ailesi ve {len(numeric_required)} "
        "kritik kaynak eşleşmesi doğrulandı."
    )


def update_comparison_sheet(excel_path: str = "komisyonlar_guncel.xlsx") -> Dict[str, int]:
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Ana Excel bulunamadı: {path}")

    print(f"[comparison] SÜRÜM: {COMPARISON_VERSION}")
    print(f"[comparison] MODÜL: {Path(__file__).resolve()}")
    print(f"[comparison] LAYOUT: {PREVIEW_LAYOUT_SIGNATURE}")
    print(f"[comparison] Ana Excel okunuyor: {path}")

    wb = load_workbook(path)
    source_ws, header_row = _find_source_sheet(wb)
    cols = _header_map(source_ws, header_row)
    rows = _read_rows(source_ws, header_row, cols)

    if not rows:
        raise RuntimeError("Karşılaştırma için uygun banka verisi bulunamadı.")

    _assert_document_statement_sources(rows)
    _assert_research_securities_sources(rows)

    old_ws = wb[COMPARISON_SHEET] if COMPARISON_SHEET in wb.sheetnames else None
    notes = _preserve_notes(old_ws)

    removed_sheets = _remove_old_comparison_sheets(wb)

    ws = wb.create_sheet(COMPARISON_SHEET)
    comparison_rows = _write_comparison(ws, rows, notes)
    # Dosya açıldığında kullanıcı doğrudan karşılaştırma sekmesini görsün.
    wb.active = wb.sheetnames.index(COMPARISON_SHEET)
    ws.sheet_state = "visible"

    # Oluşan sheet preview ile aynı yapıdan saparsa Excel'i kaydetme.
    _assert_preview_layout(ws)

    _print_transfer_audit(rows)

    # Ana Excel'i yarım/bozuk kaydetmemek için atomik karşılaştırma güncellemesi.
    tmp_path = path.with_name(path.stem + ".comparison.tmp" + path.suffix)
    wb.save(tmp_path)
    tmp_path.replace(path)

    print(
        f"[comparison] {COMPARISON_SHEET} referans-sıralı kanonik formatta güncellendi. "
        f"Kaynak={source_ws.title}, 4 banka kaynak satırı={len(rows)}, "
        f"karşılaştırma satırı={comparison_rows}, korunan not={len(notes)}"
    )
    print(
        f"[comparison] Silinen eski karşılaştırma sayfaları: "
        f"{removed_sheets or 'yok'}"
    )
    print(
        "[comparison] Görünüm doğrulandı: "
        "A=ortak masraf, B:I=4 banka Mobil/Şube (filtreli), J:L=boş, M=NOTLAR (sabit sözlük, düz font, tek tek açıklamalar)."
    )

    resolution_counts = {
        "NUMERIC": 0,
        "GENERIC_TARIFF": 0,
        "STATUS": 0,
        "PUBLICATION_STATUS": 0,
        "NOT_APPLICABLE": 0,
        "SOURCE_GAP": 0,
        "N/A": 0,
    }
    possible_cells = 0

    for kind, payload in LAYOUT:
        if kind != "ROW":
            continue
        spec = payload
        for bank in BANKS:
            channels = ("MOBIL", "SUBE") if spec.split_channel else ("GENEL",)
            for channel in channels:
                possible_cells += 1
                _, _, resolution = _resolve_cell(rows, bank, spec, channel)
                resolution_counts[resolution] = resolution_counts.get(resolution, 0) + 1

    source_gaps = resolution_counts.get("SOURCE_GAP", 0)
    true_na = resolution_counts.get("N/A", 0)
    verified_cells = possible_cells - source_gaps - true_na
    numeric_like = resolution_counts.get("NUMERIC", 0) + resolution_counts.get("GENERIC_TARIFF", 0)
    status_like = (
        resolution_counts.get("STATUS", 0)
        + resolution_counts.get("PUBLICATION_STATUS", 0)
        + resolution_counts.get("NOT_APPLICABLE", 0)
    )

    print(
        f"[comparison] KALİTE: doğrulanmış={verified_cells}/{possible_cells} "
        f"(%{(verified_cells / possible_cells * 100):.1f}) | "
        f"sayısal={resolution_counts.get('NUMERIC', 0)} | "
        f"genel_tarife={resolution_counts.get('GENERIC_TARIFF', 0)} | "
        f"resmî_durum={status_like} | "
        f"kaynak_boşluğu={source_gaps} | "
        f"belirsiz_eşleşme={len(_AMBIGUITY_LOGGED)} | N/A={true_na}"
    )


    return {
        "source_rows": len(rows),
        "comparison_rows": comparison_rows,
        "notes_preserved": len(notes),
        "matched_cells": verified_cells,
        "numeric_cells": numeric_like,
        "status_cells": status_like,
        "source_gap_cells": source_gaps,
        "ambiguous_cells": len(_AMBIGUITY_LOGGED),
        "missing_cells": true_na,
        "possible_cells": possible_cells,
    }


if __name__ == "__main__":
    print(update_comparison_sheet())
